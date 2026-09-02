"""
찌라시 마스터 v6 - 컴플라이언스형 (정직한 자동 발행)
========================================================
- bm21.com 스타일 리치 HTML 템플릿 (컬러 헤더, 칩 태그, FAQ, 이미지)
- Selenium 기반 그누보드/Cafe24 글쓰기 (정상 동작, 회피코드 없음)
- {지역} {서비스} {브랜드} 3키워드 치환
- 정직한 전화번호 표기 (난독화 없음)
- rate limit + 사이트당 1일 발행 한도로 도배 방지
※ 게시 대상 게시판에 대한 게시 권한(운영자 허락/홍보 전용 게시판)은
  이용자가 보장해야 합니다. 자동 프로그램 금지 규칙이 있는 곳에는 사용 금지.
"""

import sys, os, re, json, time, random, threading, queue, urllib.parse, secrets, hashlib, base64, copy, uuid, html as html_lib
# 콘솔 코드페이지가 cp949 등일 때 이모지 print가 UnicodeEncodeError로 죽는 것 방지.
# (Windows에서 PYTHONIOENCODING 미설정 시 startup print의 ⏰/🔁 등이 크래시 유발)
for _stream in ('stdout', 'stderr'):
    try:
        getattr(sys, _stream).reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
import urllib3; urllib3.disable_warnings()

from flask import Flask, request, jsonify, render_template_string, session, redirect, send_from_directory
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import HTTPException

# ==================== 설정 ====================
BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = BASE_DIR / 'data'; DATA_DIR.mkdir(exist_ok=True)
SITES_FILE = DATA_DIR / 'sites.json'; CONFIG_FILE = DATA_DIR / 'config.json'; LOG_FILE = DATA_DIR / 'logs.json'
HISTORY_FILE = DATA_DIR / 'history.json'   # 발행 이력 원장(엑셀/결과탭)
QUEUE_FILE = DATA_DIR / 'queue.json'       # 미완료 작업(재시작 복구용)
SCHED_FILE = DATA_DIR / 'schedules.json'   # 예약 발행 스케줄
KEYWORDS_FILE = DATA_DIR / 'keywords.json' # 키워드 풀(엑셀 랜덤 치환)
UNIQ_FILE = DATA_DIR / 'uniq.json'          # 발행 콘텐츠 중복방지(제목/본문 해시)
IMAGES_FILE = DATA_DIR / 'images.json'      # 사용자 이미지 URL 풀(본문 삽입용)
UPLOAD_DIR = DATA_DIR / 'uploads'; UPLOAD_DIR.mkdir(exist_ok=True)
IMAGE_EXTENSIONS = {'.jpg','.jpeg','.png','.gif','.webp'}
MEMBERS_FILE = DATA_DIR / 'members.json'    # 회원(고객) 관리 + 월 정산
CAND_FILE = DATA_DIR / 'candidates.json'    # 발굴 후보(승인 대기함)
REGIONS_FILE = BASE_DIR / 'regions_full.json'  # 전국 시도·시군구·읍면동
DISCO_FILE = DATA_DIR / 'discover.json'     # 발굴 상태(쿼리 커서·일일 카운트)
SIGNUP_PROFILES_FILE = DATA_DIR / 'signup_profiles.json'  # 가입 폼 측정·학습 이력
REJECTED_DOMAINS_FILE = DATA_DIR / 'rejected_domains.json'  # 영구 탈락 도메인(다음 발굴에서 제외)
AI_USAGE_FILE = DATA_DIR / 'openai_usage.json'  # 글 생성별 토큰·예상 비용 원장
WORKROOMS_FILE = DATA_DIR / 'workrooms.json'    # 키워드별 독립 작업실

IMAGES = [f'https://picsum.photos/id/{i}/800/400' for i in [1,20,26,48,60,64,76,91,96,104,152,160,175,180,185,201]]
COLORS = ['#3b1f2b','#2B8A3E','#37474f','#1a5276','#6c3483','#b7950b','#a04000']

_json_locks={}; _json_locks_guard=threading.Lock()
_signup_learn_locks=defaultdict(threading.Lock)
def _json_lock(p):
    key=str(Path(p).resolve())
    with _json_locks_guard:
        return _json_locks.setdefault(key,threading.RLock())

def load_json(p, d=None):
    if d is None: d=[]
    try:
        with _json_lock(p):
            if os.path.exists(p):
                with open(p,'r',encoding='utf-8') as f: return json.load(f)
    except: pass
    return d

def save_json(p, data):
    """같은 파일시스템의 임시 파일에 기록 후 원자 교체하여 JSON 손상을 방지."""
    p=Path(p); p.parent.mkdir(parents=True,exist_ok=True)
    tmp=p.with_name(f'.{p.name}.{os.getpid()}.{threading.get_ident()}.tmp')
    with _json_lock(p):
        try:
            with open(tmp,'w',encoding='utf-8') as f:
                json.dump(data,f,ensure_ascii=False,indent=2); f.flush(); os.fsync(f.fileno())
            os.replace(tmp,p)
        finally:
            try:
                if tmp.exists(): tmp.unlink()
            except Exception: pass

# ---- 사용자 이미지 URL 풀 (본문 삽입용) ----
def load_image_urls():
    d=load_json(IMAGES_FILE,[])
    return [u.strip() for u in d if isinstance(u,str) and u.strip().startswith('http')]
def save_image_urls(urls): save_json(IMAGES_FILE,urls)
def pick_images(n):
    """본문용 이미지 n개 선택. 사용자 URL 풀이 있으면 그걸, 없으면 기본(picsum) 폴백."""
    pool=load_image_urls() or IMAGES
    if len(pool)>=n: return random.sample(pool,n)
    return [random.choice(pool) for _ in range(n)]   # URL이 부족하면 중복 허용

def _record_openai_usage(model, usage, cfg):
    """Chat Completions 응답 usage를 저장하고 공식 토큰 단가 기준 예상비용을 계산한다."""
    inp=int((usage or {}).get('prompt_tokens') or (usage or {}).get('input_tokens') or 0)
    out=int((usage or {}).get('completion_tokens') or (usage or {}).get('output_tokens') or 0)
    cached=int((((usage or {}).get('prompt_tokens_details') or {}).get('cached_tokens')) or 0)
    pin=float(cfg.get('openai_input_price_per_million') or 0.15)
    pout=float(cfg.get('openai_output_price_per_million') or 0.60)
    pcache=float(cfg.get('openai_cached_input_price_per_million') or 0.075)
    regular=max(0,inp-cached)
    estimated=(regular*pin+cached*pcache+out*pout)/1_000_000
    rec={'time':datetime.now().astimezone().isoformat(timespec='seconds'),'model':model,
         'input_tokens':inp,'cached_input_tokens':cached,'output_tokens':out,
         'requests':1,'estimated_cost_usd':round(estimated,8)}
    with _json_lock(AI_USAGE_FILE):
        rows=load_json(AI_USAGE_FILE,[])
        if not isinstance(rows,list): rows=[]
        rows.append(rec); save_json(AI_USAGE_FILE,rows[-20000:])
    return rec

def _local_openai_usage_summary(cfg):
    rows=load_json(AI_USAGE_FILE,[]); now=datetime.now().astimezone()
    month=now.strftime('%Y-%m'); today=now.strftime('%Y-%m-%d')
    def agg(items):
        return {'requests':sum(int(x.get('requests',1) or 0) for x in items),
                'input_tokens':sum(int(x.get('input_tokens',0) or 0) for x in items),
                'cached_input_tokens':sum(int(x.get('cached_input_tokens',0) or 0) for x in items),
                'output_tokens':sum(int(x.get('output_tokens',0) or 0) for x in items),
                'estimated_cost_usd':round(sum(float(x.get('estimated_cost_usd',0) or 0) for x in items),6)}
    mr=[x for x in rows if str(x.get('time','')).startswith(month)]
    tr=[x for x in rows if str(x.get('time','')).startswith(today)]
    budget=float(cfg.get('openai_monthly_budget_usd') or 0)
    m=agg(mr); t=agg(tr)
    return {'source':'local_estimate','month':m,'today':t,'monthly_budget_usd':budget,
            'remaining_budget_usd':round(max(0,budget-m['estimated_cost_usd']),6) if budget>0 else None,
            'note':'프로젝트 키 응답의 토큰 기준 예상치'}

def _openai_admin_costs(cfg):
    """관리자 키가 있을 때만 공식 조직 Costs API로 이번 달 실제 비용을 조회한다."""
    key=(cfg.get('openai_admin_key') or '').strip()
    if not key: return None
    import requests as _rq
    now=datetime.now().astimezone(); start=int(now.replace(day=1,hour=0,minute=0,second=0,microsecond=0).timestamp())
    r=_rq.get('https://api.openai.com/v1/organization/costs',params={'start_time':start,'limit':31},
              headers={'Authorization':f'Bearer {key}','Content-Type':'application/json'},timeout=20)
    r.raise_for_status(); data=r.json(); total=0.0
    for bucket in data.get('data',[]):
        for result in bucket.get('results',[]):
            amount=result.get('amount') or {}
            if str(amount.get('currency','usd')).lower()=='usd': total+=float(amount.get('value') or 0)
    return round(total,6)

def uploaded_images():
    out=[]
    for p in sorted(UPLOAD_DIR.iterdir(),key=lambda x:x.stat().st_mtime,reverse=True):
        if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS:
            out.append({'name':p.name,'size':p.stat().st_size,
                        'url':'/media/'+urllib.parse.quote(p.name),'saved_at':datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec='seconds')})
    return out

def pick_attachment_paths(max_n=2):
    files=[UPLOAD_DIR/x['name'] for x in uploaded_images()]
    return [str(p.resolve()) for p in files[:max(0,int(max_n))]]

def attach_saved_images(d,max_n=1):
    from selenium.webdriver.common.by import By
    paths=pick_attachment_paths(max_n)
    if not paths: return 0,'저장 이미지 없음'
    inputs=[e for e in d.find_elements(By.CSS_SELECTOR,"input[type='file']") if _sel_vis(e)]
    if not inputs: return 0,'파일 첨부 입력란 없음'
    used=0
    for el,path in zip(inputs,paths):
        try: el.send_keys(path); used+=1
        except Exception: continue
    return used,(f'{used}개 첨부' if used else '첨부 입력 실패')

def _credential_fernet():
    """사이트 로그인정보용 서버측 암호화기. 세션키/전용 env에서 안정적으로 파생한다."""
    try:
        from cryptography.fernet import Fernet
        persisted=''
        try:
            kf=DATA_DIR/'secret.key'
            if kf.exists(): persisted=kf.read_text().strip()
        except Exception: pass
        raw=(os.environ.get('CHIRASHI_CREDENTIAL_KEY') or os.environ.get('CHIRASHI_SECRET') or persisted)
        if not raw: raise RuntimeError('credential key unavailable')
        raw=raw.encode()
        return Fernet(base64.urlsafe_b64encode(hashlib.sha256(raw).digest()))
    except Exception:
        return None

def _encrypt_password(value):
    if not value: return ''
    if str(value).startswith('fernet:'): return str(value)
    f=_credential_fernet()
    if not f: raise RuntimeError('로그인정보 암호화를 위해 cryptography 패키지가 필요합니다')
    return 'fernet:'+f.encrypt(str(value).encode()).decode()

def _decrypt_password(value):
    if not value: return ''
    if not str(value).startswith('fernet:'): return str(value)  # 기존 평문 데이터 마이그레이션 입력
    f=_credential_fernet()
    if not f: return ''
    try: return f.decrypt(str(value)[7:].encode()).decode()
    except Exception: return ''

def load_sites():
    sites=load_json(SITES_FILE,[])
    for s in sites:
        s['mb_pass']=_decrypt_password(s.get('mb_pass_enc') or s.get('mb_pass',''))
    return sites

def save_sites(sites):
    out=copy.deepcopy(sites)
    for s in out:
        pw=s.pop('mb_pass','') or _decrypt_password(s.get('mb_pass_enc',''))
        if pw: s['mb_pass_enc']=_encrypt_password(pw)
        elif not s.get('mb_pass_enc'): s.pop('mb_pass_enc',None)
    save_json(SITES_FILE,out)

def load_signup_profiles(): return load_json(SIGNUP_PROFILES_FILE,{}) or {}
def save_signup_profiles(d): save_json(SIGNUP_PROFILES_FILE,d)

def _signup_origin(site):
    p=urllib.parse.urlsplit(site.get('site_url',''))
    return f'{p.scheme}://{p.netloc}' if p.scheme and p.netloc else site.get('site_url','').rstrip('/')

def _signup_form_measure(site):
    """가입 폼을 제출 없이 측정한다. 약관 다음 화면 조회까지만 하며 계정 생성은 하지 않는다."""
    from html.parser import HTMLParser
    import requests as _rq
    class FormParser(HTMLParser):
        def __init__(self): super().__init__(); self.forms=[]; self.cur=None; self.text=[]
        def handle_starttag(self,tag,attrs):
            a=dict(attrs)
            if tag=='form': self.cur={'action':a.get('action',''),'method':a.get('method','get').lower(),'fields':[]}
            elif self.cur is not None and tag in ('input','select','textarea','button'):
                safe={k:v for k,v in a.items() if k in ('type','name','id','minlength','maxlength','pattern','required','autocomplete','placeholder')}
                safe['tag']=tag; self.cur['fields'].append(safe)
        def handle_endtag(self,tag):
            if tag=='form' and self.cur is not None: self.forms.append(self.cur); self.cur=None
        def handle_data(self,data):
            t=' '.join(data.split())
            if t: self.text.append(t)
    base=_signup_origin(site); platform=site.get('platform') or 'gnuboard'
    url=site.get('signup_url') or base+('/member/join.html' if platform=='cafe24' else '/bbs/register.php')
    sess=_rq.Session(); hdr={'User-Agent':'Mozilla/5.0 (signup-form-audit; administrator initiated)'}
    r=sess.get(url,timeout=20,verify=False,headers=hdr,allow_redirects=True); r.raise_for_status()
    html=r.text; measured_url=r.url
    # 그누보드 약관 화면이면 동의값을 세션에 전달해 실제 가입 폼까지만 조회한다.
    if platform!='cafe24' and not re.search(r'name=["\']mb_password["\']',html,re.I):
        form_url=urllib.parse.urljoin(r.url,'register_form.php')
        r=sess.post(form_url,data={'agree':'1','agree2':'1'},timeout=20,verify=False,headers=hdr,allow_redirects=True)
        r.raise_for_status(); html=r.text; measured_url=r.url
    p=FormParser(); p.feed(html)
    def signup_forms(parsed):
        out=[]
        for f in parsed.forms:
            keys=' '.join((x.get('name') or '')+' '+(x.get('id') or '') for x in f['fields']).lower()
            action=(f.get('action') or '').lower()
            if ('register_form_update' in action or 'member/join' in action or
                (re.search(r'(password_re|password_confirm|passwd_confirm)',keys) and re.search(r'(email|nick|name)',keys))): out.append(f)
        return out
    forms=signup_forms(p)
    # 정적 요청에서 상단 로그인폼만 보이는 사이트는 Selenium으로 약관 다음 화면까지 재측정한다.
    if not forms:
        try:
            from selenium.webdriver.common.by import By
            d=get_driver(); d.get(url); time.sleep(1.5)
            for nm in ('agree','agree2'):
                for el in d.find_elements(By.CSS_SELECTOR,f"input[name='{nm}']"):
                    try:
                        if not el.is_selected(): el.click()
                    except: pass
            submits=d.find_elements(By.CSS_SELECTOR,"form#fregister input[type='submit'],form#fregister button[type='submit'],form[name='fregister'] input[type='submit']")
            if submits: d.execute_script('arguments[0].click()',submits[0]); time.sleep(2)
            html=d.page_source; measured_url=d.current_url
            p=FormParser(); p.feed(html); forms=signup_forms(p)
        except Exception: forms=[]
    if not forms: raise RuntimeError('가입 입력 폼을 찾지 못했습니다')
    form=max(forms,key=lambda f:len(f['fields'])); fields=[]
    for x in form['fields']:
        name=x.get('name') or ''; fid=x.get('id') or ''; typ=(x.get('type') or x.get('tag') or '').lower()
        role=''
        key=(name+' '+fid).lower()
        if re.search(r'(mb_id|member_id|login_id|user.?id)',key): role='id'
        elif typ=='password' and re.search(r'(_re\b|_confirm\b|confirm_|check)',key): role='password_confirm'
        elif typ=='password': role='password'
        elif re.search(r'(mb_email|e.?mail)',key): role='email'
        elif re.search(r'(mb_nick|nickname)',key): role='nickname'
        elif re.search(r'(mb_name|user.?name|real.?name)',key): role='name'
        elif re.search(r'(captcha|recaptcha|turnstile)',key): role='captcha'
        y=dict(x); y['role']=role; y['selector']=('#'+fid if fid else ('[name="'+name+'"]' if name else ''))
        fields.append(y)
    text=' '.join(p.text); rules={}
    def attr_int(role,key,default):
        vals=[]
        for f in fields:
            if f.get('role')==role:
                try:
                    if f.get(key) is not None: vals.append(int(f[key]))
                except: pass
        return vals[0] if vals else default
    rules['id_min']=attr_int('id','minlength',3 if platform!='cafe24' else 4)
    rules['id_max']=attr_int('id','maxlength',20 if platform!='cafe24' else 16)
    rules['password_min']=attr_int('password','minlength',10)
    rules['password_max']=attr_int('password','maxlength',64)
    # 화면 안내문에서 더 구체적인 최소 길이를 찾으면 반영한다.
    for pat,key in [(r'(?:아이디|ID)[^0-9]{0,30}(\d+)\s*(?:자|글자)\s*이상','id_min'),
                    (r'(?:비밀번호|패스워드)[^0-9]{0,30}(\d+)\s*(?:자|글자)\s*이상','password_min')]:
        m=re.search(pat,text,re.I)
        if m: rules[key]=int(m.group(1))
    rules['require_special']=bool(re.search(r'비밀번호.{0,80}(특수문자|특수 문자)',text,re.I))
    captcha=bool(re.search(r'(captcha|kcaptcha|g-recaptcha|turnstile|자동등록방지)',html,re.I))
    email_verification=bool(
        re.search(r'(?:e-?mail|이메일).{0,160}(?:인증|확인).{0,100}(?:회원가입|가입|완료)',text,re.I) or
        re.search(r'(?:인증|확인).{0,100}(?:e-?mail|이메일)',text,re.I))
    signature=[(f.get('role'),f.get('name'),f.get('id'),f.get('type'),f.get('minlength'),f.get('maxlength'),f.get('pattern')) for f in fields]
    fingerprint=hashlib.sha256(json.dumps(signature,ensure_ascii=False,sort_keys=True).encode()).hexdigest()
    return {'signup_url':url,'form_url':measured_url,'form_action':urllib.parse.urljoin(measured_url,form.get('action','')),
            'form_method':form.get('method','post'),'fields':fields,'rules':rules,'captcha':captcha,
            'email_verification_required':email_verification,
            'fingerprint':fingerprint,'measured_at':datetime.now().isoformat(timespec='seconds')}

def learn_signup_profile(site,force=False):
    host=urllib.parse.urlsplit(_signup_origin(site)).netloc.lower()
    with _signup_learn_locks[host]:
        profiles=load_signup_profiles(); old=profiles.get(host) or {}
        # 올인원 실행 때마다 상대 사이트를 두드리지 않는다. 최근 정상 측정값은 30분 재사용.
        if old and not force:
            try: fresh=(datetime.now()-datetime.fromisoformat(old.get('measured_at',''))).total_seconds()<1800
            except: fresh=False
            if fresh:
                profile=dict(old); profile['cached']=True; changed=False
                site.update({'signup_profile_host':host,'signup_rules':profile['rules'],'signup_url':profile['signup_url'],
                             'signup_profile_version':profile['version'],'signup_profile_changed':False,
                             'signup_profile_measured_at':profile['measured_at'],'signup_has_captcha':profile['captcha'],
                             'signup_email_verification':bool(profile.get('email_verification_required'))})
                return profile
        measured=_signup_form_measure(site); changed=old.get('fingerprint')!=measured['fingerprint']; history=list(old.get('history') or [])
        if old and changed:
            history.append({'fingerprint':old.get('fingerprint'),'rules':old.get('rules',{}),'measured_at':old.get('measured_at')})
        history=history[-20:]
        profile={**measured,'host':host,'version':int(old.get('version',0))+(1 if changed else 0),
                 'seen_count':int(old.get('seen_count',0))+1,'success_count':int(old.get('success_count',0)),
                 'failure_count':int(old.get('failure_count',0)),'history':history,'cached':False}
        profiles[host]=profile; save_signup_profiles(profiles)
        site.update({'signup_profile_host':host,'signup_rules':profile['rules'],'signup_url':profile['signup_url'],
                     'signup_profile_version':profile['version'],'signup_profile_changed':changed,
                     'signup_profile_measured_at':profile['measured_at'],'signup_has_captcha':profile['captcha'],
                     'signup_email_verification':bool(profile.get('email_verification_required'))})
        return profile

def load_config():
    d={'brand':'인천홍마니','phone':'01082755736','phones':'','openai_key':'','model':'gpt-4o-mini',
       'openai_admin_key':'','openai_monthly_budget_usd':20.0,
       'openai_input_price_per_million':0.15,'openai_cached_input_price_per_million':0.075,
       'openai_output_price_per_million':0.60,
       'workers':2,'password':'admin1234','post_delay':30,'daily_limit':3,
       'use_gpt':False,'telegram_token':'','telegram_chat_id':'',
       'notify_done':False,'notify_fail':True,'update_token':'',
       'telegram_control':False,'backup_time':'','verify_enabled':True,'mix_keywords':True,
       'block_unpaid':True,
       'google_api_key':'','google_cx':'','brave_api_key':'','search_provider':'brave','discover_enabled':False,
       'discover_daily_target':100,'discover_query_limit':100,'discover_keywords':'',
       'discover_direct_queries':'','video_url':'','landing_url':'','post_email':'','guest_post_password':'',
       'twocaptcha_api_key':'','twocaptcha_enabled':False,
       'auto_pipeline_enabled':False,'auto_pipeline_batch':3}
    c=load_json(CONFIG_FILE,None)
    if c is None or not isinstance(c,dict): save_json(CONFIG_FILE,d); return d.copy()
    for k,v in d.items():
        if k not in c: c[k]=v
    return c

TWOCAPTCHA_CACHE_FILE=os.path.join(DATA_DIR,'twocaptcha_balance.json')

def _twocaptcha_usage_summary(cfg):
    key=(cfg.get('twocaptcha_api_key') or '').strip()
    if not key:
        return {'ok':False,'disabled':True,'error':'2captcha API 키 없음','currency':'USD','balance':0.0,'remaining_usd':0.0,'charged_since_last_check_usd':0.0,'updated_at':None}
    if not cfg.get('twocaptcha_enabled',False):
        return {'ok':False,'disabled':True,'error':'2captcha 비활성화','currency':'USD','balance':0.0,'remaining_usd':0.0,'charged_since_last_check_usd':0.0,'updated_at':None}
    try:
        r=requests.get('https://2captcha.com/res.php',params={'key':key,'action':'getbalance','json':'1'},timeout=20)
        raw=(r.text or '').strip()
        payload=None
        try:
            payload=r.json()
        except Exception:
            payload=None

        balance=None
        if isinstance(payload,dict):
            if isinstance(payload.get('request'), dict):
                value=payload['request'].get('balance')
                if value is not None: balance=float(value)
            elif payload.get('request') is not None:
                value=payload.get('request')
                if isinstance(value,(int,float,str)):
                    balance=float(str(value))
        if balance is None and raw.startswith('OK|'):
            balance=float(raw.split('|',1)[1].strip())
        if balance is None and raw.startswith('{'):
            try:
                j=json.loads(raw)
                if isinstance(j,dict):
                    req=j.get('request')
                    if isinstance(req,dict):
                        balance=float(req.get('balance',0))
                    elif isinstance(req,(int,float,str)):
                        balance=float(str(req))
            except Exception:
                pass
        if balance is None:
            raise ValueError(f'2captcha 응답 파싱 실패: {raw[:140]}')

        cache=load_json(TWOCAPTCHA_CACHE_FILE,{}) or {}
        prev=cache.get('balance')
        delta=0.0
        if prev is not None and balance < prev:
            delta=round(max(0.0, float(prev)-float(balance)), 6)
        cache={'balance':float(balance),'charged_since_last_check_usd':delta,'updated_at':datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')}
        save_json(TWOCAPTCHA_CACHE_FILE,cache)

        return {'ok':True,'source':'2captcha','currency':'USD','balance':round(float(balance),6),'remaining_usd':round(float(balance),6),'charged_since_last_check_usd':round(delta,6),'updated_at':cache['updated_at']}
    except Exception as e:
        return {'ok':False,'disabled':False,'error':str(e)[:180],'currency':'USD','balance':0.0,'remaining_usd':0.0,'charged_since_last_check_usd':0.0,'updated_at':None}

def save_config(c): save_json(CONFIG_FILE,c)
def add_log(msg):
    with _json_lock(LOG_FILE):
        logs=load_json(LOG_FILE,[])
        logs.append({'time':datetime.now().strftime('%H:%M:%S'),'msg':msg})
        if len(logs)>500: logs=logs[-500:]
        save_json(LOG_FILE,logs)

# ==================== 전화번호 표기 ====================
def format_phone(phone):
    d=re.sub(r'[^0-9]','',phone or '01082755736')
    if len(d)==11: return f'{d[:3]}-{d[3:7]}-{d[7:]}'
    if len(d)==10: return f'{d[:3]}-{d[3:6]}-{d[6:]}'
    return phone

PHONE_SEPS=['↔','●','=','~','-',' ','.','·','_','ㆍ','∼','◆','ㅡ']
TITLE_EXTRAS=['확실한','24시','검증된','재방문200%','인기','추천','친절한','예약가능','빠른안내','만족도높은']
def format_phone_random(phone):
    """제목용 번호 랜덤 변형 (매번 다른 기호·표기). 사람은 읽을 수 있게 유지."""
    d=re.sub(r'[^0-9]','',phone or '01082755736')
    if len(d)==11: a,b,c=d[:3],d[3:7],d[7:]
    elif len(d)==10: a,b,c=d[:3],d[3:6],d[6:]
    else: return phone
    s1=random.choice(PHONE_SEPS); s2=random.choice(PHONE_SEPS)
    # O/I 치환 여부(랜덤): 010->O1O / OIO 등, 사람이 읽을 수 있게 유지
    sub=random.choice([lambda x:x, lambda x:x.replace('0','O'),
                       lambda x:x.replace('0','O').replace('1','I')])
    a2,b2,c2=sub(a),sub(b),sub(c)
    style=random.random()
    if style<0.30:      out=f'[{a2}]{s1}{b2}{s2}{c2}'      # [010]↔8275↔5736
    elif style<0.55:    out=f'[{a2}{s1}{b2}{s2}{c2}]'      # [010●8275●5736]
    else:               out=f'{a2}{s1}{b2}{s2}{c2}'        # O1O=2572=3859
    return out

def get_phones(cfg):
    ph=cfg.get('phones','') or ''
    lst=[re.sub(r'\s+','',x) for x in ph.splitlines() if x.strip()]
    if not lst and cfg.get('phone'): lst=[cfg['phone']]
    return lst or ['01082755736']

def pick_phone(cfg):
    return random.choice(get_phones(cfg))

def build_title(r,s,b,cfg,raw=None):
    """제목 = 지역 키워드1 + 변형 번호 + 같은 지역 키워드2 + 기타요소 + 같은 지역 키워드3."""
    raw=raw or pick_phone(cfg)
    extra=random.choice(TITLE_EXTRAS)
    return f'{r} {format_phone_random(raw)} {s} {extra} {b}'[:140], raw

# ==================== 키워드 풀 (엑셀/CSV 랜덤 치환) ====================
REGION_ORDER=('인천','경기','서울','충남','충북','세종','전북','전남','경상','경북','강원','제주')
_region_order_cache=None

def _province_bucket(name):
    """시도명을 대표님 지정 12개 지역 그룹으로 정규화한다."""
    n=str(name or '').strip()
    rules=(
        ('인천',('인천',)),('경기',('경기',)),('서울',('서울',)),
        ('충남',('충남','충청남','대전')),('충북',('충북','충청북')),
        ('세종',('세종',)),('전북',('전북','전라북')),
        ('전남',('전남','전라남','광주')),
        ('경북',('경북','경상북')),
        ('경상',('경상','경남','경상남','부산','대구','울산')),
        ('강원',('강원',)),('제주',('제주',)),
    )
    for bucket,aliases in rules:
        if any(n.startswith(a) for a in aliases): return bucket
    return ''

def _region_order_map():
    """시·군·구·읍·면·동도 상위 시도 순서로 정렬할 수 있게 역색인을 만든다."""
    global _region_order_cache
    if _region_order_cache is not None: return _region_order_cache
    rank={name:i for i,name in enumerate(REGION_ORDER)}; out={}
    data=load_json(REGIONS_FILE,{})
    for province,districts in (data.items() if isinstance(data,dict) else []):
        bucket=_province_bucket(province); r=rank.get(bucket,99)
        names=[province,bucket]
        for district,dongs in ((districts or {}).items() if isinstance(districts,dict) else []):
            names.extend([district,re.sub(r'(시|군|구)$','',district)])
            for dong in (dongs or []): names.append(str(dong).strip())
        for name in names:
            name=str(name or '').strip()
            if name and (name not in out or r<out[name]): out[name]=r
    for alias in ('인천','경기','서울','충남','충북','세종','전북','전남','경상','경남','경북','강원','제주'):
        out[alias]=rank.get(_province_bucket(alias),99)
    _region_order_cache=out
    return out

def _keyword_region_rank(row):
    region=str((row or {}).get('지역','') or '').strip()
    bucket=_province_bucket(region)
    if bucket: return REGION_ORDER.index(bucket)
    index=_region_order_map()
    if region in index: return index[region]
    # 풀의 첫 칸이 '강남셔츠룸'처럼 지역+업종 전체 키워드인 경우도
    # 가장 긴 행정구역 접두어를 찾아 상위 시도 순서를 적용한다.
    matches=[(len(name),rank) for name,rank in index.items() if len(name)>=2 and region.startswith(name)]
    return max(matches,key=lambda x:x[0])[1] if matches else 99

def order_keywords(rows):
    """동일 지역 안에서는 사용자가 입력한 기존 순서를 그대로 보존한다."""
    return sorted(list(rows or []),key=_keyword_region_rank)

def load_keywords(): return order_keywords(load_json(KEYWORDS_FILE,[]))
def save_keywords(k): save_json(KEYWORDS_FILE,order_keywords(k))

def pool_columns(pool):
    """풀에서 지역/서비스/브랜드 열별 고유값 목록(순서보존)."""
    dd=lambda L:list(dict.fromkeys([x for x in L if x]))
    R=dd([(x.get('지역') or '').strip() for x in pool])
    S=dd([(x.get('서비스') or '').strip() for x in pool])
    B=dd([(x.get('브랜드') or '').strip() for x in pool])
    return R,S,B

def pick_keywords(pool, cfg):
    """키워드1(지역)을 먼저 뽑고, 키워드2·3은 반드시 같은 지역 행 안에서만 조합한다."""
    if not pool: return {'지역':'','서비스':'','브랜드':''}
    row=random.choice(pool)
    if cfg.get('mix_keywords',True):
        region=(row.get('지역') or '').strip()
        matched=[x for x in pool if (x.get('지역') or '').strip()==region]
        services=[(x.get('서비스') or '').strip() for x in matched if (x.get('서비스') or '').strip()]
        brands=[(x.get('브랜드') or '').strip() for x in matched if (x.get('브랜드') or '').strip()]
        return {'지역':region,
                '서비스':(random.choice(services) if services else ((row.get('서비스') or '').strip() or '지역정보')),
                '브랜드':(random.choice(brands) if brands else ((row.get('브랜드') or '').strip() or f'{region}추천'))}
    return {'지역':(row.get('지역') or ''),'서비스':(row.get('서비스') or ''),'브랜드':(row.get('브랜드') or '')}

# ==================== 콘텐츠 중복 방지 (제목/본문 항상 다르게) ====================
_uniq_lock=threading.Lock()
def _norm_text(html):
    """태그·공백 제거한 순수 텍스트(중복 판정 기준)."""
    return re.sub(r'\s+',' ',re.sub(r'<[^>]+>',' ',html or '')).strip()

def remember_if_unique(title, body, force=False):
    """제목/본문이 과거에 없던 새 콘텐츠면 기록하고 True. 이미 있으면 False."""
    th=hashlib.sha1((title or '').encode('utf-8')).hexdigest()
    bh=hashlib.sha1(_norm_text(body).encode('utf-8')).hexdigest()
    with _uniq_lock:
        u=load_json(UNIQ_FILE,{'t':[],'b':[]})
        if not isinstance(u,dict): u={'t':[],'b':[]}
        ts=set(u.get('t',[])); bs=set(u.get('b',[]))
        if not force and (th in ts or bh in bs): return False
        u['t']=(u.get('t',[])+[th])[-8000:]; u['b']=(u.get('b',[])+[bh])[-8000:]
        save_json(UNIQ_FILE,u); return True

def force_unique_html(html):
    """(최후수단) 눈에 거의 안 띄는 유니크 토큰을 붙여 강제로 다르게."""
    nonce=secrets.token_hex(5)
    return html+f'<p style="font-size:1px;line-height:1px;color:#ffffff;opacity:0.02;margin:0;">{nonce}</p>'

# ==================== 리치 HTML 생성 (bm21 스타일 · 변형/서비스특화/중복방지) ====================
# 서비스 유형별 특화 어휘 (미등록 서비스는 DEFAULT_FLAVOR 사용)
SERVICE_FLAVOR = {
    '셔츠룸':  {'mood':['깔끔하고 세련된','밝고 활기찬','트렌디한 감성의']},
    '노래방':  {'mood':['신나고 흥겨운','프라이빗하고 아늑한','최신 곡이 가득한']},
    '가라오케':{'mood':['고급스럽고 품격 있는','세련되고 정갈한','차분하고 프라이빗한']},
    '호빠':    {'mood':['활기차고 친근한','편안하고 밝은','유쾌한']},
    '룸싸롱':  {'mood':['프라이빗하고 고급스러운','조용하고 차분한','정중한 응대의']},
    '가요주점':{'mood':['정겹고 흥겨운','편안한','활기찬']},
}
DEFAULT_FLAVOR={'mood':['편안하고 세련된','밝고 활기찬','아늑한']}
RELATED_POOL=['하이퍼블릭','노래빠','쓰리노','가요광장','터치룸','노래클럽','가라오케','호빠','룸싸롱','셔츠룸','퍼블릭','비즈니스바','가요주점']

def generate_rich_html(keywords, cfg):
    r=(keywords.get('지역') or '서울').strip()
    s=(keywords.get('서비스') or '셔츠룸').strip()
    b=(keywords.get('브랜드') or cfg.get('brand') or '인천홍마니').strip()
    # 제목: 키워드1 + 랜덤변형 번호 + 키워드2 + 키워드3  / 본문: 선택된 번호의 정상 표기
    title,rawphone=build_title(r,s,b,cfg)
    p=format_phone(rawphone)
    mood=random.choice(SERVICE_FLAVOR.get(s,DEFAULT_FLAVOR)['mood'])
    imgs=pick_images(1)   # 게시물당 이미지는 항상 정확히 1개
    c1,c2,c3=random.sample(COLORS,3)
    rel=RELATED_POOL[:]; random.shuffle(rel)
    H=lambda t:f'<h2 style="color:{c1};border-bottom:3px solid {c2};padding-bottom:10px;font-size:24px;margin-top:34px;">{t}</h2>'
    # 이미지 alt = 치환키워드 맨앞(지역) 그대로 (SEO)
    IMG=lambda i,cap='':(f'<div style="text-align:center;margin:34px 0;"><img src="{imgs[i%len(imgs)]}" alt="{r}" style="max-width:100%;height:auto;border-radius:8px;" loading="lazy" />'+(f'<p style="color:#888;font-size:13px;margin-top:8px;">▲ {cap}</p>' if cap else '')+'</div>')

    intro=random.choice([
        f'{r} 지역에서 {s}를 찾고 계신가요? {mood} 분위기의 <strong>{r} {s}</strong>는 회식과 모임 장소로 꾸준히 사랑받는 곳입니다. {b}에서 위치와 이용 정보를 한눈에 정리해 드립니다.',
        f'기업 회식, 지인 모임, 특별한 자리까지 — <strong>{r} {s}</strong>는 다양한 목적에 어울리는 공간입니다. {mood} 인테리어와 좋은 접근성으로 {r} 인근에서 인기가 높습니다. 자세한 안내는 {b}가 도와드립니다.',
        f'{r} {s}를 처음 방문하신다면 어디로 가야 할지 고민되기 마련입니다. {b}는 {r} 지역의 {s} 정보를 정리해 편하게 선택하실 수 있도록 안내합니다. {mood} 공간에서 좋은 시간 보내세요.',
        f'{mood} 분위기와 넓은 공간을 갖춘 <strong>{r} {s}</strong>. 접대와 단체 모임에 적합하며 {r} 중심가에서 가까워 이동이 편리합니다. 예약과 문의는 {b}가 도와드립니다.',
    ])
    para2=random.choice([
        f'{r} {s}의 가장 큰 장점은 접근성과 공간감입니다. 대중교통과 주차 모두 이용하기 편리해 부담 없이 방문할 수 있고, 내부는 여유로운 좌석 배치로 편안함을 더했습니다.',
        f'단체 예약 시 인원과 목적에 맞춰 공간을 안내받을 수 있어, 회식이나 모임을 준비하는 분들의 만족도가 높습니다. 미리 문의하면 원하는 시간대에 맞춰 준비가 가능합니다.',
        f'{r} 지역 특성상 {s} 선택지가 다양하지만, 위치와 분위기, 응대 수준을 함께 고려하면 후회 없는 선택을 할 수 있습니다. {b}는 이 기준으로 정보를 정리합니다.',
        f'처음 방문하는 분도 편하게 이용할 수 있도록 안내가 잘 되어 있으며 재방문율이 높은 편입니다. 궁금한 점은 방문 전 미리 문의하시면 자세히 안내받을 수 있습니다.',
    ])

    feat_pool=[
        '넓은 홀과 프라이빗 룸으로 인원에 맞춘 공간 선택이 가능합니다.',
        '선명한 음향과 조명으로 분위기를 한층 끌어올립니다.',
        '대중교통·주차 접근성이 좋아 이동이 편리합니다.',
        '친절하고 세심한 응대로 편안한 이용을 돕습니다.',
        '예약 문의가 간편해 원하는 시간대를 잡기 쉽습니다.',
        '합리적인 가격 구성으로 부담을 줄였습니다.',
        '청결하게 관리되는 쾌적한 실내 환경을 유지합니다.',
        '단체 모임과 회식에 적합한 좌석 배치를 갖췄습니다.',
        f'{r} 중심가에서 가까워 약속 장소로 잡기 좋습니다.',
        '초행길에도 찾기 쉬운 위치에 자리하고 있습니다.',
        '분위기 있는 인테리어로 특별한 자리를 완성합니다.',
        '다양한 목적의 모임에 유연하게 대응합니다.',
    ]
    feats=''.join(f'<li style="margin:8px 0;line-height:1.8;">{x}</li>' for x in random.sample(feat_pool,random.randint(5,7)))

    faq_pool=[
        ('예약은 어떻게 하나요?', f'{p}로 문의 주시면 인원과 시간에 맞춰 빠르게 안내해 드립니다. 방문 전 예약을 권장합니다.'),
        ('주차가 가능한가요?','인근 주차 이용이 가능하며 대중교통 접근성도 좋아 차량·도보 모두 편리합니다.'),
        ('단체 이용도 되나요?','인원에 맞춰 공간을 안내해 드리므로 회식·모임 등 단체 이용에 적합합니다. 사전 문의 시 준비가 원활합니다.'),
        (f'{r} 어디에 있나요?', f'{r} 중심가 인근에 위치해 찾기 쉽습니다. 정확한 위치는 문의 시 안내해 드립니다.'),
        ('처음 방문인데 괜찮을까요?','처음 오시는 분도 편하게 이용하실 수 있도록 안내가 잘 되어 있습니다. 부담 없이 방문하세요.'),
        ('이용 시간은 어떻게 되나요?','이용 가능 시간대는 시기에 따라 다를 수 있어 방문 전 문의로 확인하시는 것을 권장합니다.'),
        ('분위기는 어떤가요?', f'{mood} 분위기로 편안하게 시간을 보내기 좋습니다.'),
        ('예약 없이 방문해도 되나요?','가능하나 원하는 시간대 이용을 위해서는 사전 예약을 추천드립니다.'),
        ('문의는 어디로 하나요?', f'{p}로 연락 주시면 친절하게 안내해 드립니다.'),
    ]
    faqs=''.join(f'<dt style="font-weight:bold;color:{c2};margin-top:12px;">Q. {q}</dt><dd style="margin:6px 0 12px 20px;line-height:1.7;">{a}</dd>' for q,a in random.sample(faq_pool,random.randint(4,6)))

    reco_pool=['회식·접대 장소를 찾는 직장인','지인들과 편하게 모일 공간이 필요한 분',f'{r} 인근에서 약속 장소를 정하려는 분','믿을 만한 정보로 실패 없이 고르고 싶은 분','분위기 좋은 자리를 원하는 분','접근성 좋은 위치를 선호하는 분']
    recos=''.join(f'<li style="margin:8px 0;line-height:1.8;">{x}</li>' for x in random.sample(reco_pool,random.randint(3,min(5,len(reco_pool)))))

    review=random.choice([
        f'지인 추천으로 {r} {s}에 다녀왔습니다. {mood} 분위기에 응대도 친절해서 편안하게 즐겼어요. 다음에도 또 방문하고 싶습니다.',
        f'회식 장소로 {r} {s}를 예약했는데 공간이 넓고 깔끔해서 만족스러웠습니다. {b} 안내대로 미리 문의하니 준비가 잘 되어 있었어요.',
        f'위치가 찾기 쉬워 초행인데도 헤매지 않았습니다. {r} {s} 분위기가 좋아 모임이 화기애애했어요. 추천합니다.',
        f'{b} 정보를 보고 방문했는데 기대 이상이었습니다. 가격도 합리적이고 응대도 세심해 좋은 자리가 됐습니다.',
        f'{r}에서 {s} 고민하다 방문했는데 선택 잘했다는 생각이 들었어요. {mood} 공간에서 편하게 대화 나눴습니다.',
    ])
    chips=' '.join(f'<span style="display:inline-block;padding:4px 12px;margin:3px;background:#E5E9EE;font-size:13px;color:{c3};border:1px solid #dde4f0;border-radius:4px;">{x}</span>' for x in random.sample(rel,random.randint(8,min(13,len(rel)))))

    # 중간 콘텐츠 블록들 — 순서를 매번 섞어 변형 폭을 키움(중복 방지)
    blocks=[
        H(f'{r} {s} 주요 특징')+f'<ul style="font-size:15px;padding-left:20px;color:#444;">{feats}</ul>',
        H(f'{r} {s} 이런 분께 추천합니다')+f'<ul style="font-size:15px;padding-left:20px;color:#444;">{recos}</ul>',
        H(f'{r} {s} 자주 묻는 질문')+f'<dl style="font-size:15px;margin:15px 0;">{faqs}</dl>',
        H(f'{r} {s} 방문 후기')+f'<p style="font-size:15px;line-height:1.9;">{review}</p>',
        H(f'{r} 인근 이용 가능 지역')+f'<div style="margin:12px 0;line-height:2.6;">{chips}</div>',
    ]
    random.shuffle(blocks)
    html=(f'<h1 style="font-size:22px;font-weight:bold;margin:0 0 18px 0;color:#222;">{title}</h1>'
        + IMG(0, f'{r} {s}의 {mood.split()[0]} 공간')
        + H(f'{r} {s} 안내')
        + f'<p style="font-size:16px;margin:18px 0;line-height:1.9;">{intro}</p>'
        + f'<p style="font-size:15px;margin:14px 0;line-height:1.9;color:#333;">{para2}</p>'
        + ''.join(blocks)
        + f'<div style="margin:40px 0;padding:22px;background:#f8f9fa;border-radius:8px;text-align:center;">'
          f'<p style="font-size:18px;font-weight:bold;color:{c1};">{r} {s} 문의 및 예약</p>'
          f'<p style="font-size:24px;font-weight:bold;color:{c2};margin:12px 0;">{p}</p>'
          f'<p style="font-size:14px;color:#666;">{b} · 믿을 수 있는 {r} {s} 정보</p></div>')
    return html, title

# ==================== GPT 본문 생성 (선택) ====================
def generate_post_gpt(keywords, cfg):
    """OpenAI로 키워드1 중심의 장문 HTML 본문 생성. 실패 시 템플릿으로 폴백."""
    import requests as _rq
    r=(keywords.get('지역') or '서울').strip(); s=(keywords.get('서비스') or '셔츠룸').strip()
    b=(keywords.get('브랜드') or cfg.get('brand') or '인천홍마니').strip()
    _rawph=pick_phone(cfg); p=format_phone(_rawph)
    key=cfg.get('openai_key',''); model=cfg.get('model') or 'gpt-4o-mini'
    if not key: raise RuntimeError('openai_key 없음')
    imgs=pick_images(1); c1,c2=random.sample(COLORS,2)   # 게시물당 이미지는 항상 정확히 1개
    sys_p=("너는 한국어 정보형 랜딩페이지와 지역 안내 글을 작성하는 전문 카피라이터다. "
           "세 개의 키워드 중 키워드1을 문서 전체의 명확한 메인 주제로 삼고, 키워드2와 키워드3은 "
           "메인 주제를 설명하는 보조 문맥으로만 사용한다. 읽기 쉬운 장문 콘텐츠를 쓰되 키워드 도배, "
           "과장·허위·불법·선정적 표현과 검증되지 않은 수치·후기는 만들지 않는다.")
    usr_p=(f"키워드1(메인)='{r}', 키워드2(보조)='{s}', 키워드3(보조)='{b}', 문의전화='{p}'.\n"
           f"첨부 예시처럼 길고 구조적인 정보형 게시글을 작성하라. 반드시 지킬 조건:\n"
           f"1. 문서의 검색 의도와 핵심 주제는 오직 키워드1 '{r}'이다. 키워드2·3은 '{r}'을 설명할 때만 자연스럽게 보조한다.\n"
           f"2. 첫 문단, 주요 소제목 3개 이상, 핵심 정리와 마지막 문단에 '{r}'을 자연스럽게 포함한다. "
           f"동일 문장이나 부자연스러운 반복은 금지한다.\n"
           f"3. 키워드2 '{s}'와 키워드3 '{b}'는 각각 2~4회 정도만 사용하고 메인키워드보다 눈에 띄지 않게 한다.\n"
           f"4. 순수 HTML 조각만 출력한다. 코드블록·마크다운·설명·<html><body> 태그는 금지한다.\n"
           f"5. 1,800~2,800자 분량으로 작성한다. 구성은 도입 → 핵심 안내 → 세부 항목 8~12개 → "
           f"정보 카드 4개 → 이용 팁 → FAQ 5개 → 핵심 정리 → 문의 CTA 순서로 한다.\n"
           f"6. <h2> 5~7개, 세부 항목은 <h3>과 <p>, 특징은 <ul><li>, FAQ는 <dl><dt><dd>를 사용한다. "
           f"정보 카드 4개는 테두리와 여백을 준 <div>로 만든다. 모바일에서도 읽기 쉬운 인라인 스타일을 사용한다.\n"
           f"7. 전화번호 {p}는 본문 중간에 반복하지 말고 마지막 문의 CTA에 1회만 넣는다.\n"
           f"8. 실제로 주어지지 않은 주소·가격·운영시간·후기·보장 표현은 단정하지 않는다. 매번 문장과 항목 순서를 다르게 한다.")
    resp=_rq.post("https://api.openai.com/v1/chat/completions",
        headers={"Authorization":f"Bearer {key}","Content-Type":"application/json"},
        json={"model":model,"temperature":0.85,"max_tokens":3800,
              "messages":[{"role":"system","content":sys_p},{"role":"user","content":usr_p}]},
        timeout=60)
    resp.raise_for_status()
    payload=resp.json()
    _record_openai_usage(model,payload.get('usage') or {},cfg)
    body=payload['choices'][0]['message']['content'].strip()
    if body.startswith('```'): body=re.sub(r'^```[a-zA-Z]*\n?|```$','',body).strip()
    title,_=build_title(r,s,b,cfg,_rawph)
    header=(f'<h1 style="font-size:22px;font-weight:bold;color:#222;margin:0 0 16px;">{title}</h1>'
            f'<div style="text-align:center;margin:20px 0;"><img src="{imgs[0]}" alt="{r}" style="max-width:100%;border-radius:8px;" loading="lazy"/></div>')
    cta=(f'<div style="margin:36px 0;padding:20px;background:#f8f9fa;border-radius:8px;text-align:center;">'
         f'<p style="font-size:18px;font-weight:bold;color:{c1};">{r} {s} 문의·예약</p>'
         f'<p style="font-size:24px;font-weight:bold;color:{c2};margin:10px 0;">{p}</p>'
         f'<p style="font-size:14px;color:#666;">{b}</p></div>')
    return header+body+cta, title

def _gen_once(keywords, cfg):
    if cfg.get('use_gpt') and cfg.get('openai_key'):
        try:
            return generate_post_gpt(keywords,cfg)
        except Exception as e:
            add_log(f'[GPT 실패→템플릿] {str(e)[:80]}')
    return generate_rich_html(keywords,cfg)

def generate_article(keywords, cfg, unique=True):
    """본문 생성. unique=True 면 제목/본문이 과거와 겹치지 않을 때까지 재생성(상시 다르게)."""
    if not unique:
        return _gen_once(keywords,cfg)
    html=title=None
    for _ in range(8):
        html,title=_gen_once(keywords,cfg)
        if remember_if_unique(title,html): return html,title
    # 8회 모두 충돌(사실상 불가) → 강제 유니크 토큰 부착
    html=force_unique_html(html); remember_if_unique(title,html,force=True)
    return html,title

# ==================== 텔레그램 알림 (선택) ====================
def send_telegram(cfg, msg):
    tok=cfg.get('telegram_token',''); chat=cfg.get('telegram_chat_id','')
    if not tok or not chat: return False
    try:
        import requests as _rq
        _rq.post(f"https://api.telegram.org/bot{tok}/sendMessage",
                 json={"chat_id":chat,"text":msg,"disable_web_page_preview":True},timeout=10)
        return True
    except Exception: return False

def send_telegram_doc(cfg, filename, data_bytes, caption=''):
    """텔레그램으로 파일(백업 zip/엑셀) 전송."""
    tok=cfg.get('telegram_token',''); chat=cfg.get('telegram_chat_id','')
    if not tok or not chat: return False,'텔레그램 토큰/챗ID 미설정'
    try:
        import requests as _rq
        r=_rq.post(f"https://api.telegram.org/bot{tok}/sendDocument",
                   data={'chat_id':chat,'caption':caption[:1000]},
                   files={'document':(filename,data_bytes)},timeout=40)
        j=r.json()
        return bool(j.get('ok')),('' if j.get('ok') else str(j)[:150])
    except Exception as e: return False,str(e)[:150]

# ==================== 캡차 / 보안 인증 감지 (우회 아님 — 감지해서 제외) ====================
def detect_captcha(d):
    """현재 페이지에 캡차/보안 인증이 있는지 감지. 있으면 종류, 없으면 ''.
       (정책상 자동 해석·우회 없음 — 감지되면 자동발행에서 제외한다.)"""
    try: html=d.page_source or ''
    except Exception: html=''
    low=html.lower()
    if 'cf-turnstile' in low or 'turnstile' in low: return 'turnstile'
    if 'h-captcha' in low or 'hcaptcha' in low: return 'hcaptcha'
    if 'g-recaptcha' in low or 'grecaptcha' in low or 'recaptcha' in low: return 'recaptcha'
    if ('captcha_key' in low or 'kcaptcha' in low or 'g5_captcha' in low
        or '자동등록방지' in html or '자동입력방지' in html or '보안문자' in html or 'captcha' in low):
        return 'kcaptcha'
    return ''

# CAPTCHA는 자동 해석하거나 우회하지 않는다. 폼을 모두 채운 뒤 사람이
# 입력한 값만 같은 Selenium 세션에 전달하여 등록을 계속한다.
CAPTCHA_LOCK=threading.Lock()
CAPTCHA_TASKS={}

def _captcha_public(task):
    return {k:v for k,v in task.items() if k not in ('event','value','cancelled')}

def _kcaptcha_force_load(d):
    """그누보드 kcaptcha 특화: 초기 #captcha_img는 dot.gif 플레이스홀더라서
       JS가 새로고침 전엔 실이미지가 없다. g5_captcha_url을 읽어 세션을 새로 만들고
       실제 kcaptcha_image.php 를 src에 강제 로드한 뒤 디코딩까지 기다린다.
       실이미지가 뜨면 True, 아니면 False(다른 플랫폼이면 조용히 False).
       세션 정합성: kcaptcha_session.php가 세션에 정답을 심고 kcaptcha_image.php는
       그 정답을 그릴 뿐이므로, 심기 실패 시 stale 이미지를 풀지 않도록 abort한다."""
    import time
    try:
        # 이미 진짜 이미지가 떠 있으면(dot.gif 아니고 naturalWidth>0) 세션 재생성 불필요
        already=d.execute_script("""
            var i=document.getElementById('captcha_img');
            if(!i) return false;
            var s=i.getAttribute('src')||'';
            return (s.indexOf('dot.gif')===-1 && (i.naturalWidth||0)>5);
        """)
        if already:
            return True
        cap_url=d.execute_script("return (typeof g5_captcha_url!=='undefined')?g5_captcha_url:'';") or ''
        if not cap_url:
            return False
        # 동기 XHR이 서버 지연 시 JS 스레드를 오래 막을 수 있어 스크립트 타임아웃 상한을 건다
        try: d.set_script_timeout(8)
        except Exception: pass
        # 세션에 캡차 정답 심기 → 성공(2xx)해야만 그 정답의 이미지 로드. 실패 시 False 반환.
        ok=d.execute_script("""
            var base=arguments[0], done=false;
            try{
              var xhr=new XMLHttpRequest();
              xhr.open('POST', base+'/kcaptcha_session.php', false); xhr.send();
              done=(xhr.status>=200 && xhr.status<300);
            }catch(e){ done=false; }
            if(done){
              var img=document.getElementById('captcha_img');
              if(img){ img.src = base+'/kcaptcha_image.php?t='+(new Date()).getTime(); }
            }
            return done;
        """, cap_url)
        if not ok:
            return False   # 세션 재생성 실패 → stale 이미지를 풀지 않는다
        # 이미지 디코딩 완료까지 대기(naturalWidth>0)
        for _ in range(20):
            try:
                natw=d.execute_script("var i=document.getElementById('captcha_img');return i?(i.naturalWidth||0):0;") or 0
                if natw>5: return True
            except Exception: pass
            time.sleep(0.3)
        return False
    except Exception:
        return False

def _captcha_image_data(d):
    """그누보드(kcaptcha)/XE·Rhymix/Cafe24 캡차 이미지를 base64로 반환. 못 찾으면 ''."""
    from selenium.webdriver.common.by import By
    import time

    # 0) 그누보드 kcaptcha면 실이미지를 먼저 강제 로드(dot.gif 플레이스홀더 문제 해결)
    _kcaptcha_force_load(d)

    # 넓은 폴백(크기추정/data-URI)은 캡차 INPUT이 실제 있을 때만 → 로고·아이콘 오탐 방지
    has_captcha_input=False
    try:
        has_captcha_input=bool(d.find_elements(By.CSS_SELECTOR,
            "input[name*='captcha'],input[id*='captcha'],#captcha_key,input[name='wr_key'],#secret_text"))
    except Exception:
        pass

    def _is_recaptcha(el):
        # reCAPTCHA/hCaptcha 요소는 이미지캡차 솔버(normal)로 보내면 안 됨 → 제외
        try:
            src=(el.get_attribute('src') or '').lower()
            if any(k in src for k in ('recaptcha','hcaptcha','gstatic.com/recaptcha','google.com/recaptcha')):
                return True
        except Exception: pass
        return False

    def _shot(el):
        # 어떤 엘리먼트든(img/canvas/svg/div/bg-image) 렌더된 픽셀을 그대로 캡처
        try:
            if not el or _is_recaptcha(el):
                return ''
            try:
                sz = el.size
                w, h = sz.get('width', 0), sz.get('height', 0)
            except Exception:
                w = h = 0
            # naturalWidth로 실제 디코딩 여부 확인(레이아웃상 크기 0이어도 이미지 픽셀 존재 가능)
            try:
                natw = d.execute_script("return arguments[0].naturalWidth||0;", el) or 0
            except Exception:
                natw = 0
            if w <= 0 and h <= 0 and natw <= 5:
                return ''
            try:
                d.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                time.sleep(0.25)  # 스샷 전 픽셀 페인트 대기(kcaptcha src 교체 직후 방지)
            except Exception:
                pass
            b64 = el.screenshot_as_base64
            return 'data:image/png;base64,' + b64 if b64 else ''
        except Exception:
            return ''

    # 1) 명시적 CSS 셀렉터: 가장 구체적 → 가장 일반적 (gnuboard/XE/Rhymix/Cafe24/일반)
    SELECTORS = [
        # kcaptcha src가 교체된 진짜 이미지 (dot.gif 플레이스홀더 배제)
        "img#captcha_img[src*='kcaptcha_image.php']",
        "img#captcha_img[src*='captcha.php']",
        "img#captcha_img:not([src*='dot.gif'])",
        "img#captcha_img",                          # gnuboard5 kcaptcha 정식 id
        "img#captcha_image",                        # XE/Rhymix 인라인
        "#captcha img#captcha_img", "fieldset#captcha img", "#captcha img",
        ".captcha img", "fieldset.captcha img",
        # src 특징 기반
        "img[src*='kcaptcha_image.php']", "img[src*='captcha_action=captchaImage']",
        "img[src*='captchaImage']", "img[src*='captcha.php']", "img[src*='kcaptcha']",
        "img[src*='/kcaptcha/']", "img[src*='seccode']", "img[src*='securimage']",
        "img[src*='authimg']", "img[src*='boan']", "img[src*='=captcha']",
        "img[src*='image.php']", "img[src*='vcode']", "img[src*='chkcaptcha']",
        # id/class/alt/title 기반
        "img[id*='captcha']:not([src*='dot.gif'])", "img[class*='captcha']",
        "img[alt='CAPTCHA']", "img[alt*='captcha']", "img[alt*='보안']",
        "img[title*='자동등록방지']",
        # 래퍼 스코프
        "#captchaArea img", ".captcha-box img", "td.captcha img", ".kcaptcha img",
        ".captcha_wrap img", ".captcha_box img", "#captchaWrap img", ".captchaImage img",
        # 인라인 base64 data-URI (일부 커스텀 스킨)
        "img[src^='data:image']",
        # src에 잡히는 캡차성 이미지(최후 일반)
        "img[src*='captcha']",
    ]

    # 1-pass: is_displayed()가 True인 것 우선
    for sel in SELECTORS:
        try:
            for el in d.find_elements(By.CSS_SELECTOR, sel):
                try:
                    if el.is_displayed():
                        r = _shot(el)
                        if r:
                            return r
                except Exception:
                    pass
        except Exception:
            pass

    # 2-pass: is_displayed()가 False로 잘못 잡혀도 크기(width>0,height>0)만 있으면 수용
    for sel in SELECTORS:
        try:
            for el in d.find_elements(By.CSS_SELECTOR, sel):
                try:
                    sz = el.size
                    if sz.get('width', 0) > 0 and sz.get('height', 0) > 0:
                        r = _shot(el)
                        if r:
                            return r
                except Exception:
                    pass
        except Exception:
            pass

    # 3) canvas / svg 캡차 (screenshot_as_base64는 img 아닌 엘리먼트도 캡처됨)
    for sel in ["canvas[id*='captcha']", "canvas[class*='captcha']", "#captcha canvas",
                ".captcha canvas", "svg[id*='captcha']", "svg[class*='captcha']",
                "#captcha svg", ".captcha svg"]:
        try:
            for el in d.find_elements(By.CSS_SELECTOR, sel):
                r = _shot(el)
                if r:
                    return r
        except Exception:
            pass

    # 4) background-image div/span 캡차: 계산된 스타일에서 captcha 토큰 탐지 후 그 박스를 스샷
    try:
        els = d.execute_script("""
            const out=[];
            for (const el of document.querySelectorAll("div,span,a,i,button,td,[class*='captcha'],[id*='captcha']")) {
                const bg = getComputedStyle(el).backgroundImage || '';
                if (bg && bg !== 'none' && /captcha|kcaptcha|securimage|seccode|boan|vcode/i.test(bg)) {
                    const r = el.getBoundingClientRect();
                    if (r.width > 0 && r.height > 0) out.push(el);
                }
            }
            return out;
        """) or []
        for el in els:
            r = _shot(el)
            if r:
                return r
    except Exception:
        pass

    # 5) 폴백: 캡차 INPUT을 찾고, 공통 조상을 4단계까지 거슬러 올라가 가까운 <img>를 캡처
    #    (input#captcha_key, name=captcha_key, name=wr_key, id/name*=captcha)
    for isel in ["#captcha_key", "input[name='captcha_key']", "input[name='wr_key']",
                 "input[name*='captcha']", "input[id*='captcha']", "#secret_text"]:
        try:
            for inp in d.find_elements(By.CSS_SELECTOR, isel):
                node = inp
                for _ in range(4):  # 조상 최대 4단계 상승
                    try:
                        node = node.find_element(By.XPATH, "./..")
                    except Exception:
                        break
                    try:
                        cands = []
                        for im in node.find_elements(By.TAG_NAME, "img"):
                            try:
                                sz = im.size
                                w, h = sz.get('width', 0), sz.get('height', 0)
                            except Exception:
                                w = h = 0
                            try:
                                natw = d.execute_script("return arguments[0].naturalWidth||0;", im) or 0
                            except Exception:
                                natw = 0
                            if (w > 0 and h > 0) or natw > 5:
                                cands.append((w * h if w and h else natw, im))
                        if cands:
                            # 캡차는 작은 이미지 → 가장 작은 후보 우선(스페이서/배너 회피)
                            cands.sort(key=lambda t: t[0])
                            r = _shot(cands[0][1])
                            if r:
                                return r
                    except Exception:
                        pass
        except Exception:
            pass

    # 6) 최종 폴백: 글쓰기 <form> 내부에서 캡차 크기(폭 40~260, 높이 20~120)의 보이는 <img> 첫 매치
    #    캡차 INPUT이 있을 때만 실행(없으면 로고/버튼 이미지 오탐 위험이라 스킵)
    try:
        forms = d.find_elements(By.CSS_SELECTOR,
            "form[name='fwrite'], form#fwrite, form[action*='write_update'], form") if has_captcha_input else []
        for form in forms:
            try:
                for im in form.find_elements(By.TAG_NAME, "img"):
                    try:
                        sz = im.size
                        w, h = sz.get('width', 0), sz.get('height', 0)
                    except Exception:
                        continue
                    if 40 <= w <= 260 and 20 <= h <= 120:  # 캡차 전형 크기(아이콘·배너 배제)
                        r = _shot(im)
                        if r:
                            return r
            except Exception:
                pass
    except Exception:
        pass

    return ''  # 로그: 여기 도달 시 캡차 미필요(로그인/관리자·비활성)이거나 iframe 내부일 수 있음

def solve_captcha_with_2captcha(d,site,cap_type,cfg,timeout=300):
    """2captcha API를 사용해 CAPTCHA를 자동으로 해결한다."""
    api_key=(cfg.get('twocaptcha_api_key') or '').strip()
    if not api_key or not cfg.get('twocaptcha_enabled'):
        return False,'2captcha 설정 없음','',{}
    
    try:
        from twocaptcha import TwoCaptcha
        from selenium.webdriver.common.by import By
        solver=TwoCaptcha(api_key)
        
        # recaptcha 처리
        if cap_type=='recaptcha':
            # sitekey는 여러 위치에 있을 수 있다: data-sitekey 속성 / g-recaptcha 클래스 /
            # reCAPTCHA iframe src의 ?k= 파라미터 / 페이지 소스 정규식.
            sitekey=None
            for sel in ('[data-sitekey]','.g-recaptcha[data-sitekey]','div.g-recaptcha'):
                try:
                    el=d.find_element(By.CSS_SELECTOR,sel)
                    v=el.get_attribute('data-sitekey')
                    if v: sitekey=v; break
                except Exception: pass
            if not sitekey:
                try:
                    for fr in d.find_elements(By.CSS_SELECTOR,"iframe[src*='recaptcha']"):
                        src=fr.get_attribute('src') or ''
                        mk=re.search(r'[?&]k=([A-Za-z0-9_-]{20,})',src)
                        if mk: sitekey=mk.group(1); break
                except Exception: pass
            if not sitekey:
                try:
                    mk=re.search(r'data-sitekey=["\']([A-Za-z0-9_-]{20,})["\']',d.page_source or '')
                    if mk: sitekey=mk.group(1)
                except Exception: pass
            if not sitekey:
                return False,'sitekey를 찾을 수 없음','',{}
            try:
                result=solver.recaptcha(sitekey=sitekey,url=d.current_url)
                token=result.get('code') or result.get('token') or str(result)
                # reCAPTCHA는 g-recaptcha-response 텍스트영역에 토큰을 넣어야 검증된다.
                try:
                    d.execute_script(
                        "var t=document.getElementById('g-recaptcha-response');"
                        "if(!t){t=document.querySelector('textarea[name=\"g-recaptcha-response\"]');}"
                        "if(t){t.style.display='block';t.value=arguments[0];}", token)
                except Exception: pass
                return True,'recaptcha 해결 완료',token,{'type':'recaptcha','token':token}
            except Exception as e:
                return False,f'recaptcha 해결 실패: {str(e)[:80]}','',{}
        
        # kcaptcha (이미지) 처리
        elif cap_type=='kcaptcha':
            image_data=_captcha_image_data(d)
            if not image_data:
                return False,'captcha 이미지를 찾을 수 없음','',{}
            
            import base64, tempfile
            base64_str=image_data.split(',')[1] if ',' in image_data else image_data
            image_bytes=base64.b64decode(base64_str)
            
            with tempfile.NamedTemporaryFile(suffix='.png',delete=False) as f:
                f.write(image_bytes); temp_path=f.name
            
            try:
                result=solver.normal(temp_path)
                answer=result.get('code') or str(result)
                return True,'kcaptcha 해결 완료',answer,{'type':'kcaptcha','answer':answer}
            except Exception as e:
                return False,f'kcaptcha 해결 실패: {str(e)[:80]}','',{}
            finally:
                try: os.unlink(temp_path)
                except: pass
        
        else:
            return False,f'지원하지 않는 captcha 타입: {cap_type}','',{}
            
    except ImportError:
        return False,'2captcha 라이브러리 미설치','',{}
    except Exception as e:
        return False,f'2captcha 오류: {str(e)[:80]}','',{}

def wait_for_manual_captcha(d,site,cap_type,timeout=600):
    """관리자가 CAPTCHA 값을 입력할 때까지 같은 브라우저 세션을 보존한다."""
    from selenium.webdriver.common.by import By
    tid=uuid.uuid4().hex[:12]; ev=threading.Event(); now=_kst_now()
    task={'id':tid,'site_id':site.get('id'),'site_name':site.get('name') or site.get('site_url',''),
          'captcha_type':cap_type,'image_data':_captcha_image_data(d),'status':'waiting_input',
          'message':'CAPTCHA를 입력하면 같은 작성 화면에서 자동 발행합니다.',
          'created_at':now.strftime('%Y-%m-%d %H:%M:%S'),
          'expires_at':(now+timedelta(seconds=timeout)).strftime('%Y-%m-%d %H:%M:%S'),
          'event':ev,'value':'','cancelled':False}
    with CAPTCHA_LOCK: CAPTCHA_TASKS[tid]=task
    add_log(f'[CAPTCHA 대기] {task["site_name"]} — 관리자 입력 후 자동 발행')
    if not ev.wait(timeout):
        with CAPTCHA_LOCK: task['status']='expired'; task['message']='입력 대기시간 10분 만료'
        return False,'CAPTCHA 수동 입력 대기시간 만료',tid
    with CAPTCHA_LOCK:
        value=str(task.get('value') or '').strip(); cancelled=bool(task.get('cancelled'))
    if cancelled:
        with CAPTCHA_LOCK: task['status']='cancelled'; task['message']='관리자가 취소함'
        return False,'CAPTCHA 수동 입력 취소',tid
    inputs=[]
    for sel in ["input[name='captcha_key']","#captcha_key","input[name='wr_key']","input[name*='captcha']","input[id*='captcha']"]:
        try: inputs.extend([x for x in d.find_elements(By.CSS_SELECTOR,sel) if x.is_displayed()])
        except Exception: pass
        if inputs: break
    if not inputs:
        with CAPTCHA_LOCK: task['status']='failed'; task['message']='CAPTCHA 입력칸을 찾지 못함'
        return False,'CAPTCHA 입력칸을 찾지 못함',tid
    try:
        inputs[0].clear(); inputs[0].send_keys(value)
        with CAPTCHA_LOCK: task['status']='submitting'; task['message']='입력 완료 · 자동 등록 중'
        return True,'',tid
    except Exception as e:
        with CAPTCHA_LOCK: task['status']='failed'; task['message']=str(e)[:120]
        return False,'CAPTCHA 입력 전달 실패: '+str(e)[:100],tid

def finish_captcha_task(tid,ok,message):
    if not tid: return
    with CAPTCHA_LOCK:
        t=CAPTCHA_TASKS.get(tid)
        if t:
            t['status']='done' if ok else 'failed'; t['message']=str(message)[:180]
            t['finished_at']=_kst_now().strftime('%Y-%m-%d %H:%M:%S')

def _page_is_blocked(d):
    """403/보안 차단 페이지 판별. 일반 사이트의 Cloudflare 스크립트 문자열은 차단으로 보지 않는다."""
    try:
        from selenium.webdriver.common.by import By
        html=(d.page_source or '').lower()
        title=(d.title or '').lower()
        body=(d.find_element(By.TAG_NAME,'body').text or '').lower()[:3000]
    except Exception:
        html=''; title=''; body=''
    strong=['403 forbidden','access denied','error 1020','request blocked','접근이 거부','차단되었습니다']
    if any(k in title or k in body for k in strong): return True
    # Cloudflare 도전/차단 화면은 제품명 하나가 아니라 고유 문구 조합으로 확인한다.
    if ('attention required' in title or 'just a moment' in title) and \
       ('cloudflare' in html or 'cf-chl-' in html): return True
    return False

# ==================== 실패 원인 분류 ====================
def classify_fail(msg):
    """발행 실패 메시지를 사람이 읽을 수 있는 원인으로 분류. (코드, 한글, 일시적여부)"""
    ms=str(msg or ''); low=ms.lower()
    if '캡차' in ms or 'captcha' in low or '보안 인증' in ms:
        return 'captcha','캡차/보안인증 감지',False
    if any(k in ms for k in ['타임아웃','시간 초과']) or any(k in low for k in
           ['timeout','timed out','renderer','chrome not reachable','disconnected','connection',
            'session deleted','session not created','net::err','unreachable','max retries','read timed']):
        return 'timeout','타임아웃/브라우저',True   # 일시적 → 자동 재시도
    if any(k in ms for k in ['로그인','아이디','비밀번호']) or any(k in low for k in
           ['login','mb_id','mb_password','password']):
        return 'login','로그인 실패',False
    if any(k in ms for k in ['권한','차단','금지','스팸','도배']) or any(k in low for k in
           ['blocked','forbidden','403','spam','denied','권한이 없']):
        return 'blocked','차단/권한없음',False
    if any(k in ms for k in ['게시판','에디터','셀렉터','확인 불가','글쓰기','페이지 못찾음','입력란 못찾음']) or any(k in low for k in
           ['write.php','bo_table','board_no','no such element','not found','404','wr_subject','wr_content']):
        return 'board','게시판/에디터 못찾음',False
    return 'other','기타',False

# ==================== 데이터 백업 ====================
def build_backup_zip():
    """설정·사이트·이력·예약·키워드·큐를 zip 하나로 묶음(민감정보 마스킹)."""
    import io,zipfile
    buf=io.BytesIO()
    with zipfile.ZipFile(buf,'w',zipfile.ZIP_DEFLATED) as z:
        for p in [SITES_FILE,HISTORY_FILE,SCHED_FILE,KEYWORDS_FILE,IMAGES_FILE,MEMBERS_FILE,CAND_FILE,QUEUE_FILE]:
            if os.path.exists(p): z.write(str(p),os.path.basename(str(p)))
        # config 는 비번해시·API키 제거하고 저장
        c=dict(load_config()); c.pop('password',None); c['openai_key']=''; c['telegram_token']=''
        z.writestr('config.json',json.dumps(c,ensure_ascii=False,indent=2))
    buf.seek(0); return buf.read()

def do_backup(cfg=None, reason='자동'):
    cfg=cfg or load_config()
    try:
        data=build_backup_zip()
    except Exception as e:
        add_log(f'[백업] zip 생성 실패: {str(e)[:80]}'); return False,str(e)[:120]
    ts=_kst_now().strftime('%Y%m%d_%H%M')
    cap=(f'📦 찌라시 백업 ({reason})\n{_kst_now().strftime("%Y-%m-%d %H:%M")} KST\n'
         f'사이트 {len(load_sites())}개 · 이력 {len(load_json(HISTORY_FILE,[]))}건')
    ok,err=send_telegram_doc(cfg,f'chirashi_backup_{ts}.zip',data,caption=cap)
    add_log(f'[백업] {"전송 성공" if ok else "실패: "+err} ({reason})')
    return ok,err

# ==================== 발행 검증 (게시글 생존 확인) ====================
def check_post_alive(url):
    """게시된 URL 이 아직 살아있는지 HTTP 로 확인. True=생존, False=삭제/없음, None=확인불가."""
    import requests as _rq
    try:
        r=_rq.get(url,timeout=15,verify=False,allow_redirects=True,
                  headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36'})
        if r.status_code==429:   # 과도한 요청 — Retry-After 존중하고 이번엔 판정 보류
            ra=r.headers.get('Retry-After','')
            try: time.sleep(min(int(ra),30)) if ra.isdigit() else time.sleep(5)
            except Exception: time.sleep(5)
            return None
        if r.status_code==403: return None   # 차단 — 삭제로 오판하지 않도록 보류
        if r.status_code==404 or r.status_code>=500: return False
        if r.status_code>=400: return False
        body=r.text or ''
        dead=['존재하지 않','삭제된','삭제되었','없는 게시','게시물이 없','원글이 없','잘못된 접근',
              '페이지를 찾을 수 없','not found','더 이상','권한이 없']
        if any(s in body for s in dead): return False
        return True
    except Exception:
        return None

def verify_once(limit=40):
    """이력의 성공 발행글 URL 을 재확인해 생존여부(alive)를 기록. 처리 건수 반환."""
    cfg=load_config()
    with JOB_LOCK:
        h=load_json(HISTORY_FILE,[])
    cand=[r for r in h if r.get('status')=='done' and str(r.get('result_url','')).startswith('http')]
    cand.sort(key=lambda r:(r.get('verified_at') or ''))   # 미검증·오래된 것 우선
    todo=cand[:max(1,limit)]
    results={}
    for r in todo:
        alive=check_post_alive(r['result_url'])
        if alive is not None: results[r['id']]=('yes' if alive else 'no')
        time.sleep(1)
    if results:
        with JOB_LOCK:
            h=load_json(HISTORY_FILE,[]); vat=_kst_now().strftime('%Y-%m-%d %H:%M')
            for rec in h:
                if rec.get('id') in results:
                    rec['alive']=results[rec['id']]; rec['verified_at']=vat
            save_json(HISTORY_FILE,h)
        add_log(f'[검증] {len(results)}건 확인 (생존 {sum(1 for v in results.values() if v=="yes")}·삭제 {sum(1 for v in results.values() if v=="no")})')
    return len(results)

def verify_loop():
    while True:
        try:
            if load_config().get('verify_enabled',True): verify_once(40)
        except Exception as e:
            add_log(f'[검증 오류] {str(e)[:80]}')
        time.sleep(3600)

# ==================== Selenium 드라이버 (회피코드 제거) ====================
_drivers = {}
_drv_lock = threading.Lock()

def get_driver():
    tid = threading.current_thread().name
    with _drv_lock:
        if tid in _drivers:
            try: _drivers[tid].current_url; return _drivers[tid]
            except: pass
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        opts = webdriver.ChromeOptions()
        opts.add_argument('--headless=new'); opts.add_argument('--no-sandbox')
        opts.add_argument('--disable-dev-shm-usage'); opts.add_argument('--disable-gpu')
        opts.add_argument('--window-size=1920,1080'); opts.add_argument('--log-level=3')
        # 정상 환경 일치(위조 아님): 서버가 한국(Seoul)에 있으므로 로케일/언어를 실제와 맞춤.
        # navigator.webdriver·Canvas·WebGL 등은 건드리지 않음(지문 위조·스텔스 미사용).
        opts.add_argument('--lang=ko-KR')
        opts.add_experimental_option('prefs',{'intl.accept_languages':'ko-KR,ko'})
        opts.page_load_strategy='eager'
        cb=os.environ.get('CHROME_BIN')  # 리눅스 VPS: chromium 경로 지정 가능
        if cb: opts.binary_location=cb
        svc=Service(ChromeDriverManager().install())
        d=webdriver.Chrome(service=svc,options=opts)
        d.set_page_load_timeout(25); d.implicitly_wait(3)
        _drivers[tid]=d; return d

def dismiss_alerts(d):
    for _ in range(3):
        try: d.switch_to.alert.accept(); time.sleep(0.2)
        except: break

def html_to_plain(content):
    """HTML 모드가 없는 게시판에 넣을 읽기 쉬운 일반 텍스트."""
    text=re.sub(r'(?is)<(script|style).*?>.*?</\1>','',content or '')
    text=re.sub(r'(?i)<br\s*/?>','\n',text)
    text=re.sub(r'(?i)</(?:p|div|h[1-6]|li|tr)>','\n',text)
    text=re.sub(r'(?s)<[^>]+>','',text)
    text=html_lib.unescape(text).replace('\xa0',' ')
    return re.sub(r'\n{3,}','\n\n',text).strip()

def enable_html_mode(d):
    """HTML 체크박스 또는 KBoard의 '코드' 탭을 활성화한다."""
    from selenium.webdriver.common.by import By
    for el in d.find_elements(By.CSS_SELECTOR,"input[type='checkbox']"):
        if not _sel_vis(el): continue
        blob=' '.join([_sel_attr(el,'name'),_sel_attr(el,'id'),_sel_attr(el,'value'),_sel_attr(el,'title')]).lower()
        try:
            eid=_sel_attr(el,'id')
            if eid:
                labels=d.find_elements(By.CSS_SELECTOR,"label[for='"+eid.replace("'","\\'")+"']")
                blob+=' '+' '.join((x.text or '') for x in labels).lower()
            parent=el.find_element(By.XPATH,'..')
            blob+=' '+(parent.text or '').lower()
        except Exception: pass
        if 'html' not in blob: continue
        try:
            if not el.is_selected(): d.execute_script('arguments[0].click()',el)
            return bool(el.is_selected())
        except Exception: continue
    # WordPress/KBoard 클래식 에디터는 HTML 체크박스 대신
    # '비주얼 / 코드' 탭을 제공한다. 코드 탭을 누르면 실제 제출용
    # textarea(kboard_content)가 표시되므로 그곳에 HTML 조각을 넣는다.
    try:
        kboard_textareas=d.find_elements(By.CSS_SELECTOR,"textarea#kboard_content,textarea[name='kboard_content']")
        if kboard_textareas:
            for btn in d.find_elements(By.CSS_SELECTOR,"button,a,input[type='button']"):
                if not _sel_vis(btn): continue
                label=((btn.text or '')+' '+_sel_attr(btn,'value')+' '+_sel_attr(btn,'id')+' '+_sel_attr(btn,'class')).strip().lower()
                if re.search(r'(^|\s)(코드|code|html)(\s|$)',label):
                    d.execute_script('arguments[0].click()',btn); time.sleep(0.2)
                    return True
    except Exception: pass
    return False

def editor_content_for_page(d,content_html):
    html_mode=enable_html_mode(d)
    return (content_html if html_mode else html_to_plain(content_html)),html_mode

def fill_required_post_fields(d,site):
    """빨간 별표/required 추가 필드를 의미에 맞는 설정값으로 채운다."""
    from selenium.webdriver.common.by import By
    cfg=load_config(); filled=[]; missing=[]
    brand=(cfg.get('brand') or '게시자').strip()
    guest_pw=(cfg.get('guest_post_password') or '').strip()
    video_url=(cfg.get('video_url') or '').strip()
    landing=(cfg.get('landing_url') or '').strip()
    try: page_text=(d.find_element(By.TAG_NAME,'body').text or '').lower()[:6000]
    except Exception: page_text=''
    # KBoard 비회원 글쓰기는 별표 필수항목이어도 required 속성이 없는
    # 경우가 많다. 알려진 작성자/비밀번호 필드는 선제적으로 채운다.
    for sel,value,label in [
        ("#kboard-input-member-display,input[name='member_display']",brand,'member_display'),
        ("#kboard-input-password,input[name='password']",guest_pw,'password')]:
        try:
            elems=[x for x in d.find_elements(By.CSS_SELECTOR,sel) if _sel_vis(x)]
            if not elems: continue
            el=elems[0]
            if (el.get_attribute('value') or '').strip(): continue
            if value:
                el.clear(); el.send_keys(value); filled.append(label)
            else: missing.append(label)
        except Exception: missing.append(label)
    skip={'wr_subject','subject','title','wr_content','content','captcha_key','captcha'}
    for el in d.find_elements(By.CSS_SELECTOR,"input[required],textarea[required],select[required],.required"):
        if not _sel_vis(el): continue
        typ=(_sel_attr(el,'type') or el.tag_name).lower(); name=_sel_attr(el,'name') or _sel_attr(el,'id')
        if not name or name in skip or typ in ('hidden','submit','button','checkbox','radio','file'): continue
        try:
            if (el.get_attribute('value') or '').strip(): continue
        except Exception: pass
        try: parent=(el.find_element(By.XPATH,'..').text or '').lower()
        except Exception: parent=''
        blob=(name+' '+_sel_attr(el,'id')+' '+_sel_attr(el,'placeholder')+' '+parent).lower()
        tag=el.tag_name.lower()
        # select 필수항목: 값 있는 첫 옵션을 고른다
        if tag=='select':
            try:
                from selenium.webdriver.support.ui import Select
                sel_obj=Select(el); chosen=False
                for opt in sel_obj.options:
                    ov=(opt.get_attribute('value') or '').strip()
                    if ov and ov not in ('0','-1'):
                        sel_obj.select_by_value(ov); filled.append(name); chosen=True; break
                if not chosen and len(sel_obj.options)>1:
                    sel_obj.select_by_index(1); filled.append(name); chosen=True
                if not chosen: missing.append(name)
            except Exception: missing.append(name)
            continue
        value=''
        if re.search(r'(wr_name|이름)',blob): value=brand
        elif typ=='password' or re.search(r'(password|passwd|비밀번호)',blob): value=guest_pw
        elif re.search(r'(email|e-mail|이메일)',blob): value=(cfg.get('post_email') or '')
        elif re.search(r'(tel|phone|연락처|전화|휴대)',blob): value=(cfg.get('phone') or brand)
        elif re.search(r'(youtube|youtu\.be|vimeo|동영상|영상)',blob) or (name in ('wr_link1','link1') and re.search(r'(youtube|유투브|유튜브|vimeo|비메오|동영상)',page_text)): value=video_url or landing or site.get('site_url','')
        elif re.search(r'(link|url|homepage|홈페이지|링크)',blob): value=landing or site.get('site_url','')
        else:
            # 의미를 특정 못한 필수 텍스트/텍스트영역(그누보드 wr_2/wr_5 등 커스텀 확장필드)은
            # 발행을 막지 말고 안전한 일반값으로 채운다(브랜드명). 이메일 형태면 이메일값.
            value=(cfg.get('post_email') or f'{brand}@gmail.com') if 'mail' in blob else brand
        if value:
            if _robust_fill(d,el,value): filled.append(name)
            else: missing.append(name)
        else: missing.append(name)
    return filled,list(dict.fromkeys(missing))

def reset_driver():
    """현재 워커 스레드의 크롬 드라이버를 종료 → 다음 get_driver() 에서 새로 띄움(자동 재시작)."""
    tid=threading.current_thread().name
    with _drv_lock:
        d=_drivers.pop(tid,None)
    if d:
        try: d.quit()
        except: pass

# ==================== Selenium 그누보드 글쓰기 ====================
def _robust_fill(d, el, value):
    """입력 요소에 값을 넣는다. send_keys가 실패(hidden/readonly 등 invalid element state)하면
       JS로 .value를 설정하고 input·change 이벤트를 발생시켜 에디터 연동 스크립트가 반영하게 한다.
       hidden 필드도 확실히 채워진다(많은 그누보드 커스텀 스킨 대응)."""
    try:
        if el.is_displayed() and el.is_enabled():
            el.clear(); el.send_keys(value); return True
    except Exception:
        pass
    try:
        d.execute_script(
            "var e=arguments[0],v=arguments[1];e.value=v;"
            "e.dispatchEvent(new Event('input',{bubbles:true}));"
            "e.dispatchEvent(new Event('change',{bubbles:true}));", el, value)
        return True
    except Exception:
        return False

def _verify_post_by_title(d, bbs, bo, title):
    """제출 후 확인 불가일 때 최후 검증: 게시판 목록을 다시 읽어 방금 올린 제목이
       실제로 등록됐는지 확인한다(느린 서버의 리다이렉트 지연으로 인한 오탐 방지).
       성공하면 해당 글의 뷰 URL을, 실패하면 None을 반환한다. 우회가 아니라 '읽기' 검증."""
    from selenium.webdriver.common.by import By
    def norm(s):
        # 제목 변형(010→OIO, 공백·구분기호 차이)에 견디도록 한글·영숫자만 남겨 비교
        return re.sub(r'[^0-9A-Za-z가-힣]', '', str(s or '')).lower()
    key=norm(title)
    if len(key) < 8:
        return None
    # 목록은 긴 제목을 잘라 보여주므로(행 텍스트가 제목의 접두어) 행⊆제목으로 매칭한다.
    # 또한 브랜드/전화 등 '이 글에만 있는' 특징 조각으로도 본문 전체를 대조한다.
    frag=key[-12:] if len(key)>=12 else key   # 제목 뒤쪽(브랜드/전화)이 가장 변별적
    list_url=f'{bbs}/board.php?bo_table={bo}'
    for _ in range(3):
        try:
            d.set_page_load_timeout(25)
            d.get(list_url); time.sleep(2)
        except Exception:
            pass
        # 1) wr_id 링크를 훑어 '행 텍스트가 제목의 부분(접두어)'인 것을 찾는다
        try:
            anchors=d.find_elements(By.CSS_SELECTOR, "a[href*='wr_id=']")
        except Exception:
            anchors=[]
        for a in anchors:
            try:
                txt=norm(a.text)
                if len(txt) >= 8 and (txt in key or key in txt):
                    href=a.get_attribute('href') or ''
                    if 'wr_id=' in href:
                        return href
            except Exception:
                continue
        # 2) 목록 전체 텍스트에서 변별적 조각으로 확인(스킨이 제목을 잘라 a.text가 짧을 때)
        try:
            page=norm(d.find_element(By.TAG_NAME,'body').text)
            if frag and frag in page:
                # 가능하면 가장 최근(가장 큰) wr_id 를 붙여 뷰 URL을 만든다
                try:
                    ids=[int(x) for x in re.findall(r'wr_id=(\d+)', d.page_source or '')]
                    if ids:
                        return f'{list_url}&wr_id={max(ids)}'
                except Exception:
                    pass
                return list_url
        except Exception:
            pass
        time.sleep(1.5)
    return None

def gnuboard_post(site, title, content_html):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    url=site.get('site_url','').rstrip('/')
    # site_url 이 board.php?... 형태일 수 있으므로 도메인 기준 bbs 경로 계산
    m=re.match(r'(https?://[^/]+)',url)
    base=m.group(1) if m else url
    bbs=base+'/bbs'
    bo=site.get('bo_table','m8_qna')
    mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
    d=get_driver()

    # 로그인
    if mid:
        d.get(f'{bbs}/login.php'); time.sleep(2)
        d.find_element(By.CSS_SELECTOR,"input[name='mb_id']").send_keys(mid)
        d.find_element(By.CSS_SELECTOR,"input[name='mb_password']").send_keys(mpw)
        # 제출 버튼은 로그인 폼(#login_fs/flogin/action*=login_check) 안의 것만. write.php와 동일하게
        # 헤더 검색폼(fsearchbox)의 검색 버튼이 문서상 먼저라 그냥 submit 셀렉터를 쓰면 잘못 잡힌다.
        login_btn=None
        for sel in ("form[name='flogin'] input[type='submit']","form[name='flogin'] button[type='submit']",
                    "form[action*='login_check'] input[type='submit']","form[action*='login_check'] button[type='submit']",
                    "#login_fs .btn_submit","form[name='flogin'] .btn_submit"):
            try:
                for el in d.find_elements(By.CSS_SELECTOR,sel):
                    if el.is_displayed(): login_btn=el; break
            except Exception: pass
            if login_btn: break
        if login_btn:
            try: login_btn.click()
            except Exception:
                try: d.execute_script("var f=document.forms['flogin']||document.querySelector(\"form[action*='login_check']\");if(f){if(f.requestSubmit)f.requestSubmit();else f.submit();}")
                except Exception: pass
        else:
            # 버튼을 못 찾으면 로그인 폼을 직접 제출(onsubmit 경유)
            try: d.execute_script("var f=document.forms['flogin']||document.querySelector(\"form[action*='login_check']\");if(f){if(f.requestSubmit)f.requestSubmit();else f.submit();}")
            except Exception: pass
        time.sleep(2); dismiss_alerts(d)

    # 글쓰기 페이지
    d.get(f'{bbs}/write.php?bo_table={bo}'); time.sleep(2)

    # 글쓰기가 로그인으로 튕기는 게시판: wr_subject를 찾다가 드라이버가 꼬이기 전에
    # 여기서 깔끔히 중단하고 '로그인 필요'로 반환한다(파이프라인이 자동가입으로 재시도).
    try:
        cur=(d.current_url or '').lower()
    except Exception:
        cur=''
    if ('login.php' in cur or 'login_check' in cur) or (not mid and _page_login_state(d)):
        return False,'로그인이 필요한 게시판입니다 — 비회원 글쓰기 불가'

    # 보안 차단은 즉시 중단한다. CAPTCHA는 내용을 채운 뒤 사람이 입력한다.
    if _page_is_blocked(d):
        return False,'보안 차단 페이지(403 등) — 즉시 중단'
    _cap=detect_captcha(d)

    # 제목 — 일부 스킨은 wr_subject가 hidden 입력(JS 에디터 연동)이라 send_keys가
    # 'invalid element state'로 실패한다. 실패 시 JS로 value를 넣고 input/change 이벤트 발생.
    wait=WebDriverWait(d,8)
    el=wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,"input[name='wr_subject'],input#wr_subject")))
    _robust_fill(d,el,title)
    editor_content,html_mode=editor_content_for_page(d,content_html)

    # 본문 (smarteditor2/iframe/textarea 대응)
    try:
        d.execute_script(f"if(typeof oEditors!=='undefined')oEditors.getById['wr_content'].exec('SET_IR',[arguments[0]])", editor_content)
    except: pass
    try:
        iframe=d.find_element(By.CSS_SELECTOR,"iframe.se2_input_wysiwyg,iframe[id*='editor']")
        d.switch_to.frame(iframe)
        d.execute_script("document.body.innerHTML=arguments[0]",editor_content)
        d.switch_to.default_content()
    except:
        d.switch_to.default_content()
        try:
            ed=d.find_element(By.CSS_SELECTOR,"div[contenteditable='true']")
            d.execute_script("arguments[0].innerHTML=arguments[1]",ed,editor_content)
        except:
            ta=d.find_element(By.CSS_SELECTOR,"textarea[name='wr_content']")
            _robust_fill(d,ta,editor_content)   # hidden textarea면 JS value로 폴백

    # 본문 HTML에 이미지가 이미 있으면 중복 파일 첨부하지 않는다.
    if '<img' not in (content_html or '').lower(): attach_saved_images(d,1)
    _,missing=fill_required_post_fields(d,site)
    if missing: return False,'필수항목 설정 필요: '+', '.join(missing[:6])

    captcha_tid=''
    if _cap:
        cfg=load_config()
        # 2captcha 자동 해결 시도
        success,msg,answer,info=solve_captcha_with_2captcha(d,site,_cap,cfg)
        if success:
            from selenium.webdriver.common.by import By
            for sel in ["input[name='captcha_key']","#captcha_key","input[name='wr_key']","input[name*='captcha']","input[id*='captcha']"]:
                try:
                    inp=d.find_element(By.CSS_SELECTOR,sel)
                    if inp and inp.is_displayed():
                        inp.clear(); inp.send_keys(answer); add_log(f'[2captcha] {msg}'); time.sleep(1); break
                except: pass
        else:
            # 2captcha 자동 해결 실패: 무인 자동화라 수동 대기 없이 이 게시판은 건너뛴다.
            add_log(f'[2captcha] 자동 해결 실패: {msg} → 이 게시판 건너뜀(무인)')
            return False,f'캡차 자동 해결 실패({_cap}): {msg}'

    # 등록: native click으로 정상 제출(referer/token 자연스러움) + 짧은 page_load_timeout으로
    #       렌더러 25초 블록 회피 + 상태 접근은 _safe 폴링(제출 후 current_url이 25초 멈추던 문제 해결)
    def _safe(fn,dv=None):
        try: return fn()
        except Exception: return dv
    _safe(lambda: d.execute_script("if(typeof oEditors!=='undefined')try{oEditors.getById['wr_content'].exec('UPDATE_CONTENTS_FIELD',[])}catch(e){}"))
    # 제출 버튼은 반드시 '글쓰기 폼' 안의 것을 골라야 한다. CSS 셀렉터 그룹은 문서 순서로
    # 첫 매치를 주므로 "input[type=submit]"만 쓰면 헤더 검색폼(fsearchbox)의 검색 버튼이
    # 먼저 잡혀 not-interactable → 제출 실패한다(그누보드 기본 스킨의 대표적 함정).
    def _find_submit_btn():
        # 1) 표준 그누보드5: 글쓰기 폼(#fwrite) 안의 #btn_submit / submit
        for sel in ("#fwrite #btn_submit",
                    "form[name='fwrite'] #btn_submit",
                    "#fwrite input[type='submit']",
                    "#fwrite button[type='submit']",
                    "form[action*='write_update'] input[type='submit']",
                    "form[action*='write_update'] button[type='submit']",
                    "#btn_submit"):
            for el in _safe(lambda: d.find_elements(By.CSS_SELECTOR,sel),[]) or []:
                if _safe(lambda: el.is_displayed(),False):
                    return el
        # 2) 폴백: 검색폼(fsearchbox/search.php)에 속하지 않은 표시된 submit
        for el in _safe(lambda: d.find_elements(By.CSS_SELECTOR,"input[type='submit'],button[type='submit']"),[]) or []:
            if not _safe(lambda: el.is_displayed(),False):
                continue
            in_search=_safe(lambda: d.execute_script(
                "var f=arguments[0].form;return !!(f&&((f.name||'').indexOf('search')>=0||(f.action||'').indexOf('search.php')>=0));",el),False)
            if not in_search:
                return el
        return None
    _safe(lambda: d.set_page_load_timeout(6))
    try:
        btn=_find_submit_btn()
        submitted=False
        if btn:
            submitted=_safe(lambda: (btn.click(),True)[1],False)  # navigation 트리거(6초 상한이라 무한블록 안 됨)
        if not submitted:
            # 클릭이 막히면(not interactable 등) 폼을 직접 제출 — onsubmit(fwrite_submit) 경유로
            # 캡차/금지어 검증까지 정상 수행. requestSubmit이 있으면 그것을(핸들러 실행 보장), 없으면 submit().
            _safe(lambda: d.execute_script("""
                var f=document.getElementById('fwrite')||document.forms['fwrite']
                     ||document.querySelector("form[action*='write_update']");
                if(f){ if(f.requestSubmit){f.requestSubmit();} else { if(typeof fwrite_submit==='function'){if(fwrite_submit(f)===false)return;} f.submit(); } }
            """))
        curl=''; body=''
        deadline=time.time()+14
        while time.time()<deadline:
            _safe(lambda: dismiss_alerts(d))
            curl=_safe(lambda: d.current_url,'') or ''
            # 1) 등록 성공 = 뷰/목록으로 이동
            if 'wr_id=' in curl or 'board.php' in curl:
                finish_captcha_task(captcha_tid,True,curl)
                return True,curl
            # 2) write_update.php에 머물면 에러페이지(캡차불일치·금지단어 등) → 본문 확인
            if 'write_update' in curl:
                body=_safe(lambda: d.find_element(By.TAG_NAME,'body').text[:1500],'') or ''
                if body: break
            time.sleep(0.5)
    finally:
        _safe(lambda: d.set_page_load_timeout(25))
    # 3) URL 로 판정 불가 시 본문 텍스트로 분류
    if not body:
        body=_safe(lambda: d.find_element(By.TAG_NAME,'body').text[:1500],'') or ''
    # 승인제 게시판: 이미 1회 등록되었을 수 있으므로 재시도로 중복 발행되지 않게 성공 처리
    if any(k in body for k in ['승인 대기','승인대기','관리자 확인','등록되었습니다']):
        finish_captcha_task(captcha_tid,True,'등록됨(승인 대기)')
        return True,'등록됨(승인 대기) — 게시판 승인제'
    if any(k in body for k in ['올바른 방법','잘못된 접근','비정상']):
        finish_captcha_task(captcha_tid,False,'제출 거부(referer/token) — 폼 재로드 필요')
        return False,'제출 거부(referer/token 검증 실패)'
    # 캡차 불일치는 명확한 에러문구만으로 판정('자동등록방지'는 write 폼의 캡차 라벨이라 오탐 유발)
    if any(k in body for k in ['입력 글자가 틀','횟수가 넘었','자동등록방지 숫자를 다시','보안문자가 일치']):
        finish_captcha_task(captcha_tid,False,'캡차 불일치')
        return False,'캡차 불일치 — 재시도 필요'
    if any(k in body for k in ['권한이 없','권한 없','로그인이 필요','게시가 금지','차단']):
        return False,'게시 권한 없음/로그인 필요 — 계정·게시판 권한 확인'
    # 4) 최후 검증: 느린 서버가 제출 후 뷰로 리다이렉트하는 데 시간이 걸려 위 URL/본문
    #    판정을 놓쳤을 수 있다. 게시판 목록을 다시 읽어 방금 올린 제목이 실제로
    #    등록됐는지 확인한다(오탐으로 인한 재시도→중복발행 방지).
    landed=_safe(lambda: _verify_post_by_title(d, bbs, bo, title))
    if landed:
        finish_captcha_task(captcha_tid,True,landed)
        return True,landed
    msg='등록 확인 불가 — 게시판 규칙/에디터 셀렉터/승인대기 확인'
    finish_captcha_task(captcha_tid,False,msg)
    return False,msg

# ==================== Selenium Cafe24 글쓰기 ====================
def _fill_first(d, selectors, value):
    """여러 셀렉터 후보 중 처음 찾은 입력란에 값 입력."""
    from selenium.webdriver.common.by import By
    for sel in selectors:
        try:
            el=d.find_element(By.CSS_SELECTOR,sel); el.clear(); el.send_keys(value); return True
        except Exception: continue
    return False

def _click_first(d, selectors):
    from selenium.webdriver.common.by import By
    for sel in selectors:
        try: d.find_element(By.CSS_SELECTOR,sel).click(); return True
        except Exception: continue
    return False

def cafe24_post(site, title, content_html):
    """Cafe24 쇼핑몰 게시판 글쓰기. (게시판 구조가 사이트마다 달라 셀렉터 다중 폴백)"""
    from selenium.webdriver.common.by import By
    url=site.get('site_url','').rstrip('/')
    m=re.match(r'(https?://[^/]+)',url); base=m.group(1) if m else url
    bo=str(site.get('bo_table','') or '').strip()
    mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
    d=get_driver()

    # 로그인 (Cafe24 회원 로그인 — 필드명 다양성 대응)
    if mid:
        d.get(base+'/member/login.html'); time.sleep(2)
        _fill_first(d,["input[name='member_id']","input[name='login_id']","input[name='id']",
                       "#member_id","#loginId","input#id"],mid)
        _fill_first(d,["input[name='member_passwd']","input[name='passwd']","input[name='password']",
                       "input[name='login_password']","#passwd","#loginPasswd","input[type='password']"],mpw)
        _click_first(d,["a.btnSubmit","#btnLogin","a.btnLogin",".btnEm","button[type='submit']",
                        "input[type='submit']","a[onclick*='login']"])
        time.sleep(2); dismiss_alerts(d)

    # 글쓰기 페이지 후보 (bo_table 이 숫자면 board_no, 문자면 board 경로)
    write_urls=[]
    if bo.isdigit():
        write_urls=[base+f'/board/write.html?board_no={bo}', base+f'/board/{bo}/write.html']
    elif bo:
        write_urls=[base+f'/board/{bo}/write.html', base+f'/board/write.html?board_no=1', base+f'/board/write.html?board_no={bo}']
    else:
        write_urls=[base+'/board/write.html?board_no=1', base+'/board/free/write.html']
    opened=False
    for wu in write_urls:
        try:
            d.get(wu); time.sleep(2); dismiss_alerts(d)
            if d.find_elements(By.CSS_SELECTOR,"input[name='subject'],#subject,input[name='title']"):
                opened=True; break
        except Exception: continue
    if not opened:
        return False,'Cafe24 글쓰기 페이지 못찾음 — 게시판번호(board_no) 확인'

    # 보안 차단 / 캡차 감지 → 우회하지 않고 즉시 중단
    if _page_is_blocked(d): return False,'보안 차단 페이지(403 등) — 즉시 중단'
    _cap=detect_captcha(d)
    if _cap:
        cfg=load_config()
        # 2captcha 자동 해결 시도
        success,msg,answer,info=solve_captcha_with_2captcha(d,site,_cap,cfg)
        if success:
            from selenium.webdriver.common.by import By
            for sel in ["input[name='captcha_key']","#captcha_key","input[name='wr_key']","input[name*='captcha']","input[id*='captcha']"]:
                try:
                    inp=d.find_element(By.CSS_SELECTOR,sel)
                    if inp and inp.is_displayed():
                        inp.clear(); inp.send_keys(answer); add_log(f'[2captcha] {msg}'); time.sleep(1); break
                except: pass
        else:
            # 2captcha 자동 해결 실패 시 사용자 에러 반환
            add_log(f'[2captcha] 자동 해결 실패: {msg} → Cafe24 자동발행 불가')
            return False,f'캡차 감지({_cap}) — 2captcha 자동 해결 실패: {msg}'

    # 제목
    _fill_first(d,["input[name='subject']","#subject","input[name='title']"],title)
    editor_content,html_mode=editor_content_for_page(d,content_html)

    # 본문 (Cafe24 SmartEditor iframe / CKEditor / textarea)
    filled=False
    try:
        iframe=d.find_element(By.CSS_SELECTOR,"iframe[id*='content'],iframe.cke_wysiwyg_frame,iframe[title*='Rich'],iframe[title*='편집']")
        d.switch_to.frame(iframe)
        d.execute_script("document.body.innerHTML=arguments[0]",editor_content)
        d.switch_to.default_content(); filled=True
    except Exception:
        d.switch_to.default_content()
    if not filled:
        from selenium.webdriver.common.by import By as _By
        for sel in ["textarea[name='content']","textarea#content","textarea[name='contents']",
                    "div[contenteditable='true']","textarea[name='board_content']"]:
            try:
                el=d.find_element(_By.CSS_SELECTOR,sel)
                if sel.startswith('div'):
                    d.execute_script("arguments[0].innerHTML=arguments[1]",el,editor_content)
                else:
                    el.clear(); el.send_keys(editor_content)
                filled=True; break
            except Exception: continue
    if not filled:
        return False,'Cafe24 본문 입력란 못찾음 — 에디터 셀렉터 확인'

    if '<img' not in (content_html or '').lower(): attach_saved_images(d,1)
    _,missing=fill_required_post_fields(d,site)
    if missing: return False,'필수항목 설정 필요: '+', '.join(missing[:6])

    # 등록
    if not _click_first(d,["a.btnSubmit","#btnSubmit","button.btnSubmit","a.btnEm.btnStrong",
                           "input[type='submit']","button[type='submit']","a[onclick*='submit']"]):
        try: d.find_element(By.CSS_SELECTOR,'form').submit()
        except Exception: pass
    time.sleep(3); dismiss_alerts(d)
    curl=d.current_url
    if any(k in curl for k in ['read.html','list.html','article','board_no','view.html']) and 'write.html' not in curl:
        return True,(curl or '등록 완료')
    try: body=d.find_element(By.TAG_NAME,'body').text[:1500]
    except Exception: body=''
    if any(k in body for k in ['등록되었습니다','등록 완료','승인 대기','승인대기','작성되었습니다']):
        return True,'등록됨'
    if any(k in body for k in ['로그인','권한이 없','권한 없','금지','차단','스팸']):
        return False,'Cafe24 게시 권한 없음/로그인 필요 — 계정·게시판 권한 확인'
    return False,'Cafe24 등록 확인 불가 — 게시판 설정/에디터 셀렉터 확인'

# ==================== 플랫폼 자동 감지 + 발행 디스패처 ====================
_plat_cache={}
def detect_platform(url, use_cache=True):
    """사이트 URL 로 그누보드/Cafe24/KBoard 자동 판별."""
    import requests as _rq
    m=re.match(r'(https?://[^/]+)',(url or '').rstrip('/')); base=m.group(1) if m else (url or '')
    if not base: return 'gnuboard'
    if use_cache and base in _plat_cache: return _plat_cache[base]
    UA={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36'}
    def probe(path,needles):
        try:
            r=_rq.get(base+path,timeout=10,verify=False,headers=UA,allow_redirects=True)
            if r.status_code>=400: return False
            t=(r.text or '').lower()
            return any(n in t for n in needles)
        except Exception: return False
    # 등록 URL 자체에서 WordPress KBoard 신호를 먼저 확인한다.
    try:
        rr=_rq.get(url,timeout=10,verify=False,headers=UA,allow_redirects=True)
        page=(rr.text or '').lower()
    except Exception: page=''
    if any(x in page for x in ['kboard-', 'powered by kboard', 'mod=editor', 'kboard_content']):
        res='kboard'
    # 그누보드 신호 우선 확인(가장 흔함)
    elif probe('/bbs/login.php',['mb_password','mb_id']) or probe('/bbs/',['bo_table','wr_id','gnuboard']):
        res='gnuboard'
    # Cafe24 신호
    elif probe('/member/login.html',['member_id','passwd','member_passwd']) or \
         probe('/',['cafe24','xans-','ec-base','/board/write.html','smartdesign']):
        res='cafe24'
    else:
        res='gnuboard'
    _plat_cache[base]=res
    return res

def resolve_platform(site):
    """사이트의 platform 이 auto/미지정이면 감지해서 확정값 반환."""
    plat=(site.get('platform') or '').strip().lower()
    if plat in ('gnuboard','cafe24','kboard'): return plat
    return detect_platform(site.get('site_url',''))

# ==================== 감독형 자가학습 발행 (셀렉터 자동 탐색) ====================
SUBJECT_HINTS=['subject','title','제목','wr_subject','head','tit','bo_subject']
CONTENT_HINTS=['content','contents','내용','body','wr_content','memo','board_content','desc']
SUBMIT_HINTS=['등록','작성','확인','저장','올리','완료','submit','write','save','send','regist','apply','ok','confirm']

def _sel_vis(el):
    try: return el.is_displayed()
    except Exception: return False
def _sel_attr(el,name):
    try: return (el.get_attribute(name) or '')
    except Exception: return ''
def _css_for(el):
    """요소를 다시 찾을 수 있는 안정적 CSS 셀렉터 생성(id>name>type/class>tag)."""
    i=_sel_attr(el,'id'); n=_sel_attr(el,'name')
    if i: return f"[id='{i}']"
    if n: return f"[name='{n}']"
    try: tag=el.tag_name
    except Exception: return '*'
    typ=_sel_attr(el,'type')
    if tag=='input' and typ: return f"input[type='{typ}']"
    cls=[c for c in (_sel_attr(el,'class') or '').split() if c][:2]
    if cls: return tag+'.'+'.'.join(cls)
    return tag

def _is_specific(sel):
    """id/name/type/class 로 특정되는 셀렉터인지(단순 태그명은 위험)."""
    return bool(sel) and ('[' in sel or '.' in sel)

def discover_subject(d):
    from selenium.webdriver.common.by import By
    inputs=d.find_elements(By.CSS_SELECTOR,"input[type='text'],input[type='search'],input:not([type])")
    first=None
    for el in inputs:
        if not _sel_vis(el): continue
        blob=(_sel_attr(el,'name')+' '+_sel_attr(el,'id')+' '+_sel_attr(el,'placeholder')).lower()
        if any(h in blob for h in SUBJECT_HINTS): return _css_for(el)
        if first is None: first=el
    return _css_for(first) if first is not None else None

def discover_content(d):
    """본문 입력란 탐색 → (mode, selector). mode: iframe|contenteditable|textarea."""
    from selenium.webdriver.common.by import By
    iframes=[f for f in d.find_elements(By.CSS_SELECTOR,'iframe') if _sel_vis(f)]
    for f in iframes:
        blob=(_sel_attr(f,'id')+' '+_sel_attr(f,'class')+' '+_sel_attr(f,'title')).lower()
        if any(h in blob for h in ['editor','wysiwyg','content','se2','cke','편집','rich','smart']):
            return ('iframe',_css_for(f))
    for el in d.find_elements(By.CSS_SELECTOR,"[contenteditable='true']"):
        if _sel_vis(el): return ('contenteditable',_css_for(el))
    tas=[t for t in d.find_elements(By.CSS_SELECTOR,'textarea') if _sel_vis(t)]
    for t in tas:
        blob=(_sel_attr(t,'name')+' '+_sel_attr(t,'id')).lower()
        if any(h in blob for h in CONTENT_HINTS): return ('textarea',_css_for(t))
    if tas: return ('textarea',_css_for(tas[0]))
    if iframes: return ('iframe',_css_for(iframes[0]))
    return (None,None)

def discover_submit(d):
    """등록 버튼 셀렉터 탐색. 특정 불가한 단순 태그면 None(→ form.submit() 사용)."""
    from selenium.webdriver.common.by import By
    cands=d.find_elements(By.CSS_SELECTOR,"input[type='submit'],button,a,input[type='button']")
    generic=None
    for el in cands:   # type=submit 우선
        if _sel_vis(el) and el.tag_name=='input' and _sel_attr(el,'type')=='submit':
            sel=_css_for(el)
            if _is_specific(sel): return sel
            generic=generic or sel
    for el in cands:   # 텍스트/속성 힌트
        if not _sel_vis(el): continue
        blob=((el.text or '')+' '+_sel_attr(el,'value')+' '+_sel_attr(el,'onclick')+' '+_sel_attr(el,'id')+' '+_sel_attr(el,'class')).lower()
        if any(h in blob for h in SUBMIT_HINTS):
            sel=_css_for(el)
            if _is_specific(sel): return sel
            generic=generic or sel
    return None   # 특정 셀렉터 없음 → form.submit() 폴백

def discover_login(d, base):
    """로그인 페이지 후보를 돌며 id/pw/버튼 셀렉터 탐색."""
    from selenium.webdriver.common.by import By
    for path in ['/bbs/login.php','/member/login.html','/login','/member/login','/index.php?mode=login']:
        try: d.get(base+path); time.sleep(1.5)
        except Exception: continue
        pws=[p for p in d.find_elements(By.CSS_SELECTOR,"input[type='password']") if _sel_vis(p)]
        if not pws: continue
        texts=[t for t in d.find_elements(By.CSS_SELECTOR,"input[type='text'],input[type='email'],input:not([type])") if _sel_vis(t)]
        if not texts: continue
        return {'login_url':base+path,'id_sel':_css_for(texts[0]),'pw_sel':_css_for(pws[0]),'login_btn':discover_submit(d)}
    return None

def _confirm_posted(d):
    from selenium.webdriver.common.by import By
    curl=d.current_url or ''
    head=curl.split('?')[0]
    if any(k in curl for k in ['wr_id=','board.php','read.html','list.html','view.html','article','board_no','mod=document','uid=']) and 'write' not in head and 'mod=editor' not in curl:
        return True,curl
    try: body=d.find_element(By.TAG_NAME,'body').text[:1500]
    except Exception: body=''
    if any(k in body for k in ['등록되었습니다','작성되었습니다','등록 완료','승인 대기','승인대기','완료되었']):
        return True,'등록됨'
    if any(k in body for k in ['권한이 없','권한 없','로그인이 필요','로그인 필요','게시가 금지','차단','스팸']):
        return False,'게시 권한 없음/로그인 필요'
    return False,'등록 확인 불가 — 게시판 설정 확인'

def _fill_recipe_fields(d, rec, title, content):
    """현재 write 페이지에 제목/본문 채우고 등록 클릭(네비게이션 없음)."""
    from selenium.webdriver.common.by import By
    d.find_element(By.CSS_SELECTOR,rec['subject_sel']).clear()
    d.find_element(By.CSS_SELECTOR,rec['subject_sel']).send_keys(title)
    editor_content,html_mode=editor_content_for_page(d,content)
    mode=rec.get('content_mode'); csel=rec.get('content_sel')
    # KBoard 코드 탭이 활성화되면 iframe 대신 제출 textarea를 사용한다.
    if html_mode:
        try:
            kta=[x for x in d.find_elements(By.CSS_SELECTOR,"textarea#kboard_content,textarea[name='kboard_content']") if _sel_vis(x)]
            if kta:
                mode='textarea'; csel="#kboard_content" if _sel_attr(kta[0],'id') else "textarea[name='kboard_content']"
        except Exception: pass
    if mode=='iframe':
        fr=d.find_element(By.CSS_SELECTOR,csel); d.switch_to.frame(fr)
        d.execute_script("document.body.innerHTML=arguments[0]",editor_content); d.switch_to.default_content()
    elif mode=='contenteditable':
        el=d.find_element(By.CSS_SELECTOR,csel); d.execute_script("arguments[0].innerHTML=arguments[1]",el,editor_content)
    else:
        el=d.find_element(By.CSS_SELECTOR,csel); el.clear(); el.send_keys(editor_content)
    # 스마트에디터 동기화 시도(있으면)
    try: d.execute_script("if(typeof oEditors!=='undefined')try{oEditors.getById[Object.keys(oEditors.getById)[0]].exec('UPDATE_CONTENTS_FIELD',[])}catch(e){}")
    except Exception: pass
    if '<img' not in (content or '').lower(): attach_saved_images(d,1)
    if rec.get('submit_sel'):
        try: d.find_element(By.CSS_SELECTOR,rec['submit_sel']).click()
        except Exception:
            try: d.find_element(By.CSS_SELECTOR,'form').submit()
            except Exception: pass
    else:
        try: d.find_element(By.CSS_SELECTOR,'form').submit()
        except Exception: pass
    time.sleep(3); dismiss_alerts(d)

def _apply_recipe(d, site, rec, title, content):
    """저장된 레시피로 발행(로그인→글쓰기→채우기→확인)."""
    from selenium.webdriver.common.by import By
    mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
    if mid:
        base=re.match(r'(https?://[^/]+)',rec.get('write_url','') or ''); base=base.group(1) if base else ''
        if rec.get('login_url') and rec.get('id_sel') and rec.get('pw_sel'):
            d.get(rec['login_url']); time.sleep(1.5)
            try: e=d.find_element(By.CSS_SELECTOR,rec['id_sel']); e.clear(); e.send_keys(mid)
            except Exception: pass
            try: e=d.find_element(By.CSS_SELECTOR,rec['pw_sel']); e.clear(); e.send_keys(mpw)
            except Exception: pass
            if rec.get('login_btn'):
                try: d.find_element(By.CSS_SELECTOR,rec['login_btn']).click()
                except Exception: pass
            time.sleep(2); dismiss_alerts(d)
        elif base:
            _platform_login(d,base,site)   # 저장된 상세 셀렉터 없으면 플랫폼 로그인
    d.get(rec['write_url']); time.sleep(2); dismiss_alerts(d)
    _,missing=fill_required_post_fields(d,site)
    if missing: return False,'필수항목 설정 필요: '+', '.join(missing[:6])
    _fill_recipe_fields(d, rec, title, content)
    return _confirm_posted(d)

def _page_login_state(d):
    """현재 페이지가 로그인 화면인지(본문 에디터 없음 + 비번칸/로그인안내) 판별."""
    from selenium.webdriver.common.by import By
    try: has_pw=any(_sel_vis(e) for e in d.find_elements(By.CSS_SELECTOR,"input[type='password']"))
    except Exception: has_pw=False
    has_editor=bool(discover_content(d)[1])   # 글쓰기 페이지엔 본문 에디터가 있음
    try: body=d.find_element(By.TAG_NAME,'body').text[:1000]
    except Exception: body=''
    login_notice=('로그인' in body and ('필요' in body or '회원가입' in body or '아이디' in body))
    return (not has_editor) and (has_pw or login_notice)

def _platform_login(d, base, site):
    """플랫폼별 신뢰 셀렉터로 로그인 시도(그누보드/Cafe24). 로그인 URL 반환(실패 None)."""
    from selenium.webdriver.common.by import By
    mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
    if not mid: return None
    plat=resolve_platform(site)
    tries=[('/bbs/login.php',"input[name='mb_id']","input[name='mb_password']")] if plat=='gnuboard' else \
          [('/member/login.html',"input[name='member_id'],input[name='login_id'],input[name='id']","input[name='member_passwd'],input[name='passwd'],input[name='password']")]
    tries.append(('/bbs/login.php',"input[name='mb_id']","input[name='mb_password']"))
    tries.append(('/member/login.html',"input[name='member_id'],input[name='login_id'],input[name='id']","input[name='member_passwd'],input[name='passwd'],input[name='password']"))
    for path,idsel,pwsel in tries:
        try:
            d.get(base+path); time.sleep(1.5)
            ide=d.find_elements(By.CSS_SELECTOR,idsel); pwe=d.find_elements(By.CSS_SELECTOR,pwsel)
            if not ide or not pwe: continue
            ide[0].clear(); ide[0].send_keys(mid); pwe[0].clear(); pwe[0].send_keys(mpw)
            if not _click_first(d,["input[type='submit']","button[type='submit']",".btn_submit","a.btnSubmit","#btnLogin",".btnEm"]):
                try: pwe[0].submit()
                except Exception: pass
            time.sleep(2); dismiss_alerts(d)
            return base+path
        except Exception: continue
    return None

def discover_write_page(d, base, site):
    bo=str(site.get('bo_table','') or '').strip(); plat=resolve_platform(site)
    cands=[]
    # 관리자가 등록한 게시판 화면에서 실제 글쓰기 링크를 먼저 실측한다.
    # 사이트 전체를 탐색하지 않고, 등록 URL 한 화면의 동일 출처 링크만 사용한다.
    try:
        start=(site.get('site_url') or base).strip()
        d.get(start); time.sleep(2); dismiss_alerts(d)
        from selenium.webdriver.common.by import By
        for a in d.find_elements(By.CSS_SELECTOR,'a[href]'):
            href=urllib.parse.urljoin(d.current_url or start,_sel_attr(a,'href'))
            pu=urllib.parse.urlparse(href); pb=urllib.parse.urlparse(base)
            if (pu.scheme,pu.netloc)!=(pb.scheme,pb.netloc): continue
            blob=((a.text or '')+' '+href+' '+_sel_attr(a,'class')+' '+_sel_attr(a,'id')).lower()
            if any(k in blob for k in ['글쓰기','글 쓰기','write.php','/write.html','mode=write','act=write']):
                cands.append(href)
    except Exception: pass
    if plat=='cafe24':
        if bo.isdigit(): cands += [f'/board/write.html?board_no={bo}',f'/board/{bo}/write.html']
        elif bo: cands += [f'/board/{bo}/write.html',f'/board/write.html?board_no=1']
        else: cands += ['/board/write.html?board_no=1','/board/free/write.html']
    else:
        cands += [f'/bbs/write.php?bo_table={bo or "free"}']
    cands+=[f'/bbs/write.php?bo_table={bo or "free"}',f'/board/write.html?board_no={bo if bo.isdigit() else "1"}']
    seen=set()
    for path in cands:
        target=path if str(path).startswith(('http://','https://')) else base+path
        if target in seen: continue
        seen.add(target)
        try: d.get(target); time.sleep(2); dismiss_alerts(d)
        except Exception: continue
        if discover_subject(d) and discover_content(d)[1]:
            return d.current_url or target
    return None

def _save_site_analysis(site_id, analysis, rec=None):
    """실측 결과는 성공/실패 모두 저장하고, 확실한 폼만 발행 레시피로 승격한다."""
    if not site_id: return
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')!=site_id: continue
            s['analysis']=analysis
            if rec:
                s['learned']=rec
                if rec.get('platform'): s['platform']=rec['platform']
        save_sites(sites)

def analyze_site_logic(site):
    """허용된 한 사이트의 등록 URL/로그인/글쓰기 DOM을 측정한다. 절대 제출하지 않는다."""
    from selenium.webdriver.common.by import By
    now=datetime.now().astimezone().isoformat(timespec='seconds')
    result={'measured_at':now,'mode':'read_only_no_submit','ok':False,'platform':'',
            'start_url':site.get('site_url',''),'final_url':'','write_url':'',
            'blocked':False,'captcha':'','form':{},'steps':[]}
    def step(name,ok,detail=''):
        result['steps'].append({'name':name,'ok':bool(ok),'detail':str(detail)[:300]})
    url=(site.get('site_url') or '').strip(); m=re.match(r'(https?://[^/]+)',url)
    base=m.group(1) if m else url
    if not base:
        step('등록 URL',False,'URL 없음'); return result,None
    d=get_driver(); plat=resolve_platform(site); result['platform']=plat
    step('플랫폼',True,plat)
    try:
        d.get(url or base); time.sleep(2); dismiss_alerts(d)
        result['final_url']=d.current_url or ''
        step('등록 화면',True,result['final_url'])
    except Exception as e:
        step('등록 화면',False,str(e)[:180]); return result,None
    if _page_is_blocked(d):
        result['blocked']=True; step('보안 차단',False,'403/보안 차단 감지 — 우회하지 않음'); return result,None
    cap=detect_captcha(d)
    if cap:
        result['captcha']=str(cap); step('CAPTCHA',False,str(cap)+' — 우회하지 않음'); return result,None
    step('초기 화면 보안',True,'차단/CAPTCHA 없음')
    if site.get('mb_id'):
        lu=_platform_login(d,base,site)
        step('로그인',bool(lu),lu or '저장 계정으로 로그인 페이지/필드 확인 실패')
    else:
        step('로그인',True,'계정 미설정 — 공개/비회원 글쓰기만 측정')
    # 설정된 게시판 ID의 표준 쓰기 URL을 먼저 직접 측정하여 로그인/차단 원인을 보존한다.
    bo=str(site.get('bo_table') or '').strip(); direct=''
    if bo:
        direct=(base+f'/board/write.html?board_no={bo}') if plat=='cafe24' and bo.isdigit() else \
               (base+f'/bbs/write.php?bo_table={urllib.parse.quote(bo)}')
    wu=None
    if direct:
        try:
            d.get(direct); time.sleep(2); dismiss_alerts(d)
            result['final_url']=d.current_url or direct
            if _page_is_blocked(d):
                result['write_url']=direct; result['blocked']=True
                step('설정 글쓰기 URL',False,'접근 차단: '+direct+' — 우회하지 않음'); return result,None
            direct_cap=detect_captcha(d)
            if direct_cap:
                result['write_url']=direct; result['captcha']=str(direct_cap)
                step('설정 글쓰기 URL',False,'CAPTCHA 감지: '+str(direct_cap)+' — 우회하지 않음'); return result,None
            if _page_login_state(d) and not site.get('mb_id'):
                result['write_url']=direct
                step('설정 글쓰기 URL',False,'로그인 필요: '+result['final_url']); return result,None
            if discover_subject(d) and discover_content(d)[1]: wu=result['final_url']
        except Exception as e:
            step('설정 글쓰기 URL',False,str(e)[:180])
    if not wu: wu=discover_write_page(d,base,site)
    result['final_url']=d.current_url or result['final_url']
    if not wu:
        detail='로그인 화면으로 이동됨' if _page_login_state(d) else '실제 링크 및 설정 게시판ID 후보에서 폼을 찾지 못함'
        step('글쓰기 폼',False,detail); return result,None
    result['write_url']=wu; result['final_url']=d.current_url or wu
    step('글쓰기 폼',True,wu)
    if _page_is_blocked(d):
        result['blocked']=True; step('글쓰기 보안',False,'403/보안 차단 감지 — 우회하지 않음'); return result,None
    cap=detect_captcha(d)
    if cap:
        result['captcha']=str(cap); step('CAPTCHA',False,str(cap)+' — 우회하지 않음'); return result,None
    subj=discover_subject(d); cmode,csel=discover_content(d); sub=discover_submit(d)
    form_el=None
    try:
        if subj: form_el=d.find_element(By.CSS_SELECTOR,subj).find_element(By.XPATH,'ancestor::form[1]')
    except Exception: pass
    form={'action':'','method':'','subject_sel':subj or '','content_mode':cmode or '',
          'content_sel':csel or '','submit_sel':sub or '','required_fields':[]}
    if form_el is not None:
        form['action']=urllib.parse.urljoin(result['final_url'],_sel_attr(form_el,'action'))
        form['method']=(_sel_attr(form_el,'method') or 'get').lower()
        try:
            for el in form_el.find_elements(By.CSS_SELECTOR,'input[required],textarea[required],select[required]'):
                name=_sel_attr(el,'name') or _sel_attr(el,'id')
                typ=_sel_attr(el,'type') or el.tag_name
                if name: form['required_fields'].append({'name':name,'type':typ})
        except Exception: pass
    result['form']=form
    step('제목 필드',bool(subj),subj or '없음')
    step('본문 필드',bool(csel),f'{cmode} · {csel}' if csel else '없음')
    step('등록 동작',bool(sub or form_el is not None),sub or ('form '+form['method'] if form_el is not None else '없음'))
    if not subj or not csel or form_el is None:
        step('판정',False,'발행 레시피로 저장할 신뢰도 부족'); return result,None
    rec={'platform':plat,'learned_at':now,'learned_mode':'measured_no_submit',
         'write_url':wu,'subject_sel':subj,'content_mode':cmode,
         'content_sel':csel,'submit_sel':sub,'form_action':form['action'],'form_method':form['method']}
    result['ok']=True; step('판정',True,'DOM 실측 완료 · 제출 없이 레시피 저장 가능')
    return result,rec

def discover_and_post(site, title, content):
    """DOM을 훑어 셀렉터를 스스로 찾아 발행. 성공 시 (ok,msg,레시피) 반환."""
    d=get_driver()
    url=site.get('site_url','').rstrip('/'); m=re.match(r'(https?://[^/]+)',url); base=m.group(1) if m else url
    rec={'platform':resolve_platform(site),'learned_at':datetime.now().strftime('%Y-%m-%d %H:%M')}
    # 로그인: 플랫폼별 신뢰 셀렉터 우선, 실패 시 자동 탐색
    if site.get('mb_id'):
        lu=_platform_login(d,base,site)
        if lu: rec['login_url']=lu
        else:
            login=discover_login(d,base)
            if login:
                rec.update(login)
                from selenium.webdriver.common.by import By
                try: e=d.find_element(By.CSS_SELECTOR,login['id_sel']); e.clear(); e.send_keys(site.get('mb_id',''))
                except Exception: pass
                try: e=d.find_element(By.CSS_SELECTOR,login['pw_sel']); e.clear(); e.send_keys(site.get('mb_pass',''))
                except Exception: pass
                if login.get('login_btn'):
                    try: d.find_element(By.CSS_SELECTOR,login['login_btn']).click()
                    except Exception: pass
                time.sleep(2); dismiss_alerts(d)
    wu=discover_write_page(d,base,site)
    if not wu:
        # 원인 세분화: 로그인 화면으로 튕겼는지 확인
        if _page_login_state(d):
            if not site.get('mb_id'):
                return False,'로그인이 필요한 게시판입니다 — 사이트에 아이디/비밀번호를 입력한 뒤 다시 학습하세요',None
            return False,'로그인 실패로 보입니다 — 아이디/비밀번호가 맞는지 확인하세요(캡차/보안문자 게시판일 수도)',None
        return False,'글쓰기 페이지 못찾음(학습 실패) — 게시판ID(bo_table) 확인',None
    rec['write_url']=wu
    # 보안 차단은 즉시 중단. 캡차는 감지만 해두고 제출 직전에 2captcha로 해결한다.
    if _page_is_blocked(d): return False,'보안 차단 페이지(403 등) — 즉시 중단',None
    _cap=detect_captcha(d)
    subj=discover_subject(d); cmode,csel=discover_content(d); sub=discover_submit(d)
    if not subj or not csel:
        return False,'제목/본문 입력란 못찾음(학습 실패)',None
    rec['subject_sel']=subj; rec['content_mode']=cmode; rec['content_sel']=csel; rec['submit_sel']=sub
    _,missing=fill_required_post_fields(d,site)
    if missing: return False,'필수항목 설정 필요: '+', '.join(missing[:6]),None
    try: _fill_recipe_fields(d, rec, title, content)
    except Exception as e: return False,f'입력 실패(학습): {str(e)[:80]}',None
    # 캡차 감지 시 2captcha로 자동 해결 후 입력 (실패하면 발행 중단)
    if _cap:
        cfg=load_config()
        success,msg,answer,info=solve_captcha_with_2captcha(d,site,_cap,cfg)
        if success:
            from selenium.webdriver.common.by import By
            for sel in ["input[name='captcha_key']","#captcha_key","input[name='wr_key']","input[name*='captcha']","input[id*='captcha']"]:
                try:
                    inp=d.find_element(By.CSS_SELECTOR,sel)
                    if inp and inp.is_displayed():
                        inp.clear(); inp.send_keys(answer); add_log(f'[2captcha] {msg}'); time.sleep(1); break
                except Exception: pass
        else:
            add_log(f'[2captcha] 자동 해결 실패: {msg} → 자동발행 불가')
            return False,f'캡차 감지({_cap}) — 2captcha 자동 해결 실패: {msg}',None
    ok,msg=_confirm_posted(d)
    return ok,msg,(rec if ok else None)

def dryrun_post(site, title, content_html):
    """등록 '직전'까지만 수행 — 실제 글은 올리지 않고 발행 가능 여부를 검증한다.
       로그인 → 글쓰기 페이지 → 캡차확인 → 입력란 탐색 → 값 채우기 까지. 제출 클릭 없음."""
    from selenium.webdriver.common.by import By
    steps=[]
    def st(name,ok,detail=''): steps.append({'name':name,'ok':bool(ok),'detail':str(detail)[:220]})
    url=site.get('site_url','').rstrip('/'); m=re.match(r'(https?://[^/]+)',url)
    base=m.group(1) if m else url
    d=get_driver()
    plat=resolve_platform(site); st('플랫폼 판별',True,plat)

    # 1) 로그인
    mid=site.get('mb_id','')
    if mid:
        lu=_platform_login(d,base,site)
        if lu:
            # 로그인 성공 여부: 로그아웃 링크 존재로 추정
            try: body=d.find_element(By.TAG_NAME,'body').text[:1500]
            except Exception: body=''
            logged=('로그아웃' in body) or ('logout' in (d.page_source or '').lower())
            st('로그인',logged,f'{lu} → '+('세션 확보됨' if logged else '로그인 확인 불가(비번/아이디 확인)'))
        else:
            st('로그인',False,'로그인 페이지를 못 찾음')
    else:
        st('로그인',True,'아이디 미설정 — 비회원 글쓰기로 진행')

    # 2) 글쓰기 페이지
    rec=site.get('learned') or {}
    wu=rec.get('write_url') if rec.get('write_url') else None
    if wu:
        try: d.get(wu); time.sleep(2); dismiss_alerts(d)
        except Exception: wu=None
    if not wu:
        wu=discover_write_page(d,base,site)
    if not wu:
        if _page_login_state(d):
            st('글쓰기 페이지',False,'로그인 화면으로 이동됨 — 계정 필요 또는 로그인 실패')
        else:
            st('글쓰기 페이지',False,'못 찾음 — 게시판ID(bo_table) 확인')
        return False,steps
    st('글쓰기 페이지',True,wu)

    # 3) 차단/캡차
    if _page_is_blocked(d):
        st('보안 차단 확인',False,'403/보안 차단 페이지'); return False,steps
    st('보안 차단 확인',True,'차단 없음')
    cap=detect_captcha(d)
    if cap:
        _cfg=load_config()
        if _cfg.get('twocaptcha_enabled') and (_cfg.get('twocaptcha_api_key') or '').strip():
            st('캡차 확인',True,f'{cap} 감지 — 발행 시 2captcha로 자동 해결 예정')
        else:
            st('캡차 확인',False,f'{cap} 감지 — 자동발행 부적합(2captcha 비활성화)'); return False,steps
    else:
        st('캡차 확인',True,'캡차 없음')

    # 4) 입력란 탐색
    subj=rec.get('subject_sel') or discover_subject(d)
    cmode,csel=(rec.get('content_mode'),rec.get('content_sel')) if rec.get('content_sel') else discover_content(d)
    sub=rec.get('submit_sel') if rec.get('submit_sel') else discover_submit(d)
    st('제목 입력란',bool(subj),subj or '못 찾음')
    st('본문 에디터',bool(csel),f'{cmode} · {csel}' if csel else '못 찾음')
    st('등록 버튼',True,sub or '특정 셀렉터 없음 → form.submit() 사용 예정')
    if not subj or not csel: return False,steps

    # 5) 실제로 값 채워보기 (제출은 하지 않음)
    try:
        e=d.find_element(By.CSS_SELECTOR,subj); e.clear(); e.send_keys(title)
        st('제목 입력 테스트',True,f'{len(title)}자 입력 성공')
    except Exception as ex:
        st('제목 입력 테스트',False,str(ex)[:120]); return False,steps
    try:
        editor_content,html_mode=editor_content_for_page(d,content_html)
        st('본문 형식',True,'HTML 우선' if html_mode else 'HTML 옵션 없음 → 일반 텍스트')
        if html_mode:
            try:
                kta=[x for x in d.find_elements(By.CSS_SELECTOR,"textarea#kboard_content,textarea[name='kboard_content']") if _sel_vis(x)]
                if kta:
                    cmode='textarea'; csel="#kboard_content" if _sel_attr(kta[0],'id') else "textarea[name='kboard_content']"
            except Exception: pass
        if cmode=='iframe':
            fr=d.find_element(By.CSS_SELECTOR,csel); d.switch_to.frame(fr)
            d.execute_script("document.body.innerHTML=arguments[0]",editor_content); d.switch_to.default_content()
        elif cmode=='contenteditable':
            el=d.find_element(By.CSS_SELECTOR,csel); d.execute_script("arguments[0].innerHTML=arguments[1]",el,editor_content)
        else:
            el=d.find_element(By.CSS_SELECTOR,csel); el.clear(); el.send_keys(editor_content[:2000])
        st('본문 입력 테스트',True,f'{len(editor_content)}자 '+('HTML' if html_mode else '일반 텍스트')+' 입력 성공')
    except Exception as ex:
        d.switch_to.default_content()
        st('본문 입력 테스트',False,str(ex)[:120]); return False,steps

    if '<img' in (content_html or '').lower():
        attached,attach_detail=0,'본문 HTML 이미지 1개 사용 · 중복 파일 첨부 생략'
    else:
        attached,attach_detail=attach_saved_images(d,1)
    st('이미지 파일 첨부',True,attach_detail)

    filled_required,missing_required=fill_required_post_fields(d,site)
    st('추가 필수항목',not missing_required,
       ('자동 입력: '+', '.join(filled_required) if filled_required else '추가 입력 없음')+
       ((' · 설정 필요: '+', '.join(missing_required)) if missing_required else ''))
    if missing_required: return False,steps

    # 6) 필수 추가 입력란(비회원 이름/비번 등) 확인
    try:
        reqs=[]
        for el in d.find_elements(By.CSS_SELECTOR,"input[required],input[type='password']"):
            if not _sel_vis(el): continue
            nm=_sel_attr(el,'name') or _sel_attr(el,'id')
            if nm and nm not in (subj or ''): reqs.append(nm)
        st('추가 필수 입력란',True,(', '.join(reqs[:6]) if reqs else '없음')+(' (비회원 글쓰기는 이름/비번 필요할 수 있음)' if reqs else ''))
    except Exception: pass

    st('종합',True,'✅ 등록 직전까지 모두 성공 — 실제 발행 가능 상태 (글은 올리지 않았습니다)')
    return True,steps

def save_learned(site_id, rec):
    """학습된 셀렉터 레시피를 사이트에 영구 저장."""
    if not site_id: return
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')==site_id:
                s['learned']=rec
                if rec.get('platform'): s['platform']=rec['platform']
        save_sites(sites)

def set_site_flag(site_id, **fields):
    """사이트에 플래그/상태 저장(예: has_captcha)."""
    if not site_id: return
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')==site_id: s.update(fields)
        save_sites(sites)

def do_post(site, title, content_html):
    """발행 라우팅(하이브리드+자가학습): 학습레시피→플랫폼기본→자동학습."""
    rec=site.get('learned')
    # 1) 저장된 학습 레시피 우선
    if rec and rec.get('write_url') and rec.get('subject_sel') and rec.get('content_sel'):
        try:
            ok,msg=_apply_recipe(get_driver(),site,rec,title,content_html)
            if ok: return True,msg
            add_log(f'[학습레시피 실패→폴백] {site.get("name","")}')
        except Exception as e:
            add_log(f'[학습레시피 오류→폴백] {str(e)[:60]}')
    # 2) 플랫폼별 기본 발행기
    plat=resolve_platform(site)
    if plat=='kboard':
        try:
            ok,msg,newrec=discover_and_post(site,title,content_html)
            if ok and newrec: save_learned(site.get('id'),newrec)
            return ok,msg
        except Exception as e:
            return False,'KBoard 발행 오류: '+str(e)[:120]
    try:
        ok,msg=(cafe24_post if plat=='cafe24' else gnuboard_post)(site,title,content_html)
    except Exception as e:
        ok,msg=False,str(e)
    if ok: return True,msg
    # 3) 구조적 실패(일시적·로그인·차단 제외)면 자동 학습(디스커버리) 발행 후 레시피 저장
    reason,_,_=classify_fail(msg)
    if reason in ('board','other') or rec is not None:
        try:
            ok2,msg2,newrec=discover_and_post(site,title,content_html)
            if ok2 and newrec:
                save_learned(site.get('id'),newrec)
                add_log(f'[자동학습 성공] {site.get("name","")} — 셀렉터 저장됨')
                return True,'(자가학습) '+str(msg2)
            if ok2: return True,str(msg2)
            return False,f'{msg} · 학습시도:{msg2}'
        except Exception as e:
            return False,f'{msg} · 학습실패:{str(e)[:50]}'
    return False,msg

# ==================== 큐 & 워커 & 이력 ====================
post_queue=queue.Queue()
wk_active=False
wk_paused=False
wk_stats={'success':0,'fail':0,'queued':0,'total':0,'done':0,'skipped':0,'retry':0}
STATS_LOCK=threading.Lock()
POST_LOCK=threading.Lock()
JOB_LOCK=threading.Lock()   # history.json / queue.json 동시성 보호
BULK_LOCK=threading.Lock()
BULK_TASKS={}

# ---- 일시적 실패 자동 재시도(지연 재큐) ----
RETRY_DELAY=300      # 5분 뒤 재시도
RETRY_MAX=2          # 지연 재시도 최대 횟수
_retry_jobs=[]
_retry_lock=threading.Lock()
def schedule_retry(job, delay=None):
    if delay is None: delay=RETRY_DELAY
    with _retry_lock: _retry_jobs.append({'due':time.time()+delay,'job':job})

def retry_loop():
    """지연 재시도 큐를 감시해 시간이 되면 발행 큐로 되돌림."""
    while True:
        try:
            now=time.time(); ready=[]
            with _retry_lock:
                keep=[]
                for it in _retry_jobs:
                    (ready if it['due']<=now else keep).append(it)
                _retry_jobs[:]=keep
            for it in ready:
                post_queue.put(it['job'])
                with STATS_LOCK: wk_stats['queued']=post_queue.qsize()
                if not wk_active: start_workers(load_config().get('workers',2))
        except Exception as e:
            add_log(f'[재시도 루프 오류] {str(e)[:80]}')
        time.sleep(30)

# ---- 발행 이력 원장 (엑셀/결과탭) ----
def history_add(rec):
    with JOB_LOCK:
        h=load_json(HISTORY_FILE,[]); h.append(rec)
        if len(h)>5000: h=h[-5000:]
        save_json(HISTORY_FILE,h)

def history_update(hid,**fields):
    with JOB_LOCK:
        h=load_json(HISTORY_FILE,[])
        for rec in h:
            if rec.get('id')==hid:
                rec.update(fields); rec['updated']=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                break
        save_json(HISTORY_FILE,h)

# ---- 미완료 작업 영속화 (재시작 복구) ----
def _persist_add(job):
    with JOB_LOCK:
        q=load_json(QUEUE_FILE,[]); q.append(job); save_json(QUEUE_FILE,q)

def _persist_remove(job_id):
    with JOB_LOCK:
        q=[j for j in load_json(QUEUE_FILE,[]) if j.get('job_id')!=job_id]
        save_json(QUEUE_FILE,q)

def recover_queue():
    """재시작 시 유효한 미완료 작업만 복구하고 차단·유령 큐는 스킵 처리."""
    q=load_json(QUEUE_FILE,[])
    n=0; keep=[]; queued_hist=set()
    for job in q:
        hid=job.get('hist_id'); site=job.get('site') or {}
        current=next((s for s in load_sites() if s.get('id')==site.get('id')),None)
        if job.get('site') and job.get('content') and current and is_publishable(current):
            job['site']=current; post_queue.put(job); keep.append(job); n+=1
            if hid: queued_hist.add(hid)
        else:
            why='재시작 시 발행조건 미충족'
            if hid: history_update(hid,status='skipped',message=why)
    save_json(QUEUE_FILE,keep)
    # queue.json에 실제 작업이 없는 queued 이력은 더 이상 진행되지 않는 유령 표시다.
    h=load_json(HISTORY_FILE,[]); changed=False
    for rec in h:
        if rec.get('status')=='queued' and rec.get('id') not in queued_hist:
            rec.update({'status':'skipped','message':'재시작 후 실행 큐 없음'}); changed=True
    if changed: save_json(HISTORY_FILE,h)
    if n:
        with STATS_LOCK:
            wk_stats['total']+=n; wk_stats['queued']=post_queue.qsize()
        add_log(f'[복구] 미완료 작업 {n}건 복구됨 — 워커 시작 시 이어서 발행')
    return n

def site_daily_limit(site,cfg):
    try: return max(0,int(site.get('daily_limit',cfg.get('daily_limit',3)) or 0))
    except Exception: return 3

def site_min_interval(site):
    try: return max(0,int(site.get('min_interval_minutes',60) or 0))
    except Exception: return 60

def _fresh_site(site):
    return next((s for s in load_sites() if s.get('id')==site.get('id')),site)

def under_daily_limit(site,cfg):
    """사이트당 1일 발행 한도 확인 (0 = 무제한). 도배 방지."""
    site=_fresh_site(site)
    limit=site_daily_limit(site,cfg)
    if limit<=0: return True
    today=datetime.now().strftime('%Y-%m-%d')
    for s in load_sites():
        if s.get('id')==site.get('id'):
            if s.get('posted_date')!=today: return True
            return s.get('posted_today',0)<limit
    return True

def under_min_interval(site):
    """마지막 성공 발행 후 사이트별 최소 간격 준수 여부와 남은 초 반환."""
    site=_fresh_site(site); mins=site_min_interval(site)
    if mins<=0 or not site.get('last_post_at'): return True,0
    try:
        last=datetime.strptime(site['last_post_at'],'%Y-%m-%d %H:%M:%S')
        remain=int(mins*60-(datetime.now()-last).total_seconds())
        return remain<=0,max(0,remain)
    except Exception: return True,0

FAIL_STREAK_DROP=3   # 연속 실패 이 횟수 이상이면 자동 탈락(도배·헛발행 방지)

def finalize_post(site,ok,fail_reason=''):
    """상태 갱신 + 성공 시 오늘 발행 카운트 증가 + 연속 실패 카운트(fail_streak) 추적.
       (한 번의 락으로 처리)"""
    today=datetime.now().strftime('%Y-%m-%d')
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')==site.get('id'):
                s['status']='done' if ok else 'failed'
                if ok:
                    if s.get('posted_date')!=today: s['posted_date']=today; s['posted_today']=0
                    s['posted_today']=s.get('posted_today',0)+1
                    s['last_post_at']=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    s['fail_streak']=0; s.pop('last_fail_reason',None)
                else:
                    s['fail_streak']=int(s.get('fail_streak',0) or 0)+1
                    if fail_reason: s['last_fail_reason']=str(fail_reason)[:120]
                break
        save_sites(sites)
    if not ok:
        reconcile_sites()   # 실패 직후 자동 탈락 조건 재평가(목록 최신화)

def _site_permanent_block(s):
    """이 사이트가 '영구히 발행 불가'인지 판정. 사유 문자열 반환(아니면 '')."""
    if s.get('signup_email_verification'): return '이메일 인증 필요(자동가입 불가)'
    tbr=str(s.get('technical_block_reason') or '')
    if any(k in tbr for k in ('폼을 찾지 못','글쓰기 폼','게시판ID')): return '글쓰기 폼 없음: '+tbr[:60]
    lfr=str(s.get('last_fail_reason') or '')
    if any(k in lfr for k in ('보안 차단','403','게시가 금지','권한이 없','권한 없','차단')): return '차단/권한없음: '+lfr[:50]
    if any(k in lfr for k in ('제출 거부','referer','token')): return '제출 거부(referer/token)'
    if int(s.get('fail_streak',0) or 0)>=FAIL_STREAK_DROP:
        return f'연속 실패 {s.get("fail_streak")}회'
    return ''

def reconcile_sites():
    """사이트 목록 상시 최신화: 발행이 막힌 사이트를 자동 탈락(permission 해제)시킨다.
       - 실제게시 검증 완료(verified_post_url 있음) 사이트는 보호(오탈락 방지).
       - 영구 불가 조건에 해당하면 rejected + permission=False + 사유 저장."""
    changed=0; dropped_doms=[]; now=_kst_now().strftime('%Y-%m-%d %H:%M')
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('status')=='rejected': continue
            # 실게시 검증된 사이트는 일시 실패로 함부로 탈락시키지 않는다
            if str(s.get('verified_post_url') or '').startswith(('http://','https://')):
                # 단, 검증됐어도 연속 실패가 크게 쌓이면(서버 사망 등) 발행만 잠근다
                if int(s.get('fail_streak',0) or 0)>=FAIL_STREAK_DROP+2:
                    if s.get('permission'):
                        s['permission']=False; s['auto_drop_reason']=f'검증됨이나 연속 실패 {s.get("fail_streak")}회 — 발행 잠금'
                        s['auto_dropped_at']=now; changed+=1
                continue
            reason=_site_permanent_block(s)
            if reason:
                s.update({'status':'rejected','permission':False,
                          'auto_drop_reason':reason,'auto_dropped_at':now})
                changed+=1; dropped_doms.append(_domain_of(s.get('site_url','')))
        if changed: save_sites(sites)
    if dropped_doms:   # 자동 탈락 사이트 도메인도 영구 목록에 기록(재발굴 방지)
        add_rejected_domains(dropped_doms,'사이트 자동탈락')
    if changed: add_log(f'[자동정리] {changed}개 사이트 자동 탈락(발행 불가)')
    return changed

def start_workers(n=2):
    global wk_active,wk_paused
    if wk_active: return
    n=1  # 동일 사이트 동시 발행에 의한 한도/간격 우회 방지
    wk_active=True; wk_paused=False; add_log(f'[워커] {n}개 시작 (사이트 한도 보호)')
    for i in range(n):
        t=threading.Thread(target=worker_loop,name=f'W-{i+1}',daemon=True); t.start()

def stop_workers():
    global wk_active; wk_active=False; add_log('[워커] 정지')

def pause_workers():
    global wk_paused; wk_paused=True; add_log('[워커] 일시정지')

def resume_workers():
    global wk_paused; wk_paused=False; add_log('[워커] 재개')

def worker_loop():
    while wk_active:
        # 일시정지: 큐에서 꺼내지 않고 대기 (작업 보존)
        if wk_paused: time.sleep(1); continue
        try: job=post_queue.get(timeout=1)
        except: continue
        site=job['site']; title=job['title']; content=job['content']; cfg=load_config()
        job_id=job.get('job_id'); hid=job.get('hist_id')
        name=site.get('name') or site.get('site_url','')[:24]

        # 큐 등록 뒤 사이트가 삭제/미허용/캡차 상태로 바뀌어도 발행되지 않도록 현재 상태 재검증.
        current=next((s for s in load_sites() if s.get('id')==site.get('id')),None)
        if not current or not is_publishable(current):
            why='사이트 삭제됨' if not current else '관리자 허용 또는 실게시 검증 조건 미충족'
            add_log(f'[스킵] {name} {why}')
            if hid: history_update(hid,status='skipped',message=why)
            if job_id: _persist_remove(job_id)
            with STATS_LOCK:
                wk_stats['skipped']+=1; wk_stats['done']+=1; wk_stats['queued']=post_queue.qsize()
            continue
        site=current; job['site']=current

        # 1일 한도 확인 (도배 방지)
        if not under_daily_limit(site,cfg):
            add_log(f'[스킵] {name} 일일 발행 한도 도달')
            if hid: history_update(hid,status='skipped',message='일일 발행 한도 도달')
            if job_id: _persist_remove(job_id)
            with STATS_LOCK:
                wk_stats['skipped']+=1; wk_stats['done']+=1; wk_stats['queued']=post_queue.qsize()
            continue

        interval_ok,remain=under_min_interval(site)
        if not interval_ok:
            mins=max(1,(remain+59)//60)
            add_log(f'[스킵] {name} 최소 발행 간격 미충족 ({mins}분 남음)')
            if hid: history_update(hid,status='skipped',message=f'사이트 최소 발행 간격 미충족 ({mins}분 남음)')
            if job_id: _persist_remove(job_id)
            with STATS_LOCK:
                wk_stats['skipped']+=1; wk_stats['done']+=1; wk_stats['queued']=post_queue.qsize()
            continue

        if hid: history_update(hid,status='posting')
        # 발행 (실패 시 3회 재시도, 지수 백오프 + 드라이버 자동 재시작)
        ok=False; msg=''
        for attempt in range(1,4):
            try:
                ok,msg=do_post(site,title,content)
                if ok: break
                add_log(f'[재시도 {attempt}/3] {name} - {msg}')
                reset_driver()   # 실패 시 드라이버 새로 띄워 세션 꼬임 방지
            except Exception as e:
                msg=str(e); add_log(f'[재시도 {attempt}/3] {name} - {msg[:60]}')
                reset_driver()   # 크롬 죽었을 때 자동 재시작
            if not wk_active: break
            if attempt<3 and not wk_paused: time.sleep(min(5*attempt,15))

        # 실패 원인 분류 + 일시적 실패는 지연 후 자동 재시도
        reason=reason_ko=''; transient=False
        if not ok:
            reason,reason_ko,is_temp=classify_fail(msg)
            requeues=job.get('requeues',0)
            transient=is_temp and requeues<RETRY_MAX and wk_active
            if transient:
                job['requeues']=requeues+1
        if transient:
            with STATS_LOCK:
                wk_stats['retry']+=1; wk_stats['queued']=post_queue.qsize()
            if hid: history_update(hid,status='retry',fail_reason=reason,fail_reason_ko=reason_ko,
                                   message=f'{reason_ko} → {RETRY_DELAY//60}분 후 자동 재시도 ({job["requeues"]}/{RETRY_MAX}): {str(msg)[:150]}')
            schedule_retry(job)   # queue.json 은 그대로 유지(재시작 시 복구)
            add_log(f'[재시도 예약 {job["requeues"]}/{RETRY_MAX}] {name} - {reason_ko}')
            if cfg.get('notify_fail'): send_telegram(cfg,f'🔄 일시적 실패({reason_ko}) 재시도 예약: {name}')
            continue
        with STATS_LOCK:
            if ok: wk_stats['success']+=1
            else: wk_stats['fail']+=1
            wk_stats['done']+=1; wk_stats['queued']=post_queue.qsize()
        finalize_post(site,ok,fail_reason=('' if ok else str(msg)))
        # CAPTCHA 값은 관리자 수동 입력만 허용하며 자동 판독/우회하지 않는다.
        if hid: history_update(hid,status='done' if ok else 'failed',
                               result_url=(msg if ok and str(msg).startswith('http') else ''),
                               fail_reason=('' if ok else reason),
                               fail_reason_ko=('' if ok else reason_ko),
                               alive=('yes' if ok and str(msg).startswith('http') else ''),
                               message=str(msg)[:300])
        if job_id: _persist_remove(job_id)
        add_log(f'[{"성공" if ok else "실패:"+reason_ko}] {name}')
        # 텔레그램 알림 (설정에 따라)
        if ok and cfg.get('notify_done'): send_telegram(cfg,f'✅ 발행 성공: {name}\n{title[:60]}')
        if (not ok) and cfg.get('notify_fail'): send_telegram(cfg,f'❌ 발행 실패({reason_ko}): {name}\n{str(msg)[:120]}')

        # rate limit: 포스트 간 지연 (도배 방지)
        delay=int(cfg.get('post_delay',30) or 0)
        if delay>0 and wk_active and not wk_paused: time.sleep(delay)

def is_permitted(site):
    """사이트 관리에서 등록 상태로 전환된 사이트만 발행."""
    return bool(site.get('permission')) and site.get('registration_source') in ('manual_admin','admin_bulk','legacy_admin','candidate_registered','verified_test')

def is_autopostable(site):
    """자동발행 대상 = 허용 + 실제 게시 성공 URL 검증 완료.
       (CAPTCHA는 2captcha로 자동 해결 시도 → 실패 시 수동 대기)"""
    verified_url=str(site.get('verified_post_url') or '')
    return (is_permitted(site)
            and site.get('status')!='rejected'
            and site.get('write_test_status')=='passed'
            and verified_url.startswith(('http://','https://')))

def is_assisted_postable(site):
    """보조발행 대상 = 허용 + 실제 게시 성공 URL 검증 완료."""
    verified_url=str(site.get('verified_post_url') or '')
    return (is_permitted(site)
            and site.get('status')!='rejected'
            and site.get('write_test_status')=='passed'
            and verified_url.startswith(('http://','https://')))

def is_publishable(site):
    return is_autopostable(site)

def enqueue(sites,title,content,meta=None):
    meta=meta or {}
    allowed=[s for s in sites if is_publishable(s)]
    blocked=[s for s in sites if not is_publishable(s)]
    now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for s in allowed:
        jid=secrets.token_hex(8)
        history_add({'id':jid,'time':now,'updated':now,'site_id':s.get('id'),
                     'site_name':s.get('name') or s.get('site_url',''),'site_url':s.get('site_url',''),
                     'bo_table':s.get('bo_table',''),'title':title,
                     'region':meta.get('region',''),'service':meta.get('service',''),
                     'status':'queued','result_url':'','message':'','attempts':0})
        job={'job_id':jid,'hist_id':jid,'site':s,'title':title,'content':content}
        post_queue.put(job); _persist_add(job)
    for s in blocked:
        _why=('미허용 도메인' if not is_permitted(s) else '제외')
        add_log(f'[차단:{_why}] 발행 스킵: {s.get("name") or (s.get("site_url","") or "")[:30]}')
    with STATS_LOCK:
        wk_stats['total']+=len(allowed); wk_stats['queued']=post_queue.qsize()
    return len(allowed),len(blocked)

def enqueue_generated(sites, keywords, cfg, meta=None):
    """허용 사이트마다 '각각 다른' 유니크 제목·본문을 새로 생성해 큐 등록.
       → 같은 키워드라도 사이트마다 글이 달라져 중복 발행을 방지."""
    meta=meta or {}
    allowed=[s for s in sites if is_publishable(s)]
    blocked=[s for s in sites if not is_publishable(s)]
    now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for s in allowed:
        html,title=generate_article(keywords,cfg)   # 사이트마다 새로 생성(유니크)
        jid=secrets.token_hex(8)
        history_add({'id':jid,'time':now,'updated':now,'site_id':s.get('id'),
                     'site_name':s.get('name') or s.get('site_url',''),'site_url':s.get('site_url',''),
                     'bo_table':s.get('bo_table',''),'title':title,
                     'region':meta.get('region',''),'service':meta.get('service',''),
                     'workroom_id':meta.get('workroom_id',''),'workroom_name':meta.get('workroom_name',''),
                     'member':meta.get('member',''),
                     'status':'queued','result_url':'','message':'','attempts':0})
        job={'job_id':jid,'hist_id':jid,'site':s,'title':title,'content':html}
        post_queue.put(job); _persist_add(job)
    for s in blocked:
        _why=('미허용 도메인' if not is_permitted(s) else '제외')
        add_log(f'[차단:{_why}] 발행 스킵: {s.get("name") or (s.get("site_url","") or "")[:30]}')
    with STATS_LOCK:
        wk_stats['total']+=len(allowed); wk_stats['queued']=post_queue.qsize()
    return len(allowed),len(blocked)

# ==================== 예약 발행 스케줄러 ====================
def load_scheds(): return load_json(SCHED_FILE,[])
def save_scheds(s): save_json(SCHED_FILE,s)

def schedule_keyword_key(kw):
    """예약별 키워드 진행률을 재시작 후에도 동일하게 식별한다."""
    return '|'.join(str((kw or {}).get(k,'') or '').strip() for k in ('지역','서비스','브랜드'))

def _kst_now():
    """서버가 UTC여도 한국시간 기준으로 계산."""
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo('Asia/Seoul'))
    except Exception:
        return datetime.utcfromtimestamp(time.time()+9*3600)

# ==================== 회원(고객) 관리 + 월 정산 ====================
def load_members(): return load_json(MEMBERS_FILE,[])
def save_members(m): save_json(MEMBERS_FILE,m)
def _cur_month(): return _kst_now().strftime('%Y-%m')

def member_fee(m):
    """월 청구액 = 기본료 + (추가 광고수 × 추가단가)."""
    base=int(m.get('plan_fee',30000) or 0)
    addons=int(m.get('addons',0) or 0)
    afee=int(m.get('addon_fee',10000) or 0)
    return base+addons*afee

def member_view(m):
    """회원 1건을 정산정보 포함해 표시용으로 반환."""
    cm=_cur_month(); pays=m.get('payments',{}) or {}
    pm=pays.get(cm,{}) if isinstance(pays,dict) else {}
    fee=member_fee(m)
    return {**m,'fee':fee,'this_month':cm,
            'paid':bool(pm.get('paid')),'paid_at':pm.get('paid_at',''),
            'unpaid_months':[k for k,v in pays.items() if not (v or {}).get('paid')] if isinstance(pays,dict) else []}

# ==================== 도메인 발굴 (구글 검색 → 자동검수 → 승인 대기) ====================
def load_cands():
    """후보 데이터 로드. 더 이상 사용하지 않는 연락처 정보는 영구 제거한다."""
    cands=load_json(CAND_FILE,[])
    changed=False
    for cand in cands:
        if cand.pop('emails',None) is not None: changed=True
    if changed: save_json(CAND_FILE,cands)
    return cands
def save_cands(c): save_json(CAND_FILE,c)
_cand_lock=threading.Lock()

# ---- 영구 탈락 도메인(다음 발굴에서 제외) ----
def load_rejected_domains():
    """영구 탈락 도메인 집합. 탈락 후보를 삭제해도 재수집되지 않게 한다."""
    d=load_json(REJECTED_DOMAINS_FILE,{})
    if isinstance(d,list): d={'domains':d}
    return set(x.lower() for x in (d.get('domains') or []))

def add_rejected_domains(domains, reason=''):
    """도메인들을 영구 탈락 목록에 추가(사유·시각 기록)."""
    if isinstance(domains,str): domains=[domains]
    doms=[(_domain_of(x) if x.startswith('http') else x).lower().replace('www.','') for x in domains if x]
    doms=[d for d in doms if d]
    if not doms: return 0
    with _cand_lock:
        raw=load_json(REJECTED_DOMAINS_FILE,{})
        if isinstance(raw,list): raw={'domains':raw,'log':[]}
        cur=set(x.lower() for x in (raw.get('domains') or []))
        log=raw.get('log') or []
        now=_kst_now().strftime('%Y-%m-%d %H:%M')
        added=0
        for d in doms:
            if d not in cur:
                cur.add(d); log.append({'domain':d,'reason':str(reason)[:80],'at':now}); added+=1
        raw['domains']=sorted(cur); raw['log']=log[-2000:]
        save_json(REJECTED_DOMAINS_FILE,raw)
    return added

# 쿼리 조합 — 플랫폼 흔적 × 홍보 의도
GNU_PATTERNS=['inurl:bbs/board.php bo_table=promotion','inurl:bbs/board.php bo_table=hongbo',
              'inurl:bbs/board.php bo_table=ad','inurl:bbs/board.php bo_table=link',
              'inurl:bbs/board.php bo_table=partner','inurl:bbs/board.php bo_table=banner',
              'inurl:bbs/write.php bo_table=promotion','inurl:bbs/board.php "홍보게시판"']
CAFE_PATTERNS=['inurl:/board/ list.html "홍보"','inurl:/board/free/ "홍보"','"cafe24" inurl:/board/ "제휴"']
INTENT=['"홍보게시판"','"자유홍보"','"홍보 환영"','"홍보 가능"','"링크등록"','"제휴문의"','"상호등록"','"업체등록"']
# 제외 도메인(포털·정부·언론·대형)
BLACK_DOMAINS=['naver.com','daum.net','google.','youtube.','facebook.','instagram.','tistory.com',
               'blog.','cafe.','.go.kr','.or.kr','.ac.kr','.mil.kr','wikipedia.','namu.wiki',
               'chosun.com','donga.com','joins.com','hani.co.kr','mk.co.kr','news','gov',
               'coupang.com','11st.co.kr','gmarket.co.kr','auction.co.kr','interpark',
               'twseo.kr','marketingmonster.kr']
AD_BAN_WORDS=['광고 금지','광고금지','홍보 금지','홍보금지','상업적 게시물','상업적게시물','광고성 글 삭제',
              '광고글 삭제','도배 금지','스팸 금지','영리 목적','상업적 목적 금지','무단 홍보']
# 주차/만료 도메인 (검색엔진엔 남아있지만 실제론 껍데기)
PARKED_WORDS=['resources and information','this domain','domain is for sale','도메인 판매',
              '이 도메인은','buy this domain','parked','sedoparking','afternic','dan.com',
              'hugedomains','도메인이 만료','관련 검색어','sponsored listings','related searches']
# 불법·도박·성인 사이트 (후보 부적합 — 제휴 대상 아님)
ILLEGAL_WORDS=['카지노','바카라','슬롯','토토','먹튀','배팅','베팅','도박','홀덤','파워볼',
               '사설','환전','꽁머니','livecasino','casino','baccarat','betting']
PROMO_WORDS=['홍보게시판','자유홍보','홍보 환영','홍보가능','홍보 가능','제휴문의','제휴 문의','링크등록',
             '업체등록','상호등록','광고게시판','홍보하기','업체홍보','파트너 모집']

def _domain_of(url):
    m=re.match(r'https?://([^/]+)',url or '')
    return (m.group(1).lower() if m else '').replace('www.','')

def _is_blacklisted(url):
    u=(url or '').lower()
    return any(b in u for b in BLACK_DOMAINS)

# 그누보드 홍보/자유 게시판에서 흔한 bo_table 값 (URL 조각으로 직접 검색)
GNU_BO_TABLES=['promotion','promotion1','hongbo','hongbo1','ad','link','partner','banner',
               'free','free1','guest','company','pr','event','notice_pr']
# Brave에 잘 먹히는 '평문 URL조각' — inurl: 대신 실제 경로 문자열을 그대로 검색
BRAVE_URL_FRAGMENTS=['bbs/board.php bo_table=promotion','bbs/board.php bo_table=hongbo',
                     'bbs/board.php bo_table=link','bbs/board.php bo_table=partner',
                     'bbs/write.php bo_table=promotion','bbs/board.php bo_table=free 홍보']

def _board_finder_queries(provider):
    """플랫폼(그누보드/카페24) 홍보·자유 게시판을 '찾기 위한' 검색어.
       사용자의 지역×업종 목록(=경쟁사 검색)과 무관하게 항상 앞에 실행한다."""
    qs=[]
    if provider=='brave':
        # 그누보드: 평문 URL조각 + 홍보의도 (라이브 검증: 실제 홍보게시판 10/10 적중)
        for frag in BRAVE_URL_FRAGMENTS:
            qs.append(frag)
            for i in INTENT[:3]:
                qs.append(f'{frag} {i}')
        for bo in GNU_BO_TABLES:
            qs.append(f'bbs/board.php bo_table={bo} 홍보')
        for i in INTENT:
            qs.append(f'{i} bbs board.php 글쓰기')
            qs.append(f'{i} 그누보드 게시판')
        # 카페24: 평문 게시판 경로 조각(그누보드보다 약하지만 커버)
        for frag in ['board/free list.html 홍보','board_no 자유게시판 cafe24',
                     'board 홍보게시판 write.html','cafe24 게시판 홍보 환영']:
            qs.append(frag)
    else:
        # Google: inurl: 연산자가 강력
        for p in GNU_PATTERNS+CAFE_PATTERNS:
            qs.append(p)
            for i in INTENT[:4]:
                qs.append(f'{p} {i}')
        for i in INTENT:
            qs.append(f'{i} inurl:bbs')
            qs.append(f'{i} inurl:board')
    return qs

def build_queries(cfg):
    """플랫폼 흔적 × 홍보 의도 × (선택)업종 키워드 조합 생성.
       provider가 brave면 inurl: 연산자가 안 먹으므로 '평문 URL조각' 쿼리를 쓴다.
       그누보드/카페24 홍보게시판 '찾기' 검색어를 항상 먼저 실행하고, 그 뒤에
       사용자가 저장한 목록(지역×업종 등)을 붙인다."""
    provider=(cfg.get('search_provider') or 'brave').lower()
    extra=[x.strip() for x in (cfg.get('discover_keywords','') or '').splitlines() if x.strip()]
    direct=[x.strip() for x in (cfg.get('discover_direct_queries','') or '').splitlines()
            if x.strip() and not x.lstrip().startswith('#')]
    # 1) 항상 먼저: 플랫폼 홍보게시판 찾기 검색어
    finder=_board_finder_queries(provider)
    # 2) 업종/지역 키워드가 있으면 게시판 조각과 곱해 보강
    if provider=='brave':
        for e in extra:
            for i in INTENT[:3]:
                finder.append(f'{e} {i}')
            finder.append(f'{e} bbs/board.php bo_table=promotion')
            finder.append(f'{e} 홍보게시판 글쓰기')
    else:
        for e in extra:
            for i in INTENT[:5]:
                finder.append(f'{e} {i}')
            finder.append(f'{e} inurl:bbs/board.php')
    # 3) 사용자가 저장한 목록은 뒤에 이어붙인다(보존). finder가 앞이라 쿼리한도 안에서 우선 실행됨.
    return list(dict.fromkeys(finder+direct))

def google_search(cfg, query, start=1, num=10):
    """Google Custom Search JSON API. (공식 API — 결과 직접 스크래핑 안 함)"""
    key=cfg.get('google_api_key',''); cx=cfg.get('google_cx','')
    if not key or not cx: raise RuntimeError('구글 API 키/검색엔진ID(cx) 미설정')
    import requests as _rq
    r=_rq.get('https://www.googleapis.com/customsearch/v1',
              params={'key':key,'cx':cx,'q':query,'start':start,'num':num,'hl':'ko','lr':'lang_ko'},
              timeout=20)
    if r.status_code==429: raise RuntimeError('구글 API 일일 한도 초과(429)')
    if r.status_code>=400: raise RuntimeError(f'구글 API 오류 {r.status_code}: {r.text[:120]}')
    j=r.json()
    return [{'url':it.get('link',''),'title':it.get('title',''),'snippet':it.get('snippet','')}
            for it in (j.get('items') or [])]

def brave_search(cfg, query, start=1, num=10):
    """Brave Search 공식 Web API."""
    key=(cfg.get('brave_api_key') or '').strip()
    if not key: raise RuntimeError('Brave Search API 키 미설정 — 설정 탭에서 입력하세요')
    import requests as _rq
    offset=max(0,(int(start or 1)-1)//max(1,int(num or 10)))
    r=_rq.get('https://api.search.brave.com/res/v1/web/search',
              headers={'Accept':'application/json','X-Subscription-Token':key},
              params={'q':query,'count':min(20,max(1,int(num or 10))),'offset':min(9,offset),
                      'country':'KR','search_lang':'ko','safesearch':'moderate'},timeout=20)
    if r.status_code==429: raise RuntimeError('Brave Search API 사용량/속도 한도 초과(429)')
    if r.status_code in (401,403): raise RuntimeError('Brave Search API 키 또는 구독 상태 확인 필요')
    if r.status_code>=400: raise RuntimeError(f'Brave Search API 오류 {r.status_code}: {r.text[:120]}')
    rows=((r.json().get('web') or {}).get('results') or [])
    return [{'url':it.get('url',''),'title':it.get('title',''),'snippet':it.get('description','')}
            for it in rows if it.get('url')]

def web_search(cfg, query, start=1, num=10):
    provider=(cfg.get('search_provider') or 'brave').lower()
    return google_search(cfg,query,start,num) if provider=='google' else brave_search(cfg,query,start,num)

def screen_candidate(url, cfg=None):
    """HTTP 1~3회로 후보 자동 검수. 반환: 검수 결과 dict."""
    import requests as _rq
    UA={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36'}
    m=re.match(r'(https?://[^/]+)',url or ''); base=m.group(1) if m else url
    res={'base':base,'domain':_domain_of(url),'platform':'unknown','board_name':'','bo_table':'',
         'write_form':False,'captcha':'','login_required':False,'ad_banned':False,'promo_hint':False,
         'last_post_days':None,'reachable':False,'note':'',
         'parked':False,'illegal':False}
    def get(u):
        try: return _rq.get(u,timeout=12,verify=False,headers=UA,allow_redirects=True)
        except Exception: return None
    # bo_table 추출
    bm=re.search(r'bo_table=([A-Za-z0-9_]+)',url or '')
    if bm: res['bo_table']=bm.group(1)
    r=get(url)
    if not r or r.status_code>=400:
        res['note']=f'접속 실패({r.status_code if r else "timeout"})'; return res
    res['reachable']=True
    html=r.text or ''; low=html.lower()
    # 플랫폼
    if 'bo_table' in low or 'gnuboard' in low or '/bbs/' in low: res['platform']='gnuboard'
    elif 'cafe24' in low or 'xans-' in low or '/board/write.html' in low: res['platform']='cafe24'
    # 게시판명 (title/h1)
    tm=re.search(r'<title[^>]*>(.*?)</title>',html,re.S|re.I)
    title=re.sub(r'\s+',' ',re.sub(r'&nbsp;',' ',(tm.group(1) if tm else '')))
    res['board_name']=title[:80]
    # 주차/만료 도메인 판별 (검색엔진엔 남아있지만 실제론 껍데기)
    tl=title.lower()
    res['parked']=(any(w in low or w in tl for w in PARKED_WORDS) or len(html)<800)
    # 불법·도박 사이트 판별 (제휴 대상 부적합)
    res['illegal']=any(w in low for w in ILLEGAL_WORDS)
    # 광고 금지 / 홍보 허용 흔적 (본문 + URL/게시판ID/타이틀까지 함께 판단)
    res['ad_banned']=any(w in html for w in AD_BAN_WORDS)
    hint_blob=html+' '+title+' '+(url or '')+' '+res['bo_table']
    res['promo_hint']=(any(w in hint_blob for w in PROMO_WORDS)
                       or bool(re.search(r'bo_table=(promotion|hongbo|ad|link|partner|banner)',url or '',re.I)))
    # 최근 글(날짜 패턴에서 가장 최근)
    try:
        ds=re.findall(r'(20\d{2})[-.](\d{1,2})[-.](\d{1,2})',html)
        if ds:
            from datetime import date
            latest=max(date(int(y),int(mo),int(d)) for y,mo,d in ds
                       if 1<=int(mo)<=12 and 1<=int(d)<=31)
            res['last_post_days']=max(0,(_kst_now().date()-latest).days)
    except Exception: pass
    # 글쓰기 페이지: 실제 쓰기 링크(write.php)만 신뢰한다.
    # ※ board.php?wr_id= (글 조회) 페이지에도 wr_content가 있지만 그건 '댓글 폼'이라
    #    글쓰기 폼으로 오인하면 안 된다(오판 시 write.php에서 게시판 못찾음으로 실패).
    from urllib.parse import urljoin
    wurls=[]
    url_low=(url or '').lower()
    is_view = 'wr_id=' in url_low  # 글 조회 URL이면 그 페이지의 폼은 댓글일 가능성 → 신뢰 안 함
    if (not is_view) and 'write.php' in url_low and \
       ('wr_subject' in low or 'name="subject"' in low) and ('wr_content' in low or 'name="content"' in low):
        wurls.append(url)
    for href in re.findall(r'href=["\']([^"\']+)["\']',html,re.I):
        hl=href.lower()
        if ('write.php' in hl and 'bo_table=' in hl) or ('/board/write.html' in hl and 'board_no=' in hl):
            wurls.append(urljoin(r.url,href))
    if res['platform']=='cafe24':
        wurls.append(base+'/board/write.html?board_no=1')
    else:
        bo=res['bo_table'] or 'free'
        wurls.append(base+f'/bbs/write.php?bo_table={bo}')
    wurls=list(dict.fromkeys(wurls))[:6]
    for wu in wurls:
        wr=get(wu)
        if not wr or wr.status_code>=400: continue
        wh=wr.text or ''; wl=wh.lower()
        has_subject=('wr_subject' in wl or 'name="subject"' in wl or "name='subject'" in wl)
        has_content=('wr_content' in wl or 'name="content"' in wl or "name='content'" in wl)
        if has_subject and has_content:
            res['write_form']=True; res['write_url']=wr.url
        if any(k in wl for k in ['captcha_key','kcaptcha','g-recaptcha','h-captcha','cf-turnstile']) or '자동등록방지' in wh:
            res['captcha']='감지'
        if not res['write_form'] and ('mb_password' in wl or '로그인' in wh):
            res['login_required']=True
        if any(w in wh for w in AD_BAN_WORDS): res['ad_banned']=True
        if res['write_form']: break
    return res

def score_candidate(c):
    """점수화 — 위에서부터 100개만 보면 되게."""
    s=0
    name=(c.get('board_name','')+' '+c.get('bo_table',''))
    if c.get('promo_hint') or any(w in name for w in ['홍보','제휴','광고','링크','업체']): s+=30
    if c.get('write_form') and not c.get('login_required'): s+=20
    lp=c.get('last_post_days')
    if lp is not None:
        if lp<=7: s+=15
        elif lp<=30: s+=8
        elif lp>365: s-=10
    if not c.get('captcha'): s+=10
    if c.get('platform') in ('gnuboard','cafe24'): s+=5
    if c.get('ad_banned'): s-=50
    if c.get('captcha'): s-=30
    if c.get('login_required'): s-=20
    if c.get('parked'): s-=100      # 주차/만료 도메인 = 껍데기
    if c.get('illegal'): s-=100     # 도박·불법 사이트 = 제휴 부적합
    if not c.get('reachable'): s-=100
    return s

def precheck_search_result(url):
    """검색 후보 저장 전 실제 접근 및 HTML title 숫자 개수를 가볍게 확인."""
    import requests as _rq
    try:
        r=_rq.get(url,timeout=12,verify=False,allow_redirects=True,headers={
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0 Safari/537.36'})
        if r.status_code>=400: return {'reachable':False,'title':'','digits':0}
        html=r.text or ''
        tm=re.search(r'<title[^>]*>(.*?)</title>',html,re.S|re.I)
        title=re.sub(r'\s+',' ',re.sub(r'<[^>]+>',' ',tm.group(1) if tm else '')).strip()
        return {'reachable':True,'title':title[:200],'digits':len(re.findall(r'\d',title))}
    except Exception:
        return {'reachable':False,'title':'','digits':0}

def add_candidates_from(items, cfg, source='search'):
    """검색 결과 → 접근·title 숫자·글쓰기 폼까지 확인 → 후보 등록."""
    with _cand_lock:
        cands=load_cands()
        known_dom={c.get('domain') for c in cands}
        site_dom={_domain_of(s.get('site_url','')) for s in load_sites()}
        rejected_dom=load_rejected_domains()   # 영구 탈락: 재수집 안 함
        added=0; blocked_unreachable=0; blocked_title=0; blocked_write=0; blocked_rejected=0
        for it in items:
            url=it.get('url') if isinstance(it,dict) else str(it)
            if not url or not url.startswith('http'): continue
            if _is_blacklisted(url): continue
            dom=_domain_of(url)
            if not dom or dom in known_dom or dom in site_dom: continue
            if dom.replace('www.','') in rejected_dom:   # 이전에 탈락한 도메인은 건너뜀
                blocked_rejected+=1; continue
            # URL 자체가 게시판 경로(그누보드/카페24)면 '제목 숫자 8개' 규칙을 면제한다.
            # (홍보게시판은 제목이 '홍보게시판'처럼 숫자가 없을 때가 많아 오탈락하던 문제)
            url_is_board=bool(re.search(r'(bbs/board\.php|bo_table=|/board/.*list\.html|board_no=)',url.lower()))
            check=precheck_search_result(url) if source!='manual' else {'reachable':True,'title':'','digits':8}
            if not check.get('reachable'):
                blocked_unreachable+=1; continue
            if int(check.get('digits',0))<8 and not url_is_board:
                blocked_title+=1; continue
            form_check=screen_candidate(url,cfg) if source!='manual' else {}
            if source!='manual' and not form_check.get('write_form'):
                blocked_write+=1; continue
            if source!='manual':
                form_check['screened']=True
                form_check['score']=score_candidate(form_check)
                if form_check.get('parked'): form_check['status']='rejected'; form_check['reject_reason']='주차/만료 도메인'
                elif form_check.get('illegal'): form_check['status']='rejected'; form_check['reject_reason']='도박·불법 사이트'
                else: form_check['status']='ready'
            known_dom.add(dom)
            rec={'id':secrets.token_hex(6),'url':url,'domain':dom,
                 'title':check.get('title') or (it.get('title','') if isinstance(it,dict) else ''),
                 'title_digit_count':check.get('digits',0),'precheck_reachable':True,
                 'snippet':(it.get('snippet','') if isinstance(it,dict) else ''),
                 'found_at':_kst_now().strftime('%Y-%m-%d %H:%M'),'source':source,
                 'query':(it.get('query','') if isinstance(it,dict) else ''),
                 'status':'new','score':0,'screened':False}
            rec.update(form_check)
            cands.append(rec)
            added+=1
        save_cands(cands)
    if source!='manual' and (blocked_unreachable or blocked_title or blocked_write or blocked_rejected):
        add_log(f'[후보 사전필터] 접속불가 {blocked_unreachable} · 제목숫자<8 {blocked_title} · 글쓰기폼없음 {blocked_write} · 영구탈락 {blocked_rejected} 제외')
    return added

def screen_pending(limit=30):
    """미검수 후보를 검수·점수화. 처리 건수 반환."""
    cfg=load_config()
    with _cand_lock:
        cands=load_cands()
        todo=[c for c in cands if not c.get('screened')][:max(1,limit)]
        ids=[c['id'] for c in todo]
    results={}
    for c in todo:
        try:
            r=screen_candidate(c['url'],cfg)
        except Exception as e:
            r={'note':str(e)[:100],'reachable':False}
        r['screened']=True
        r['score']=score_candidate(r)
        # 자동 탈락 사유
        if not r.get('reachable'): r['status']='ready'; r['reject_reason']='현재 접속 불가 — 후보 유지·재검수 가능'
        elif r.get('parked'): r['status']='rejected'; r['reject_reason']='주차/만료 도메인 (실제 게시판 아님)'
        elif r.get('illegal'): r['status']='rejected'; r['reject_reason']='도박·불법 사이트 (제휴 부적합)'
        elif r.get('ad_banned'): r['status']='rejected'; r['reject_reason']='광고 금지 명시'
        elif r.get('captcha'): r['status']='ready'; r['reject_reason']='캡차 있음 — 2captcha 자동해결 시도 예정'
        elif not r.get('write_form'):
            # 글쓰기 폼이 없다: 그누보드/카페24면 로그인 후 쓰기 가능성 있어 ready 유지(자동가입 대상).
            # 그 외(디렉토리·전화번호검색·경쟁업체 랜딩 등 게시판 아님)는 자동 탈락시켜 목록을 깨끗이.
            if r.get('platform') in ('gnuboard','cafe24') or r.get('login_required'):
                r['status']='ready'; r['reject_reason']='글쓰기 폼 미확인 — 로그인 필요할 수 있음(자동가입 시도)'
            else:
                r['status']='rejected'; r['reject_reason']='게시판 글쓰기 폼 없음(디렉토리/랜딩 페이지 — 홍보 대상 아님)'
        else: r['status']='ready'
        results[c['id']]=r
        time.sleep(1)   # 요청 속도 관리
    rejected_now=[]
    with _cand_lock:
        cands=load_cands()
        for c in cands:
            if c['id'] in results:
                c.update(results[c['id']])
                if results[c['id']].get('status')=='rejected':
                    rejected_now.append(c.get('domain') or _domain_of(c.get('url','')))
        save_cands(cands)
    if rejected_now:   # 검수에서 탈락한 도메인은 영구 목록에 기록(재수집 방지)
        add_rejected_domains(rejected_now,'검수 탈락')
    return len(results)

def discover_once(cfg=None, max_queries=10):
    """구글 검색 1배치 실행 → 후보 등록 → 검수. (검색량 제어)"""
    cfg=cfg or load_config()
    st=load_json(DISCO_FILE,{})
    today=_kst_now().strftime('%Y-%m-%d')
    if st.get('date')!=today: st={'date':today,'queries':0,'found':0,'cursor':0}
    target=int(cfg.get('discover_daily_target',100) or 100)
    qlimit=int(cfg.get('discover_query_limit',100) or 100)   # 구글 무료 하루 100
    queries=build_queries(cfg)
    if not queries: return {'ok':False,'error':'쿼리 없음'}
    added=0; used=0; errs=[]
    for _ in range(max_queries):
        if st['queries']>=qlimit: errs.append('일일 쿼리 한도 도달'); break
        if st['found']>=target: errs.append('일일 후보 목표 달성'); break
        q=queries[st['cursor']%len(queries)]; st['cursor']+=1
        try:
            items=web_search(cfg,q)
            for it in items: it['query']=q
            n=add_candidates_from(items,cfg); added+=n; st['found']+=n
        except Exception as e:
            errs.append(str(e)[:120]); break
        finally:
            st['queries']+=1; used+=1
        time.sleep(1)
    save_json(DISCO_FILE,st)
    screened=0
    try: screened=screen_pending(20)
    except Exception as e: errs.append('검수 오류 '+str(e)[:80])
    add_log(f'[발굴] 쿼리 {used}회 · 신규 {added}개 · 검수 {screened}건'+(' · '+errs[0] if errs else ''))
    return {'ok':True,'queries':used,'added':added,'screened':screened,
            'today_queries':st['queries'],'today_found':st['found'],'errors':errs}

# ==================== 임시메일(mail.tm) — 이메일 인증 자동 처리 ====================
TEMPMAIL_API='https://api.mail.tm'

def _tempmail_req(path, method='GET', data=None, token=None):
    hdr={'Content-Type':'application/json'}
    if token: hdr['Authorization']='Bearer '+token
    body=json.dumps(data).encode() if data else None
    r=urllib.request.Request(TEMPMAIL_API+path, data=body, headers=hdr, method=method)
    try:
        with urllib.request.urlopen(r, timeout=15) as resp:
            return resp.status, json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        try: return e.code, json.loads(e.read().decode())
        except Exception: return e.code, {}
    except Exception as e:
        return 0, {'error': str(e)[:80]}

def tempmail_create():
    """임시메일 주소를 발급하고 (address, password, token)을 반환. 실패 시 (None,...)."""
    st, doms = _tempmail_req('/domains')
    if st != 200 or not isinstance(doms, dict): return None, None, None
    members = doms.get('hydra:member') or []
    if not members: return None, None, None
    domain = members[0].get('domain')
    addr = f'twseo{secrets.token_hex(5)}@{domain}'; pw = secrets.token_hex(10)
    st, _ = _tempmail_req('/accounts', 'POST', {'address': addr, 'password': pw})
    if st not in (200, 201): return None, None, None
    st, tok = _tempmail_req('/token', 'POST', {'address': addr, 'password': pw})
    token = tok.get('token') if (st == 200 and isinstance(tok, dict)) else None
    return (addr, pw, token) if token else (None, None, None)

def tempmail_wait_verify_link(token, timeout=120):
    """받은편지함을 폴링해 인증 메일의 인증 링크(또는 6자리 코드)를 찾아 반환.
       반환: {'link':url} 또는 {'code':'123456'} 또는 None."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        st, msgs = _tempmail_req('/messages', token=token)
        items = (msgs.get('hydra:member') or []) if isinstance(msgs, dict) else []
        for m in items:
            mid = m.get('id')
            st2, full = _tempmail_req(f'/messages/{mid}', token=token)
            if st2 != 200 or not isinstance(full, dict): continue
            text = (full.get('text') or '') + ' '
            html = ''
            h = full.get('html')
            if isinstance(h, list): html = ' '.join(h)
            elif isinstance(h, str): html = h
            blob = text + ' ' + html
            # 1) 인증 링크(그누보드: register_email_check / auth / confirm / verify)
            for murl in re.findall(r'https?://[^\s"\'<>]+', blob):
                if re.search(r'(email_check|auth|confirm|verify|activate|register|certify|인증)', murl, re.I):
                    return {'link': murl}
            # 2) 6자리 인증 코드
            mcode = re.search(r'\b(\d{6})\b', text)
            if mcode:
                return {'code': mcode.group(1)}
        time.sleep(4)
    return None

def auto_signup(site, submit=True):
    """로그인 필요 사이트에 앱이 스스로 회원가입한다(2captcha로 캡차 해결).
       submit=False면 제출 직전까지만(폼 입력·캡차해결) 수행하고 실제 가입은 안 함(검증용).
       반환: (ok, msg). 성공 시 site에 mb_id/mb_pass가 저장돼 있다.
       이메일 인증 필요·지원불가 캡차는 실패로 반환(가입 대상 제외)."""
    from selenium.webdriver.common.by import By
    cfg=load_config()
    # 1) 가입폼 측정(캐시 30분) — 이메일 인증 필요하면 즉시 제외
    try:
        profile=learn_signup_profile(site,force=False)
    except Exception as e:
        return False,f'가입폼 측정 실패: {str(e)[:100]}'
    if not profile:
        return False,'가입폼을 찾지 못함'
    fields=profile.get('fields') or []
    form_url=profile.get('form_url') or profile.get('signup_url') or site.get('signup_url','')
    if not form_url:
        return False,'가입 폼 URL 불명'
    # 2) 자격증명 생성 (사이트 규칙 반영)
    mid,pw=_signup_credentials(site,profile.get('rules'))
    # 이메일 인증이 필요한 게시판이면 임시메일(mail.tm)로 실제 수신 가능한 주소를 쓴다.
    need_email_verify=bool(profile.get('email_verification_required'))
    tm_addr=tm_pw=tm_token=None
    if need_email_verify:
        tm_addr,tm_pw,tm_token=tempmail_create()
        if not tm_token:
            return False,'이메일 인증 필요 — 임시메일 발급 실패'
        email=tm_addr
        add_log(f'[자동가입] {site.get("name") or site.get("site_url","")} 이메일 인증 → 임시메일 {tm_addr}')
    else:
        email_local=re.sub(r'[^a-z0-9]','',mid.lower())[:20] or ('u'+secrets.token_hex(4))
        email=f'{email_local}@gmail.com'
    nick=(cfg.get('brand') or 'user')+secrets.token_hex(2)
    name=cfg.get('brand') or '홍길동'
    vals_by_role={'id':mid,'password':pw,'password_confirm':pw,'email':email,
                  'nickname':nick,'name':name}
    d=get_driver()
    # 그누보드 가입은 register.php(약관) → 동의 → register_form.php(실제 폼) 2단계다.
    # register_form.php를 GET으로 직접 열면 약관 화면이라 필드가 없다 → 약관부터 진행.
    signup_url=profile.get('signup_url') or site.get('signup_url') or form_url
    def _has_pw_field():
        return bool(_safe_find(d,"input[name='mb_password'],#reg_mb_password,input[type='password']"))
    try:
        d.set_page_load_timeout(25); d.get(signup_url); time.sleep(2); dismiss_alerts(d)
    except Exception:
        pass
    # 약관 동의 체크(있으면 전부 체크) 후 '동의' 제출 버튼으로 실제 폼 진입
    for cb in _safe_find(d,"input[type='checkbox']"):
        try:
            nm=(cb.get_attribute('name') or '').lower()
            if any(k in nm for k in ('agree','약관','provision','privacy','all')) or not nm:
                if not cb.is_selected(): d.execute_script('arguments[0].click()',cb)
        except Exception: pass
    if not _has_pw_field():
        # 약관 폼(fregister_form / register.php)을 제출해 register_form.php로 이동
        submitted_agree=False
        for sel in ("form[name='fregister'] input[type='submit']","#fregister input[type='submit']",
                    "form[action*='register_form'] input[type='submit']","form[action*='register_form'] button",
                    "input[type='submit']","button[type='submit']"):
            for el in _safe_find(d,sel):
                try:
                    if el.is_displayed(): d.execute_script('arguments[0].click()',el); submitted_agree=True; break
                except Exception: pass
            if submitted_agree: break
        time.sleep(2); dismiss_alerts(d)
    if not _has_pw_field():
        # 그래도 없으면 register_form.php를 직접 열되 약관값을 붙여서 접근 시도
        try:
            base_o=_signup_origin(site)
            d.get(base_o+'/bbs/register_form.php?agree=1&agree2=1'); time.sleep(2); dismiss_alerts(d)
        except Exception: pass
    if not _has_pw_field():
        return False,'가입 폼(비밀번호 입력칸)에 도달 실패 — 약관/인증 단계 확인'
    # 3) role별 필드 자동 입력
    filled=[]
    for f in fields:
        role=f.get('role') or ''; sel=f.get('selector') or ''
        if not sel or role in ('','captcha'): continue
        val=vals_by_role.get(role)
        if not val: continue
        for el in _safe_find(d,sel):
            try:
                if el.is_displayed(): el.clear(); el.send_keys(val); filled.append(role); break
            except Exception: pass
    if 'id' not in filled or 'password' not in filled:
        return False,f'가입 필수필드 입력 실패(입력됨: {",".join(filled) or "없음"})'
    # 4) 캡차 있으면 2captcha로 해결 후 입력
    cap=detect_captcha(d)
    if cap:
        ok,cmsg,answer,info=solve_captcha_with_2captcha(d,site,cap,cfg)
        if not ok:
            return False,f'가입 캡차({cap}) 해결 실패: {cmsg}'
        if cap in ('kcaptcha',) and answer:
            done=False
            for csel in ["input[name='captcha_key']","#captcha_key","input[name*='captcha']","input[id*='captcha']"]:
                for el in _safe_find(d,csel):
                    try:
                        if el.is_displayed(): el.clear(); el.send_keys(answer); done=True; break
                    except Exception: pass
                if done: break
        add_log(f'[자동가입 캡차] {site.get("name") or site.get("site_url","")} — {cmsg}')
    if not submit:
        return True,f'가입 직전까지 성공(입력: {",".join(filled)}{" +캡차" if cap else ""}) — 실제 가입 안 함'
    # 5) 제출
    submitted=False
    for sel in ("form#fregister input[type='submit']","form[name='fregister'] input[type='submit']",
                "form#fregister button[type='submit']","#register_form input[type='submit']",
                "form[action*='register_form_update'] input[type='submit']",
                "form[action*='register_form_update'] button[type='submit']"):
        for el in _safe_find(d,sel):
            try:
                if el.is_displayed(): d.execute_script('arguments[0].click()',el); submitted=True; break
            except Exception: pass
        if submitted: break
    if not submitted:
        _safe_js(d,"var f=document.getElementById('fregister')||document.forms['fregister']||document.querySelector(\"form[action*='register_form_update']\");if(f){if(f.requestSubmit)f.requestSubmit();else f.submit();}")
    time.sleep(3); dismiss_alerts(d)
    # 5.5) 이메일 인증: 임시메일 받은편지함을 폴링해 인증 링크/코드를 처리
    if need_email_verify and tm_token:
        add_log(f'[자동가입] {site.get("name") or site.get("site_url","")} 인증메일 대기 중...')
        verify=tempmail_wait_verify_link(tm_token, timeout=120)
        if not verify:
            return False,'이메일 인증 실패 — 인증메일이 오지 않음(임시메일 차단 가능)'
        if verify.get('link'):
            try: d.get(verify['link']); time.sleep(2); dismiss_alerts(d)
            except Exception: pass
            add_log(f'[자동가입] 인증링크 방문 완료')
        elif verify.get('code'):
            # 코드 입력형: 인증코드 입력칸을 찾아 넣고 제출
            for csel in ("input[name*='auth']","input[name*='cert']","input[name*='code']","input[id*='auth']","input[id*='code']"):
                done=False
                for el in _safe_find(d,csel):
                    try:
                        if el.is_displayed(): el.clear(); el.send_keys(verify['code']); done=True; break
                    except Exception: pass
                if done:
                    _safe_js(d,"var b=document.querySelector(\"input[type='submit'],button[type='submit']\");if(b)b.click();")
                    time.sleep(2); break
            add_log(f'[자동가입] 인증코드 입력 완료')
    # 6) 가입 성공 검증 = 로그인 시도해서 로그아웃 링크 확인
    raw=site.get('site_url',''); m=re.match(r'(https?://[^/]+)',raw); base=m.group(1) if m else raw
    try:
        d.get(f'{base}/bbs/login.php'); time.sleep(2)
        for s2 in (("input[name='mb_id']",mid),("input[name='mb_password']",pw)):
            for el in _safe_find(d,s2[0]):
                try:
                    if el.is_displayed(): el.clear(); el.send_keys(s2[1]); break
                except Exception: pass
        for lsel in ("form[name='flogin'] input[type='submit']","form[action*='login_check'] input[type='submit']","#login_fs .btn_submit"):
            for el in _safe_find(d,lsel):
                try:
                    if el.is_displayed(): el.click(); break
                except Exception: pass
        time.sleep(2); dismiss_alerts(d)
        body=''
        try: body=d.find_element(By.TAG_NAME,'body').text[:1500]
        except Exception: pass
        logged=('로그아웃' in body) or ('logout' in (d.page_source or '').lower())
    except Exception:
        logged=False
    if logged:
        set_site_flag(site.get('id'),mb_id=mid,mb_pass=pw,signup_status='complete',
                      login_saved=True,signup_updated_at=datetime.now().isoformat(timespec='seconds'))
        add_log(f'[자동가입 성공] {site.get("name") or site.get("site_url","")} — {mid}')
        return True,f'자동가입 성공 — {mid}'
    return False,'가입 제출했으나 로그인 확인 실패(가입 규칙·중복ID·승인제 가능)'

def _safe_find(d, sel):
    try: return d.find_elements(__import__('selenium.webdriver.common.by',fromlist=['By']).By.CSS_SELECTOR, sel) or []
    except Exception: return []

def _safe_js(d, js):
    try: return d.execute_script(js)
    except Exception: return None

def _promote_candidate_to_site(cand, result_url, write_url='', bo='', permission=True):
    """실제 발행 성공한 후보를 사이트 목록으로 승격(api_cand_verified 로직 재사용).
       permission=True면 즉시 발행 허용 ON(완전 자동)."""
    domain=(cand.get('domain') or _domain_of(cand.get('url',''))).lower()
    m=re.match(r'(https?://[^/]+)',cand.get('url','')); base=m.group(1) if m else cand.get('url','')
    bo=bo or cand.get('bo_table') or 'free'
    now=_kst_now().strftime('%Y-%m-%d %H:%M')
    with POST_LOCK:
        sites=load_sites(); site=next((s for s in sites if _domain_of(s.get('site_url',''))==domain),None)
        created=site is None
        if created:
            site={'id':secrets.token_hex(6),'site_url':base,'platform':cand.get('platform','gnuboard'),
                  'mb_id':'','mb_pass':'','bo_table':bo,'name':cand.get('board_name') or domain,
                  'daily_limit':3,'min_interval_minutes':60,'status':'idle',
                  'added':_kst_now().strftime('%m/%d %H:%M')}
            sites.append(site)
        site.update({'bo_table':bo,'write_url':write_url,'verified_post_url':result_url,
                     'write_test_status':'passed','verified_at':now,'last_structure_check':now,
                     'registration_source':'verified_test'})
        if permission:
            site.update({'permission':True,'permission_note':'자동 파이프라인: 실게시 검증 완료',
                         'permission_date':_kst_now().strftime('%Y-%m-%d')})
        else:
            site.setdefault('permission',False)
        save_sites(sites)
    with _cand_lock:
        cands=load_cands()
        for c in cands:
            if c.get('domain','').lower()==domain:
                c.update({'status':'approved','verified_at':now,'verified_post_url':result_url,'site_id':site['id']})
        save_cands(cands)
    add_log(f'[자동등록] {domain} · {bo} · '+('발행허용 ON' if permission else '발행잠금'))
    return site

def auto_pipeline_once(limit=5):
    """완전 자동 파이프라인: ready 후보 → (필요시)자동가입 → 실제 글1건 발행 → 성공시 자동등록.
       배치당 limit개만 처리(부하·탐지 회피). 반환: 요약 dict."""
    cfg=load_config()
    if not cfg.get('auto_pipeline_enabled'):
        return {'ok':False,'error':'auto_pipeline 비활성화'}
    site_domains={_domain_of(s.get('site_url','')) for s in load_sites()}
    with _cand_lock:
        cands=load_cands()
    # 대상: 검수완료(ready) + 아직 사이트 미등록 + 자동탈락 아님 + '실제 글쓰기 경로'가 있는 것만.
    # (114/맵 등 전화번호·디렉토리 사이트는 write_form도 bo_table도 없어 가입/발행이 불가 → 제외)
    def _has_write_path(c):
        if c.get('write_form'): return True
        return c.get('platform') in ('gnuboard','cafe24') and bool((c.get('bo_table') or '').strip())
    pend=[c for c in cands
          if c.get('screened') and c.get('status')=='ready'
          and not c.get('parked') and not c.get('illegal') and not c.get('ad_banned')
          and (c.get('domain') or '').lower() not in site_domains
          and c.get('reachable') and _has_write_path(c)]
    # 비회원 글쓰기 가능(로그인 불필요) 게시판을 먼저 처리한다. 로그인 필요 게시판은
    # 이메일 인증 등으로 자동가입이 막히는 경우가 많아 배치 슬롯을 낭비하기 쉽다.
    def _prio(c):
        direct = c.get('write_form') and not c.get('login_required')  # 바로 발행 가능
        no_cap = not c.get('captcha')                                  # 캡차 없으면 더 빠름
        return (1 if direct else 0, 1 if no_cap else 0, c.get('score',0))
    pend.sort(key=_prio, reverse=True)
    pend=pend[:max(1,limit)]
    done=0; registered=0; signed=0; results=[]
    for c in pend:
        name=c.get('board_name') or c.get('domain') or c.get('url','')[:30]
        # 임시 site dict(발행 함수는 site 형태를 기대) — 후보 정보로 구성
        m=re.match(r'(https?://[^/]+)',c.get('url','')); base=m.group(1) if m else c.get('url','')
        tmp={'id':'cand_'+c.get('id',''),'site_url':base,'platform':c.get('platform','gnuboard'),
             'bo_table':c.get('bo_table') or 'free','name':name,'mb_id':'','mb_pass':''}
        try:
            # 1) 로그인 필요(=write_form 미확인)면 자동가입 먼저
            if c.get('login_required') or not c.get('write_form'):
                ok_su,msg_su=auto_signup(tmp,submit=True)
                if ok_su: signed+=1
                else:
                    _cand_set(c['id'],status='rejected',reject_reason=f'자동가입 실패: {msg_su[:80]}')
                    results.append({'name':name,'stage':'signup','ok':False,'msg':msg_su})
                    continue  # done 증가·드라이버 리셋은 finally에서 처리
            # 2) 실제 글 1건 발행 (기존 do_post; false-negative 수정 반영됨)
            kw={'지역':'인천','서비스':'셔츠룸','브랜드':cfg.get('brand','') or '테스트'}
            html,title=generate_article(kw,cfg,unique=True)
            ok,msg=do_post(tmp,title,html)
            # 검수는 비회원 글쓰기로 봤지만 실제 write.php가 로그인으로 튕기는 게시판이 있다.
            # 이 경우 자동가입 후 1회 재시도(gjsec처럼 login_required 오판된 케이스 구제).
            if (not ok) and (not tmp.get('mb_id')) and re.search(r'(로그인이 필요|로그인 실패|로그인 화면|권한이 없|권한 없)',str(msg)):
                reset_driver(); time.sleep(1)
                add_log(f'[파이프라인] {name} 로그인필요 → 자동가입 시도')
                ok_su,msg_su=auto_signup(tmp,submit=True)
                if ok_su:
                    signed+=1; add_log(f'[파이프라인] {name} 자동가입 성공 → 재발행')
                    reset_driver(); time.sleep(1)
                    ok,msg=do_post(tmp,title,html)
                else:
                    add_log(f'[파이프라인] {name} 자동가입 실패: {str(msg_su)[:80]}')
                    msg=f'{msg} · 자동가입 실패: {str(msg_su)[:60]}'
            result_url=msg if (ok and str(msg).startswith(('http://','https://'))) else ''
            if ok and result_url:
                _promote_candidate_to_site(c,result_url,write_url=tmp.get('learned',{}).get('write_url','') if isinstance(tmp.get('learned'),dict) else '',
                                           bo=tmp.get('bo_table'),permission=True)
                # 가입정보가 생겼으면 등록 사이트에 반영
                if tmp.get('mb_id'):
                    set_site_flag(_promoted_site_id(c),mb_id=tmp.get('mb_id'),mb_pass=tmp.get('mb_pass'))
                registered+=1
                results.append({'name':name,'stage':'post','ok':True,'url':result_url})
            elif ok:
                _cand_set(c['id'],status='rejected',reject_reason='발행됨(결과 URL 확인 불가)')
                results.append({'name':name,'stage':'post','ok':False,'msg':'결과 URL 없음'})
            else:
                reason,_,is_temp=classify_fail(msg)
                # 일시적 실패라도 무한 재시도로 배치 슬롯을 소모하지 않게 상한(5회)을 둔다.
                attempts=int(c.get('pipeline_attempts',0) or 0)+1
                if is_temp and attempts<5:
                    _cand_set(c['id'],status='ready',reject_reason=f'일시적 실패({attempts}/5): {str(msg)[:60]}',pipeline_attempts=attempts)
                else:
                    _cand_set(c['id'],status='rejected',reject_reason=(str(msg)[:90] if not is_temp else f'재시도 {attempts}회 초과: {str(msg)[:60]}'),pipeline_attempts=attempts)
                results.append({'name':name,'stage':'post','ok':False,'msg':str(msg)[:90]})
        except Exception as e:
            results.append({'name':name,'stage':'error','ok':False,'msg':str(e)[:100]})
        finally:
            done+=1
            reset_driver(); time.sleep(3)
    dropped=reconcile_sites()   # 파이프라인 후 사이트 목록 최신화(막힌 곳 자동 탈락)
    add_log(f'[자동파이프라인] 처리 {done} · 가입 {signed} · 등록 {registered}'+(f' · 자동탈락 {dropped}' if dropped else ''))
    return {'ok':True,'processed':done,'signed_up':signed,'registered':registered,'dropped':dropped,'results':results}

def _cand_set(cid, **fields):
    rejected_dom=None
    with _cand_lock:
        cands=load_cands()
        for c in cands:
            if c.get('id')==cid:
                c.update(fields)
                # 탈락 처리되면 도메인을 영구 탈락 목록에 기록(다음 발굴에서 제외)
                if fields.get('status')=='rejected':
                    rejected_dom=(c.get('domain') or _domain_of(c.get('url','')))
        save_cands(cands)
    if rejected_dom:
        add_rejected_domains(rejected_dom, fields.get('reject_reason',''))

def _promoted_site_id(cand):
    dom=(cand.get('domain') or _domain_of(cand.get('url',''))).lower()
    for s in load_sites():
        if _domain_of(s.get('site_url',''))==dom: return s.get('id')
    return None

def discover_loop():
    """24시간 자동 발굴 — 목표치까지 천천히 채우고 남는 시간엔 검수."""
    while True:
        try:
            cfg=load_config()
            provider=(cfg.get('search_provider') or 'brave').lower()
            ready=bool(cfg.get('brave_api_key')) if provider=='brave' else bool(cfg.get('google_api_key') and cfg.get('google_cx'))
            if cfg.get('discover_enabled') and ready:
                discover_once(cfg,max_queries=4)
            else:
                # 발굴 꺼져 있어도 미검수 후보는 계속 처리
                if any(not c.get('screened') for c in load_cands()): screen_pending(10)
            # 완전 자동 파이프라인: 검수완료 후보를 자동가입→실발행→자동등록까지 처리
            if cfg.get('auto_pipeline_enabled'):
                try:
                    limit=int(cfg.get('auto_pipeline_batch',3) or 3)
                    auto_pipeline_once(limit=limit)
                except Exception as e:
                    add_log(f'[자동파이프라인 오류] {str(e)[:100]}')
            else:
                # 파이프라인 꺼져 있어도 사이트 목록은 상시 최신화(막힌 곳 자동 탈락)
                try: reconcile_sites()
                except Exception as e: add_log(f'[자동정리 오류] {str(e)[:80]}')
        except Exception as e:
            add_log(f'[발굴 루프 오류] {str(e)[:100]}')
        time.sleep(900)   # 15분마다

def member_paid_now(m):
    """이번 달 납부 완료 여부."""
    pays=m.get('payments',{}) or {}
    return bool((pays.get(_cur_month(),{}) or {}).get('paid'))

def member_runnable(m,cfg):
    """이 회원의 스케줄을 지금 돌려도 되는가. (사유, 가능여부)"""
    if m.get('status','active')!='active': return False,'정지 회원'
    if not m.get('sched_enabled'): return False,'스케줄 꺼짐'
    if not (m.get('sched_times') or []): return False,'시간대 미설정'
    if cfg.get('block_unpaid') and not member_paid_now(m): return False,'미납 — 자동 정지'
    return True,''

def member_sites(m):
    """회원에게 배정된 발행 가능 사이트(미배정이면 전체 허용 사이트)."""
    ids=m.get('site_ids') or []
    sites=[s for s in load_sites() if (not ids or s.get('id') in ids)]
    return [s for s in sites if is_publishable(s)]

def member_keywords(m):
    """회원 전용 키워드 풀(없으면 공용 풀)."""
    kw=m.get('keywords') or []
    return kw if kw else load_keywords()

def run_member_job(mid, minute_key):
    """회원 1명의 스케줄 1회 실행 — 지터 대기 후 발행 큐 등록."""
    try:
        cfg=load_config()
        m=next((x for x in load_members() if x.get('id')==mid),None)
        if not m: return
        jitter=int(m.get('jitter',0) or 0)
        if jitter>0:
            time.sleep(random.randint(0,jitter*60))   # 시간 분산: 동시 폭주 방지
        m=next((x for x in load_members() if x.get('id')==mid),None)  # 대기 중 변경 반영
        if not m: return
        ok,why=member_runnable(m,cfg)
        if not ok:
            add_log(f'[회원스케줄:{m.get("name") or m.get("biz")}] 건너뜀 — {why}'); return
        sites=member_sites(m); pool=member_keywords(m)
        nm=m.get('name') or m.get('biz') or mid
        if not sites: add_log(f'[회원스케줄:{nm}] 발행 가능 사이트 없음'); return
        if not pool: add_log(f'[회원스케줄:{nm}] 키워드 없음'); return
        cnt=max(1,int(m.get('per_run',1) or 1))
        total=0
        for _ in range(cnt):
            kw=pick_keywords(pool,cfg)
            total+=enqueue_generated(sites,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),
                                            '브랜드':kw.get('브랜드','') or cfg.get('brand','')},cfg,
                                     {'region':kw.get('지역',''),'service':kw.get('서비스',''),'member':nm})[0]
        # 실행 기록
        mem=load_members()
        for x in mem:
            if x.get('id')==mid:
                x['last_run']=_kst_now().strftime('%Y-%m-%d %H:%M')
                x['run_count']=int(x.get('run_count',0) or 0)+1
                x['last_run_min']=minute_key
        save_members(mem)
        add_log(f'[회원스케줄:{nm}] {total}건 큐 등록 (사이트 {len(sites)}개 × {cnt}회)')
        if total and not wk_active: start_workers(cfg.get('workers',2))
        if cfg.get('notify_done'): send_telegram(cfg,f'⏰ {nm} 자동발행 {total}건 등록')
    except Exception as e:
        add_log(f'[회원스케줄 오류] {str(e)[:100]}')

def member_scheduler_loop():
    """jump 방식: 회원마다 설정한 시간대에 서버가 24시간 자동 구동."""
    last_min=None
    while True:
        try:
            now=_kst_now(); hm=now.strftime('%H:%M'); wd=now.weekday()
            minute_key=now.strftime('%Y-%m-%d %H:%M')
            if minute_key!=last_min:
                last_min=minute_key
                cfg=load_config()
                for m in load_members():
                    if hm not in (m.get('sched_times') or []): continue
                    days=m.get('sched_days') or []
                    if days and wd not in days: continue
                    if m.get('last_run_min')==minute_key: continue
                    ok,why=member_runnable(m,cfg)
                    if not ok:
                        add_log(f'[회원스케줄:{m.get("name") or m.get("biz")}] 시간 도달했으나 건너뜀 — {why}')
                        continue
                    # 즉시 선점 기록(중복 실행 방지) 후 백그라운드 실행
                    mem=load_members()
                    for x in mem:
                        if x.get('id')==m.get('id'): x['last_run_min']=minute_key
                    save_members(mem)
                    threading.Thread(target=run_member_job,args=(m.get('id'),minute_key),daemon=True).start()
        except Exception as e:
            add_log(f'[회원스케줄러 오류] {str(e)[:100]}')
        time.sleep(20)

def settle_summary():
    """이번 달 정산 요약(활성 회원 기준)."""
    cm=_cur_month(); mem=load_members()
    active=[m for m in mem if m.get('status','active')=='active']
    billed=sum(member_fee(m) for m in active)
    paid=sum(member_fee(m) for m in active if ((m.get('payments',{}) or {}).get(cm,{}) or {}).get('paid'))
    unpaid_list=[{'id':m.get('id'),'name':m.get('name') or m.get('biz',''),'biz':m.get('biz',''),
                  'fee':member_fee(m)} for m in active
                 if not ((m.get('payments',{}) or {}).get(cm,{}) or {}).get('paid')]
    return {'month':cm,'members':len(mem),'active':len(active),
            'billed':billed,'paid':paid,'unpaid':billed-paid,
            'unpaid_count':len(unpaid_list),'unpaid_list':unpaid_list}

sched_active=True
def scheduler_loop():
    """1분 단위로 스케줄 확인 → 조건 맞으면 발행 큐 등록. days: 0=월..6=일, 빈 리스트=매일."""
    last_min=None
    while sched_active:
        try:
            now=_kst_now(); hm=now.strftime('%H:%M'); wd=now.weekday(); today=now.strftime('%Y-%m-%d')
            minute_key=now.strftime('%Y-%m-%d %H:%M')
            if minute_key!=last_min:
                last_min=minute_key
                cfg=load_config(); changed=False
                # ---- 매일 자동 백업(텔레그램) ----
                bt=(cfg.get('backup_time') or '').strip()
                if bt and hm==bt:
                    st=load_json(DATA_DIR/'backup.state',{})
                    if st.get('date')!=today:
                        save_json(DATA_DIR/'backup.state',{'date':today})
                        try: do_backup(cfg,'자동')
                        except Exception as e: add_log(f'[백업 오류] {str(e)[:80]}')
                scheds=load_scheds()
                for sc in scheds:
                    if not sc.get('enabled',True): continue
                    if hm not in (sc.get('times') or []): continue
                    days=sc.get('days') or []
                    if days and wd not in days: continue
                    if sc.get('last_run_min')==minute_key: continue
                    sc['last_run_min']=minute_key; sc['last_run']=now.strftime('%Y-%m-%d %H:%M'); changed=True
                    # 예약에 키워드를 직접 넣지 않으면 공용 키워드 풀을 사용한다.
                    # 완료 키는 schedules.json에 저장하므로 재시작해도 이미 사용한 키워드를 반복하지 않는다.
                    source_sets=sc.get('keyword_sets') or load_keywords()
                    completed=set(sc.get('completed_keys') or [])
                    remaining=[k for k in source_sets if schedule_keyword_key(k) not in completed]
                    if not remaining:
                        sc['enabled']=False; sc['completed_at']=now.strftime('%Y-%m-%d %H:%M')
                        changed=True
                        add_log(f'[예약 완료:{sc.get("name")}] 모든 키워드 사용 완료 — 스케줄 자동 종료')
                        if cfg.get('notify_done'):
                            send_telegram(cfg,f'🏁 {sc.get("name")} 모든 키워드 완료 · 스케줄 자동 종료')
                        continue
                    cnt=max(1,int(sc.get('count',1) or 1))
                    ksets=remaining[:cnt]
                    sids=sc.get('site_ids') or []
                    sites=[s for s in load_sites() if not sids or s.get('id') in sids]
                    allowed=[s for s in sites if is_permitted(s)]
                    if not allowed:
                        add_log(f'[예약:{sc.get("name")}] 허용 사이트 없음 — 건너뜀'); continue
                    n=0
                    for ks in ksets:
                        kw={'지역':ks.get('지역',''),'서비스':ks.get('서비스',''),'브랜드':ks.get('브랜드','')}
                        # 사이트마다 유니크 본문 생성(중복 방지)
                        added=enqueue_generated(allowed,kw,cfg,{'region':kw['지역'],'service':kw['서비스']})[0]
                        n+=added
                        if added:
                            completed.add(schedule_keyword_key(ks))
                    sc['completed_keys']=sorted(completed); changed=True
                    left=sum(1 for k in source_sets if schedule_keyword_key(k) not in completed)
                    if left==0:
                        sc['enabled']=False; sc['completed_at']=now.strftime('%Y-%m-%d %H:%M')
                        add_log(f'[예약 완료:{sc.get("name")}] 모든 키워드 큐 등록 완료 — 스케줄 자동 종료')
                        if cfg.get('notify_done'):
                            send_telegram(cfg,f'🏁 {sc.get("name")} 모든 키워드 완료 · 스케줄 자동 종료')
                    add_log(f'[예약 실행:{sc.get("name")}] {n}건 큐 등록')
                    if n and not wk_active: start_workers(cfg.get('workers',2))
                if changed: save_scheds(scheds)
        except Exception as e:
            add_log(f'[스케줄러 오류] {str(e)[:80]}')
        time.sleep(20)

# ==================== 텔레그램 폰 제어 (명령 수신) ====================
TG_HELP=('📱 찌라시 봇 명령어\n'
         '/상태 — 워커·큐 현황\n'
         '/오늘 — 오늘 발행 통계\n'
         '/발행 — 키워드 풀에서 랜덤 뽑아 전체 발행(원클릭)\n'
         '/발행 지역,서비스[,브랜드] — 지정 키워드로 즉시 발행\n'
         '/정지 · /재개 — 워커 일시정지/재개\n'
         '/백업 — 지금 백업 파일 전송\n'
         '/검증 — 발행글 생존 확인 실행')

def handle_tg_command(cfg,text):
    t=(text or '').strip(); low=t.lower()
    def reply(m): send_telegram(cfg,m)
    if low in ('/help','/start','/도움말') or t in ('도움말','명령어'):
        reply(TG_HELP); return
    if t.startswith('/상태') or low.startswith('/status'):
        reply(f'📊 상태\n워커: {"ON" if wk_active else "OFF"}{" (일시정지)" if wk_paused else ""}\n'
              f'큐: {post_queue.qsize()} · 재시도대기: {len(_retry_jobs)}\n'
              f'성공 {wk_stats["success"]} · 실패 {wk_stats["fail"]} · 스킵 {wk_stats["skipped"]}'); return
    if t.startswith('/오늘') or low.startswith('/today'):
        today=_kst_now().strftime('%Y-%m-%d'); h=load_json(HISTORY_FILE,[])
        th=[x for x in h if (x.get('time') or '').startswith(today)]
        ok=sum(1 for x in th if x.get('status')=='done'); fl=sum(1 for x in th if x.get('status')=='failed')
        reply(f'📅 오늘({today})\n성공 {ok} · 실패 {fl} · 전체 {len(th)}건'); return
    if t.startswith('/정지') or low.startswith('/pause'):
        pause_workers(); reply('⏸ 워커 일시정지'); return
    if t.startswith('/재개') or low.startswith('/resume'):
        resume_workers()
        if not wk_active: start_workers(cfg.get('workers',2))
        reply('▶️ 워커 재개'); return
    if t.startswith('/백업') or low.startswith('/backup'):
        reply('📦 백업 생성 중...'); okb,err=do_backup(cfg,'텔레그램'); reply('✅ 백업 전송 완료' if okb else '❌ 실패: '+err); return
    if t.startswith('/검증') or low.startswith('/verify'):
        reply('🔎 생존 확인 실행...(백그라운드)')
        threading.Thread(target=verify_once,kwargs={'limit':40},daemon=True).start(); return
    if t.startswith('/발행') or low.startswith('/post'):
        parts_split=t.split(None,1)
        arg=(parts_split[1] if len(parts_split)>1 else '').replace('，',',')
        parts=[x.strip() for x in arg.split(',') if x.strip()]
        sites=[s for s in load_sites() if is_publishable(s)]
        if not sites: reply('⚠️ 발행 가능한 사이트가 없습니다(허용·캡차없음)'); return
        # 인자 없으면 키워드 풀에서 랜덤(폰 원클릭)
        if len(parts)<2:
            pool=load_keywords()
            if not pool: reply('형식: /발행 지역,서비스[,브랜드]\n또는 키워드 풀을 등록하면 /발행 만으로 랜덤 발행'); return
            kw=pick_keywords(pool,cfg)
        else:
            kw={'지역':parts[0],'서비스':parts[1],'브랜드':(parts[2] if len(parts)>2 and parts[2] else cfg.get('brand',''))}
        try:
            n=enqueue_generated(sites,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),'브랜드':kw.get('브랜드','')},cfg,{'region':kw.get('지역',''),'service':kw.get('서비스','')})[0]
            if n and not wk_active: start_workers(cfg.get('workers',2))
            reply(f'📤 {n}건 큐 등록 (사이트별 유니크)\n키워드: {kw.get("지역","")} {kw.get("서비스","")}')
        except Exception as e:
            reply('❌ 생성 실패: '+str(e)[:100])
        return
    reply('알 수 없는 명령입니다. /help')

_tg_offset={'v':0}
def telegram_loop():
    """텔레그램 getUpdates 롱폴링. 등록된 chat_id 만 명령 처리."""
    import time as _t
    while True:
        cfg=load_config()
        if not (cfg.get('telegram_control') and cfg.get('telegram_token') and cfg.get('telegram_chat_id')):
            _t.sleep(5); continue
        tok=cfg['telegram_token']; chat=str(cfg['telegram_chat_id'])
        try:
            import requests as _rq
            r=_rq.get(f"https://api.telegram.org/bot{tok}/getUpdates",
                      params={'offset':_tg_offset['v']+1,'timeout':30},timeout=40)
            j=r.json()
            for up in j.get('result',[]):
                _tg_offset['v']=up['update_id']
                msg=up.get('message') or up.get('edited_message') or {}
                frm=str((msg.get('chat') or {}).get('id',''))
                text=(msg.get('text') or '').strip()
                if not text or frm!=chat: continue   # 인증: 등록된 챗만
                try: handle_tg_command(cfg,text)
                except Exception as e: add_log(f'[텔레그램 처리 오류] {str(e)[:80]}')
        except Exception:
            _t.sleep(5)

# ==================== 사이트 헬스체크 (HTTP) ====================
def site_health(site):
    """로그인/글쓰기 페이지가 살아있는지 가벼운 HTTP 점검(셀레니움 없이)."""
    import requests as _rq
    url=site.get('site_url','').rstrip('/')
    m=re.match(r'(https?://[^/]+)',url); base=m.group(1) if m else url
    bo=site.get('bo_table','free')
    out={'reachable':False,'login_form':False,'write_page':False,'note':''}
    try:
        r=_rq.get(base,timeout=12,verify=False,headers={'User-Agent':'Mozilla/5.0'})
        out['reachable']=r.status_code<500
        try:
            lp=_rq.get(base+'/bbs/login.php',timeout=12,verify=False,headers={'User-Agent':'Mozilla/5.0'})
            out['login_form']=("mb_id" in lp.text and "mb_password" in lp.text)
        except Exception: pass
        try:
            wp=_rq.get(base+f'/bbs/write.php?bo_table={bo}',timeout=12,verify=False,headers={'User-Agent':'Mozilla/5.0'})
            out['write_page']=(wp.status_code<500 and ('wr_subject' in wp.text or 'wr_content' in wp.text or '로그인' in wp.text))
        except Exception: pass
    except Exception as e:
        out['note']=str(e)[:80]
    out['ok']=out['reachable'] and (out['login_form'] or out['write_page'])
    return out

# ==================== 통계 집계 ====================
def compute_stats():
    h=load_json(HISTORY_FILE,[])
    total=len(h); ok=sum(1 for x in h if x.get('status')=='done'); fail=sum(1 for x in h if x.get('status')=='failed')
    skip=sum(1 for x in h if x.get('status')=='skipped')
    alive=sum(1 for x in h if x.get('alive')=='yes'); dead=sum(1 for x in h if x.get('alive')=='no')
    alive_rate=round(alive/(alive+dead)*100,1) if (alive+dead) else 0.0
    reasons={}   # 실패 원인 분류 집계
    by_site={}; by_day={}
    for x in h:
        sn=x.get('site_name') or '(미상)'; d=(x.get('time') or '')[:10]; st=x.get('status')
        bs=by_site.setdefault(sn,{'done':0,'failed':0,'other':0,'alive':0,'dead':0})
        bs['done' if st=='done' else 'failed' if st=='failed' else 'other']+=1
        if x.get('alive')=='yes': bs['alive']+=1
        elif x.get('alive')=='no': bs['dead']+=1
        bd=by_day.setdefault(d,{'done':0,'failed':0})
        if st in ('done','failed'): bd[st]+=1
        if st=='failed':
            rk=x.get('fail_reason_ko') or '기타'; reasons[rk]=reasons.get(rk,0)+1
    rate=round(ok/(ok+fail)*100,1) if (ok+fail) else 0.0
    days=sorted(by_day.items())[-14:]
    top_sites=sorted(by_site.items(),key=lambda kv:-(kv[1]['done']+kv[1]['failed']))[:12]
    return {'total':total,'ok':ok,'fail':fail,'skip':skip,'rate':rate,
            'alive':alive,'dead':dead,'alive_rate':alive_rate,
            'reasons':[{'reason':k,'n':v} for k,v in sorted(reasons.items(),key=lambda kv:-kv[1])],
            'by_day':[{'day':d,'done':v['done'],'failed':v['failed']} for d,v in days],
            'by_site':[{'site':s,**v} for s,v in top_sites]}

# ==================== Flask ====================
app=Flask(__name__)

def _get_secret():
    """세션 서명키 — 하드코딩 대신 env 또는 persist 파일에서 로드 (없으면 랜덤 생성)."""
    env=os.environ.get('CHIRASHI_SECRET')
    if env: return env
    kf=DATA_DIR/'secret.key'
    try:
        if kf.exists(): return kf.read_text().strip()
        s=secrets.token_hex(32); kf.write_text(s); return s
    except: return secrets.token_hex(32)
app.secret_key=_get_secret()
# 세션 쿠키 보안: JS 접근 차단·HTTPS 전용·크로스사이트 제한
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE='Lax', SESSION_COOKIE_SECURE=True)

# ---- 크롤/스크래핑/색인/임베드 차단 ----
BLOCK_UA=['bot','crawler','spider','scrapy','ahrefs','semrush','mj12','dotbot','python-requests',
          'httpclient','curl','wget','headless','phantom','slurp','baiduspider','yandex','censys','zgrab']
@app.after_request
def _sec_headers(resp):
    # 검색엔진 색인·아카이브 금지 + 임베드(아이프레임) 차단 + 스니핑·캐시 방지
    resp.headers['X-Robots-Tag']='noindex, nofollow, noarchive, nosnippet, noimageindex'
    resp.headers['X-Frame-Options']='DENY'
    resp.headers['Content-Security-Policy']="frame-ancestors 'none'"
    resp.headers['X-Content-Type-Options']='nosniff'
    resp.headers['Referrer-Policy']='no-referrer'
    resp.headers['Cache-Control']='no-store'
    resp.headers['Permissions-Policy']='geolocation=(), camera=(), microphone=()'
    return resp

@app.errorhandler(Exception)
def _backend_error(err):
    """API 예외는 HTML 오류 페이지 대신 일관된 JSON으로 기록·반환."""
    if isinstance(err,HTTPException): return err
    try: add_log(f'[API 오류] {request.method} {request.path} - {type(err).__name__}: {str(err)[:160]}')
    except Exception: pass
    if request.path.startswith('/api/'):
        return jsonify({'ok':False,'error':'서버 처리 중 오류가 발생했습니다'}),500
    return ('Internal Server Error',500)

def auth():
    return session.get('login',False)

# ---- 로그인 브루트포스 방어 (IP당 5분 내 6회 실패 시 잠금) ----
_login_hits=defaultdict(list); _login_lock=threading.Lock()
def _client_ip():
    xff=request.headers.get('X-Forwarded-For','')
    return (xff.split(',')[0].strip() if xff else request.remote_addr) or '?'
def login_allowed(ip):
    now=time.time()
    with _login_lock:
        arr=[t for t in _login_hits[ip] if now-t<300]; _login_hits[ip]=arr
        return len(arr)<6
def login_fail(ip):
    with _login_lock: _login_hits[ip].append(time.time())
def check_pw(pw):
    """평문/해시 모두 지원 + env(CHIRASHI_PASSWORD) 우선."""
    stored=os.environ.get('CHIRASHI_PASSWORD') or load_config().get('password','admin1234')
    if isinstance(stored,str) and (stored.startswith('pbkdf2:') or stored.startswith('scrypt:')):
        try: return check_password_hash(stored,pw)
        except: return False
    return pw==stored

@app.route('/robots.txt')
def robots():
    from flask import Response
    return Response("User-agent: *\nDisallow: /\n",mimetype='text/plain')

@app.before_request
def chk():
    # 알려진 크롤러/스크래퍼 User-Agent 즉시 차단(로그인·업데이트 제외)
    if request.path not in ['/robots.txt','/api/admin/update']:
        ua=(request.headers.get('User-Agent','') or '').lower()
        if not ua or any(b in ua for b in BLOCK_UA):
            return ('Forbidden',403)
    if request.path=='/robots.txt': return
    if request.path.startswith('/static'): return
    if request.path in ['/login','/logout']: return
    if request.path=='/api/admin/update': return  # 자체 토큰 인증
    if not auth():
        if request.path.startswith('/api/') or request.is_json:
            return jsonify({'ok':False,'error':'로그인이 만료되었습니다'}),401
        return redirect('/login')

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        ip=_client_ip()
        if not login_allowed(ip):
            return R('로그인',error='로그인 시도 초과 - 잠시 후 다시 시도하세요')
        if check_pw(request.form.get('pw','')):
            session['login']=True; return redirect('/')
        login_fail(ip)
        return R('로그인',error='비밀번호 틀림')
    return R('로그인',error='')

@app.route('/logout')
def logout():
    session.clear(); return redirect('/login')

@app.route('/')
def index():
    sites=load_sites()
    return R('대시보드',cfg=load_config(),sites=sites,
             publish_sites=[s for s in sites if is_publishable(s)],wk=wk_stats,wk_on=wk_active)

# API
@app.route('/api/generate',methods=['POST'])
def api_gen():
    d=request.get_json(silent=True) or {}; kw=d.get('keywords',{}); cfg=load_config()
    html,title=generate_article(kw,cfg)
    return jsonify({'ok':True,'title':title,'content':html})

# ---- 이미지 URL 풀 (본문 삽입용) ----
@app.route('/api/images',methods=['GET','POST','DELETE'])
def api_images():
    if request.method=='POST':
        d=request.get_json() or {}
        urls=d.get('urls')
        if urls is None:
            urls=[x.strip() for x in (d.get('text','') or '').splitlines() if x.strip()]
        urls=[u for u in urls if u.startswith('http')]
        if d.get('append'):
            urls=load_image_urls()+urls
        urls=list(dict.fromkeys(urls))   # 중복 제거(순서보존)
        save_image_urls(urls); return jsonify({'ok':True,'count':len(urls)})
    if request.method=='DELETE':
        save_image_urls([]); return jsonify({'ok':True})
    return jsonify(load_image_urls())

@app.route('/media/<path:filename>')
def image_media(filename):
    return send_from_directory(UPLOAD_DIR,filename,conditional=True,max_age=86400)

@app.route('/api/images/upload',methods=['POST'])
def api_images_upload():
    files=request.files.getlist('files')
    if not files: return jsonify({'ok':False,'error':'이미지 파일을 선택하세요'}),400
    saved=[]
    for f in files[:20]:
        original=secure_filename(f.filename or '')
        ext=Path(original).suffix.lower()
        if ext not in IMAGE_EXTENSIONS: continue
        f.stream.seek(0,2); size=f.stream.tell(); f.stream.seek(0)
        if size<=0 or size>10*1024*1024: continue
        stem=secure_filename(Path(original).stem)[:60] or 'image'
        name=f'{datetime.now().strftime("%Y%m%d_%H%M%S")}_{secrets.token_hex(4)}_{stem}{ext}'
        f.save(UPLOAD_DIR/name); saved.append(name)
    if not saved: return jsonify({'ok':False,'error':'JPG·PNG·GIF·WEBP만 가능하며 파일당 최대 10MB입니다'}),400
    return jsonify({'ok':True,'count':len(saved),'files':saved})

@app.route('/api/images/file',methods=['DELETE'])
def api_images_file_delete():
    name=Path((request.get_json(silent=True) or {}).get('name','')).name
    p=UPLOAD_DIR/name
    if not name or p.suffix.lower() not in IMAGE_EXTENSIONS or not p.exists():
        return jsonify({'ok':False,'error':'파일을 찾을 수 없습니다'}),404
    p.unlink(); return jsonify({'ok':True})

@app.route('/api/images/files')
def api_images_files(): return jsonify(uploaded_images())

# ---- 도메인 발굴 후보 (승인해야만 사이트 목록에 투입) ----
@app.route('/api/candidates',methods=['GET','DELETE'])
def api_candidates():
    if request.method=='DELETE':
        d=request.get_json() or {}
        if d.get('id'):
            save_cands([c for c in load_cands() if c.get('id')!=d['id']])
        elif d.get('clear')=='rejected':
            save_cands([c for c in load_cands() if c.get('status')!='rejected'])
        else:
            save_cands([])
        return jsonify({'ok':True})
    cands=load_cands()
    # 제거된 메일초안/문의함 기능의 기존 데이터도 함께 정리한다.
    legacy_changed=False
    for c in cands:
        if c.pop('mail_draft',None) is not None: legacy_changed=True
        if c.get('status')=='contacted': c['status']='ready'; legacy_changed=True
    if legacy_changed: save_cands(cands)
    order={'ready':0,'new':1,'approved':2,'rejected':3}
    cands.sort(key=lambda c:(order.get(c.get('status'),9),-int(c.get('score',0) or 0)))
    st=load_json(DISCO_FILE,{})
    summary={'total':len(cands),
             'ready':sum(1 for c in cands if c.get('status')=='ready'),
             'new':sum(1 for c in cands if c.get('status')=='new'),
             'approved':sum(1 for c in cands if c.get('status')=='approved'),
             'rejected':sum(1 for c in cands if c.get('status')=='rejected'),
             'today_queries':st.get('queries',0),'today_found':st.get('found',0),
             'date':st.get('date','')}
    return jsonify({'candidates':cands[:400],'summary':summary})

@app.route('/api/candidates/discover',methods=['POST'])
def api_cand_discover():
    d=request.get_json() or {}; cfg=load_config()
    provider=(cfg.get('search_provider') or 'brave').lower()
    if provider=='brave' and not cfg.get('brave_api_key'):
        return jsonify({'ok':False,'error':'설정 탭에서 Brave Search API 키를 먼저 입력하세요'})
    if provider=='google' and not (cfg.get('google_api_key') and cfg.get('google_cx')):
        return jsonify({'ok':False,'error':'Google API 키와 검색엔진ID(cx)가 필요합니다'})
    try:
        return jsonify(discover_once(cfg,max_queries=int(d.get('queries',5) or 5)))
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:200]})

@app.route('/api/candidates/manual',methods=['POST'])
def api_cand_manual():
    """수동 URL 붙여넣기 → 후보 등록 + 검수 (API 없이도 사용 가능)."""
    d=request.get_json() or {}; cfg=load_config()
    urls=[x.strip() for x in (d.get('urls','') or '').splitlines() if x.strip().startswith('http')]
    if not urls: return jsonify({'ok':False,'error':'http로 시작하는 URL을 한 줄에 하나씩 넣으세요'})
    n=add_candidates_from([{'url':u} for u in urls],cfg,source='manual')
    threading.Thread(target=screen_pending,kwargs={'limit':len(urls)},daemon=True).start()
    return jsonify({'ok':True,'added':n,'screening':True})

@app.route('/api/candidates/screen',methods=['POST'])
def api_cand_screen():
    d=request.get_json() or {}
    if d.get('id') or d.get('rescreen'):
        with _cand_lock:
            cands=load_cands()
            for c in cands:
                if d.get('rescreen') and c.get('status')!='approved': c['screened']=False
                elif c.get('id')==d.get('id'): c['screened']=False
            save_cands(cands)
    n=screen_pending(1 if d.get('id') else int(d.get('limit',20) or 20))
    return jsonify({'ok':True,'screened':n})

@app.route('/api/candidates/status',methods=['POST'])
def api_cand_status():
    d=request.get_json() or {}
    with _cand_lock:
        cands=load_cands()
        for c in cands:
            if c.get('id')==d.get('id'):
                if 'status' in d: c['status']=d['status']
                if 'note' in d: c['note']=d['note']
        save_cands(cands)
    return jsonify({'ok':True})

@app.route('/api/candidates/approve/<cid>',methods=['POST'])
def api_cand_approve(cid):
    """후보를 사이트 목록에 직접 등록. 근거 메모는 선택사항."""
    d=request.get_json() or {}
    note=(d.get('permission_note','') or '').strip() or '후보 목록에서 직접 등록'
    c=next((x for x in load_cands() if x.get('id')==cid),None)
    if not c: return jsonify({'ok':False,'error':'후보 없음'})
    with POST_LOCK:
        sites=load_sites()
        if any(_domain_of(s.get('site_url',''))==c.get('domain') for s in sites):
            return jsonify({'ok':False,'error':'이미 등록된 도메인'})
        sites.append({'id':secrets.token_hex(6),'site_url':(c.get('base') or c.get('url','')).rstrip('/'),
                  'platform':c.get('platform','auto') if c.get('platform') in ('gnuboard','cafe24') else 'auto',
                  'mb_id':d.get('mb_id',''),'mb_pass':d.get('mb_pass',''),
                  'bo_table':(d.get('bo_table') or c.get('bo_table') or 'free'),
                  'name':(d.get('name') or c.get('domain','')),
                  'permission':True,'permission_note':note,
                  'registration_source':'candidate_registered','daily_limit':3,'min_interval_minutes':60,
                  'permission_date':_kst_now().strftime('%Y-%m-%d'),
                  'has_captcha':bool(c.get('captcha')),
                      'status':'idle','added':_kst_now().strftime('%m/%d %H:%M')})
        save_sites(sites)
    with _cand_lock:
        cands=load_cands()
        for x in cands:
            if x.get('id')==cid: x['status']='approved'; x['approved_at']=_kst_now().strftime('%Y-%m-%d %H:%M')
        save_cands(cands)
    add_log(f'[후보 등록] {c.get("domain")} → 사이트 목록 등록')
    return jsonify({'ok':True})

@app.route('/api/candidates/verified',methods=['POST'])
def api_cand_verified():
    """실제 게시까지 검증된 후보를 즉시 사이트 목록으로 승격하고 변동 규칙을 저장."""
    d=request.get_json(silent=True) or {}
    result_url=(d.get('result_url') or '').strip(); write_url=(d.get('write_url') or '').strip()
    # 제출 버튼이 눌렸다는 사실만으로 성공 처리하지 않는다. 결과 URL이 없거나
    # 게시물 후속 검색에서 발견되지 않으면 후보/사이트를 즉시 탈락시킨다.
    post_found=d.get('post_found',True) is not False
    if not result_url.startswith(('http://','https://')) or not post_found:
        source_url=write_url or (d.get('url') or '').strip()
        domain=urllib.parse.urlsplit(source_url).netloc.lower() if source_url else (d.get('domain') or '').lower()
        now=_kst_now().strftime('%Y-%m-%d %H:%M')
        reason='결과 URL 없음' if not result_url.startswith(('http://','https://')) else '게시물 검색 결과 없음'
        with _cand_lock:
            cands=load_cands()
            for c in cands:
                if (domain and c.get('domain','').lower()==domain) or (d.get('candidate_id') and c.get('id')==d.get('candidate_id')):
                    c.update({'status':'rejected','reject_reason':reason,'write_test_status':'failed',
                              'verified_at':now,'verified_post_url':''})
            save_cands(cands)
        if domain:
            with POST_LOCK:
                sites=load_sites()
                for site in sites:
                    if _domain_of(site.get('site_url',''))==domain:
                        site.update({'status':'rejected','permission':False,'write_test_status':'failed',
                                     'verification_fail_reason':'','verified_post_url':'',
                                     'last_structure_check':now})
                save_sites(sites)
        add_log(f'[실게시 검증 탈락] {domain or "도메인 미확인"} · {reason}')
        return jsonify({'ok':False,'rejected':True,'error':reason,'domain':domain}),409
    rp=urllib.parse.urlsplit(result_url); wp=urllib.parse.urlsplit(write_url or result_url)
    if not rp.netloc or (write_url and rp.netloc.lower()!=wp.netloc.lower()):
        return jsonify({'ok':False,'error':'글쓰기 URL과 결과 URL의 도메인이 다릅니다'}),400
    qs=urllib.parse.parse_qs(rp.query); wqs=urllib.parse.parse_qs(wp.query)
    bo=(d.get('bo_table') or (wqs.get('bo_table') or qs.get('bo_table') or ['free'])[0]).strip()
    domain=rp.netloc.lower(); base=f'{rp.scheme}://{rp.netloc}'
    permission_note=(d.get('permission_note') or '').strip()
    caps=d.get('capabilities') if isinstance(d.get('capabilities'),dict) else {}
    now=_kst_now().strftime('%Y-%m-%d %H:%M')
    with POST_LOCK:
        sites=load_sites(); site=next((s for s in sites if _domain_of(s.get('site_url',''))==domain),None)
        created=site is None
        if created:
            site={'id':secrets.token_hex(6),'site_url':base,'platform':d.get('platform','gnuboard'),
                  'mb_id':'','mb_pass':'','bo_table':bo,'name':d.get('name') or domain,
                  'daily_limit':3,'min_interval_minutes':60,'status':'idle','added':_kst_now().strftime('%m/%d %H:%M')}
            sites.append(site)
        site.update({'bo_table':bo,'write_url':write_url,'verified_post_url':result_url,
                     'write_test_status':'passed','verified_at':now,'capabilities':caps,
                     'last_structure_check':now,'registration_source':'verified_test'})
        if permission_note:
            site.update({'permission':True,'permission_note':permission_note,
                         'permission_date':_kst_now().strftime('%Y-%m-%d')})
        else:
            site.setdefault('permission',False)
            site.setdefault('permission_note','실게시 검증 완료 · 홍보 허용 근거 확인 전 발행 잠금')
        save_sites(sites)
    with _cand_lock:
        cands=load_cands()
        for c in cands:
            if c.get('domain','').lower()==domain:
                c.update({'status':'approved','verified_at':now,'verified_post_url':result_url,
                          'capabilities':caps,'site_id':site['id']})
        save_cands(cands)
    add_log(f'[실게시 검증→자동등록] {domain} · {bo} · '+('발행 허용' if permission_note else '허용근거 대기'))
    return jsonify({'ok':True,'created':created,'site_id':site['id'],'permission':bool(site.get('permission'))})

@app.route('/api/candidates/export',methods=['GET'])
def api_cand_export():
    from flask import Response
    import io
    cands=load_cands()
    cols=[('domain','도메인'),('url','URL'),('platform','플랫폼'),('board_name','게시판명'),
          ('bo_table','게시판ID'),('score','점수'),('status','상태'),('promo_hint','홍보허용흔적'),
          ('parked','주차도메인'),('illegal','도박불법'),
          ('ad_banned','광고금지'),('captcha','캡차'),('login_required','로그인필요'),
          ('write_form','글쓰기폼'),('last_post_days','최근글(일)'),
          ('reject_reason','탈락사유'),('found_at','발견일')]
    try:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title='발굴후보'
        ws.append([c[1] for c in cols])
        for r in cands:
            ws.append([', '.join(r.get(c[0])) if isinstance(r.get(c[0]),list) else str(r.get(c[0],'') or '') for c in cols])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition':'attachment; filename=candidates.xlsx'})
    except Exception:
        import csv
        sio=io.StringIO(); w=csv.writer(sio); w.writerow([c[1] for c in cols])
        for r in cands: w.writerow([', '.join(r.get(c[0])) if isinstance(r.get(c[0]),list) else str(r.get(c[0],'') or '') for c in cols])
        return Response('﻿'+sio.getvalue(),mimetype='text/csv',
                        headers={'Content-Disposition':'attachment; filename=candidates.csv'})

# ---- 회원(고객) 관리 + 월 정산 ----
@app.route('/api/members',methods=['GET','POST','DELETE'])
def api_members():
    if request.method=='POST':
        d=request.get_json() or {}; mem=load_members()
        mid=d.get('id') or secrets.token_hex(6)
        ex=next((m for m in mem if m.get('id')==mid),None)
        rec=ex or {'id':mid,'payments':{},'join_date':_kst_now().strftime('%Y-%m-%d')}
        for k in ['name','biz','phone','memo','status']:
            if k in d: rec[k]=d[k]
        for k in ['plan_fee','addons','addon_fee','settle_day','jitter','per_run']:
            if k in d:
                try: rec[k]=int(d[k])
                except Exception: rec[k]=0
        # ---- 회원별 스케줄 (jump 방식: 계정당 개별 시간대) ----
        if 'sched_enabled' in d: rec['sched_enabled']=bool(d['sched_enabled'])
        if 'sched_times' in d:
            raw=d['sched_times']
            if isinstance(raw,str): raw=[x.strip() for x in raw.replace('，',',').split(',')]
            rec['sched_times']=sorted({t for t in (raw or []) if re.match(r'^\d{1,2}:\d{2}$',str(t).strip())})
        if 'sched_days' in d:
            try: rec['sched_days']=[int(x) for x in (d['sched_days'] or []) if str(x).isdigit()]
            except Exception: rec['sched_days']=[]
        if 'site_ids' in d: rec['site_ids']=list(d['site_ids'] or [])
        if 'keywords_csv' in d:
            rows=[]
            for line in (d.get('keywords_csv','') or '').splitlines():
                p=[x.strip() for x in line.split(',')]
                if len(p)>=2 and p[0] and p[1]:
                    rows.append({'지역':p[0],'서비스':p[1],'브랜드':(p[2] if len(p)>2 else '')})
            rec['keywords']=rows
        rec.setdefault('status','active'); rec.setdefault('plan_fee',30000)
        rec.setdefault('addon_fee',10000); rec.setdefault('addons',0); rec.setdefault('settle_day',1)
        rec.setdefault('sched_enabled',False); rec.setdefault('sched_times',[])
        rec.setdefault('sched_days',[]); rec.setdefault('site_ids',[])
        rec.setdefault('keywords',[]); rec.setdefault('jitter',5); rec.setdefault('per_run',1)
        if not ex: mem.append(rec)
        save_members(mem); return jsonify({'ok':True,'id':mid})
    if request.method=='DELETE':
        d=request.get_json() or {}
        save_members([m for m in load_members() if m.get('id')!=d.get('id')])
        return jsonify({'ok':True})
    return jsonify({'members':[member_view(m) for m in load_members()],'summary':settle_summary()})

@app.route('/api/members/pay',methods=['POST'])
def api_member_pay():
    """특정 회원의 특정 월 납부 상태 토글/설정."""
    d=request.get_json() or {}; mid=d.get('id'); month=d.get('month') or _cur_month()
    paid=bool(d.get('paid',True)); mem=load_members()
    for m in mem:
        if m.get('id')==mid:
            pays=m.get('payments') or {}
            if not isinstance(pays,dict): pays={}
            if paid:
                pays[month]={'paid':True,'paid_at':_kst_now().strftime('%Y-%m-%d %H:%M'),'amount':member_fee(m)}
            else:
                pays[month]={'paid':False}
            m['payments']=pays; save_members(mem)
            return jsonify({'ok':True})
    return jsonify({'ok':False,'error':'회원 없음'})

@app.route('/api/members/run/<mid>',methods=['POST'])
def api_member_run(mid):
    """회원 스케줄 지금 즉시 1회 실행(테스트)."""
    m=next((x for x in load_members() if x.get('id')==mid),None)
    if not m: return jsonify({'ok':False,'error':'회원 없음'})
    cfg=load_config()
    sites=member_sites(m); pool=member_keywords(m)
    if not sites: return jsonify({'ok':False,'error':'배정된 발행 가능 사이트가 없습니다'})
    if not pool: return jsonify({'ok':False,'error':'회원 전용 키워드도 공용 키워드도 비어있습니다'})
    if cfg.get('block_unpaid') and not member_paid_now(m) and m.get('status','active')=='active':
        return jsonify({'ok':False,'error':'미납 회원 — 설정에서 미납 자동정지를 끄거나 납부 처리 후 실행'})
    nm=m.get('name') or m.get('biz') or mid
    cnt=max(1,int(m.get('per_run',1) or 1)); total=0
    for _ in range(cnt):
        kw=pick_keywords(pool,cfg)
        total+=enqueue_generated(sites,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),
                                        '브랜드':kw.get('브랜드','') or cfg.get('brand','')},cfg,
                                 {'region':kw.get('지역',''),'service':kw.get('서비스',''),'member':nm})[0]
    mem=load_members()
    for x in mem:
        if x.get('id')==mid:
            x['last_run']=_kst_now().strftime('%Y-%m-%d %H:%M'); x['run_count']=int(x.get('run_count',0) or 0)+1
    save_members(mem)
    if total and not wk_active: start_workers(cfg.get('workers',2))
    return jsonify({'ok':True,'generated':total,'sites':len(sites),'runs':cnt})

@app.route('/api/members/export',methods=['GET'])
def api_members_export():
    from flask import Response
    import io
    mem=load_members(); cm=_cur_month()
    cols=[('name','이름/담당'),('biz','업소'),('phone','연락처'),('status','상태'),
          ('plan_fee','기본료'),('addons','추가광고'),('addon_fee','추가단가'),('fee','월청구'),
          ('paid','이번달납부'),('join_date','가입일'),('memo','메모')]
    rows=[member_view(m) for m in mem]
    try:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title=f'회원정산_{cm}'
        ws.append([c[1] for c in cols])
        for r in rows:
            ws.append([('O' if (c[0]=='paid' and r.get('paid')) else ('X' if c[0]=='paid' else str(r.get(c[0],'') or ''))) for c in cols])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition':f'attachment; filename=members_{cm}.xlsx'})
    except Exception:
        import csv
        sio=io.StringIO(); w=csv.writer(sio); w.writerow([c[1] for c in cols])
        for r in rows: w.writerow([('O' if (c[0]=='paid' and r.get('paid')) else ('X' if c[0]=='paid' else str(r.get(c[0],'') or ''))) for c in cols])
        return Response('﻿'+sio.getvalue(),mimetype='text/csv',
                        headers={'Content-Disposition':f'attachment; filename=members_{cm}.csv'})

# ---- 키워드 풀 (엑셀/CSV 저장 → 랜덤 치환) ----
@app.route('/api/keywords',methods=['GET','POST','DELETE'])
def api_keywords():
    if request.method=='POST':
        d=request.get_json() or {}
        rows=d.get('rows')
        if rows is None:
            rows=[]
            for line in (d.get('csv','') or '').splitlines():
                line=line.strip()
                if not line: continue
                p=[x.strip() for x in line.split(',')]
                if len(p)>=2 and p[0] and p[1]:
                    rows.append({'지역':p[0],'서비스':p[1],'브랜드':(p[2] if len(p)>2 else '')})
        if d.get('append'): rows=load_keywords()+rows
        save_keywords(rows); return jsonify({'ok':True,'count':len(rows)})
    if request.method=='DELETE':
        save_keywords([]); return jsonify({'ok':True})
    return jsonify(load_keywords())

@app.route('/api/workrooms',methods=['GET','POST','DELETE'])
def api_workrooms():
    """키워드 작업실 CRUD. 키워드 조합과 대상 사이트를 서로 독립 저장한다."""
    rooms=load_json(WORKROOMS_FILE,[])
    if request.method=='POST':
        d=request.get_json(silent=True) or {}; rid=str(d.get('id') or secrets.token_hex(6))
        room=next((x for x in rooms if x.get('id')==rid),None)
        if room is None:
            room={'id':rid,'created_at':_kst_now().strftime('%Y-%m-%d %H:%M')}; rooms.append(room)
        name=(d.get('name') or '').strip() or f'{len(rooms)}번 작업실'
        keyword_csv=str(d.get('keyword_csv') or '')
        rows=[]
        for line in keyword_csv.splitlines():
            p=[x.strip() for x in line.split(',')]
            if len(p)>=3 and all(p[:3]): rows.append(','.join(p[:3]))
        room.update({'name':name,'keyword_csv':'\n'.join(rows),'site_id':str(d.get('site_id') or ''),
                     'updated_at':_kst_now().strftime('%Y-%m-%d %H:%M')})
        save_json(WORKROOMS_FILE,rooms)
        return jsonify({'ok':True,'id':rid,'name':name,'count':len(rows)})
    if request.method=='DELETE':
        d=request.get_json(silent=True) or {}; rid=str(d.get('id') or '')
        save_json(WORKROOMS_FILE,[x for x in rooms if x.get('id')!=rid])
        return jsonify({'ok':True})
    return jsonify(rooms)

@app.route('/api/regions',methods=['GET'])
def api_regions():
    """전국 시도 → 시군구 → 읍면동 계층. 키워드 일괄 생성 도구에서 사용."""
    data=load_json(REGIONS_FILE,{})
    if not data: return jsonify({'error':'행정구역 데이터 없음'}),503
    # 현재 명칭 보정 및 단층 구조인 세종특별자치시 추가
    if '강원도' in data: data['강원특별자치도']=data.pop('강원도')
    if '전라북도' in data: data['전북특별자치도']=data.pop('전라북도')
    data.setdefault('세종특별자치시',{})['세종시']=[
        '조치원읍','연기면','연동면','부강면','금남면','장군면','연서면','전의면','전동면','소정면',
        '한솔동','새롬동','나성동','도담동','어진동','아름동','종촌동','고운동','보람동','대평동',
        '소담동','반곡동','해밀동','다정동','집현동']
    return jsonify(data)

@app.route('/api/keywords/upload',methods=['POST'])
def api_keywords_upload():
    f=request.files.get('file')
    if not f: return jsonify({'ok':False,'error':'파일 없음'}),400
    rows=[]
    try:
        from openpyxl import load_workbook
        import io
        wb=load_workbook(io.BytesIO(f.read()),read_only=True,data_only=True)
        ws=wb.active
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            vals=[('' if c is None else str(c)).strip() for c in row]
            if not vals or not vals[0]: continue
            if i==0 and ('지역' in vals[0] or '키워드' in vals[0]): continue  # 헤더 스킵
            if len(vals)>=2 and vals[0] and vals[1]:
                rows.append({'지역':vals[0],'서비스':vals[1],'브랜드':(vals[2] if len(vals)>2 else '')})
    except Exception as e:
        return jsonify({'ok':False,'error':f'엑셀 파싱 실패: {str(e)[:80]}'}),400
    if request.form.get('append')=='1': rows=load_keywords()+rows
    save_keywords(rows); return jsonify({'ok':True,'count':len(rows)})

@app.route('/api/generate/random',methods=['POST'])
def api_gen_random():
    d=request.get_json() or {}; cfg=load_config(); pool=load_keywords()
    if not pool: return jsonify({'ok':False,'error':'키워드 풀이 비어있습니다 (엑셀/CSV 등록)'})
    site_ids=d.get('site_ids',[]); count=max(1,int(d.get('count',1) or 1))
    # 발행 모드: site_ids(또는 count>1) 지정 시 랜덤 N개를 허용 사이트 큐에 등록
    if site_ids or count>1:
        sites=[s for s in load_sites() if not site_ids or s.get('id') in site_ids]
        allowed=[s for s in sites if is_permitted(s)]; blocked=len(sites)-len(allowed)
        if not allowed:
            return jsonify({'ok':False,'error':f'홍보 허용된 사이트가 없습니다 (미허용 {blocked}개 제외)'})
        total=0
        for _ in range(count):
            kw=pick_keywords(pool,cfg)
            # 사이트마다 유니크 본문 생성(중복 방지)
            total+=enqueue_generated(allowed,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),'브랜드':kw.get('브랜드','')},cfg,{'region':kw.get('지역',''),'service':kw.get('서비스','')})[0]
        if not wk_active: start_workers(cfg.get('workers',2))
        return jsonify({'ok':True,'generated':total,'blocked':blocked,'picks':count,'queued':wk_stats['queued']})
    # 미리보기 모드: 1개 생성해서 결과창에 표시
    kw=pick_keywords(pool,cfg)
    html,title=generate_article({'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),'브랜드':kw.get('브랜드','')},cfg)
    return jsonify({'ok':True,'title':title,'content':html,'picked':kw})

@app.route('/api/backup/now',methods=['POST'])
def api_backup_now():
    ok,err=do_backup(load_config(),'수동')
    return jsonify({'ok':ok,'error':('' if ok else err)})

@app.route('/api/diag',methods=['GET'])
def api_diag():
    """서버 환경 자가진단: 크롬·드라이버·실제 페이지 로드·리소스."""
    import shutil,subprocess,platform
    out={'time':_kst_now().strftime('%Y-%m-%d %H:%M:%S'),'python':sys.version.split()[0],
         'platform':platform.platform()[:60],'steps':[]}
    def step(name,ok,detail=''):
        out['steps'].append({'name':name,'ok':bool(ok),'detail':str(detail)[:200]})
    # 1) 크롬 바이너리
    cb=os.environ.get('CHROME_BIN') or ''
    cands=[cb,'/usr/bin/google-chrome','/usr/bin/chromium','/usr/bin/chromium-browser','/snap/bin/chromium']
    found=''
    for p in cands:
        if p and os.path.exists(p): found=p; break
    if not found:
        w=shutil.which('google-chrome') or shutil.which('chromium') or shutil.which('chromium-browser')
        if w: found=w
    ver=''
    if found:
        try: ver=subprocess.run([found,'--version'],capture_output=True,text=True,timeout=15).stdout.strip()
        except Exception as e: ver='버전확인 실패: '+str(e)[:60]
    step('크롬 설치',bool(found),f'{found} {ver}' if found else '크롬 없음 — 설치 필요')
    out['chrome_bin']=found
    # 2) selenium / webdriver_manager 모듈
    try:
        import selenium; step('selenium 모듈',True,'v'+getattr(selenium,'__version__','?'))
    except Exception as e: step('selenium 모듈',False,str(e)[:80])
    try:
        import webdriver_manager; step('webdriver_manager',True,'설치됨')
    except Exception as e: step('webdriver_manager',False,str(e)[:80])
    # 3) 실제 드라이버 기동 + 페이지 로드
    d=None
    if found:
        try:
            if not os.environ.get('CHROME_BIN'): os.environ['CHROME_BIN']=found
            t0=time.time(); d=get_driver()
            step('크롬 드라이버 기동',True,f'{round(time.time()-t0,1)}초')
            try:
                t1=time.time(); d.get('https://example.com'); time.sleep(1)
                ttl=(d.title or '')[:60]
                step('페이지 로드 테스트',bool(ttl),f'title="{ttl}" ({round(time.time()-t1,1)}초)')
            except Exception as e: step('페이지 로드 테스트',False,str(e)[:150])
        except Exception as e:
            step('크롬 드라이버 기동',False,str(e)[:200])
    else:
        step('크롬 드라이버 기동',False,'크롬 미설치로 건너뜀')
    # 4) 리소스
    try:
        du=shutil.disk_usage('/'); step('디스크',du.free>500*1024*1024,
            f'여유 {round(du.free/1024/1024/1024,1)}GB / 전체 {round(du.total/1024/1024/1024,1)}GB')
    except Exception as e: step('디스크',False,str(e)[:60])
    try:
        mt=open('/proc/meminfo').read()
        tot=int(re.search(r'MemTotal:\s+(\d+)',mt).group(1))//1024
        av=int(re.search(r'MemAvailable:\s+(\d+)',mt).group(1))//1024
        step('메모리',av>200,f'가용 {av}MB / 전체 {tot}MB')
    except Exception as e: step('메모리',False,str(e)[:60])
    # 5) 데이터 상태
    step('허용 사이트',len([s for s in load_sites() if is_publishable(s)])>0,
         f'발행가능 {len([s for s in load_sites() if is_publishable(s)])}개 / 전체 {len(load_sites())}개')
    step('키워드 풀',len(load_keywords())>0,f'{len(load_keywords())}개')
    step('이미지 URL',True,f'{len(load_image_urls())}개'+(' (기본 이미지 사용)' if not load_image_urls() else ''))
    out['ok']=all(s['ok'] for s in out['steps'] if s['name'] in ('크롬 설치','크롬 드라이버 기동','페이지 로드 테스트'))
    return jsonify(out)

@app.route('/api/verify/now',methods=['POST'])
def api_verify_now():
    threading.Thread(target=verify_once,kwargs={'limit':40},daemon=True).start()
    return jsonify({'ok':True,'started':True})

@app.route('/api/oneclick',methods=['POST'])
def api_oneclick():
    """원클릭 실시간 발행: 키워드 풀 랜덤 추출 → 사이트마다 유니크 생성 → 허용·캡차없는 사이트 전체 발행."""
    d=request.get_json() or {}; cfg=load_config(); pool=load_keywords()
    if not pool: return jsonify({'ok':False,'error':'키워드 풀이 비어있습니다 (글 생성 탭에서 등록)'})
    sites=load_sites(); allowed=[s for s in sites if is_publishable(s)]
    if not allowed:
        cap=sum(1 for s in sites if is_permitted(s) and s.get('has_captcha'))
        return jsonify({'ok':False,'error':f'발행 가능한 사이트가 없습니다 (허용·캡차없음 0개'+(f' · 캡차제외 {cap}개' if cap else '')+')'})
    count=max(1,min(50,int(d.get('count',1) or 1)))
    total=0
    for _ in range(count):
        kw=pick_keywords(pool,cfg)
        total+=enqueue_generated(allowed,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),'브랜드':kw.get('브랜드','')},cfg,{'region':kw.get('지역',''),'service':kw.get('서비스','')})[0]
    if total and not wk_active: start_workers(cfg.get('workers',2))
    return jsonify({'ok':True,'generated':total,'sites':len(allowed),'picks':count})

@app.route('/api/post',methods=['POST'])
def api_post():
    d=request.get_json(silent=True) or {}; cfg=load_config()
    ids=d.get('site_ids',[]); title=d.get('title',''); content=d.get('content','')
    region=d.get('region',''); service=d.get('service',''); brand=d.get('brand','')
    meta={'region':region,'service':service}
    sites=[s for s in load_sites() if s.get('id') in ids]
    allowed=[s for s in sites if is_permitted(s)]
    # 2개 이상 사이트 + 키워드 있으면 사이트마다 유니크 재생성(중복 방지). 단일 사이트는 검토한 원문 그대로.
    if len(allowed)>1 and region and service:
        q,blocked=enqueue_generated(sites,{'지역':region,'서비스':service,'브랜드':brand or cfg.get('brand','')},cfg,meta)
        note='사이트별 유니크 본문 재생성'
    else:
        if title and content: remember_if_unique(title,content)   # 원문도 중복DB에 기록
        q,blocked=enqueue(sites,title,content,meta); note=''
    if q and not wk_active: start_workers(cfg.get('workers',2))
    return jsonify({'ok':True,'queued':q,'blocked':blocked,'note':note})

@app.route('/api/bulk',methods=['POST'])
def api_bulk():
    d=request.get_json(silent=True) or {}; cfg=load_config()
    keyword_sets=d.get('keyword_sets',[]); site_ids=d.get('site_ids',[])
    workroom_id=str(d.get('workroom_id') or ''); workroom_name=str(d.get('workroom_name') or '직접 입력')[:80]
    sites=[s for s in load_sites() if not site_ids or s.get('id') in site_ids]
    if not sites: return jsonify({'ok':False,'error':'사이트 없음'})
    allowed=[s for s in sites if is_publishable(s)]; blocked=len(sites)-len(allowed)
    if not allowed:
        return jsonify({'ok':False,'error':f'실게시 검증을 통과한 허용 사이트가 없습니다 ({blocked}개 제외)'})
    keyword_sets=[x for x in keyword_sets if isinstance(x,dict) and x.get('지역') and x.get('서비스') and x.get('브랜드')]
    if not keyword_sets: return jsonify({'ok':False,'error':'유효한 키워드 조합이 없습니다'})
    # 즉시 발행은 오늘 남은 한도와 최소 간격을 넘기지 않는다. 0=무제한이어도 한 번의
    # 요청에서 과도한 API 생성이 일어나지 않도록 10개까지만 준비한다.
    capacities=[]
    for s in allowed:
        lim=site_daily_limit(s,cfg)
        today=_kst_now().strftime('%Y-%m-%d')
        used=int(s.get('posted_today',0) or 0) if s.get('posted_date')==today else 0
        cap=max(0,lim-used) if lim>0 else 10
        interval_ok,_=under_min_interval(s)
        if not interval_ok: cap=0
        elif site_min_interval(s)>0: cap=min(cap,1)
        capacities.append(cap)
    accepted=min(len(keyword_sets),min(capacities) if capacities else 0,10)
    if accepted<=0:
        return jsonify({'ok':False,'error':'오늘 한도 또는 최소 발행 간격 때문에 지금 실행 가능한 작업이 없습니다. 다음 가능 시간에 다시 실행하세요.'})
    with BULK_LOCK:
        active=next((x for x in BULK_TASKS.values() if x.get('status') in ('preparing','running')),None)
        if active:
            return jsonify({'ok':False,'error':'이미 목록 생성 작업이 진행 중입니다. 중복 실행하지 않았습니다.','task_id':active.get('id')})
        tid=secrets.token_hex(8)
        task={'id':tid,'status':'preparing','total':accepted,'done':0,'queued':0,
              'workroom_id':workroom_id,'workroom_name':workroom_name,
              'requested':len(keyword_sets),'remaining':len(keyword_sets)-accepted,'error':'','created_at':_kst_now().strftime('%Y-%m-%d %H:%M:%S')}
        BULK_TASKS[tid]=task
    def prepare_bulk():
        global wk_active
        try:
            task['status']='running'
            for ks in keyword_sets[:accepted]:
                kw={'지역':ks.get('지역',''),'서비스':ks.get('서비스',''),'브랜드':ks.get('브랜드','')}
                task['queued']+=enqueue_generated(allowed,kw,cfg,{'region':kw['지역'],'service':kw['서비스'],
                    'workroom_id':workroom_id,'workroom_name':workroom_name})[0]
                task['done']+=1
            if task['queued'] and not wk_active: start_workers(cfg.get('workers',2))
            task['status']='done'; task['finished_at']=_kst_now().strftime('%Y-%m-%d %H:%M:%S')
        except Exception as e:
            task['status']='failed'; task['error']=str(e)[:180]
    threading.Thread(target=prepare_bulk,name='BULK-PREP',daemon=True).start()
    return jsonify({'ok':True,'task_id':tid,'accepted':accepted,'requested':len(keyword_sets),
                    'remaining':len(keyword_sets)-accepted,'blocked':blocked})

@app.route('/api/bulk/status/<tid>',methods=['GET'])
def api_bulk_status(tid):
    with BULK_LOCK: task=BULK_TASKS.get(tid)
    if not task: return jsonify({'ok':False,'error':'작업 상태 없음'}),404
    return jsonify({'ok':True,**task})

@app.route('/api/worker-log',methods=['GET'])
def api_worker_log():
    """작업실별 준비 작업과 실제 워커 발행 이력을 한 화면에 제공한다."""
    rid=str(request.args.get('workroom_id') or '')
    history=list(reversed(load_json(HISTORY_FILE,[])))
    if rid: history=[x for x in history if str(x.get('workroom_id') or '')==rid]
    with BULK_LOCK: tasks=list(BULK_TASKS.values())
    if rid: tasks=[x for x in tasks if str(x.get('workroom_id') or '')==rid]
    tasks.sort(key=lambda x:x.get('created_at',''),reverse=True)
    sites=load_sites(); publishable=[s for s in sites if is_publishable(s)]
    assisted=[s for s in sites if is_assisted_postable(s)]
    captcha=[s.get('name') or s.get('site_url','') for s in sites if s.get('has_captcha')]
    return jsonify({'ok':True,'tasks':tasks[:100],'history':history[:500],
                    'publishable_count':len(publishable),'assisted_count':len(assisted),'captcha_sites':captcha,
                    'workers':{**wk_stats,'active':wk_active,'paused':wk_paused}})

@app.route('/api/manual-checks',methods=['GET'])
def api_captcha_tasks():
    with CAPTCHA_LOCK:
        tasks=[_captcha_public(x) for x in CAPTCHA_TASKS.values()]
    tasks.sort(key=lambda x:x.get('created_at',''),reverse=True)
    return jsonify({'ok':True,'tasks':tasks[:20]})

@app.route('/api/manual-checks/<tid>/submit',methods=['POST'])
def api_captcha_submit(tid):
    value=str((request.get_json(silent=True) or {}).get('value') or '').strip()
    if not value or len(value)>64:
        return jsonify({'ok':False,'error':'CAPTCHA 값을 입력하세요'}),400
    with CAPTCHA_LOCK:
        task=CAPTCHA_TASKS.get(tid)
        if not task or task.get('status')!='waiting_input':
            return jsonify({'ok':False,'error':'대기 중인 CAPTCHA 작업이 아닙니다'}),409
        task['value']=value; task['status']='input_received'; task['message']='입력값 전달 중'
        task['event'].set()
    return jsonify({'ok':True,'message':'입력 완료 · 자동 발행을 계속합니다'})

@app.route('/api/manual-checks/<tid>/cancel',methods=['POST'])
def api_captcha_cancel(tid):
    with CAPTCHA_LOCK:
        task=CAPTCHA_TASKS.get(tid)
        if not task: return jsonify({'ok':False,'error':'작업 없음'}),404
        task['cancelled']=True; task['event'].set()
    return jsonify({'ok':True})

@app.route('/api/sites',methods=['GET','POST','DELETE'])
def api_sites():
    if request.method=='POST':
        d=request.get_json(silent=True) or {}
        if not str(d.get('site_url','')).strip().startswith(('http://','https://')):
            return jsonify({'ok':False,'error':'올바른 http(s) 사이트 URL이 필요합니다'}),400
        _plat=(d.get('platform','auto') or 'auto').strip().lower()
        if _plat not in ('gnuboard','cafe24'):   # auto/미지정 → 자동 감지
            try: _plat=detect_platform(d.get('site_url','').strip())
            except Exception: _plat='gnuboard'
        site={'id':d.get('id',str(int(time.time()*1000))),'site_url':d.get('site_url','').strip().rstrip('/'),
              'platform':_plat,'mb_id':d.get('mb_id',''),'mb_pass':d.get('mb_pass',''),
              'bo_table':d.get('bo_table','').strip() or 'm8_qna','name':d.get('name',''),
              'permission':bool(d.get('permission',False)),
              'registration_source':'manual_admin',
              'daily_limit':max(0,int(d.get('daily_limit',3) or 0)),
              'min_interval_minutes':max(0,int(d.get('min_interval_minutes',60) or 0)),
              'permission_note':(d.get('permission_note','') or '').strip(),
              'permission_date':(datetime.now().strftime('%Y-%m-%d') if d.get('permission') else ''),
              'status':'idle','added':datetime.now().strftime('%m/%d %H:%M')}
        with POST_LOCK:
            sites=load_sites(); ex=[s for s in sites if s.get('id')==site['id']]
            if ex:
                # 기존 발행 카운트 등은 보존
                merged=sites[sites.index(ex[0])]
                if not site.get('mb_pass'): site['mb_pass']=merged.get('mb_pass','')
                merged.update(site)
            else: sites.append(site)
            save_sites(sites)
        return jsonify({'ok':True})
    elif request.method=='DELETE':
        d=request.get_json(silent=True) or {}
        with POST_LOCK:
            sites=[s for s in load_sites() if s.get('id')!=d.get('id','')]
            save_sites(sites)
        return jsonify({'ok':True})
    # 비밀번호는 브라우저/API로 되돌려 보내지 않는다. 저장 여부만 표시한다.
    public=[]
    for s in load_sites():
        x=dict(s); x['login_saved']=bool(x.pop('mb_pass','') or x.get('mb_pass_enc'))
        x.pop('mb_pass_enc',None); public.append(x)
    return jsonify(public)

@app.route('/api/sites/limits',methods=['POST'])
def api_sites_limits():
    """사이트 목록에서 하루 발행 한도와 최소 간격만 안전하게 즉시 수정한다."""
    d=request.get_json(silent=True) or {}; sid=str(d.get('id','')).strip()
    try:
        daily=max(0,min(10000,int(d.get('daily_limit',3))))
        interval=max(0,min(10080,int(d.get('min_interval_minutes',60))))
    except (TypeError,ValueError):
        return jsonify({'ok':False,'error':'건수와 간격은 0 이상의 숫자로 입력하세요'}),400
    found=False
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')==sid:
                s['daily_limit']=daily; s['min_interval_minutes']=interval
                s['limits_updated_at']=_kst_now().strftime('%Y-%m-%d %H:%M')
                found=True; break
        if found: save_sites(sites)
    if not found: return jsonify({'ok':False,'error':'사이트 없음'}),404
    return jsonify({'ok':True,'daily_limit':daily,'min_interval_minutes':interval})

def _signup_credentials(site, rules=None):
    """사이트별 제약을 반영한 충돌 가능성이 낮은 가입정보를 생성한다."""
    profile={'id_min':4,'id_max':16,'password_min':10,'require_special':False} if site.get('platform')=='cafe24' \
        else {'id_min':3,'id_max':20,'password_min':10,'require_special':False}
    profile.update(site.get('signup_rules') or {}); profile.update(rules or {}); rules=profile
    min_id=max(3,min(20,int(rules.get('id_min',3) or 3)))
    max_id=max(min_id,min(30,int(rules.get('id_max',20) or 20)))
    prefix=re.sub(r'[^a-z0-9_]','',str(rules.get('id_prefix') or 'twseo').lower()) or 'twseo'
    suffix=datetime.now().strftime('%m%d')+secrets.token_hex(2)
    mid=(prefix+'_'+suffix)[:max_id]
    if len(mid)<min_id: mid=(mid+secrets.token_hex(8))[:min_id]
    plen=max(8,min(64,int(rules.get('password_min',10) or 10)))
    special=str(rules.get('password_specials') or '!@#$%')
    alphabet='abcdefghjkmnpqrstuvwxyzABCDEFGHJKMNPQRSTUVWXYZ23456789'
    chars=[secrets.choice('abcdefghjkmnpqrstuvwxyz'),secrets.choice('ABCDEFGHJKMNPQRSTUVWXYZ'),
           secrets.choice('23456789')]
    if rules.get('require_special',False) and special: chars.append(secrets.choice(special))
    pool=alphabet+(special if rules.get('require_special',False) else '')
    chars += [secrets.choice(pool) for _ in range(max(0,plen-len(chars)))]
    random.SystemRandom().shuffle(chars)
    return mid,''.join(chars)

@app.route('/api/sites/signup-prepare/<sid>',methods=['POST'])
def api_site_signup_prepare(sid):
    d=request.get_json(silent=True) or {}; now=datetime.now().isoformat(timespec='seconds')
    sites=load_sites(); site=next((s for s in sites if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'}),404
    learn_error=''; profile=None
    try: profile=learn_signup_profile(site,force=False)
    except Exception as e: learn_error=str(e)[:180]
    with POST_LOCK:
        current=load_sites(); saved=next((s for s in current if s.get('id')==sid),None)
        if not saved: return jsonify({'ok':False,'error':'사이트 없음'}),404
        if profile:
            for k in ['signup_profile_host','signup_rules','signup_url','signup_profile_version','signup_profile_changed','signup_profile_measured_at','signup_has_captcha','signup_email_verification']:
                saved[k]=site.get(k)
        site=saved
        if site.get('signup_email_verification'):
            site.update({'signup_status':'rejected','signup_eligible':False,'signup_reject_reason':'이메일 인증 필요',
                         'signup_updated_at':now})
            save_sites(current)
            return jsonify({'ok':False,'rejected':True,'error':'이메일 인증이 필요한 사이트라 가입 대상에서 제외했습니다'}),409
        rules=d.get('rules') or site.get('signup_rules') or {}
        mid,pw=_signup_credentials(site,rules)
        raw_url=site.get('site_url',''); parts=urllib.parse.urlsplit(raw_url)
        base=f'{parts.scheme}://{parts.netloc}' if parts.scheme and parts.netloc else raw_url.rstrip('/')
        default_signup=base+('/member/join.html' if site.get('platform')=='cafe24' else '/bbs/register.php')
        site.update({'mb_id':mid,'mb_pass':pw,'signup_rules':rules,'signup_status':'prepared',
                     'signup_updated_at':now,'signup_url':d.get('signup_url') or site.get('signup_url') or
                         default_signup})
        save_sites(current)
    # 비밀번호는 생성 직후 한 번만 전달한다. 이후 API에서는 마스킹한다.
    return jsonify({'ok':True,'mb_id':mid,'mb_pass':pw,'signup_url':site['signup_url'],
                    'status':'prepared','learned':bool(profile),'learn_error':learn_error,
                    'profile_version':site.get('signup_profile_version',0),'rules':site.get('signup_rules',{}),
                    'captcha':bool(site.get('signup_has_captcha')),
                    'notice':'CAPTCHA와 이메일 인증은 직접 완료하세요'})

@app.route('/api/sites/signup-learn/<sid>',methods=['POST'])
def api_site_signup_learn(sid):
    sites=load_sites(); site=next((s for s in sites if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'}),404
    try: profile=learn_signup_profile(site,force=True)
    except Exception as e: return jsonify({'ok':False,'error':str(e)[:200]}),400
    with POST_LOCK:
        current=load_sites(); saved=next((s for s in current if s.get('id')==sid),None)
        if saved:
            for k in ['signup_profile_host','signup_rules','signup_url','signup_profile_version','signup_profile_changed','signup_profile_measured_at','signup_has_captcha','signup_email_verification']:
                saved[k]=site.get(k)
            if site.get('signup_email_verification'):
                saved.update({'signup_status':'rejected','signup_eligible':False,'signup_reject_reason':'이메일 인증 필요',
                              'signup_updated_at':datetime.now().isoformat(timespec='seconds')})
            save_sites(current)
    return jsonify({'ok':True,'version':profile['version'],'changed':site.get('signup_profile_changed',False),
                    'seen_count':profile['seen_count'],'rules':profile['rules'],'captcha':profile['captcha'],
                    'email_verification_required':bool(profile.get('email_verification_required')),
                    'field_count':len(profile['fields']),'measured_at':profile['measured_at']})

@app.route('/api/sites/signup-status/<sid>',methods=['POST'])
def api_site_signup_status(sid):
    d=request.get_json(silent=True) or {}; status=str(d.get('status','')).strip()
    allowed={'prepared','captcha_wait','email_wait','complete','failed'}
    if status not in allowed: return jsonify({'ok':False,'error':'올바르지 않은 가입 상태'}),400
    with POST_LOCK:
        sites=load_sites(); site=next((s for s in sites if s.get('id')==sid),None)
        if not site: return jsonify({'ok':False,'error':'사이트 없음'}),404
        if status=='complete' and (not site.get('mb_id') or not site.get('mb_pass')):
            return jsonify({'ok':False,'error':'저장된 로그인정보가 없습니다'}),400
        site['signup_status']=status; site['signup_updated_at']=datetime.now().isoformat(timespec='seconds')
        site['login_saved']=bool(site.get('mb_id') and site.get('mb_pass'))
        save_sites(sites)
        host=site.get('signup_profile_host')
        if host:
            profiles=load_signup_profiles(); profile=profiles.get(host)
            if profile:
                key='success_count' if status=='complete' else ('failure_count' if status=='failed' else '')
                if key: profile[key]=int(profile.get(key,0))+1
                profile['last_outcome']=status; profile['last_outcome_at']=site['signup_updated_at']; save_signup_profiles(profiles)
    return jsonify({'ok':True,'status':status,'login_saved':site.get('login_saved',False)})

@app.route('/api/workers/start',methods=['POST'])
def api_wk_start():
    d=request.get_json() or {}; start_workers(d.get('n',load_config().get('workers',2)))
    return jsonify({'ok':True})

@app.route('/api/workers/stop',methods=['POST'])
def api_wk_stop():
    stop_workers(); return jsonify({'ok':True})

@app.route('/api/workers/pause',methods=['POST'])
def api_wk_pause():
    pause_workers(); return jsonify({'ok':True})

@app.route('/api/workers/resume',methods=['POST'])
def api_wk_resume():
    resume_workers(); return jsonify({'ok':True})

@app.route('/api/workers/stats',methods=['GET'])
def api_wk_stats():
    return jsonify({**wk_stats,'active':wk_active,'paused':wk_paused})

@app.route('/api/workers/reset',methods=['POST'])
def api_wk_reset():
    with STATS_LOCK:
        for k in ['success','fail','total','done','skipped','retry']: wk_stats[k]=0
        wk_stats['queued']=post_queue.qsize()
    return jsonify({'ok':True})

# ---- 완전 자동 파이프라인 (발굴→자동가입→실발행→자동등록) ----
PIPELINE_STATE={'running':False,'last_result':None,'started_at':'','finished_at':''}
_pipeline_lock=threading.Lock()

@app.route('/api/pipeline/run',methods=['POST'])
def api_pipeline_run():
    """검수완료 후보를 자동가입→실발행→자동등록까지 즉시 1배치 실행(백그라운드)."""
    d=request.get_json(silent=True) or {}
    limit=int(d.get('limit', load_config().get('auto_pipeline_batch',3)) or 3)
    with _pipeline_lock:
        if PIPELINE_STATE['running']:
            return jsonify({'ok':False,'error':'파이프라인이 이미 실행 중입니다'}),409
        PIPELINE_STATE.update({'running':True,'started_at':_kst_now().strftime('%Y-%m-%d %H:%M:%S'),
                               'finished_at':'','last_result':None})
    def _run():
        try:
            # 수동 실행 시에는 설정 토글과 무관하게 이번 1회는 강제로 돌린다
            cfg=load_config()
            if not cfg.get('auto_pipeline_enabled'):
                cfg2=dict(cfg); cfg2['auto_pipeline_enabled']=True; save_config(cfg2)
                res=auto_pipeline_once(limit=limit); save_config(cfg)  # 원래 설정 복구
            else:
                res=auto_pipeline_once(limit=limit)
        except Exception as e:
            res={'ok':False,'error':str(e)[:200]}
        with _pipeline_lock:
            PIPELINE_STATE.update({'running':False,'last_result':res,
                                   'finished_at':_kst_now().strftime('%Y-%m-%d %H:%M:%S')})
    threading.Thread(target=_run,name='PIPELINE',daemon=True).start()
    return jsonify({'ok':True,'started':True,'limit':limit})

@app.route('/api/pipeline/status',methods=['GET'])
def api_pipeline_status():
    with _pipeline_lock:
        return jsonify(dict(PIPELINE_STATE))

@app.route('/api/sites/reconcile',methods=['POST'])
def api_sites_reconcile():
    """사이트 목록 즉시 최신화: 발행이 막힌 사이트를 자동 탈락시킨다."""
    try:
        n=reconcile_sites()
        return jsonify({'ok':True,'dropped':n})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:150]}),500

@app.route('/api/logs',methods=['GET'])
def api_logs():
    """최근 실행 로그(add_log)를 최신순으로 반환. 파이프라인 진행 표시용."""
    try: n=max(1,min(200,int(request.args.get('n',60))))
    except Exception: n=60
    logs=load_json(LOG_FILE,[])
    return jsonify({'ok':True,'logs':list(reversed(logs))[:n]})

# ---- 발행 이력 (결과 탭) ----
@app.route('/api/history',methods=['GET'])
def api_history():
    h=load_json(HISTORY_FILE,[])
    return jsonify(list(reversed(h))[:500])   # 최신순 500건

@app.route('/api/history/clear',methods=['POST'])
def api_history_clear():
    save_json(HISTORY_FILE,[]); return jsonify({'ok':True})

@app.route('/api/history/export',methods=['GET'])
def api_history_export():
    from flask import Response
    import io
    h=load_json(HISTORY_FILE,[])
    cols=[('time','시간'),('workroom_name','작업실'),('site_name','사이트'),('site_url','URL'),('bo_table','게시판'),
          ('member','회원'),('region','지역'),('service','서비스'),('title','제목'),('status','상태'),
          ('fail_reason_ko','실패원인'),('alive','생존'),('verified_at','검증시각'),
          ('result_url','결과URL'),('message','메시지')]
    try:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title='발행이력'
        ws.append([c[1] for c in cols])
        for rec in h:
            ws.append([str(rec.get(c[0],'') or '') for c in cols])
        for i,c in enumerate(cols,1):
            ws.column_dimensions[chr(64+i)].width=[18,16,20,34,12,14,10,12,40,10,14,8,16,30,40][i-1]
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition':'attachment; filename=chirashi_history.xlsx'})
    except Exception:
        # openpyxl 미설치 시 CSV 폴백 (엑셀에서 열림)
        import csv
        sio=io.StringIO(); w=csv.writer(sio); w.writerow([c[1] for c in cols])
        for rec in h: w.writerow([str(rec.get(c[0],'') or '') for c in cols])
        data='﻿'+sio.getvalue()   # BOM: 엑셀 한글 깨짐 방지
        return Response(data,mimetype='text/csv',
                        headers={'Content-Disposition':'attachment; filename=chirashi_history.csv'})

@app.route('/api/config',methods=['GET','POST'])
def api_cfg():
    if request.method=='POST':
        d=request.get_json(silent=True) or {}; cfg=load_config()
        old_search=(cfg.get('discover_keywords',''),cfg.get('discover_direct_queries',''))
        for k in ['brand','phone','phones','openai_key','openai_admin_key','openai_monthly_budget_usd',
                  'openai_input_price_per_million','openai_cached_input_price_per_million','openai_output_price_per_million',
                  'model','workers','post_delay','daily_limit',
                  'use_gpt','telegram_token','telegram_chat_id','notify_done','notify_fail','update_token',
                  'telegram_control','backup_time','verify_enabled','mix_keywords','block_unpaid',
                  'google_api_key','google_cx','brave_api_key','search_provider','discover_enabled','discover_daily_target',
                  'discover_query_limit','discover_keywords','discover_direct_queries',
                  'video_url','landing_url','post_email','guest_post_password',
                  'twocaptcha_api_key','twocaptcha_enabled',
                  'auto_pipeline_enabled','auto_pipeline_batch']:
            if k in d:
                if k in ('openai_key','openai_admin_key','telegram_token','google_api_key','brave_api_key','guest_post_password','twocaptcha_api_key') and d[k]=='***설정됨***': continue  # 마스크 값은 무시(기존 유지)
                cfg[k]=d[k]
        if d.get('password'): cfg['password']=generate_password_hash(d['password'])  # 해시 저장
        # 완전 자동화: 필수 키(Brave 발굴 + 2captcha)가 채워지면 발굴·파이프라인을 자동 ON.
        # (키를 넣는 행위 = 자동 운영 동의로 간주. 원치 않으면 아래 토글을 수동 OFF 가능)
        if (cfg.get('brave_api_key') or '').strip():
            cfg['discover_enabled']=True
        if (cfg.get('brave_api_key') or '').strip() and (cfg.get('twocaptcha_api_key') or '').strip() and cfg.get('twocaptcha_enabled'):
            cfg['auto_pipeline_enabled']=True
        save_config(cfg)
        # 검색 조건을 바꾸면 다음 검색부터 새 조건의 첫 줄이 즉시 실행되도록 커서를 초기화한다.
        new_search=(cfg.get('discover_keywords',''),cfg.get('discover_direct_queries',''))
        if new_search!=old_search:
            st=load_json(DISCO_FILE,{}) or {}; st['cursor']=0; save_json(DISCO_FILE,st)
        return jsonify({'ok':True,'search_cursor_reset':new_search!=old_search})
    c=dict(load_config())
    if c.get('openai_key'): c['openai_key']='***설정됨***'   # 키 노출 방지
    if c.get('openai_admin_key'): c['openai_admin_key']='***설정됨***'
    if c.get('telegram_token'): c['telegram_token']='***설정됨***'
    if c.get('google_api_key'): c['google_api_key']='***설정됨***'
    if c.get('brave_api_key'): c['brave_api_key']='***설정됨***'
    if c.get('guest_post_password'): c['guest_post_password']='***설정됨***'
    if c.get('twocaptcha_api_key'): c['twocaptcha_api_key']='***설정됨***'
    c.pop('password',None)
    return jsonify(c)

@app.route('/api/openai/usage',methods=['GET'])
def api_openai_usage():
    cfg=load_config(); summary=_local_openai_usage_summary(cfg)
    try:
        actual=_openai_admin_costs(cfg)
        if actual is not None:
            summary['source']='official_costs_api'; summary['actual_month_cost_usd']=actual
            budget=float(cfg.get('openai_monthly_budget_usd') or 0)
            summary['remaining_budget_usd']=round(max(0,budget-actual),6) if budget>0 else None
            summary['note']='OpenAI 조직 Costs API 실제 비용'
    except Exception as e:
        summary['admin_error']=str(e)[:180]
        summary['note']='관리자 키 조회 실패 · 프로젝트 키 토큰 예상치 표시'
    return jsonify({'ok':True,**summary})

@app.route('/api/twocaptcha/usage',methods=['GET'])
def api_twocaptcha_usage():
    cfg=load_config();
    summary=_twocaptcha_usage_summary(cfg)
    return jsonify({'ok':True,**summary})

@app.route('/api/test/<sid>',methods=['POST'])
def api_test(sid):
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    if not is_permitted(site):
        return jsonify({'ok':False,'error':'미허용 도메인 — 테스트도 실제 발행이므로 홍보 허용(✔) 설정 후 이용하세요'})
    try:
        cfg=load_config()
        if not under_daily_limit(site,cfg):
            return jsonify({'ok':False,'error':'사이트 일일 발행 한도에 도달했습니다'})
        interval_ok,remain=under_min_interval(site)
        if not interval_ok:
            return jsonify({'ok':False,'error':f'사이트 최소 발행 간격 미충족 ({max(1,(remain+59)//60)}분 남음)'})
        html,title=generate_article({'지역':'테스트','서비스':'테스트'},cfg)
        ok,msg=do_post(site,title,html)
        finalize_post(site,ok,fail_reason=('' if ok else str(msg)))
        result_url=msg if ok and str(msg).startswith(('http://','https://')) else ''
        if ok and result_url:
            set_site_flag(sid,write_test_status='passed',verified_at=_kst_now().strftime('%Y-%m-%d %H:%M'),
                          verified_post_url=result_url,last_structure_check=_kst_now().strftime('%Y-%m-%d %H:%M'))
        elif ok:
            # 성공 응답처럼 보여도 결과 URL을 확인할 수 없으면 발행 가능 사이트에서 제외한다.
            now=_kst_now().strftime('%Y-%m-%d %H:%M')
            set_site_flag(sid,status='rejected',permission=False,write_test_status='failed',
                          verification_fail_reason='결과 URL/게시물 검색 결과 없음',
                          verified_post_url='',last_structure_check=now)
            domain=_domain_of(site.get('site_url',''))
            with _cand_lock:
                cands=load_cands()
                for c in cands:
                    if c.get('domain','').lower()==domain:
                        c.update({'status':'rejected','reject_reason':'결과 URL/게시물 검색 결과 없음',
                                  'write_test_status':'failed','verified_at':now,'verified_post_url':''})
                save_cands(cands)
            return jsonify({'ok':False,'rejected':True,'error':'결과 URL/게시물 검색 결과 없음',
                            'platform':resolve_platform(site)}),409
        return jsonify({'ok':ok,'message':msg,'platform':resolve_platform(site)})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)})

# ---- 사이트 대량등록 (CSV: url,이름,게시판,아이디,비번,허용) ----
@app.route('/api/sites/bulk',methods=['POST'])
def api_sites_bulk():
    d=request.get_json(silent=True) or {}; text=d.get('csv',''); default_perm=bool(d.get('permission',False))
    added=0; new_sites=[]
    for line in (text or '').splitlines():
        line=line.strip()
        if not line or line.startswith('#'): continue
        parts=[x.strip() for x in line.split(',')]
        url=parts[0] if parts else ''
        if not url or not url.startswith('http'): continue
        nm=parts[1] if len(parts)>1 else ''
        bo=parts[2] if len(parts)>2 else 'free'
        mid=parts[3] if len(parts)>3 else ''
        mpw=parts[4] if len(parts)>4 else ''
        perm=default_perm
        if len(parts)>5: perm=parts[5] in ('1','true','y','Y','o','O','허용','true')
        new_sites.append({'id':secrets.token_hex(6),'site_url':url.rstrip('/'),'platform':'auto',
                      'mb_id':mid,'mb_pass':mpw,'bo_table':bo or 'free','name':nm,
                      'permission':perm,'permission_note':'CSV 일괄등록',
                      'registration_source':'admin_bulk','daily_limit':3,'min_interval_minutes':60,
                      'permission_date':(datetime.now().strftime('%Y-%m-%d') if perm else ''),
                      'status':'idle','added':datetime.now().strftime('%m/%d %H:%M')})
        added+=1
    if new_sites:
        with POST_LOCK:
            sites=load_sites(); sites.extend(new_sites); save_sites(sites)
    return jsonify({'ok':True,'added':added})

# ---- 허용상태 일괄 토글 ----
@app.route('/api/sites/permission',methods=['POST'])
def api_sites_permission():
    d=request.get_json(silent=True) or {}; ids=set(d.get('ids',[])); val=bool(d.get('permission',False))
    note=(d.get('permission_note') or '').strip()
    if val and len(note)<5:
        return jsonify({'ok':False,'error':'허용 동의 근거를 5자 이상 입력하세요'}),400
    n=0
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id') in ids:
                s['permission']=val
                if val:
                    s['permission_note']=note
                    s['permission_date']=datetime.now().strftime('%Y-%m-%d')
                    s['permission_checked_at']=_kst_now().strftime('%Y-%m-%d %H:%M')
                    s['permission_checked_by']='관리자 직접 확인'
                else:
                    s['permission_revoked_at']=_kst_now().strftime('%Y-%m-%d %H:%M')
                n+=1
        save_sites(sites)
    return jsonify({'ok':True,'changed':n})

# ---- 사이트 헬스체크 ----
@app.route('/api/sites/health/<sid>',methods=['POST'])
def api_site_health(sid):
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    h=site_health(site)
    try: plat=detect_platform(site.get('site_url',''),use_cache=False)   # 점검 시 플랫폼도 재감지
    except Exception: plat=site.get('platform','gnuboard')
    h['platform']=plat
    set_site_flag(sid,health=('ok' if h.get('ok') else 'bad'),
                  health_at=datetime.now().strftime('%m/%d %H:%M'),platform=plat)
    return jsonify({'ok':True,'health':h})

@app.route('/api/sites/detect/<sid>',methods=['POST'])
def api_site_detect(sid):
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    try: plat=detect_platform(site.get('site_url',''),use_cache=False)
    except Exception as e: return jsonify({'ok':False,'error':str(e)[:100]})
    set_site_flag(sid,platform=plat)
    return jsonify({'ok':True,'platform':plat})

@app.route('/api/sites/debug/<sid>',methods=['POST'])
def api_site_debug(sid):
    """심층 디버그: 로그인·글쓰기 페이지에서 실제로 무엇이 보이는지 그대로 덤프."""
    from selenium.webdriver.common.by import By
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    url=site.get('site_url','').rstrip('/'); m=re.match(r'(https?://[^/]+)',url)
    base=m.group(1) if m else url
    out={'base':base,'bo_table':site.get('bo_table',''),'mb_id':site.get('mb_id',''),'phases':[]}
    def snap(label):
        try: body=d.find_element(By.TAG_NAME,'body').text[:600]
        except Exception: body=''
        try:
            fields=[]
            for el in d.find_elements(By.CSS_SELECTOR,'input,textarea,select'):
                if not _sel_vis(el): continue
                fields.append(f"{el.tag_name}[{_sel_attr(el,'type') or ''}] name={_sel_attr(el,'name')} id={_sel_attr(el,'id')}")
            fields=fields[:14]
        except Exception: fields=[]
        out['phases'].append({'label':label,'url':(d.current_url or '')[:200],
                              'title':(d.title or '')[:100],'body':body,'fields':fields})
    try:
        d=get_driver()
        # 1) 로그인 페이지
        d.get(base+'/bbs/login.php'); time.sleep(2); snap('로그인 페이지')
        # 2) 로그인 시도
        mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
        if mid:
            try:
                ide=d.find_elements(By.CSS_SELECTOR,"input[name='mb_id']")
                pwe=d.find_elements(By.CSS_SELECTOR,"input[name='mb_password']")
                out['login_fields_found']={'mb_id':len(ide),'mb_password':len(pwe)}
                if ide and pwe:
                    ide[0].clear(); ide[0].send_keys(mid); pwe[0].clear(); pwe[0].send_keys(mpw)
                    clicked=_click_first(d,["input[type='submit']","button[type='submit']",".btn_submit","#btn_submit"])
                    out['login_submit_clicked']=clicked
                    if not clicked:
                        try: pwe[0].submit()
                        except Exception: pass
                    time.sleep(3); dismiss_alerts(d); snap('로그인 시도 후')
            except Exception as e: out['login_error']=str(e)[:150]
        # 3) 글쓰기 페이지 후보들
        bo=site.get('bo_table','free')
        for path in [f'/bbs/write.php?bo_table={bo}', f'/bbs/board.php?bo_table={bo}']:
            try:
                d.get(base+path); time.sleep(2); dismiss_alerts(d); snap('시도: '+path)
            except Exception as e:
                out['phases'].append({'label':'시도: '+path,'url':'','title':'','body':'ERROR '+str(e)[:100],'fields':[]})
    except Exception as e:
        out['error']=str(e)[:200]
    return jsonify({'ok':True,**out})

@app.route('/api/sites/dryrun/<sid>',methods=['POST'])
def api_site_dryrun(sid):
    """드라이런: 실제 글을 올리지 않고 등록 직전까지 검증."""
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    cfg=load_config()
    kw={'지역':'인천','서비스':'셔츠룸','브랜드':cfg.get('brand','') or '테스트'}
    html,title=generate_article(kw,cfg,unique=False)   # 중복DB 오염 방지
    try:
        ok,steps=dryrun_post(site,title,html)
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:200],'steps':[]})
    # 캡차가 감지되면 플래그 저장
    if any((not s['ok']) and '캡차' in s['name'] for s in steps):
        set_site_flag(sid,has_captcha=True,captcha_note='드라이런 감지')
    if ok:
        set_site_flag(sid,dryrun_status='passed',dryrun_at=_kst_now().strftime('%Y-%m-%d %H:%M'),
                      last_structure_check=_kst_now().strftime('%Y-%m-%d %H:%M'))
    return jsonify({'ok':ok,'steps':steps,'title':title})

@app.route('/api/sites/learn/<sid>',methods=['POST'])
def api_site_learn(sid):
    """비제출 실측학습: 등록된 사이트 한 곳의 DOM/폼을 측정하고 근거와 레시피를 저장."""
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    if not is_permitted(site):
        return jsonify({'ok':False,'error':'미허용 사이트 — 관리자가 직접 등록하고 홍보 허용한 사이트만 실측 가능'})
    try:
        analysis,rec=analyze_site_logic(site)
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:150]})
    _save_site_analysis(sid,analysis,rec)
    if analysis.get('captcha'):
        set_site_flag(sid,has_captcha=True,captcha_note=('실측: '+analysis['captcha'])[:80])
    elif analysis.get('ok'):
        set_site_flag(sid,has_captcha=False,captcha_note='',status='idle',technical_block_reason='')
    else:
        reason='글쓰기 폼을 찾지 못함'
        for step in analysis.get('steps',[]):
            if step.get('name')=='글쓰기 폼' and step.get('detail'): reason=str(step['detail'])[:160]
        set_site_flag(sid,status='failed',write_test_status='failed',verified_post_url='',
                      technical_block_reason=reason,last_structure_check=_kst_now().strftime('%Y-%m-%d %H:%M'))
    learned=None if not rec else {k:rec.get(k) for k in ['platform','write_url','subject_sel','content_mode','content_sel','submit_sel','form_action','form_method','learned_mode']}
    return jsonify({'ok':bool(analysis.get('ok')),'message':('실측 레시피 저장 완료' if rec else '실측 완료 — 발행 폼을 확정하지 못함'),
                    'captcha':bool(analysis.get('captcha')),'blocked':bool(analysis.get('blocked')),
                    'analysis':analysis,'learned':learned})

# ---- 예약 스케줄 CRUD ----
@app.route('/api/schedules',methods=['GET','POST','DELETE'])
def api_scheds():
    if request.method=='POST':
        d=request.get_json(silent=True) or {}; scheds=load_scheds()
        sc={'id':d.get('id') or secrets.token_hex(6),'name':d.get('name','예약'),
            'keyword_sets':d.get('keyword_sets',[]),'site_ids':d.get('site_ids',[]),
            'times':d.get('times',[]),'days':d.get('days',[]),'enabled':bool(d.get('enabled',True)),
            'count':max(1,int(d.get('count',1) or 1)),'last_run':'','completed_keys':[],'completed_at':''}
        ex=[x for x in scheds if x.get('id')==sc['id']]
        if ex: scheds[scheds.index(ex[0])].update(sc)
        else: scheds.append(sc)
        save_scheds(scheds); return jsonify({'ok':True,'id':sc['id']})
    if request.method=='DELETE':
        d=request.get_json(silent=True) or {}; scheds=[x for x in load_scheds() if x.get('id')!=d.get('id')]
        save_scheds(scheds); return jsonify({'ok':True})
    return jsonify(load_scheds())

@app.route('/api/schedules/toggle',methods=['POST'])
def api_sched_toggle():
    d=request.get_json(silent=True) or {}; scheds=load_scheds()
    for x in scheds:
        if x.get('id')==d.get('id'): x['enabled']=not x.get('enabled',True)
    save_scheds(scheds); return jsonify({'ok':True})

# ---- 통계 ----
@app.route('/api/stats',methods=['GET'])
def api_stats():
    return jsonify(compute_stats())

# ---- 텔레그램 테스트 발송 ----
@app.route('/api/telegram/test',methods=['POST'])
def api_tg_test():
    ok=send_telegram(load_config(),'🔔 찌라시 마스터 텔레그램 연결 테스트')
    return jsonify({'ok':ok,'error':None if ok else '토큰/챗ID 확인'})

# ---- 원격 자가 업데이트 (로그인 불필요, 토큰 인증) ----
_uploads={}; _upload_lock=threading.Lock()
def _apply_update(newcode):
    import py_compile,base64 as _b64
    if b'def main' not in newcode or len(newcode)<1000:
        return {'ok':False,'error':'app.py 형식 이상'},400
    target=os.path.join(BASE_DIR,'app.py'); cur=b''
    try:
        with open(target,'rb') as f: cur=f.read()
        with open(os.path.join(BASE_DIR,'app.py.bak'),'wb') as f: f.write(cur)
    except Exception: pass
    with open(target,'wb') as f: f.write(newcode)
    try: py_compile.compile(target,doraise=True)
    except Exception as e:
        if cur:
            with open(target,'wb') as f: f.write(cur)
        return {'ok':False,'error':f'문법오류 롤백: {str(e)[:120]}'},400
    def _restart(): time.sleep(1.0); os._exit(0)
    threading.Thread(target=_restart,daemon=True).start()
    return {'ok':True,'bytes':len(newcode),'restart':'1초 후 재기동'},200

@app.route('/api/admin/update',methods=['POST'])
def api_admin_update():
    d=request.get_json(silent=True) or {}
    tok=d.get('token') or request.headers.get('X-Update-Token','')
    want=os.environ.get('CHIRASHI_UPDATE_TOKEN') or load_config().get('update_token','')
    if not want:
        return jsonify({'ok':False,'error':'원격 업데이트 비활성화됨'}),503
    if not secrets.compare_digest(str(tok),str(want)):
        return jsonify({'ok':False,'error':'토큰 불일치'}),403
    import base64 as _b64
    # 청크 업로드 (WAF 우회: 작은 조각으로 나눠 전송)
    if 'chunk' in d or 'finalize' in d:
        uid=d.get('upload_id','')
        if not uid: return jsonify({'ok':False,'error':'upload_id 필요'}),400
        with _upload_lock:
            buf=_uploads.setdefault(uid,{})
            if 'chunk' in d: buf[int(d.get('seq',0))]=d['chunk']
            if d.get('finalize'):
                total=int(d.get('total',len(buf)))
                if len(buf)<total:
                    return jsonify({'ok':False,'error':f'조각 부족 {len(buf)}/{total}'}),400
                b64=''.join(buf[i] for i in sorted(buf)); _uploads.pop(uid,None)
                try: newcode=_b64.b64decode(b64)
                except Exception as e: return jsonify({'ok':False,'error':'b64 디코드 실패'}),400
                res,code=_apply_update(newcode); return jsonify(res),code
        return jsonify({'ok':True,'received':len(buf)})
    # 단건 업로드
    b64=d.get('app_b64','')
    if not b64: return jsonify({'ok':False,'error':'app_b64 없음'}),400
    try:
        newcode=_b64.b64decode(b64)
        res,code=_apply_update(newcode); return jsonify(res),code
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:150]}),500

# ==================== HTML ====================
HTML=r'''<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>찌라시 마스터 v6</title>
<style>:root{--bg:#060913;--c:#0d1117;--b:#1a1f2e;--t:#c9d1d9;--d:#6b7280;--p:#3b82f6;--g:#22c55e;--r:#ef4444;--y:#f59e0b;--v:#8b5cf6}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--t);min-height:100vh;font-size:13px}
::-webkit-scrollbar{width:4px}::-webkit-scrollbar-track{background:var(--bg)}::-webkit-scrollbar-thumb{background:var(--b)}
header{background:var(--c);border-bottom:1px solid var(--b);padding:10px 16px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}
.logo{font-size:15px;font-weight:700}.logo s{color:var(--p);text-decoration:none}
.stats{display:flex;gap:12px;font-size:11px;color:var(--d)}.stats b{color:var(--t)}
.tabs{display:flex;gap:2px;padding:10px 14px 0;flex-wrap:wrap}
.tab{padding:7px 14px;border:1px solid transparent;border-radius:8px 8px 0 0;cursor:pointer;font-size:11px;color:var(--d);background:transparent}
.tab.on{background:var(--c);border-color:var(--b);border-bottom-color:var(--c);color:var(--t);font-weight:600}
.wrap{max-width:1200px;margin:0 auto;padding:10px 14px}
.panel{display:none;padding:10px 0}.panel.on{display:block}
.card{background:var(--c);border:1px solid var(--b);border-radius:10px;padding:14px;margin-bottom:10px}
.card h3{font-size:10px;color:var(--d);text-transform:uppercase;letter-spacing:1px;margin-bottom:10px}
.row{display:flex;gap:6px;align-items:center;flex-wrap:wrap;margin-bottom:6px}
input,select,textarea{padding:8px 10px;border:1px solid var(--b);border-radius:6px;background:var(--bg);color:var(--t);font-size:12px;outline:none;font-family:inherit;width:100%}
input:focus,textarea:focus{border-color:var(--p)}
textarea{resize:vertical;line-height:1.5}
.btn{padding:7px 14px;border:none;border-radius:6px;font-size:11px;cursor:pointer;font-weight:600;color:#fff;transition:.2s}
.btn-p{background:var(--p)}.btn-g{background:#166534}.btn-r{background:#991b1b}.btn-y{background:#92400e}.btn-v{background:#5b21b6}.btn-d{background:var(--b);color:var(--d)}
.btn:disabled{opacity:.4}.btn-xs{padding:3px 7px;font-size:10px}
table{width:100%;border-collapse:collapse;font-size:11px}
th{text-align:left;padding:5px 8px;color:var(--d);font-weight:500;border-bottom:2px solid var(--b);font-size:10px}
td{padding:5px 8px;border-bottom:1px solid #111827}
.st{display:inline-block;padding:2px 6px;border-radius:10px;font-size:9px;font-weight:600}
.st-i{background:#1e293b;color:var(--d)}.st-ok{background:#052e16;color:var(--g)}.st-f{background:#450a0a;color:var(--r)}.st-y{background:#422006;color:var(--y)}
.toast{position:fixed;top:12px;right:12px;z-index:999;padding:8px 14px;border-radius:6px;font-size:11px;font-weight:500;animation:in .3s}
@keyframes in{from{transform:translateX(100%);opacity:0}to{transform:translateX(0);opacity:1}}
.toast-ok{background:#052e16;color:var(--g)}.toast-er{background:#450a0a;color:var(--r)}
input[type=checkbox]{accent-color:var(--p)}
.twocap-shell{background:linear-gradient(180deg,#0f1729 0%,#0b1322 100%);border:1px solid var(--b);border-radius:12px;padding:10px 10px 8px}
.twocap-header{display:flex;justify-content:space-between;align-items:center;font-size:11px;color:var(--d);margin-bottom:8px}
.twocap-chip{display:inline-flex;align-items:center;gap:6px;padding:3px 7px;border-radius:999px;font-size:10px;font-weight:700}
.twocap-chip.ok{background:rgba(34,197,94,.12);color:var(--g);border:1px solid rgba(34,197,94,.35)}
.twocap-chip.warn{background:rgba(245,158,11,.12);color:var(--y);border:1px solid rgba(245,158,11,.35)}
.twocap-metrics{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:8px}
.twocap-metric{background:rgba(255,255,255,.02);border:1px solid rgba(148,163,184,.15);border-radius:8px;padding:8px 7px;text-align:center;min-height:64px}
.twocap-metric .label{font-size:9px;color:var(--d);margin-bottom:5px;letter-spacing:.02em}
.twocap-metric .value{font-size:18px;font-weight:700;line-height:1.2}
.twocap-metric .value.green{color:var(--g)}
.twocap-metric .value.blue{color:var(--p)}
.twocap-metric .value.amber{color:var(--y)}
.twocap-footer{display:flex;justify-content:space-between;gap:10px;margin-top:9px;padding-top:8px;border-top:1px solid rgba(148,163,184,.12);font-size:9px;color:var(--d)}
.prog{background:var(--b);border-radius:6px;height:14px;overflow:hidden}
.prog>div{height:100%;width:0;background:var(--g);transition:width .3s}
.note{background:#0b1a2e;border:1px solid #1a3a5e;border-radius:8px;padding:10px 12px;font-size:11px;color:#9db8d8;line-height:1.6;margin-bottom:10px}
</style></head><body>'''

LOGIN_HTML=r'''<div style="display:flex;align-items:center;justify-content:center;min-height:100vh">
<div style="background:var(--c);border:1px solid var(--b);border-radius:14px;padding:40px;width:340px;text-align:center">
<h2 style="margin-bottom:4px;font-size:18px">찌라시 마스터 v6</h2>
<p style="color:var(--d);font-size:12px;margin-bottom:24px">정직한 자동 발행 (회피코드 없음)</p>
<form method="POST">
{% if error %}<p style="color:var(--r);font-size:11px;margin-bottom:8px">{{ error }}</p>{% endif %}
<input type="password" name="pw" placeholder="비밀번호" autofocus style="margin-bottom:12px">
<button type="submit" class="btn btn-p" style="width:100%">로그인</button>
</form></div></div>'''

DASH_HTML=r'''<header><div class="logo">찌라시 <s>마스터 v6</s></div>
<div class="stats" id="live"><span>큐:<b id="q">0</b></span><span>성공:<b id="ok" style="color:var(--g)">0</b></span><span>실패:<b id="fl" style="color:var(--r)">0</b></span><span>스킵:<b id="sk" style="color:var(--y)">0</b></span><span>워커:<b id="ws" style="color:{{'var(--g)' if wk_on else 'var(--d)'}}">{{'ON' if wk_on else 'OFF'}}</b></span></div>
<a href="/logout" class="btn-xs" style="background:var(--b);color:var(--d);text-decoration:none">로그아웃</a></header>

<div class="tabs"><button class="tab on" onclick="T('gen')">글 생성</button><button class="tab" onclick="T('wlog')">워커 실행로그</button><button class="tab" onclick="T('images')">이미지 저장</button><button class="tab" onclick="T('sites')">사이트 (<span id="siteTabCount">{{sites|length}}</span>)</button><button class="tab" onclick="T('res')">결과</button><button class="tab" onclick="T('disco')">발굴</button><button class="tab" onclick="T('mem')">회원·정산</button><button class="tab" onclick="T('stats')">통계</button><button class="tab" onclick="T('set')">설정</button></div>
<div class="wrap"><div id="toasts"></div>
<div id="pvOverlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.75);z-index:500;padding:20px" onclick="if(event.target===this)closePreview()">
<div style="max-width:820px;margin:0 auto;background:#fff;color:#222;border-radius:10px;max-height:90vh;overflow:auto">
<div style="position:sticky;top:0;background:#0d1117;color:#fff;padding:10px 14px;display:flex;align-items:center;gap:10px"><b style="flex:1;font-size:13px">🔍 발행 미리보기 (실제 게시판에 보일 모양)</b><span id="pvTitle" style="font-size:11px;color:#9aa"></span><button class="btn btn-r btn-xs" onclick="closePreview()">닫기</button></div>
<iframe id="pvFrame" style="width:100%;height:70vh;border:0;background:#fff"></iframe></div></div>

<div id="p-gen" class="panel on">
<div class="note">✔ 현재 적용 규칙 — <b style="color:var(--p)">메인 키워드1 + 같은 지역 키워드2·3</b>으로 OpenAI 장문 HTML을 생성합니다. 이미지는 저장소에서 1개만 사용하고 ALT에는 키워드1을 넣습니다. 지역 순서는 인천→경기→서울→충남→충북→세종→전북→전남→경상→경북→강원→제주이며, 모든 키워드를 사용하면 예약이 자동 종료됩니다. <b style="color:var(--g)">허용 동의가 기록된 사이트만 발행</b>됩니다.</div>

<div class="card" id="progCard" style="display:none"><h3>발행 진행률</h3>
<div class="prog"><div id="progBar"></div></div>
<div style="font-size:10px;color:var(--d);margin-top:5px" id="progText">0 / 0</div></div>

<div class="card" style="border-color:#334155"><h3>🗂 키워드 작업실 — 키워드 묶음별 독립 저장</h3>
<div class="row">
<select id="wrSelect" onchange="showWorkroom()" style="min-width:180px"><option value="">작업실 선택</option></select>
<input id="wrName" placeholder="예: 1번 작업실" style="max-width:220px">
<button class="btn btn-p" onclick="newWorkroom()">+ 작업실 추가</button>
<button class="btn btn-g" onclick="saveWorkroom()">작업실 저장</button>
<button class="btn btn-r btn-xs" onclick="deleteWorkroom()">삭제</button>
<span id="wrSaved" style="color:var(--d);font-size:10px"></span></div>
<div class="row" style="margin-top:8px">
<select id="wrProvince" style="width:auto"><option value="">전국</option></select>
<label><input type="checkbox" id="wrCity" checked style="width:auto"> 시·도</label>
<label><input type="checkbox" id="wrGu" checked style="width:auto"> 시·구·군</label>
<label><input type="checkbox" id="wrDong" checked style="width:auto"> 읍·면·동</label>
<select id="wrJoin" style="width:auto"><option value="">붙여쓰기</option><option value=" ">띄어쓰기</option></select>
<span style="color:var(--d);font-size:10px">지역명은 짧게 조합됩니다: 서울+키워드 · 강남+키워드 · 호암직동+키워드</span></div>
<textarea id="wrBases" rows="3" placeholder="키워드 종류를 한 줄에 하나씩 3개 이상 입력&#10;출장마사지&#10;마사지&#10;홍보게시판" style="margin-top:7px"></textarea>
<div style="color:var(--d);font-size:10px;margin-top:4px">지역 순서는 고정하고, 지역마다 전체 키워드 중 서로 다른 3개를 새로 랜덤 선택합니다. 첫 번째가 해당 글의 메인 키워드입니다.</div>
<div class="row" style="margin-top:6px"><button class="btn btn-d" onclick="previewWorkroomRegional()">생성 개수 확인</button><button class="btn btn-v" onclick="applyWorkroomRegional(false)">작업실 목록에 추가</button><button class="btn btn-y" onclick="applyWorkroomRegional(true)">작업실 목록 교체</button><span id="wrRegionCount" style="color:var(--d);font-size:10px">0개</span></div>
<textarea id="wrKeywords" rows="5" placeholder="키워드1,키워드2,키워드3 — 작업실마다 따로 저장됩니다" style="margin-top:7px"></textarea>
<div class="row" style="margin-top:6px"><select id="wrSite" style="width:auto"><option value="">전체 실게시 검증 사이트</option>{% for s in publish_sites %}<option value="{{s.id}}">{{s.name or s.site_url[:20]}}</option>{% endfor %}</select><button class="btn btn-p" onclick="copyWorkroomToBulk()">이 작업실을 아래 발행 목록에 적용</button><span style="color:var(--d);font-size:10px">실제 글 발행과 결과 URL 확인을 통과한 허용 사이트만 선택됩니다.</span></div>
</div>

<div class="card"><h3>키워드 조합 대량 입력 (CSV: 키워드1,키워드2,키워드3)</h3>
<textarea id="kwlist" rows="4" placeholder="인천셔츠룸,인천노래방,인천가라오케&#10;서울셔츠룸,서울노래방,서울가라오케"></textarea>
<div class="row" style="margin-top:6px"><button class="btn btn-p" id="bulkRunBtn" onclick="genFromList()">목록 생성 작업 시작</button>
<span style="color:var(--d);font-size:10px" id="kwCount">0줄</span>
<span style="color:var(--p);font-size:10px" id="bulkState"></span>
<span style="flex:1"></span>
<select id="kwSiteFilter" style="width:auto"><option value="">전체 실게시 검증 사이트</option>{% for s in publish_sites %}<option value="{{s.id}}">{{s.name or s.site_url[:20]}}</option>{% endfor %}</select></div></div>

<div class="card"><h3>키워드 풀 — 지역순서 적용 및 사용 완료 추적</h3>
<div style="font-size:10px;color:var(--d);margin-bottom:6px">한 줄에 <b style="color:var(--p)">키워드1,키워드2,키워드3</b>을 입력합니다. 2·3은 키워드1과 같은 지역이어야 하며, 예약은 지정 지역순서대로 진행하고 전부 사용하면 자동 종료됩니다.</div>
<textarea id="poolCsv" rows="5" placeholder="키워드1,키워드2,키워드3 — 한 줄에 한 조합&#10;인천셔츠룸,인천노래방,인천가라오케&#10;부천셔츠룸,부천노래방,부천가라오케&#10;서울셔츠룸,서울노래방,서울가라오케"></textarea>
<div class="row" style="margin-top:6px">
<button class="btn btn-p" onclick="savePool()">풀 저장(덮어쓰기)</button>
<button class="btn btn-v" onclick="savePool(true)">풀에 추가</button>
<label class="btn btn-d" style="cursor:pointer">엑셀(.xlsx) 업로드<input type="file" id="poolXlsx" accept=".xlsx" style="display:none" onchange="uploadXlsx()"></label>
<span style="flex:1"></span>
<span style="color:var(--d);font-size:10px" id="poolCount">0개</span>
<button class="btn btn-r btn-xs" onclick="if(confirm('키워드 풀 전체 삭제?'))clearPool()">비우기</button></div>
<div class="row" style="margin-top:6px">
<select id="poolSiteFilter" style="width:auto"><option value="">전체 실게시 검증 사이트</option>{% for s in publish_sites %}<option value="{{s.id}}">{{s.name or s.site_url[:20]}}</option>{% endfor %}</select>
<input type="number" id="poolN" value="1" min="1" max="50" style="width:70px" title="랜덤 뽑을 개수">
<button class="btn btn-g" onclick="genRandom()">랜덤 생성+발행</button>
<span style="color:var(--d);font-size:10px">풀에서 N개 랜덤 추출 → 허용 사이트 발행</span></div></div>

<div class="card"><h3>생성 결과</h3>
<div class="row"><input type="text" id="gTitle" placeholder="제목" style="font-weight:600"></div>
<textarea id="gContent" rows="10" placeholder="리치HTML 본문..."></textarea>
<div class="row" style="margin-top:6px"><span style="color:var(--d);font-size:10px" id="gLen">0자</span><span style="flex:1"></span>
<button class="btn btn-d" onclick="previewPost()">미리보기</button>
<button class="btn btn-g" onclick="postSel()">선택 사이트 발행</button>
<button class="btn btn-y" onclick="postAll()">전체 사이트 발행</button></div></div></div>

<div id="p-images" class="panel"><div class="card"><h3>이미지 파일 저장</h3>
<div style="font-size:10px;color:var(--d);margin-bottom:8px">JPG·PNG·GIF·WEBP, 파일당 최대 10MB. 저장된 이미지는 파일 첨부란이 있는 게시판에 최대 2개까지 자동으로 들어갑니다.</div>
<div class="row"><input type="file" id="imgFiles" accept="image/jpeg,image/png,image/gif,image/webp" multiple style="flex:1"><button class="btn btn-p" onclick="uploadImages()">선택 이미지 저장</button></div>
<div id="imgGallery" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:10px;margin-top:12px"></div></div>
<div class="card"><h3>외부 이미지 URL 풀 (본문 삽입 — 한 줄에 하나)</h3>
<div style="font-size:10px;color:var(--d);margin-bottom:6px">여기에 넣은 이미지 주소에서 <b style="color:var(--p)">매번 랜덤으로</b> 골라 본문에 삽입합니다(alt=지역 자동). <b style="color:var(--y)">비워두면 기본 이미지</b>가 쓰입니다.</div>
<textarea id="imgUrls" rows="5" placeholder="https://내사이트.kr/img/room1.jpg&#10;https://내사이트.kr/img/room2.jpg"></textarea>
<div class="row" style="margin-top:6px"><button class="btn btn-p" onclick="saveImages(false)">저장(덮어쓰기)</button><button class="btn btn-v" onclick="saveImages(true)">추가</button><span style="flex:1"></span><span style="color:var(--d);font-size:10px" id="imgCount">0개</span><button class="btn btn-r btn-xs" onclick="if(confirm('이미지 URL 전체 삭제?'))clearImages()">비우기</button></div></div></div>

<div id="p-sites" class="panel">
<div class="card"><h3>사이트 추가</h3>
<div class="row"><input type="text" id="sUrl" placeholder="https://사이트.com (또는 board.php URL)" style="flex:1"><select id="sPlat" style="width:110px"><option value="auto">자동감지</option><option value="gnuboard">그누보드</option><option value="cafe24">Cafe24</option></select></div>
<div class="row"><input type="text" id="sName" placeholder="이름" style="flex:1"><input type="text" id="sBo" placeholder="게시판ID (bo_table) 예:free" style="width:170px"></div>
<div class="row"><input type="number" id="sDaily" min="0" value="3" placeholder="하루 발행 한도" title="0은 무제한" style="width:150px"><input type="number" id="sInterval" min="0" value="60" placeholder="최소 간격(분)" title="사이트별 최소 발행 간격(분)" style="width:150px"><span style="color:var(--d);font-size:10px">사이트별 하루 한도 / 최소 간격(분)</span></div>
<div class="row"><input type="text" id="sId" placeholder="아이디" style="width:130px"><input type="password" id="sPw" placeholder="비밀번호" style="width:130px"><button class="btn btn-p" id="addBtn" onclick="addSite()">추가</button><button class="btn btn-d btn-xs" id="editCancel" style="display:none" onclick="cancelEdit()">취소</button></div>
<div class="row" style="background:#0b1a2e;border:1px solid #1a3a5e;border-radius:6px;padding:8px"><label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:11px;white-space:nowrap"><input type="checkbox" id="sPerm" style="width:auto">✔ 홍보 허용 확인됨</label><input type="text" id="sPermNote" placeholder="허용 근거 (운영자 메일/캡처 링크·홍보게시판명 등)" style="flex:1"></div>
<div style="font-size:10px;color:var(--d)">게시판ID(bo_table)는 글이 올라갈 게시판 식별자입니다. 예) 자유게시판 free, 홍보게시판 promotion · <b style="color:var(--y)">홍보 허용을 체크한 사이트만 발행/테스트됩니다.</b></div></div>
<div class="card"><h3>CSV 대량 등록 (한 줄에 하나: URL,이름,게시판,아이디,비번,허용여부)</h3>
<textarea id="bulkCsv" rows="4" placeholder="https://a.kr,에이,free,id1,pw1,1&#10;https://b.kr,비,promotion,id2,pw2,0"></textarea>
<div class="row" style="margin-top:6px"><label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:11px"><input type="checkbox" id="bulkPerm" style="width:auto">허용여부 미기재 시 기본 허용</label><span style="flex:1"></span><button class="btn btn-p" onclick="bulkAdd()">대량 등록</button></div>
<div style="font-size:10px;color:var(--d)">허용여부: 1/0 (마지막 칸). 미기재면 위 체크박스 기본값. 등록 후에도 목록에서 선택→허용 일괄 변경 가능.</div></div>
<div class="card"><h3>사이트 목록 (실시간 갱신)</h3>
<div class="row" style="margin-bottom:6px"><button class="btn btn-g btn-xs" onclick="bulkPermSet(true)">선택 허용✔</button><button class="btn btn-r btn-xs" onclick="bulkPermSet(false)">선택 미허용</button><button class="btn btn-d btn-xs" onclick="healthAll()">선택 상태점검</button><span style="flex:1"></span><span style="color:var(--d);font-size:10px">체크박스로 선택 후 사용</span></div>
<div style="max-height:400px;overflow-y:auto" id="siteList"></div></div></div>

<div id="p-res" class="panel">
<div class="card"><h3>발행 결과 이력</h3>
<div class="row" style="margin-bottom:8px">
<button class="btn btn-g" onclick="location='/api/history/export'">엑셀 내보내기</button>
<button class="btn btn-d" onclick="renderHistory()">새로고침</button>
<span style="flex:1"></span>
<span style="color:var(--d);font-size:10px" id="histCount">0건</span>
<button class="btn btn-r btn-xs" onclick="if(confirm('이력 전체 삭제?'))api('/history/clear','POST').then(()=>{toast('이력 삭제됨');renderHistory()})">이력 비우기</button></div>
<div style="max-height:520px;overflow-y:auto" id="histList"></div></div></div>

<div id="p-wlog" class="panel">
<div class="note">작업실에서 시작한 글 생성 준비와 실제 워커 발행 상태를 작업실별로 확인합니다. 준비 완료 뒤에는 큐→발행 중→성공/실패→결과 URL 순서로 기록됩니다.</div>
<div class="card"><h3>워커 실행로그</h3>
<div class="row"><select id="wlogRoom" style="width:auto" onchange="renderWorkerLog()"><option value="">전체 작업실</option></select><button class="btn btn-d" onclick="renderWorkerLog()">새로고침</button><span style="flex:1"></span><span id="wlogWorker" style="color:var(--d);font-size:10px"></span></div>
<div id="wlogBlock" style="margin:8px 0"></div>
<div id="captchaTasks" style="margin:8px 0"></div>
<div id="wlogTasks" style="margin:8px 0"></div>
<div style="max-height:520px;overflow-y:auto" id="wlogList"></div></div></div>

<div id="p-disco" class="panel">
<div class="note">🔎 Brave 검색 결과 중 실제 페이지 접속 성공 → HTML 제목 숫자 8개 이상 → 게시판 글쓰기 제목·본문 폼 확인까지 모두 통과한 사이트만 후보에 추가합니다. CAPTCHA·광고 금지 여부는 현재 후보 등록 조건에서 제외합니다.</div>
<div id="dcSummary" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px"></div>
<div class="card"><h3>후보 수집</h3>
<div class="row"><button class="btn btn-p" onclick="discoverNow()">Brave 검색 실행 (5키워드)</button>
<button class="btn btn-d" onclick="screenNow()">미검수 일괄 검수</button>
<button class="btn btn-d" onclick="if(confirm('모든 후보를 다시 검수할까요?'))rescreenAll()">전체 재검수</button>
<button class="btn btn-g" onclick="location='/api/candidates/export'">엑셀 내보내기</button>
<span style="flex:1"></span>
<button class="btn btn-r btn-xs" onclick="if(confirm('탈락 후보만 삭제할까요?'))clearRejected()">탈락 정리</button></div>
<div class="row" style="margin-top:8px;padding-top:8px;border-top:1px solid var(--bd)">
<button class="btn btn-v" id="pipeRunBtn" onclick="pipelineRun()">🤖 완전 자동 파이프라인 실행 (가입→실발행→등록)</button>
<label style="display:flex;align-items:center;gap:4px;font-size:11px">개수 <input type="number" id="pipeBatch" value="1" min="1" max="20" style="width:52px"></label>
<span style="flex:1"></span><span id="pipeStatus" style="font-size:11px;color:var(--d)"></span></div>
<div style="font-size:10px;color:var(--y);margin:4px 0">⚠️ 실제 게시판에 글 1건을 올려 검증합니다(되돌리기 어려움). 로그인 필요 사이트는 자동 회원가입(2captcha)까지 시도합니다.</div>
<div id="pipeLog" style="display:none;max-height:220px;overflow-y:auto;background:#0a0a12;border:1px solid var(--bd);border-radius:6px;padding:8px;font-size:11px;line-height:1.5;font-family:monospace;margin-bottom:6px"></div>
<div style="font-size:10px;color:var(--d);margin-bottom:6px">API 키가 없어도 아래에 URL을 직접 붙여넣으면 접속·글쓰기 가능 여부와 점수가 자동 검수됩니다.</div>
<textarea id="dcUrls" rows="3" placeholder="URL 직접 추가 (한 줄에 하나)&#10;https://example.kr/bbs/board.php?bo_table=promotion"></textarea>
<div class="row" style="margin-top:6px"><button class="btn btn-v" onclick="addManual()">URL 추가 + 검수</button></div></div>
<div class="card"><h3>후보 목록 (점수순 — 위에서부터 검토)</h3>
<div class="row" style="margin-bottom:6px"><select id="dcFilter" style="width:auto" onchange="renderCands()"><option value="">전체</option><option value="ready">검수완료</option><option value="approved">사이트 등록됨</option><option value="rejected">제외</option><option value="new">미검수</option></select><span style="flex:1"></span><span style="color:var(--d);font-size:10px" id="dcCount">0개</span></div>
<div style="max-height:520px;overflow-y:auto" id="dcList"></div></div></div>

<div id="p-mem" class="panel">
<div class="note">👥 회원·정산 + ⏰ 계정당 개별 자동발행 — 회원마다 <b style="color:var(--p)">원하는 시간대</b>를 지정하면 서버가 24시간 그 시간에 자동 실행합니다(PC 불필요). 회원별 전용 키워드·담당 사이트를 따로 배정할 수 있고, 시간분산으로 동시 폭주를 막습니다. 월 청구 = 기본료 + (추가 광고수 × 추가단가).</div>
<div id="memSummary" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px"></div>
<div class="card"><h3>회원 추가 / 수정</h3>
<div class="row"><input type="text" id="mName" placeholder="담당자/이름" style="flex:1"><input type="text" id="mBiz" placeholder="업소명" style="flex:1"><input type="text" id="mPhone" placeholder="연락처" style="width:150px"></div>
<div class="row"><span style="color:var(--d);font-size:11px">기본료(월)</span><input type="number" id="mFee" value="30000" style="width:110px">
<span style="color:var(--d);font-size:11px">추가광고수</span><input type="number" id="mAddons" value="0" min="0" style="width:80px">
<span style="color:var(--d);font-size:11px">추가단가</span><input type="number" id="mAddonFee" value="10000" style="width:100px">
<span style="color:var(--d);font-size:11px">정산일</span><input type="number" id="mDay" value="1" min="1" max="31" style="width:70px">
<select id="mStatus" style="width:auto"><option value="active">활성</option><option value="paused">정지</option></select></div>
<div class="row"><input type="text" id="mMemo" placeholder="메모 (선택)" style="flex:1"></div>
<div style="border-top:1px solid var(--line);margin:10px 0;padding-top:12px">
<div style="font-size:11px;color:var(--p2);font-weight:700;margin-bottom:8px">⏰ 이 회원의 자동발행 스케줄 (계정당 개별 관리)</div>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:8px"><input type="checkbox" id="mSchedOn" style="width:auto">자동발행 켜기 — 서버가 24시간 이 시간에 자동 실행</label>
<div class="row"><input type="text" id="mTimes" placeholder="점프 시간대 HH:MM, 쉼표로 여러개 (예: 09:00,14:00,20:00)" style="flex:1"></div>
<div class="row"><span style="color:var(--d);font-size:11px">요일(미선택=매일):</span>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="0" style="width:auto">월</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="1" style="width:auto">화</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="2" style="width:auto">수</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="3" style="width:auto">목</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="4" style="width:auto">금</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="5" style="width:auto">토</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="6" style="width:auto">일</label>
<span style="flex:1"></span>
<span style="color:var(--d);font-size:11px">1회당 발행수</span><input type="number" id="mPerRun" value="1" min="1" max="20" style="width:70px">
<span style="color:var(--d);font-size:11px">시간분산(분)</span><input type="number" id="mJitter" value="5" min="0" max="60" style="width:70px" title="설정 시각에서 0~N분 랜덤 지연 — 동시 폭주 방지"></div>
<div class="row"><textarea id="mKw" rows="3" placeholder="이 회원 전용 키워드 (지역,서비스,브랜드 — 한 줄에 하나)&#10;비우면 공용 키워드 풀 사용&#10;인천,셔츠룸,인천홍마니"></textarea></div>
<div class="row"><span style="color:var(--d);font-size:11px">담당 사이트(미선택=전체 허용 사이트):</span><div id="mSiteBox" style="display:flex;gap:8px;flex-wrap:wrap;font-size:11px"></div></div>
</div>
<div class="row"><button class="btn btn-p" id="memBtn" onclick="addMember()">회원 추가</button><button class="btn btn-d btn-xs" id="memCancel" style="display:none" onclick="cancelMember()">취소</button></div></div>
<div class="card"><h3>회원 목록 (이번 달 정산)</h3>
<div class="row" style="margin-bottom:6px"><button class="btn btn-g" onclick="location='/api/members/export'">엑셀 내보내기</button><button class="btn btn-d" onclick="renderMembers()">새로고침</button><span style="flex:1"></span><span style="color:var(--d);font-size:10px" id="memCount">0명</span></div>
<div style="max-height:460px;overflow-y:auto" id="memList"></div></div></div>

<div id="p-stats" class="panel">
<div class="card"><h3>발행 통계</h3>
<div class="row" style="margin-bottom:8px"><button class="btn btn-d btn-xs" onclick="renderStats()">새로고침</button><button class="btn btn-v btn-xs" onclick="api('/verify/now','POST').then(()=>toast('생존 확인 시작 (1~2분 후 갱신)'))">지금 생존확인</button></div>
<div id="statTop" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px"></div>
<div id="statReasons"></div>
<div id="statSurvival"></div>
<div style="font-size:11px;color:var(--d);margin:6px 0">최근 14일 (초록=성공, 빨강=실패)</div>
<div id="statDays"></div>
<div style="font-size:11px;color:var(--d);margin:14px 0 6px">사이트별 (상위 12)</div>
<div id="statSites"></div></div></div>

<div id="p-set" class="panel">
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
<div class="card"><h3>브랜드/전화</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">브랜드명</small><input id="cBrand" value="{{cfg.brand}}"></div>
<div style="margin-bottom:6px"><small style="color:var(--d)">대표 전화번호</small><input id="cPhone" value="{{cfg.phone}}"></div>
<small style="color:var(--d)">전화번호(여러 개, 한 줄에 하나 — 제목에 랜덤 사용)</small>
<textarea id="cPhones" rows="3" placeholder="01082755736&#10;01021636400&#10;01053505892"></textarea>
<div style="margin-top:8px"><small style="color:var(--d)">기본 동영상 URL (필수 영상란 자동 입력)</small><input id="cVideoUrl" placeholder="https://www.youtube.com/watch?v=..."></div>
<div style="margin-top:6px"><small style="color:var(--d)">홍보/랜딩 URL (필수 링크란 자동 입력)</small><input id="cLandingUrl" placeholder="https://내사이트.kr/"></div>
<div style="margin-top:6px"><small style="color:var(--d)">게시용 이메일 (필수일 때만)</small><input id="cPostEmail" type="email" placeholder="name@example.com"></div>
<div style="margin-top:6px"><small style="color:var(--d)">비회원 글 비밀번호 (필수일 때만)</small><input id="cGuestPw" type="password" placeholder="변경시에만 입력"></div>
<div style="font-size:10px;color:var(--d)">제목의 번호는 매번 <b style="color:var(--p)">[010]↔8275↔5736 · O1O=2572=3859 · [OIO-5350-5892]</b> 처럼 랜덤 기호로 변형됩니다. 여러 개면 그 중 하나를 랜덤 선택. 비우면 위 대표 전화번호 사용.</div></div>
<div class="card"><h3>워커/비번</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">워커 수</small><input type="number" id="cWorkers" value="{{cfg.workers}}" min="1" max="10"></div>
<small style="color:var(--d)">비밀번호</small><input type="password" id="cPw" placeholder="변경시 입력"></div>
<div class="card"><h3>레이트 리밋 (도배 방지)</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">포스트 간 지연 (초)</small><input type="number" id="cDelay" value="{{cfg.post_delay}}" min="0"></div>
<small style="color:var(--d)">사이트당 1일 발행 한도 (0=무제한)</small><input type="number" id="cDaily" value="{{cfg.daily_limit}}" min="0"></div>
<div class="card"><h3>GPT 본문 생성</h3>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cUseGpt" style="width:auto">GPT로 본문 생성(키 필요)</label>
<div style="font-size:10px;color:var(--d);margin-bottom:6px">키워드1을 메인 주제로 인식해 1,800~2,800자 장문을 작성하고, 키워드2·3은 같은 지역의 보조 키워드로만 사용합니다.</div>
<div style="margin-bottom:6px"><small style="color:var(--d)">OpenAI API 키</small><input type="password" id="cOpenai" placeholder="변경시만 입력 (sk-...)"></div>
<div style="margin-bottom:6px"><small style="color:var(--d)">모델</small><input id="cModel" value="{{cfg.model}}" placeholder="gpt-4o-mini"></div>
<details style="margin-top:8px;border-top:1px solid var(--bd);padding-top:8px"><summary style="cursor:pointer;color:var(--p);font-size:11px;font-weight:700">사용량·비용 상세 설정</summary>
<div style="margin-top:7px"><small style="color:var(--d)">조직 관리자 키 (선택 · 실제 Costs API 조회용)</small><input type="password" id="cOpenaiAdmin" placeholder="관리자 키 없으면 로컬 예상비용 사용"></div>
<div class="row" style="margin-top:6px"><div style="flex:1"><small style="color:var(--d)">월 예산 USD</small><input type="number" id="cOpenaiBudget" min="0" step="0.01" value="20"></div>
<div style="flex:1"><small style="color:var(--d)">입력 $/1M</small><input type="number" id="cOpenaiInPrice" min="0" step="0.001" value="0.15"></div>
<div style="flex:1"><small style="color:var(--d)">출력 $/1M</small><input type="number" id="cOpenaiOutPrice" min="0" step="0.001" value="0.60"></div></div></details>
<div id="openaiUsage" style="margin-top:9px;padding:9px;background:#0b1322;border:1px solid var(--bd);border-radius:7px;font-size:10px;color:var(--d)">사용량 불러오는 중...</div>
<button class="btn btn-d btn-xs" style="margin-top:6px" onclick="loadOpenAIUsage()">사용량 새로고침</button></div>
<div class="card"><h3>텔레그램 알림</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">봇 토큰</small><input type="password" id="cTgTok" placeholder="변경시만 입력"></div>
<div style="margin-bottom:6px"><small style="color:var(--d)">챗 ID</small><input id="cTgChat" placeholder="예: 123456789"></div>
<div class="row" style="font-size:11px;color:var(--d)"><label style="display:flex;align-items:center;gap:4px"><input type="checkbox" id="cNotifyDone" style="width:auto">성공알림</label><label style="display:flex;align-items:center;gap:4px"><input type="checkbox" id="cNotifyFail" style="width:auto">실패알림</label><button class="btn btn-d btn-xs" onclick="api('/telegram/test','POST').then(r=>toast(r&&r.ok?'전송됨':'실패: '+(r&&r.error||''),r&&r.ok?'ok':'er'))">테스트 전송</button></div></div>
<div class="card"><h3>🤖 CAPTCHA 자동 해결 (2captcha)</h3>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cTwocaptchaEn" style="width:auto">2captcha 자동 해결 활성화 (reCAPTCHA, hCaptcha, Turnstile, kCaptcha)</label>
<div style="margin-bottom:6px"><small style="color:var(--d)">2captcha API 키</small><input type="password" id="cTwocaptchaKey" placeholder="변경시만 입력 (https://2captcha.com)"></div>
<div style="font-size:10px;color:var(--d)"><a href="https://2captcha.com/" target="_blank" rel="noopener" style="color:var(--p)">2captcha 계정 관리</a> · 키는 화면과 API 응답에 노출되지 않습니다. 자동 해결 실패 시 수동 입력으로 자동 폴백됩니다.</div>
<div id="twocaptchaUsage" style="margin-top:9px;padding:9px;background:#0b1322;border:1px solid var(--bd);border-radius:7px;font-size:10px;color:var(--d)">2captcha 상태 확인 중...</div></div>
<div class="card"><h3>🔎 도메인 발굴 (Brave Search API)</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">Brave Search API 키</small><input type="password" id="cBraveKey" placeholder="변경시에만 입력"></div>
<div style="font-size:10px;color:var(--d);margin-bottom:6px"><a href="https://api-dashboard.search.brave.com/" target="_blank" rel="noopener" style="color:var(--p)">Brave API 키 관리</a> · 키는 화면과 API 응답에 노출되지 않습니다.</div>
<div class="row" style="margin-bottom:6px"><span style="color:var(--d);font-size:11px">하루 후보 목표</span><input type="number" id="cDTarget" value="100" min="10" max="1000" style="width:90px">
<span style="color:var(--d);font-size:11px">하루 쿼리 한도</span><input type="number" id="cDQuery" value="100" min="1" max="10000" style="width:90px" title="Brave API 플랜 한도 안에서 사용"></div>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cDiscoOn" style="width:auto">24시간 자동 발굴 켜기 (15분마다 조금씩 수집·검수)</label>
<small style="color:var(--d)">검색 키워드 목록 (한 줄에 하나 — 입력 그대로 Brave 검색, #으로 시작하면 메모)</small>
<textarea id="cDDirect" rows="6" placeholder="&quot;홍보게시판&quot; 마사지&#10;inurl:bbs/board.php &quot;업체등록&quot;&#10;인천 광고 가능한 게시판"></textarea>
<div style="font-size:10px;color:var(--g);margin-top:4px">설정 저장을 누르면 서버에 영구 저장되며, 위 목록만 입력 순서대로 검색합니다.</div>
<details style="margin-top:10px;border:1px solid var(--bd);border-radius:8px;padding:9px"><summary style="cursor:pointer;color:var(--p);font-weight:700">전국 시·구·동 + 키워드 일괄 생성</summary>
<div style="font-size:10px;color:var(--d);margin:8px 0">전체 주소가 아닌 단계별 짧은 지역명으로 조합합니다: 서울+키워드 · 강남+키워드 · 호암직동+키워드. 같은 검색어는 자동으로 중복 제거합니다.</div>
<div class="row" style="margin-bottom:6px"><select id="rgProvince" style="width:auto"><option value="">전국</option></select>
<label style="font-size:11px"><input type="checkbox" id="rgCity" checked style="width:auto"> 시·도</label>
<label style="font-size:11px"><input type="checkbox" id="rgGu" checked style="width:auto"> 시·구·군</label>
<label style="font-size:11px"><input type="checkbox" id="rgDong" checked style="width:auto"> 읍·면·동</label>
<select id="rgJoin" style="width:auto"><option value="">붙여쓰기</option><option value=" ">띄어쓰기</option></select></div>
<textarea id="rgKeywords" rows="3" placeholder="출장마사지&#10;마사지&#10;홍보게시판"></textarea>
<div class="row" style="margin-top:6px"><button class="btn btn-d btn-xs" onclick="previewRegionalKeywords()">생성 개수 확인</button>
<button class="btn btn-v btn-xs" onclick="applyRegionalKeywords(false)">기존 목록에 추가</button>
<button class="btn btn-y btn-xs" onclick="applyRegionalKeywords(true)">목록 교체</button>
<span id="rgCount" style="font-size:10px;color:var(--g)"></span></div></details>
<div style="font-size:10px;color:var(--d);margin-top:6px">후보 목록에서 필요한 사이트를 선택해 사이트 관리 목록에 등록할 수 있습니다.</div></div>
<div class="card"><h3>자동 백업 · 봇 제어 · 발행 검증</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">매일 자동 백업 시각 (HH:MM, KST · 비우면 끔)</small><input id="cBackupTime" placeholder="예: 04:00"></div>
<div class="row" style="margin-bottom:6px"><button class="btn btn-g btn-xs" onclick="api('/backup/now','POST').then(r=>toast(r&&r.ok?'📦 백업 전송됨':'실패: '+(r&&r.error||'토큰 확인'),r&&r.ok?'ok':'er'))">지금 백업 전송</button><span style="color:var(--d);font-size:10px">텔레그램으로 zip 전송(설정·사이트·이력·예약·키워드)</span></div>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cTgControl" style="width:auto">텔레그램 폰 제어 (봇에게 /상태 /오늘 /발행 /정지 /재개 /백업 /검증)</label>
<label style="display:flex;align-items:center;gap:6px;color:var(--v);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cVerify" style="width:auto">발행글 생존 자동 검증 (1시간마다 URL 재확인)</label>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cMixKw" style="width:auto">지역 연동 키워드 혼합 (같은 지역의 서비스·브랜드만 조합)</label>
<label style="display:flex;align-items:center;gap:6px;color:var(--y);font-size:12px"><input type="checkbox" id="cBlockUnpaid" style="width:auto">미납 회원 자동 정지 (이번 달 미납이면 그 회원 자동발행 중단)</label>
<div style="font-size:10px;color:var(--d);margin-top:6px">폰 제어를 켜면 봇 채팅창에 명령을 보내 PC·패널 없이 조작할 수 있습니다(등록된 챗ID만 허용). 백업·검증은 텔레그램 토큰/챗ID가 필요합니다.<br><b style="color:var(--g)">중복 방지 상시 작동</b> — 모든 제목·본문은 과거와 겹치지 않게 매번 새로 생성되며, 같은 키워드라도 사이트마다 글이 다릅니다. 본문 끝에는 alt=지역(맨앞 키워드) 이미지가 삽입됩니다.</div></div>
</div>
<div style="display:flex;gap:8px;margin-bottom:10px">
<button class="btn btn-p" onclick="saveCfg()">설정 저장</button>
<button class="btn btn-g" onclick="api('/workers/start','POST',{n:parseInt(document.getElementById('cWorkers').value)||2}).then(()=>toast('워커 시작'))">워커 시작</button>
<button class="btn btn-y" onclick="api('/workers/pause','POST').then(()=>toast('일시정지'))">일시정지</button>
<button class="btn btn-v" onclick="api('/workers/resume','POST').then(()=>toast('재개'))">재개</button>
<button class="btn btn-r" onclick="api('/workers/stop','POST').then(()=>toast('워커 정지'))">워커 정지</button>
<button class="btn btn-d" onclick="if(confirm('통계 초기화?'))api('/workers/reset','POST').then(()=>toast('초기화됨'))">통계 초기화</button></div>
<div class="note">일시정지는 진행 중인 큐를 지우지 않고 멈춥니다(재개 시 이어서). 패널을 껐다 켜도 미완료 작업은 자동 복구됩니다.</div>
<div class="card"><h3>🩺 서버 자가진단 (실제 발행 가능 여부)</h3>
<div class="row"><button class="btn btn-v" onclick="runDiag()">진단 실행 (최대 60초)</button><span style="color:var(--d);font-size:10px">크롬 설치·드라이버 기동·페이지 로드·메모리·데이터 상태를 점검합니다</span></div>
<div id="diagOut" style="margin-top:10px"></div></div></div>

</div><!-- wrap -->

<script>
const $=id=>document.getElementById(id);
function T(n){document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));document.querySelectorAll('.panel').forEach(p=>p.classList.remove('on'));document.querySelector(`[onclick="T('${n}')"]`).classList.add('on');$('p-'+n).classList.add('on');if(n==='res')renderHistory();if(n==='wlog'){renderWorkerLog();renderCaptchaTasks()}if(n==='stats')renderStats();if(n==='set'){loadCfgUI();loadRegionTool()}if(n==='gen'){loadPool();loadImages();loadWorkrooms();loadRegionTool()}if(n==='mem'){renderMembers();if(!document.querySelector('.mSite'))fillSiteBox([])}if(n==='disco')renderCands()}
function toast(m,c='ok'){const d=$('toasts');const e=document.createElement('div');e.className='toast toast-'+c;e.textContent=m;d.appendChild(e);setTimeout(()=>e.remove(),2500)}
function esc(s){return String(s==null?'':s).replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
async function api(p,m,b){try{const o={method:m,headers:{'Content-Type':'application/json'}};if(b)o.body=JSON.stringify(b);const r=await fetch('/api'+p,o);if(r.status===401){location='/login';return null}return await r.json()}catch(e){toast(e.message,'er');return null}}
// (kv/gen/genBulk 죽은 JS 제거됨 — k1/k2/k3 입력칸이 없어 호출 불가였음)
function parseList(){return $('kwlist').value.split('\n').map(l=>l.trim()).filter(Boolean).map(l=>{const p=l.split(',');return{지역:(p[0]||'').trim(),서비스:(p[1]||'').trim(),브랜드:(p[2]||'').trim()}}).filter(k=>k.지역&&k.서비스&&k.브랜드)}
let _bulkPolling=false;
async function genFromList(){const sets=parseList();if(!sets.length){toast('키워드 없음','er');return}const sid=$('kwSiteFilter').value;const wid=$('kwlist').dataset.workroomId||'',wn=$('kwlist').dataset.workroomName||'직접 입력';if(!confirm('['+wn+'] '+sets.length+'개 키워드 조합의 생성 작업을 시작할까요?\n사이트 한도와 최소 간격에 따라 지금 가능한 수만 실행됩니다.'))return;const b=$('bulkRunBtn');b.disabled=true;b.textContent='작업 준비 중...';$('bulkState').textContent='서버에 작업을 전달하는 중';toast('['+wn+'] 목록 작업 준비를 시작했습니다');const r=await api('/bulk','POST',{keyword_sets:sets,site_ids:sid?[sid]:[],workroom_id:wid,workroom_name:wn});if(!r||!r.ok){b.disabled=false;b.textContent='목록 생성 작업 시작';$('bulkState').textContent='';if(r)toast(r.error||'실패','er');return}const rem=r.remaining?(' · 나머지 '+r.remaining+'개는 한도/간격상 미실행'):'';$('bulkState').textContent='['+wn+'] 준비 '+r.accepted+'/'+r.requested+rem;toast('즉시 실행 가능 '+r.accepted+'개 준비 시작'+rem,'ok');pollBulkTask(r.task_id)}
async function pollBulkTask(id){if(_bulkPolling)return;_bulkPolling=true;const b=$('bulkRunBtn');for(let i=0;i<600;i++){const r=await api('/bulk/status/'+id,'GET');if(!r||!r.ok)break;$('bulkState').textContent='글 생성 '+r.done+'/'+r.total+' · 큐 '+r.queued+(r.remaining?' · 대기 필요 '+r.remaining:'');if(r.status==='done'){toast('준비 완료 · '+r.queued+'건이 발행 큐에 등록됨','ok');break}if(r.status==='failed'){toast('준비 실패: '+(r.error||''),'er');break}await new Promise(x=>setTimeout(x,1500))}_bulkPolling=false;b.disabled=false;b.textContent='목록 생성 작업 시작'}
function getSiteIds(){return Array.from(document.querySelectorAll('.cb:checked')).map(c=>c.dataset.id)}
function getAllSiteIds(){return Array.from(document.querySelectorAll('#siteList tr[data-id]')).map(r=>r.dataset.id)}
async function postSel(){const t=$('gTitle').value.trim();const c=$('gContent').value.trim();if(!t||!c){toast('제목/본문 입력','er');return}const ids=getSiteIds();if(!ids.length){toast('사이트 선택','er');return}const r=await api('/post','POST',{site_ids:ids,title:t,content:c,region:$('k1').value.trim(),service:$('k2').value.trim(),brand:$('k3').value.trim()});if(r&&r.ok)toast(r.queued+'개 큐 등록'+(r.blocked?` · 미허용 ${r.blocked}개 제외`:''),r.queued?'ok':'er')}
async function postAll(){const t=$('gTitle').value.trim();const c=$('gContent').value.trim();if(!t||!c){toast('제목/본문 입력','er');return}const ids=getAllSiteIds();if(!ids.length){toast('사이트 없음','er');return}const r=await api('/post','POST',{site_ids:ids,title:t,content:c,region:$('k1').value.trim(),service:$('k2').value.trim(),brand:$('k3').value.trim()});if(r&&r.ok)toast(r.queued+'개 큐 등록'+(r.blocked?` · 미허용 ${r.blocked}개 제외`:''),r.queued?'ok':'er')}
// ---- 발행 이력 렌더 ----
function histRow(h){const stc=h.status==='done'?'ok':(h.status==='failed'?'f':(h.status==='retry'?'y':'i'));const rurl=h.result_url?`<a href="${esc(h.result_url)}" target="_blank" style="color:var(--p)">열기</a>`:'';
const stt=h.status==='retry'?'재시도':(h.status||'');
const av=h.alive==='yes'?'<span class="st st-ok">생존</span>':(h.alive==='no'?'<span class="st st-f">삭제</span>':(h.status==='done'?'<span style="color:var(--d)">-</span>':''));
const fr=h.fail_reason_ko?`<span style="color:var(--r)">${esc(h.fail_reason_ko)}</span>`:'';
return `<tr><td style="color:var(--d);white-space:nowrap">${esc((h.time||'').slice(5,16))}</td><td>${esc(h.site_name||'')}</td><td style="color:var(--d)">${esc(h.region||'')} ${esc(h.service||'')}</td><td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${esc(h.title||'')}">${esc(h.title||'')}</td><td><span class="st st-${stc}">${esc(stt)}</span></td><td>${fr}</td><td>${av}</td><td>${rurl}</td><td style="color:var(--d);max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${esc(h.message||'')}">${esc(h.message||'')}</td></tr>`}
async function renderHistory(){const h=await api('/history','GET');if(!Array.isArray(h))return;$('histCount').textContent=h.length+'건';
if(!h.length){$('histList').innerHTML='<p style="color:var(--d);padding:30px;text-align:center">아직 발행 이력이 없습니다</p>';return}
$('histList').innerHTML='<table><thead><tr><th>시간</th><th>사이트</th><th>지역/서비스</th><th>제목</th><th>상태</th><th>실패원인</th><th>생존</th><th>결과</th><th>메시지</th></tr></thead><tbody>'+h.map(histRow).join('')+'</tbody></table>'}
async function submitCaptcha(id){const el=$('cap_'+id);const value=(el&&el.value||'').trim();if(!value){toast('CAPTCHA 값을 입력하세요','er');return}const r=await api('/manual-checks/'+id+'/submit','POST',{value});if(r&&r.ok){toast('입력 완료 · 자동 발행을 계속합니다');renderCaptchaTasks()}else toast((r&&r.error)||'전달 실패','er')}
async function cancelCaptcha(id){const r=await api('/manual-checks/'+id+'/cancel','POST',{});if(r&&r.ok){toast('CAPTCHA 작업 취소됨');renderCaptchaTasks()}}
async function renderCaptchaTasks(){const box=$('captchaTasks');if(!box)return;const r=await api('/manual-checks','GET');if(!r||!r.ok)return;const waiting=(r.tasks||[]).filter(t=>['waiting_input','input_received','submitting'].includes(t.status));if(!waiting.length){box.innerHTML='';return}box.innerHTML=waiting.map(t=>'<div class="card" style="border-color:#a16207;background:#171205"><h3 style="color:var(--y)">🧩 '+esc(t.site_name)+' · CAPTCHA 입력 후 자동 발행</h3><div class="row">'+(t.image_data?'<img src="'+t.image_data+'" alt="CAPTCHA" style="max-width:240px;max-height:90px;background:#fff;border-radius:4px;padding:4px">':'<span style="color:var(--y)">이미지 캡처 실패 — 사이트 화면에서 CAPTCHA를 확인하세요.</span>')+'<input id="cap_'+esc(t.id)+'" autocomplete="off" placeholder="보이는 문자를 직접 입력" style="width:220px" '+(t.status!=='waiting_input'?'disabled':'')+'><button class="btn btn-g" onclick="submitCaptcha(\''+esc(t.id)+'\')" '+(t.status!=='waiting_input'?'disabled':'')+'>입력 후 자동 발행</button><button class="btn btn-r btn-xs" onclick="cancelCaptcha(\''+esc(t.id)+'\')">취소</button></div><div style="color:var(--d);font-size:10px;margin-top:6px">'+esc(t.message||'')+' · 만료 '+esc(t.expires_at||'')+'</div></div>').join('')}
async function renderWorkerLog(){const roomSel=$('wlogRoom');if(!roomSel.dataset.loaded){const rooms=await api('/workrooms','GET');if(Array.isArray(rooms)){roomSel.innerHTML='<option value="">전체 작업실</option>'+rooms.map(r=>'<option value="'+esc(r.id)+'">'+esc(r.name)+'</option>').join('');roomSel.dataset.loaded='1'}}const rid=roomSel.value;const r=await api('/worker-log'+(rid?'?workroom_id='+encodeURIComponent(rid):''),'GET');if(!r||!r.ok)return;const w=r.workers||{};$('wlogWorker').textContent='워커 '+(w.active?(w.paused?'일시정지':'실행 중'):'정지')+' · 큐 '+(w.queued||0)+' · 성공 '+(w.success||0)+' · 실패 '+(w.fail||0)+' · 스킵 '+(w.skipped||0);$('wlogBlock').innerHTML=r.publishable_count?'<div class="note" style="border-color:#166534;color:var(--g)">발행 가능 검증 사이트 '+r.publishable_count+'곳</div>':'<div class="note" style="border-color:#991b1b;color:var(--r)">⛔ 현재 발행 가능 사이트 0곳'+((r.captcha_sites||[]).length?' · CAPTCHA 감지: '+esc(r.captcha_sites.join(', ')):'')+' — CAPTCHA를 우회하지 않으며 사람이 처리하고 실게시 재검증하기 전까지 자동 발행하지 않습니다.</div>';const sm={preparing:'준비',running:'글 생성 중',done:'준비 완료',failed:'준비 실패'};$('wlogTasks').innerHTML=(r.tasks||[]).length?'<table><thead><tr><th>작업실</th><th>시작</th><th>준비 진행</th><th>큐 등록</th><th>대기 필요</th><th>상태</th></tr></thead><tbody>'+r.tasks.map(t=>'<tr><td><b>'+esc(t.workroom_name||'직접 입력')+'</b></td><td>'+esc(t.created_at||'')+'</td><td>'+esc(t.done||0)+'/'+esc(t.total||0)+'</td><td>'+esc(t.queued||0)+'</td><td>'+esc(t.remaining||0)+'</td><td><span class="st st-'+(t.status==='done'?'ok':t.status==='failed'?'f':'y')+'">'+esc(sm[t.status]||t.status||'')+'</span> '+esc(t.error||'')+'</td></tr>').join('')+'</tbody></table>':'<p style="color:var(--d);padding:12px">선택한 작업실의 준비 작업이 없습니다.</p>';const h=r.history||[];$('wlogList').innerHTML=h.length?'<table><thead><tr><th>작업실</th><th>시간</th><th>키워드</th><th>사이트</th><th>상태</th><th>결과 URL</th><th>메시지</th></tr></thead><tbody>'+h.map(x=>'<tr><td><b>'+esc(x.workroom_name||'직접 입력')+'</b></td><td>'+esc((x.time||'').slice(5,16))+'</td><td>'+esc(x.region||'')+' / '+esc(x.service||'')+'</td><td>'+esc(x.site_name||'')+'</td><td><span class="st st-'+(x.status==='done'?'ok':x.status==='failed'?'f':x.status==='skipped'?'y':'i')+'">'+esc(x.status||'')+'</span></td><td>'+(x.result_url?'<a href="'+esc(x.result_url)+'" target="_blank" style="color:var(--p)">열기</a>':'-')+'</td><td style="color:var(--d)">'+esc(x.fail_reason_ko||x.message||'')+'</td></tr>').join('')+'</tbody></table>':'<p style="color:var(--d);padding:30px;text-align:center">아직 이 작업실의 워커 발행 이력이 없습니다.</p>'}
let _editId=null;
async function runDiag(){$('diagOut').innerHTML='<p style="color:var(--d);padding:14px">🩺 진단 중... 크롬을 실제로 띄워보는 중이라 최대 60초 걸립니다.</p>';const r=await api('/diag','GET');if(!r){$('diagOut').innerHTML='<p style="color:var(--r)">진단 실패</p>';return}
const rows=(r.steps||[]).map(s=>`<tr><td>${s.ok?'<span style="color:var(--g)">✅</span>':'<span style="color:var(--r)">❌</span>'}</td><td><b>${esc(s.name)}</b></td><td style="color:var(--d)">${esc(s.detail)}</td></tr>`).join('');
const hdr=r.ok?'<span style="color:var(--g)">✅ 발행 가능 — 크롬 정상 작동</span>':'<span style="color:var(--r)">❌ 발행 불가 — 아래 ❌ 항목 확인</span>';
$('diagOut').innerHTML=`<div style="margin-bottom:8px;font-weight:700">${hdr}</div><table>${rows}</table><div style="font-size:10px;color:var(--d);margin-top:8px">${esc(r.platform||'')} · Python ${esc(r.python||'')}</div>`}
// ---- 도메인 발굴 ----
let _cands=[];
async function renderCands(){const r=await api('/candidates','GET');if(!r||!r.candidates)return;_cands=r.candidates;const s=r.summary||{};
$('dcSummary').innerHTML=tile('전체',s.total||0,'var(--t)')+tile('검수완료',s.ready||0,'var(--g)')+tile('사이트 등록',s.approved||0,'var(--p)')+tile('제외',s.rejected||0,'var(--r)')+tile('오늘 쿼리',(s.today_queries||0)+'/100','var(--v)');
const f=$('dcFilter').value;const list=f?_cands.filter(c=>c.status===f):_cands;
$('dcCount').textContent=list.length+'개';
if(!list.length){$('dcList').innerHTML='<p style="color:var(--d);padding:30px;text-align:center">후보가 없습니다. Brave 키워드 검색을 실행하거나 URL을 직접 추가하세요.</p>';return}
$('dcList').innerHTML='<table><thead><tr><th>점수</th><th>도메인/게시판</th><th>판정</th><th>상태</th><th>동작</th></tr></thead><tbody>'+list.map(c=>{
const sc=c.score||0;const scc=sc>=50?'var(--g)':(sc>=20?'var(--y)':'var(--r)');
const flags=[];
if(c.promo_hint)flags.push('<span class="st st-ok">홍보허용흔적</span>');
if(c.parked)flags.push('<span class="st st-f">주차도메인</span>');
if(c.illegal)flags.push('<span class="st st-f">도박·불법</span>');
if(c.ad_banned)flags.push('<span class="st st-f">광고금지</span>');
if(c.captcha)flags.push('<span class="st st-y">🧩캡차</span>');
if(c.login_required)flags.push('<span class="st st-i">로그인필요</span>');
if(c.write_form)flags.push('<span class="st st-ok">글쓰기폼</span>');
if(c.last_post_days!=null)flags.push('<span class="st st-i">최근글 '+c.last_post_days+'일</span>');
const stmap={ready:'<span class="st st-ok">검수완료</span>',new:'<span class="st st-i">미검수</span>',approved:'<span class="st st-ok">사이트등록</span>',rejected:'<span class="st st-f">제외</span>'};
const acts=(c.status==='approved')?'':(
 '<button class="btn btn-g btn-xs" onclick="approveCand(\''+esc(c.id)+'\')">사이트 등록</button> '+
 '<button class="btn btn-r btn-xs" onclick="setCand(\''+esc(c.id)+'\',\'rejected\')">탈락</button>');
return '<tr><td style="color:'+scc+';font-weight:700;font-size:15px">'+sc+'</td>'+
'<td><a href="'+esc(c.url)+'" target="_blank" style="color:var(--p)"><b>'+esc(c.domain||'')+'</b></a><br><span style="color:var(--d);font-size:10px">'+esc((c.board_name||c.title||'').slice(0,44))+'</span></td>'+
'<td style="max-width:230px">'+flags.join(' ')+(c.reject_reason?'<br><span style="color:var(--r);font-size:10px">'+esc(c.reject_reason)+'</span>':'')+'</td>'+
'<td>'+(stmap[c.status]||'')+'</td><td style="white-space:nowrap">'+acts+'</td></tr>'}).join('')+'</tbody></table>'}
async function discoverNow(){toast('🔎 Brave 키워드 검색 중...(약 10초)');const r=await api('/candidates/discover','POST',{queries:5});if(r&&r.ok){toast('신규 '+r.added+'개 · 검수 '+r.screened+'건 (오늘 검색 '+r.today_queries+'회)');renderCands()}else toast((r&&r.error)||'실패','er')}
let _regionData=null;
async function loadRegionTool(){if(!_regionData){const r=await api('/regions','GET');if(!r||r.error){toast((r&&r.error)||'행정구역 로드 실패','er');return null}_regionData=r}['rgProvince','wrProvince'].forEach(id=>{const s=$(id);if(!s||s.dataset.loaded)return;Object.keys(_regionData).forEach(x=>{const o=document.createElement('option');o.value=x;o.textContent=x;s.appendChild(o)});s.dataset.loaded='1'});return _regionData}
function shortProvince(x){return x.replace(/특별자치시$|특별자치도$|특별시$|광역시$|자치도$|도$/,'')}
function shortDistrict(x){const last=x.trim().split(/\s+/).pop();return last.replace(/시$|군$|구$/,'')}
let _workrooms=[];
async function loadWorkrooms(){const r=await api('/workrooms','GET');if(!Array.isArray(r))return;_workrooms=r;const s=$('wrSelect');const keep=s.value;s.innerHTML='<option value="">작업실 선택</option>'+r.map(x=>'<option value="'+esc(x.id)+'">'+esc(x.name)+'</option>').join('');if(r.some(x=>x.id===keep))s.value=keep;else if(r.length)s.value=r[0].id;showWorkroom()}
function showWorkroom(){const r=_workrooms.find(x=>x.id===$('wrSelect').value);$('wrName').value=r?r.name:'';$('wrKeywords').value=r?r.keyword_csv:'';$('wrSite').value=r?r.site_id:'';$('wrSaved').textContent=r?('저장 '+(r.updated_at||'')):''}
async function newWorkroom(){const name=(prompt('새 작업실 이름',(_workrooms.length+1)+'번 작업실')||'').trim();if(!name)return;const r=await api('/workrooms','POST',{name:name,keyword_csv:'',site_id:''});if(r&&r.ok){await loadWorkrooms();$('wrSelect').value=r.id;showWorkroom();toast(name+' 추가됨','ok')}}
async function saveWorkroom(){const id=$('wrSelect').value;if(!id){toast('먼저 작업실을 추가하세요','er');return}const r=await api('/workrooms','POST',{id:id,name:$('wrName').value.trim(),keyword_csv:$('wrKeywords').value,site_id:$('wrSite').value});if(r&&r.ok){toast('작업실 저장 완료 · '+r.count+'개 조합','ok');await loadWorkrooms();$('wrSelect').value=id;showWorkroom()}else toast((r&&r.error)||'저장 실패','er')}
async function deleteWorkroom(){const id=$('wrSelect').value;if(!id)return;if(!confirm('선택한 작업실과 키워드 목록을 삭제할까요?'))return;await api('/workrooms','DELETE',{id:id});await loadWorkrooms();toast('작업실 삭제됨')}
function workroomProvinceRank(p){const x=shortProvince(p);if(x==='인천')return 0;if(x==='경기')return 1;if(x==='서울')return 2;if(x==='충청남'||x==='충남'||x==='대전')return 3;if(x==='충청북'||x==='충북')return 4;if(x==='세종')return 5;if(x==='전북'||x==='전라북')return 6;if(x==='전남'||x==='전라남'||x==='광주')return 7;if(['부산','대구','울산','경남','경상남'].includes(x))return 8;if(x==='경북'||x==='경상북')return 9;if(x==='강원')return 10;if(x==='제주')return 11;return 99}
function randomThree(items){const a=[...new Set(items)];for(let i=a.length-1;i>0;i--){const j=Math.floor(Math.random()*(i+1));[a[i],a[j]]=[a[j],a[i]]}return a.slice(0,3)}
async function makeWorkroomRegional(){const data=await loadRegionTool();if(!data)return[];const bases=[...new Set($('wrBases').value.split(/\r?\n/).map(x=>x.trim()).filter(x=>x&&!x.startsWith('#')))];if(bases.length<3){toast('서로 다른 키워드를 한 줄에 하나씩 최소 3개 입력하세요','er');return[]}const only=$('wrProvince').value,join=$('wrJoin').value,regions=[];Object.entries(data).sort((a,b)=>workroomProvinceRank(a[0])-workroomProvinceRank(b[0])).forEach(([province,districts])=>{if(only&&province!==only)return;if($('wrCity').checked)regions.push(shortProvince(province));Object.entries(districts||{}).forEach(([district,dongs])=>{if($('wrGu').checked)regions.push(shortDistrict(district));if($('wrDong').checked)(dongs||[]).forEach(d=>regions.push(d))})});const out=[];regions.forEach(region=>{const picked=randomThree(bases);out.push(picked.map(base=>region+join+base).join(','))});return out}
async function previewWorkroomRegional(){const rows=await makeWorkroomRegional();$('wrRegionCount').textContent=rows.length.toLocaleString()+'개 생성 예정'}
async function applyWorkroomRegional(replace){const rows=await makeWorkroomRegional();if(!rows.length)return;if(rows.length>50000){toast('5만 개를 초과합니다. 지역 범위를 줄여주세요','er');return}const current=replace?[]:$('wrKeywords').value.split(/\r?\n/).map(x=>x.trim()).filter(Boolean);const merged=[...new Set(current.concat(rows))];$('wrKeywords').value=merged.join('\n');$('wrRegionCount').textContent=rows.length.toLocaleString()+'개 생성 · 작업실 전체 '+merged.length.toLocaleString()+'개';toast('작업실 목록에 반영됨 · 작업실 저장을 눌러주세요','ok')}
function copyWorkroomToBulk(){const rows=$('wrKeywords').value.trim();if(!rows){toast('작업실 키워드가 없습니다','er');return}const room=_workrooms.find(x=>x.id===$('wrSelect').value);$('kwlist').value=rows;$('kwlist').dataset.workroomId=room?room.id:'';$('kwlist').dataset.workroomName=room?room.name:'직접 입력';$('kwSiteFilter').value=$('wrSite').value;$('kwCount').textContent=(room?'['+room.name+'] ':'')+rows.split(/\r?\n/).filter(Boolean).length+'줄';toast((room?'['+room.name+'] ':'')+'발행 목록에 적용됨','ok');$('kwlist').scrollIntoView({behavior:'smooth',block:'center'})}
async function makeRegionalKeywords(){const data=await loadRegionTool();if(!data)return[];const bases=$('rgKeywords').value.split(/\r?\n/).map(x=>x.trim()).filter(x=>x&&!x.startsWith('#'));if(!bases.length){toast('조합할 키워드를 한 줄에 하나씩 입력하세요','er');return[]}const only=$('rgProvince').value;const join=$('rgJoin').value;const regions=[];Object.entries(data).forEach(([province,districts])=>{if(only&&province!==only)return;if($('rgCity').checked)regions.push(shortProvince(province));Object.entries(districts||{}).forEach(([district,dongs])=>{if($('rgGu').checked)regions.push(shortDistrict(district));if($('rgDong').checked)(dongs||[]).forEach(d=>regions.push(d))})});const out=[];const seen=new Set();regions.forEach(region=>bases.forEach(base=>{const q=region+join+base;if(!seen.has(q)){seen.add(q);out.push(q)}}));return out}
async function previewRegionalKeywords(){const rows=await makeRegionalKeywords();$('rgCount').textContent=rows.length.toLocaleString()+'개 생성 예정'}
async function applyRegionalKeywords(replace){const rows=await makeRegionalKeywords();if(!rows.length)return;if(rows.length>50000){toast('5만 개를 초과합니다. 지역 또는 단계를 줄여주세요','er');return}const current=replace?[]:$('cDDirect').value.split(/\r?\n/).map(x=>x.trim()).filter(Boolean);const merged=[...new Set(current.concat(rows))];$('cDDirect').value=merged.join('\n');$('rgCount').textContent=rows.length.toLocaleString()+'개 생성 · 전체 '+merged.length.toLocaleString()+'개';toast('목록에 반영됨 · 설정 저장을 눌러주세요','ok')}
async function screenNow(){toast('검수 중...(최대 1분)');const r=await api('/candidates/screen','POST',{limit:20});if(r&&r.ok){toast(r.screened+'건 검수 완료');renderCands()}}
async function pipelineRun(){
  // 이 브라우저는 네이티브 confirm/prompt가 막혀 있으므로 사용하지 않는다.
  // 배치 개수는 입력칸(pipeBatch)에서 읽고, 실행 확인은 버튼 클릭 자체로 갈음.
  const el=document.getElementById('pipeBatch');
  const n=Math.max(1,parseInt((el&&el.value)||'1',10)||1);
  const btn=document.getElementById('pipeRunBtn');
  if(btn){btn.disabled=true;btn.textContent='🤖 실행 중...';}
  const restore=()=>{if(btn){btn.disabled=false;btn.textContent='🤖 완전 자동 파이프라인 실행 (가입→실발행→등록)';}};
  // 실행 로그 영역 준비
  const logBox=document.getElementById('pipeLog');
  const startTs=Date.now();
  function pLog(line,cls){ if(!logBox)return; logBox.style.display='block';
    const color=cls==='ok'?'var(--g)':cls==='er'?'var(--r)':cls==='hi'?'var(--p)':'var(--d)';
    logBox.innerHTML+=`<div style="color:${color}">${line}</div>`; logBox.scrollTop=logBox.scrollHeight; }
  if(logBox){logBox.innerHTML='';logBox.style.display='block';}
  pLog('▶ 파이프라인 시작 — 후보 '+n+'건 처리','hi');
  $('pipeStatus').textContent='시작 — 후보 '+n+'건 처리 중...';
  toast('🤖 파이프라인 시작 ('+n+'건)');
  let seen=0;  // 이미 표시한 로그 개수(중복 방지)
  async function pumpLogs(){
    try{
      const lr=await api('/logs?n=40','GET');
      if(lr&&lr.ok&&Array.isArray(lr.logs)){
        // 오래된→최신 순으로 뒤집고, 파이프라인/발행 관련만
        const rel=lr.logs.slice().reverse().filter(e=>/자동|파이프라인|발행|가입|등록|캡차|2captcha|탈락|스킵|성공|실패|재시도/.test(e.msg||''));
        for(let k=seen;k<rel.length;k++){ const e=rel[k];
          const cls=/성공|완료|등록|해결/.test(e.msg)?'ok':/실패|오류|탈락|거부|불가/.test(e.msg)?'er':'';
          pLog('· '+(e.time||'')+' '+(e.msg||''),cls); }
        seen=Math.max(seen,rel.length);
      }
    }catch(_){}
  }
  try{
    const r=await api('/pipeline/run','POST',{limit:n});
    if(!r||!r.ok){pLog('✖ 실행 실패: '+((r&&r.error)||''),'er');toast((r&&r.error)||'실행 실패','er');$('pipeStatus').textContent=(r&&r.error)||'실행 실패';return}
    $('pipeStatus').textContent='실행 중...';
    for(let i=0;i<120;i++){
      await new Promise(s=>setTimeout(s,2500));
      await pumpLogs();
      const st=await api('/pipeline/status','GET');
      if(st&&!st.running&&st.finished_at&&(new Date(st.finished_at.replace(' ','T')).getTime()>=startTs-3000||st.last_result)){
        await pumpLogs();
        const R=st.last_result||{};
        // 사이트별 결과 상세
        pLog('──────────','');
        (R.results||[]).forEach(x=>{
          if(x.ok) pLog('✅ 등록: '+(x.name||'').slice(0,40)+' → '+(x.url||''),'ok');
          else pLog('❌ '+(x.stage||'')+' 실패: '+(x.name||'').slice(0,30)+' — '+(x.msg||''),'er');
        });
        pLog('■ 완료 — 처리 '+(R.processed||0)+' · 자동가입 '+(R.signed_up||0)+' · 등록 '+(R.registered||0)+(R.dropped?' · 자동탈락 '+R.dropped:''),'hi');
        $('pipeStatus').textContent='완료: 처리 '+(R.processed||0)+' · 가입 '+(R.signed_up||0)+' · 등록 '+(R.registered||0);
        toast('파이프라인 완료 — 등록 '+(R.registered||0)+'개 / 처리 '+(R.processed||0)+'개', (R.registered>0?'ok':''));
        renderCands(); if(typeof renderSites==='function')renderSites();
        return;
      }
    }
    pLog('⏱ 시간 초과 — 백그라운드에서 계속됩니다','er');
    $('pipeStatus').textContent='시간 초과(백그라운드 계속)';
  }finally{ restore(); }
}
async function rescreenAll(){toast('전체 재검수 중...(최대 2분)');const r=await api('/candidates/screen','POST',{rescreen:true,limit:40});if(r&&r.ok){toast(r.screened+'건 재검수 완료');renderCands()}}
async function addManual(){const u=$('dcUrls').value;if(!u.trim()){toast('URL 입력','er');return}const r=await api('/candidates/manual','POST',{urls:u});if(r&&r.ok){toast(r.added+'개 추가 · 검수 시작(잠시 후 새로고침)');$('dcUrls').value='';setTimeout(renderCands,3000);renderCands()}else toast((r&&r.error)||'실패','er')}
async function setCand(id,st){await api('/candidates/status','POST',{id:id,status:st});renderCands()}
async function clearRejected(){await api('/candidates','DELETE',{clear:'rejected'});renderCands()}
async function approveCand(id){const c=_cands.find(x=>x.id===id);if(!c)return;
if(c.ad_banned&&!confirm('⚠️ 이 사이트는 "광고 금지" 문구가 감지되었습니다. 그래도 사이트 목록에 등록할까요?'))return;
const bo=prompt('게시판ID(bo_table)',c.bo_table||'free')||'free';
const mid=prompt('로그인 아이디 (없으면 비워두세요)','')||'';
const mpw=mid?(prompt('비밀번호','')||''):'';
const r=await api('/candidates/approve/'+id,'POST',{bo_table:bo,mb_id:mid,mb_pass:mpw});
if(r&&r.ok){toast('✅ 사이트 목록에 등록됨');renderCands();renderSites()}else toast((r&&r.error)||'실패','er')}
// ---- 회원·정산 ----
function won(n){return (n||0).toLocaleString()+'원'}
let _memEdit=null;
async function renderMembers(){const r=await api('/members','GET');if(!r||!r.members)return;const s=r.summary||{};
$('memSummary').innerHTML=tile('회원',s.members||0,'var(--t)')+tile('활성',s.active||0,'var(--g)')+tile('이번달 청구',won(s.billed),'var(--p)')+tile('납부완료',won(s.paid),'var(--g)')+tile('미납',won(s.unpaid),'var(--r)')+tile('미납 회원',s.unpaid_count||0,'var(--y)');
$('memCount').textContent=(r.members.length)+'명';
if(!r.members.length){$('memList').innerHTML='<p style="color:var(--d);padding:30px;text-align:center">등록된 회원이 없습니다</p>';return}
const DW=['월','화','수','목','금','토','일'];
$('memList').innerHTML='<table><thead><tr><th>이름/업소</th><th>상태</th><th>월청구</th><th>이번달</th><th>스케줄</th><th>최근실행</th><th>동작</th></tr></thead><tbody>'+r.members.map(m=>{
const st=m.status==='active'?'<span class="st st-ok">활성</span>':'<span class="st st-i">정지</span>';
const pay=m.paid?'<button class="btn btn-g btn-xs" onclick="togglePay(\''+esc(m.id)+'\',false)" title="'+esc(m.paid_at||'')+'">✔ 납부</button>':'<button class="btn btn-r btn-xs" onclick="togglePay(\''+esc(m.id)+'\',true)">미납</button>';
const addon=m.addons?(' <span style="color:var(--d);font-size:9px">+광고'+m.addons+'</span>'):'';
const times=(m.sched_times||[]);
const days=(m.sched_days&&m.sched_days.length)?m.sched_days.map(d=>DW[d]).join(''):'매일';
const sched=m.sched_enabled&&times.length
  ? '<span class="st st-ok">ON</span> <span style="color:var(--p);font-size:10px">'+esc(times.join(', '))+'</span><br><span style="color:var(--d);font-size:9px">'+days+' · '+(m.per_run||1)+'건/회 · ±'+(m.jitter==null?5:m.jitter)+'분'+((m.keywords||[]).length?' · 전용키워드'+m.keywords.length:'')+((m.site_ids||[]).length?' · 사이트'+m.site_ids.length:'')+'</span>'
  : '<span class="st st-i">OFF</span>';
return '<tr><td><b>'+esc(m.name||'')+'</b><br><span style="color:var(--d);font-size:10px">'+esc(m.biz||'')+' '+esc(m.phone||'')+'</span></td><td>'+st+'</td><td style="color:var(--p)">'+won(m.fee)+addon+'</td><td>'+pay+'</td><td>'+sched+'</td><td style="color:var(--d);font-size:10px">'+esc(m.last_run||'-')+(m.run_count?'<br>총 '+m.run_count+'회':'')+'</td><td><button class="btn btn-g btn-xs" onclick="runMember(\''+esc(m.id)+'\')" title="지금 1회 실행">실행</button> <button class="btn btn-p btn-xs" onclick=\'editMember('+JSON.stringify(m)+')\'>편집</button> <button class="btn btn-r btn-xs" onclick="delMember(\''+esc(m.id)+'\')">삭제</button></td></tr>'}).join('')+'</tbody></table>'}
function mDays(){return Array.from(document.querySelectorAll('.mDay:checked')).map(c=>parseInt(c.value))}
function mSiteIds(){return Array.from(document.querySelectorAll('.mSite:checked')).map(c=>c.dataset.id)}
async function fillSiteBox(sel){const sites=await api('/sites','GET');if(!Array.isArray(sites))return;const set=new Set(sel||[]);
$('mSiteBox').innerHTML=sites.map(s=>`<label style="display:flex;align-items:center;gap:3px"><input type="checkbox" class="mSite" data-id="${esc(s.id)}" style="width:auto" ${set.has(s.id)?'checked':''}>${esc(s.name||(s.site_url||'').slice(0,18))}${s.has_captcha?'🧩':''}</label>`).join('')}
async function addMember(){const d={name:$('mName').value.trim(),biz:$('mBiz').value.trim(),phone:$('mPhone').value.trim(),plan_fee:parseInt($('mFee').value)||0,addons:parseInt($('mAddons').value)||0,addon_fee:parseInt($('mAddonFee').value)||0,settle_day:parseInt($('mDay').value)||1,status:$('mStatus').value,memo:$('mMemo').value.trim(),
sched_enabled:$('mSchedOn').checked,sched_times:$('mTimes').value,sched_days:mDays(),site_ids:mSiteIds(),keywords_csv:$('mKw').value,jitter:parseInt($('mJitter').value)||0,per_run:parseInt($('mPerRun').value)||1};
if(!d.name&&!d.biz){toast('이름 또는 업소명 입력','er');return}if(_memEdit)d.id=_memEdit;const r=await api('/members','POST',d);if(r&&r.ok){toast(_memEdit?'수정됨':'회원 추가됨');cancelMember();renderMembers()}}
function editMember(m){_memEdit=m.id;$('mName').value=m.name||'';$('mBiz').value=m.biz||'';$('mPhone').value=m.phone||'';$('mFee').value=m.plan_fee||0;$('mAddons').value=m.addons||0;$('mAddonFee').value=m.addon_fee||0;$('mDay').value=m.settle_day||1;$('mStatus').value=m.status||'active';$('mMemo').value=m.memo||'';
$('mSchedOn').checked=!!m.sched_enabled;$('mTimes').value=(m.sched_times||[]).join(', ');$('mPerRun').value=m.per_run||1;$('mJitter').value=(m.jitter==null?5:m.jitter);
document.querySelectorAll('.mDay').forEach(c=>c.checked=(m.sched_days||[]).includes(parseInt(c.value)));
$('mKw').value=(m.keywords||[]).map(k=>[k.지역||'',k.서비스||'',k.브랜드||''].join(',')).join('\n');
fillSiteBox(m.site_ids||[]);
$('memBtn').textContent='수정 저장';$('memBtn').classList.add('btn-y');$('memCancel').style.display='';$('mName').scrollIntoView({behavior:'smooth',block:'center'})}
function cancelMember(){_memEdit=null;['mName','mBiz','mPhone','mMemo','mTimes','mKw'].forEach(i=>$(i).value='');$('mFee').value=30000;$('mAddons').value=0;$('mAddonFee').value=10000;$('mDay').value=1;$('mStatus').value='active';$('mSchedOn').checked=false;$('mPerRun').value=1;$('mJitter').value=5;document.querySelectorAll('.mDay').forEach(c=>c.checked=false);fillSiteBox([]);$('memBtn').textContent='회원 추가';$('memBtn').classList.remove('btn-y');$('memCancel').style.display='none'}
async function runMember(id){if(!confirm('이 회원의 스케줄을 지금 1회 실행합니다.\n배정 사이트에 실제로 발행됩니다. 진행할까요?'))return;toast('⏰ 실행중...');const r=await api('/members/run/'+id,'POST');if(r&&r.ok)toast('✅ '+r.generated+'건 큐 등록 (사이트 '+r.sites+'개)');else toast((r&&r.error)||'실패','er');renderMembers()}
async function delMember(id){if(!confirm('회원을 삭제할까요? (정산 기록도 삭제)'))return;await api('/members','DELETE',{id});renderMembers()}
async function togglePay(id,paid){await api('/members/pay','POST',{id:id,paid:paid});renderMembers()}
async function addSite(){const d={site_url:$('sUrl').value.trim(),platform:$('sPlat').value,name:$('sName').value.trim(),bo_table:$('sBo').value.trim(),mb_id:$('sId').value.trim(),mb_pass:$('sPw').value,permission:$('sPerm').checked,permission_note:$('sPermNote').value.trim(),daily_limit:Math.max(0,parseInt($('sDaily').value)||0),min_interval_minutes:Math.max(0,parseInt($('sInterval').value)||0)};if(!d.site_url){toast('URL 입력','er');return}if(_editId)d.id=_editId;const r=await api('/sites','POST',d);if(r&&r.ok){toast(_editId?'수정됨':(d.permission?'추가됨 (홍보 허용)':'추가됨 (미검증 — 발행 제외)'));cancelEdit();renderSites()}}
async function editSite(id){const sites=await api('/sites','GET');if(!Array.isArray(sites))return;const s=sites.find(x=>x.id===id);if(!s){toast('사이트 없음','er');return}
$('sUrl').value=s.site_url||'';$('sPlat').value=s.platform||'auto';$('sName').value=s.name||'';$('sBo').value=s.bo_table||'';$('sId').value=s.mb_id||'';$('sPw').value='';$('sPw').placeholder=s.login_saved?'저장됨 · 변경시에만 입력':'비밀번호';$('sPerm').checked=!!s.permission;$('sPermNote').value=s.permission_note||'';$('sDaily').value=(s.daily_limit==null?3:s.daily_limit);$('sInterval').value=(s.min_interval_minutes==null?60:s.min_interval_minutes);
_editId=id;$('addBtn').textContent='수정 저장';$('addBtn').classList.add('btn-y');$('editCancel').style.display='';
$('sUrl').scrollIntoView({behavior:'smooth',block:'center'});toast('편집 모드 — 값을 고치고 "수정 저장"')}
function cancelEdit(){_editId=null;['sUrl','sName','sBo','sId','sPw','sPermNote'].forEach(i=>$(i).value='');$('sDaily').value=3;$('sInterval').value=60;$('sPerm').checked=false;$('sPlat').value='auto';$('addBtn').textContent='추가';$('addBtn').classList.remove('btn-y');$('editCancel').style.display='none'}
async function prepareSignup(id){
if(!confirm('사이트 조건에 맞는 가입용 아이디·비밀번호를 생성해 암호화 저장합니다.\nCAPTCHA와 이메일 인증, 최종 가입은 직접 진행합니다. 계속할까요?'))return;
const r=await api('/sites/signup-prepare/'+id,'POST',{});if(!r||!r.ok){toast((r&&r.error)||'생성 실패','er');return}
const rules=r.rules||{};const v='아이디: '+r.mb_id+'\n비밀번호: '+r.mb_pass+'\n가입주소: '+r.signup_url+'\n\n학습 규칙: ID '+(rules.id_min||'?')+'~'+(rules.id_max||'?')+'자 / 비밀번호 '+(rules.password_min||'?')+'자 이상'+(r.captcha?' / CAPTCHA 있음':'');
prompt('가입정보가 생성·암호화 저장되었습니다. 지금 복사하세요(비밀번호는 다시 표시되지 않습니다).',v);
window.open(r.signup_url,'_blank','noopener');renderSites();}
async function learnSignup(id){toast('🧠 가입 폼을 제출 없이 측정·학습 중...');const r=await api('/sites/signup-learn/'+id,'POST',{});if(r&&r.ok){const x=r.rules||{};toast('🧠 학습 v'+r.version+' · 필드 '+r.field_count+'개 · ID '+x.id_min+'~'+x.id_max+' · PW '+x.password_min+'+'+(r.changed?' · 폼 변경 감지':''),'ok');renderSites()}else toast('가입 폼 학습 실패: '+(r&&r.error||''),'er')}
async function signupDone(id){if(!confirm('CAPTCHA와 이메일 인증까지 끝나 실제 회원가입이 완료됐습니까?'))return;const r=await api('/sites/signup-status/'+id,'POST',{status:'complete'});if(r&&r.ok){toast('✅ 가입 완료 · 자동 로그인정보 저장됨');renderSites()}else toast((r&&r.error)||'처리 실패','er')}
async function signupAll(id,status){if(status==='rejected'){toast('⛔ 이메일 인증 필요 사이트 — 가입 대상 제외','er');return}if(status==='complete'){toast('✅ 회원가입 및 로그인정보 저장 완료');return}if(['prepared','captcha_wait','email_wait'].includes(status)){await signupDone(id);return}await prepareSignup(id)}
// (oneClick 죽은 JS 제거됨 — ocN 입력칸이 없어 호출 불가였음)
function previewPost(){const c=$('gContent').value.trim();if(!c){toast('먼저 글을 생성하세요','er');return}$('pvTitle').textContent=$('gTitle').value||'';const doc='<!DOCTYPE html><html lang=ko><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><style>body{font-family:-apple-system,sans-serif;max-width:760px;margin:0 auto;padding:16px;color:#222;line-height:1.7}img{max-width:100%}</style></head><body>'+c+'</body></html>';$('pvFrame').srcdoc=doc;$('pvOverlay').style.display='block'}
function closePreview(){$('pvOverlay').style.display='none';$('pvFrame').srcdoc=''}
async function delSite(id){if(!confirm('삭제?'))return;await api('/sites','DELETE',{id});renderSites()}
async function testSite(id){toast('Selenium 테스트 중...');const r=await api('/test/'+id,'POST');if(r&&r.ok)toast('✅ 테스트 성공!'+(r.platform?' ['+(r.platform==='cafe24'?'Cafe24':'그누보드')+']':'')+' '+(r.message||''));else toast('실패: '+(r?.error||r?.message||''),'er')}
async function saveCfg(){const d={brand:$('cBrand').value.trim(),phone:$('cPhone').value.trim(),phones:$('cPhones').value,video_url:$('cVideoUrl').value.trim(),landing_url:$('cLandingUrl').value.trim(),post_email:$('cPostEmail').value.trim(),workers:parseInt($('cWorkers').value)||2,post_delay:parseInt($('cDelay').value)||0,daily_limit:parseInt($('cDaily').value)||0,use_gpt:$('cUseGpt').checked,model:$('cModel').value.trim()||'gpt-4o-mini',openai_monthly_budget_usd:parseFloat($('cOpenaiBudget').value)||0,openai_input_price_per_million:parseFloat($('cOpenaiInPrice').value)||0,openai_output_price_per_million:parseFloat($('cOpenaiOutPrice').value)||0,telegram_chat_id:$('cTgChat').value.trim(),notify_done:$('cNotifyDone').checked,notify_fail:$('cNotifyFail').checked,backup_time:$('cBackupTime').value.trim(),telegram_control:$('cTgControl').checked,verify_enabled:$('cVerify').checked,mix_keywords:$('cMixKw').checked,block_unpaid:$('cBlockUnpaid').checked,search_provider:'brave',discover_enabled:$('cDiscoOn').checked,discover_daily_target:parseInt($('cDTarget').value)||100,discover_query_limit:parseInt($('cDQuery').value)||100,discover_keywords:'',discover_direct_queries:$('cDDirect').value,twocaptcha_enabled:$('cTwocaptchaEn').checked};
const bk=$('cBraveKey').value.trim();if(bk)d.brave_api_key=bk;
const pw=$('cPw').value.trim();if(pw)d.password=pw;const gp=$('cGuestPw').value.trim();if(gp)d.guest_post_password=gp;const ok=$('cOpenai').value.trim();if(ok)d.openai_key=ok;const oa=$('cOpenaiAdmin').value.trim();if(oa)d.openai_admin_key=oa;const tg=$('cTgTok').value.trim();if(tg)d.telegram_token=tg;const tc=$('cTwocaptchaKey').value.trim();if(tc)d.twocaptcha_api_key=tc;const r=await api('/config','POST',d);if(r&&r.ok){toast('저장 완료');$('cPw').value='';$('cGuestPw').value='';$('cOpenai').value='';$('cOpenaiAdmin').value='';$('cTgTok').value='';$('cTwocaptchaKey').value='';loadOpenAIUsage()}}
async function loadCfgUI(){const c=await api('/config','GET');if(!c)return;$('cVideoUrl').value=c.video_url||'';$('cLandingUrl').value=c.landing_url||'';$('cPostEmail').value=c.post_email||'';$('cGuestPw').placeholder=(c.guest_post_password==='***설정됨***')?'설정됨 · 변경시에만 입력':'변경시에만 입력';$('cUseGpt').checked=!!c.use_gpt;$('cNotifyDone').checked=!!c.notify_done;$('cNotifyFail').checked=!!c.notify_fail;$('cTgControl').checked=!!c.telegram_control;$('cVerify').checked=(c.verify_enabled!==false);$('cMixKw').checked=(c.mix_keywords!==false);$('cBlockUnpaid').checked=(c.block_unpaid!==false);$('cDiscoOn').checked=!!c.discover_enabled;if(c.discover_daily_target)$('cDTarget').value=c.discover_daily_target;if(c.discover_query_limit)$('cDQuery').value=c.discover_query_limit;if(typeof c.discover_direct_queries==='string')$('cDDirect').value=c.discover_direct_queries;$('cBraveKey').placeholder=(c.brave_api_key==='***설정됨***')?'설정됨 · 변경시에만 입력':'Brave API 키 입력';if(c.backup_time)$('cBackupTime').value=c.backup_time;if(c.model)$('cModel').value=c.model;if(c.telegram_chat_id)$('cTgChat').value=c.telegram_chat_id;if(typeof c.phones==='string')$('cPhones').value=c.phones;$('cOpenai').placeholder=(c.openai_key==='***설정됨***')?'설정됨 · 변경시만 입력':'sk-... (변경시만)';$('cOpenaiAdmin').placeholder=(c.openai_admin_key==='***설정됨***')?'관리자 키 설정됨 · 변경시만 입력':'관리자 키 없으면 로컬 예상비용 사용';$('cOpenaiBudget').value=c.openai_monthly_budget_usd==null?20:c.openai_monthly_budget_usd;$('cOpenaiInPrice').value=c.openai_input_price_per_million==null?0.15:c.openai_input_price_per_million;$('cOpenaiOutPrice').value=c.openai_output_price_per_million==null?0.60:c.openai_output_price_per_million;$('cTgTok').placeholder=(c.telegram_token==='***설정됨***')?'설정됨 · 변경시만 입력':'변경시만 입력';$('cTwocaptchaEn').checked=!!c.twocaptcha_enabled;$('cTwocaptchaKey').placeholder=(c.twocaptcha_api_key==='***설정됨***')?'설정됨 · 변경시만 입력':'변경시만 입력';loadOpenAIUsage()}
async function loadOpenAIUsage(){
  const r=await api('/openai/usage','GET');
  const c=await api('/twocaptcha/usage','GET');
  if(r&&r.ok){
    const m=r.month||{},t=r.today||{};const actual=r.actual_month_cost_usd;const cost=(actual==null?m.estimated_cost_usd:actual)||0;const remain=r.remaining_budget_usd;const src=r.source==='official_costs_api'?'공식 Costs API':'토큰 기준 예상';
    $('openaiUsage').innerHTML='<b style="color:var(--p)">이번 달 $'+Number(cost).toFixed(4)+'</b> · 남은 예산 '+(remain==null?'예산 미설정':'$'+Number(remain).toFixed(4))+'<br>오늘 요청 '+(t.requests||0)+'회 · 입력 '+Number(t.input_tokens||0).toLocaleString()+' · 출력 '+Number(t.output_tokens||0).toLocaleString()+' 토큰<br><span style="color:var(--g)">'+src+'</span> · '+esc(r.note||'')+(r.admin_error?'<br><span style="color:var(--y)">관리자 조회: '+esc(r.admin_error)+'</span>':'');
  } else {
    $('openaiUsage').innerHTML='<span style="color:var(--y)">OpenAI 상태를 불러오지 못했습니다.</span>';
  }
  if(c&&c.ok){
    const bal=(typeof c.balance==='number')?c.balance:0;const rem=(typeof c.remaining_usd==='number')?c.remaining_usd:bal;const delta=(typeof c.charged_since_last_check_usd==='number')?c.charged_since_last_check_usd:0;
    $('twocaptchaUsage').innerHTML = `
      <div class="twocap-shell">
        <div class="twocap-header">
          <span>2captcha</span>
          <span class="twocap-chip ok">정상 연결</span>
        </div>
        <div class="twocap-metrics">
          <div class="twocap-metric"><div class="label">잔액</div><div class="value green">$${Number(bal).toFixed(4)}</div></div>
          <div class="twocap-metric"><div class="label">남은 금액</div><div class="value blue">$${Number(rem).toFixed(4)}</div></div>
          <div class="twocap-metric"><div class="label">실시간 차감</div><div class="value amber">$${Number(delta).toFixed(4)}</div></div>
        </div>
        <div class="twocap-footer">
          <span>업데이트 ${esc(c.updated_at||'최근')}</span>
          <span>${esc(c.error||'2captcha 정상 연결')}</span>
        </div>
      </div>
    `;
  } else {
    $('twocaptchaUsage').innerHTML = `
      <div class="twocap-shell">
        <div class="twocap-header">
          <span>2captcha</span>
          <span class="twocap-chip warn">미연결</span>
        </div>
        <div class="twocap-metrics">
          <div class="twocap-metric"><div class="label">잔액</div><div class="value green">$0.0000</div></div>
          <div class="twocap-metric"><div class="label">남은 금액</div><div class="value blue">$0.0000</div></div>
          <div class="twocap-metric"><div class="label">실시간 차감</div><div class="value amber">$0.0000</div></div>
        </div>
        <div class="twocap-footer">
          <span>상태</span>
          <span>${esc((c&&c.error)||'API 키/활성화 상태를 확인하세요')}</span>
        </div>
      </div>
    `;
  }
}
async function loadPool(){const p=await api('/keywords','GET');if(!Array.isArray(p))return;$('poolCount').textContent=p.length+'개';$('poolCsv').value=p.map(k=>[k.지역||'',k.서비스||'',k.브랜드||''].join(',')).join('\n')}
async function savePool(append){const csv=$('poolCsv').value;const r=await api('/keywords','POST',{csv:csv,append:!!append});if(r&&r.ok){toast('풀 저장: '+r.count+'개');loadPool()}else if(r)toast('실패','er')}
async function clearPool(){const r=await api('/keywords','DELETE');if(r&&r.ok){toast('풀 비움');loadPool()}}
async function loadImages(){const p=await api('/images','GET');if(!Array.isArray(p))return;$('imgCount').textContent=p.length+'개';$('imgUrls').value=p.join('\n')}
async function saveImages(append){const text=$('imgUrls').value;const r=await api('/images','POST',{text:text,append:!!append});if(r&&r.ok){toast('이미지 URL: '+r.count+'개 저장');loadImages()}else if(r)toast('실패','er')}
async function clearImages(){const r=await api('/images','DELETE');if(r&&r.ok){toast('이미지 URL 비움');loadImages()}}
async function loadImageFiles(){const rows=await api('/images/files','GET');if(!Array.isArray(rows))return;const g=$('imgGallery');g.innerHTML=rows.length?rows.map(x=>`<div style="background:#0b1322;border:1px solid var(--line);border-radius:8px;padding:7px"><img src="${x.url}" alt="${esc(x.name)}" style="width:100%;height:105px;object-fit:cover;border-radius:5px"><div style="font-size:9px;color:var(--d);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;margin:5px 0" title="${esc(x.name)}">${esc(x.name)}</div><div class="row"><span style="font-size:9px;color:var(--d)">${Math.ceil(x.size/1024)}KB</span><span style="flex:1"></span><button class="btn btn-r btn-xs" onclick="deleteImageFile('${esc(x.name)}')">삭제</button></div></div>`).join(''):'<div style="color:var(--d);font-size:11px">저장된 이미지가 없습니다.</div>'}
async function uploadImages(){const f=$('imgFiles').files;if(!f.length)return toast('이미지를 선택하세요','er');const fd=new FormData();[...f].forEach(x=>fd.append('files',x));try{const r=await(await fetch('/api/images/upload',{method:'POST',body:fd})).json();if(r.ok){toast(r.count+'개 이미지 저장됨','ok');$('imgFiles').value='';loadImageFiles()}else toast(r.error||'업로드 실패','er')}catch(e){toast(e.message,'er')}}
async function deleteImageFile(name){if(!confirm('이 이미지를 삭제할까요?'))return;const r=await api('/images/file','DELETE',{name});if(r&&r.ok){toast('이미지 삭제됨','ok');loadImageFiles()}else toast(r&&r.error||'삭제 실패','er')}
async function uploadXlsx(){const f=$('poolXlsx').files[0];if(!f)return;const fd=new FormData();fd.append('file',f);try{const r=await(await fetch('/api/keywords/upload',{method:'POST',body:fd})).json();if(r&&r.ok){toast('엑셀 업로드: '+r.count+'개');loadPool()}else toast(r&&r.error||'업로드 실패','er')}catch(e){toast(e.message,'er')}$('poolXlsx').value=''}
async function genRandom(){const sid=$('poolSiteFilter').value;const n=parseInt($('poolN').value)||1;const r=await api('/generate/random','POST',{site_ids:sid?[sid]:[],count:n});if(r&&r.ok){if(r.generated!=null)toast(r.generated+'건 큐 등록 (랜덤 '+r.picks+'개)'+(r.blocked?` · 미허용 ${r.blocked} 제외`:''));else{$('gTitle').value=r.title;$('gContent').value=r.content;$('gLen').textContent=(r.content||'').length.toLocaleString()+'자';toast('랜덤 미리보기 생성')}loadOpenAIUsage()}else if(r)toast(r.error||'실패','er')}
// ---- 사이트 대량/허용/헬스 ----
async function bulkAdd(){const csv=$('bulkCsv').value.trim();if(!csv){toast('CSV 입력','er');return}const r=await api('/sites/bulk','POST',{csv:csv,permission:$('bulkPerm').checked});if(r&&r.ok){toast(r.added+'개 등록');$('bulkCsv').value='';renderSites()}}
async function bulkPermSet(v){const ids=getSiteIds();if(!ids.length){toast('사이트 선택','er');return}let note='';if(v){note=(prompt('운영자에게 받은 자동 게시 허용 근거를 입력하세요.\n예: 운영자 이메일 2026-08-31, 홍보게시판 이용정책 URL')||'').trim();if(note.length<5){toast('허용 근거를 5자 이상 입력해야 합니다','er');return}}const r=await api('/sites/permission','POST',{ids:ids,permission:v,permission_note:note});if(r&&r.ok){toast(r.changed+'개 '+(v?'허용 동의 기록':'미허용'));renderSites()}else if(r)toast(r.error||'변경 실패','er')}
async function toggleSitePermission(id,on){let note='';if(on){note=(prompt('운영자에게 받은 자동 게시 허용 근거를 입력하세요.')||'').trim();if(note.length<5){toast('허용 근거를 5자 이상 입력해야 합니다','er');renderSites();return}}const r=await api('/sites/permission','POST',{ids:[id],permission:on,permission_note:note});if(r&&r.ok){toast(on?'✅ 허용 동의 기록됨':'자동발행 잠금됨');renderSites()}else{toast((r&&r.error)||'변경 실패','er');renderSites()}}
async function saveSiteLimits(id){const daily=parseInt($('limD_'+id).value);const mins=parseInt($('limM_'+id).value);if(!Number.isFinite(daily)||daily<0||!Number.isFinite(mins)||mins<0){toast('건수와 간격은 0 이상의 숫자로 입력하세요','er');return}const r=await api('/sites/limits','POST',{id:id,daily_limit:daily,min_interval_minutes:mins});if(r&&r.ok){toast('✅ 하루 '+r.daily_limit+'건 · '+r.min_interval_minutes+'분 저장');renderSites()}else toast((r&&r.error)||'저장 실패','er')}
async function healthAll(){const ids=getSiteIds();if(!ids.length){toast('사이트 선택','er');return}toast(ids.length+'개 점검중...');for(const id of ids){await api('/sites/health/'+id,'POST')}renderSites();toast('점검 완료')}
// (예약 스케줄 UI 제거됨 — 회원별 스케줄러가 대체. 죽은 JS 정리)
// ---- 통계 ----
function tile(label,val,color){return `<div class="card" style="flex:1;min-width:110px;text-align:center;margin:0"><div style="font-size:22px;font-weight:700;color:${color}">${val}</div><div style="font-size:10px;color:var(--d)">${label}</div></div>`}
async function renderStats(){const s=await api('/stats','GET');if(!s)return;
$('statTop').innerHTML=tile('전체',s.total,'var(--t)')+tile('성공',s.ok,'var(--g)')+tile('실패',s.fail,'var(--r)')+tile('스킵',s.skip,'var(--y)')+tile('성공률',s.rate+'%','var(--p)')+tile('생존율',(s.alive_rate||0)+'%','var(--v)');
const rz=(s.reasons||[]);
$('statReasons').innerHTML=rz.length?('<div style="font-size:11px;color:var(--d);margin:10px 0 4px">실패 원인 분류</div><div style="display:flex;gap:6px;flex-wrap:wrap">'+rz.map(r=>`<span class="st st-f">${esc(r.reason)} ${r.n}</span>`).join('')+'</div>'):'';
$('statSurvival').innerHTML=(s.alive+s.dead)?`<div style="font-size:11px;color:var(--d);margin:12px 0 4px">발행글 생존 (검증됨 ${s.alive+s.dead}건)</div><span class="st st-ok">생존 ${s.alive}</span> <span class="st st-f">삭제 ${s.dead}</span>`:'';
const mx=Math.max(1,...s.by_day.map(d=>d.done+d.failed));
$('statDays').innerHTML='<div style="display:flex;align-items:flex-end;gap:4px;height:120px">'+s.by_day.map(d=>{const h=Math.round((d.done+d.failed)/mx*100);const go=d.done+d.failed?Math.round(d.done/(d.done+d.failed)*h):0;return `<div style="flex:1;text-align:center" title="${d.day}: 성공${d.done}/실패${d.failed}"><div style="height:${h}px;display:flex;flex-direction:column-reverse;border-radius:3px;overflow:hidden;background:var(--b)"><div style="height:${h-go}px;background:var(--r)"></div><div style="height:${go}px;background:var(--g)"></div></div><div style="font-size:8px;color:var(--d);margin-top:2px">${(d.day||'').slice(5)}</div></div>`}).join('')+'</div>';
$('statSites').innerHTML='<table><thead><tr><th>사이트</th><th>성공</th><th>실패</th><th>생존</th><th>삭제</th></tr></thead><tbody>'+s.by_site.map(x=>`<tr><td>${esc(x.site)}</td><td style="color:var(--g)">${x.done}</td><td style="color:var(--r)">${x.failed}</td><td style="color:var(--v)">${x.alive||0}</td><td style="color:var(--d)">${x.dead||0}</td></tr>`).join('')+'</tbody></table>'}

// ---- 사이트 목록 실시간 렌더 (새로고침 없이) ----
function siteRow(s){const st=s.status==='done'?'ok':s.status==='failed'?'f':'i';const nm=esc(s.name||(s.site_url||'').slice(0,20));const today=(s.posted_today||0);
const adminSource=['manual_admin','admin_bulk','legacy_admin','candidate_registered','verified_test'].includes(s.registration_source);const permitted=!!(s.permission&&adminSource);const perm='<label title="'+esc(s.permission_note||'허용 근거 미입력')+'" style="display:flex;align-items:center;gap:5px;white-space:nowrap;color:'+(permitted?'var(--g)':'var(--r)')+'"><input type="checkbox" style="width:auto" '+(permitted?'checked':'')+' onchange="toggleSitePermission(\''+esc(s.id)+'\',this.checked)">'+(permitted?'허용 동의됨':'발행 불가')+'</label>';
const lb=(s.permission&&adminSource)?'':'border-left:3px solid var(--r)';
const hdot=s.health?('<span title="상태점검: '+esc(s.health)+' ('+esc(s.health_at||'')+')" style="color:'+(s.health==='ok'?'var(--g)':'var(--r)')+'">●</span> '):'';
const pl=(s.platform||'auto');const plname=pl==='cafe24'?'Cafe24':(pl==='gnuboard'?'그누보드':'자동');const plcolor=pl==='cafe24'?'var(--v)':(pl==='gnuboard'?'var(--p)':'var(--d)');
const learned=s.learned&&s.learned.write_url;const lbadge=learned?' <span title="자가학습 셀렉터 저장됨" style="color:var(--g)">🎓</span>':'';
const cbadge=s.has_captcha?' <span title="캡차 감지 — 자동발행 제외('+esc(s.captcha_note||'')+')" style="color:var(--y)">🧩캡차</span>':'';
const pltag='<span style="font-size:9px;color:'+plcolor+'" title="발행 방식">'+plname+'</span>'+lbadge+cbadge;
const ss=s.signup_status||'';const signup=ss==='complete'?'<span class="st st-ok">가입완료·로그인저장</span>':(ss==='rejected'?'<span class="st st-f" title="'+esc(s.signup_reject_reason||'')+'">가입 제외 · 이메일인증</span>':(ss==='prepared'?'<span class="st st-y">가입정보만 준비됨</span>':(ss?'<span class="st st-y">가입 '+esc(ss)+'</span>':'')));const pv=s.signup_profile_version?'<span class="st st-i" title="최근측정 '+esc(s.signup_profile_measured_at||'')+'">가입학습 v'+s.signup_profile_version+(s.signup_profile_changed?' 변경':'')+'</span>':'';
const allLabel=ss==='rejected'?'가입 제외':(ss==='complete'?'✓ 로그인 저장됨':(['prepared','captcha_wait','email_wait'].includes(ss)?'실제 가입완료 확인':'올인원 가입'));const allClass=(ss==='rejected'||ss==='complete')?'btn-d':(['prepared','captcha_wait','email_wait'].includes(ss)?'btn-g':'btn-v');
const menu=`<details style="display:inline-block;position:relative"><summary class="btn btn-d btn-xs" style="list-style:none;cursor:pointer">관리 ▾</summary><div style="position:absolute;right:0;z-index:20;background:#101a2c;border:1px solid #33425f;border-radius:8px;padding:7px;min-width:125px;display:grid;gap:5px;box-shadow:0 8px 24px #0008"><button class="btn btn-p btn-xs" onclick="editSite('${esc(s.id)}')">편집</button><button class="btn ${allClass} btn-xs" onclick="signupAll('${esc(s.id)}','${esc(ss)}')">${allLabel}</button><button class="btn btn-d btn-xs" onclick="learnSignup('${esc(s.id)}')">가입폼 재학습</button><button class="btn btn-y btn-xs" onclick="dryRun('${esc(s.id)}')">발행 드라이런</button><button class="btn btn-d btn-xs" onclick="detectSite('${esc(s.id)}')">플랫폼 감지</button><button class="btn btn-v btn-xs" onclick="learnSite('${esc(s.id)}')">글쓰기 학습</button><button class="btn btn-d btn-xs" onclick="healthSite('${esc(s.id)}')">상태 점검</button><button class="btn btn-g btn-xs" onclick="testSite('${esc(s.id)}')">발행 테스트</button><button class="btn btn-r btn-xs" onclick="delSite('${esc(s.id)}')">삭제</button></div></details>`;
return `<tr data-id="${esc(s.id)}" style="${lb}"><td><input type="checkbox" class="cb" data-id="${esc(s.id)}"></td><td>${hdot}<b>${nm}</b><br>${signup} ${pv}</td><td>${perm}</td><td style="min-width:240px;max-width:340px"><a href="${esc(s.site_url||'#')}" target="_blank" rel="noopener" title="${esc(s.site_url||'')}" style="display:block;color:var(--p);word-break:break-all;line-height:1.45">${esc(s.site_url||'-')}</a><a href="${esc(s.site_url||'#')}" target="_blank" rel="noopener" class="btn btn-p btn-xs" style="display:inline-block;margin-top:5px">🔗 링크 열기</a></td><td style="color:var(--p)">${esc(s.bo_table||'')}<br>${pltag}</td><td style="color:var(--d);white-space:nowrap"><div>${today}/<input id="limD_${esc(s.id)}" type="number" min="0" max="10000" value="${s.daily_limit==null?3:s.daily_limit}" style="width:58px;padding:3px 5px">건</div><div style="margin-top:3px"><input id="limM_${esc(s.id)}" type="number" min="0" max="10080" value="${s.min_interval_minutes==null?60:s.min_interval_minutes}" style="width:58px;padding:3px 5px">분 <button class="btn btn-p btn-xs" onclick="saveSiteLimits('${esc(s.id)}')">저장</button></div></td><td><span class="st st-${st}" title="${esc(s.technical_block_reason||s.verification_fail_reason||'')}">${esc(s.status||'idle')}</span>${s.technical_block_reason?'<br><span style="color:var(--r);font-size:9px">'+esc(s.technical_block_reason)+'</span>':''}</td><td style="white-space:nowrap">${menu}</td></tr>`}
async function healthSite(id){toast('점검중...');const r=await api('/sites/health/'+id,'POST');if(r&&r.ok){const h=r.health;const pn=h.platform==='cafe24'?'Cafe24':'그누보드';toast((h.ok?'✅ 정상':'⚠️ 확인필요')+` [${pn}] 접속:${h.reachable?'O':'X'} 로그인폼:${h.login_form?'O':'X'} 글쓰기:${h.write_page?'O':'X'}`,h.ok?'ok':'er');renderSites()}else toast('실패','er')}
async function dryRun(id){toast('🧪 드라이런 실행중... (글은 올리지 않습니다, 최대 60초)');const r=await api('/sites/dryrun/'+id,'POST');if(!r){toast('실패','er');return}
const rows=(r.steps||[]).map(s=>`<tr><td>${s.ok?'<span style="color:var(--g)">✅</span>':'<span style="color:var(--r)">❌</span>'}</td><td><b>${esc(s.name)}</b></td><td style="color:var(--d)">${esc(s.detail)}</td></tr>`).join('');
const hdr=r.ok?'<span style="color:var(--g)">✅ 발행 가능 — 등록 직전까지 모두 통과</span>':'<span style="color:var(--r)">❌ 발행 불가 — 아래 ❌ 지점에서 막힘</span>';
$('pvTitle').textContent='드라이런 결과';
$('pvFrame').srcdoc='<!DOCTYPE html><html lang=ko><head><meta charset=utf-8><style>body{font-family:-apple-system,sans-serif;padding:18px;color:#222}table{width:100%;border-collapse:collapse;font-size:13px}td{padding:8px 6px;border-bottom:1px solid #eee;vertical-align:top}h3{margin-bottom:12px;font-size:16px}</style></head><body><h3>'+hdr+'</h3><table>'+rows+'</table><p style="color:#888;font-size:12px;margin-top:14px">※ 실제 게시글은 등록되지 않았습니다. 제목/본문을 채워보기만 하고 중단했습니다.</p></body></html>';
$('pvOverlay').style.display='block';renderSites()}
async function detectSite(id){toast('플랫폼 감지중...');const r=await api('/sites/detect/'+id,'POST');if(r&&r.ok){toast('감지됨: '+(r.platform==='cafe24'?'Cafe24':'그누보드'));renderSites()}else toast('실패: '+(r&&r.error||''),'er')}
async function learnSite(id){if(!confirm('실측 학습을 실행합니다.\n등록된 사이트 한 곳의 글쓰기 DOM과 폼을 확인해 저장하며 글은 제출하지 않습니다. 진행할까요?'))return;toast('🎓 비제출 실측 학습 중...(최대 60초)');const r=await api('/sites/learn/'+id,'POST');if(r&&r.ok&&r.learned){toast('🎓 실측 성공! 비제출 레시피 저장됨 ('+(r.learned.content_mode||'')+')');renderSites()}else if(r&&r.captcha){toast('🧩 CAPTCHA 감지 — 우회하지 않고 제외 상태를 저장했습니다','er');renderSites()}else if(r&&r.blocked){toast('⛔ 보안 차단 감지 — 우회하지 않고 측정 결과를 저장했습니다','er');renderSites()}else toast('실측 완료: '+((r&&(r.message||r.error))||'폼 미확정'),'er')}
async function renderSites(){const sites=await api('/sites','GET');if(!Array.isArray(sites))return;
$('siteTabCount').textContent=sites.length;
// 발행용 드롭다운: 실제 게시 성공 URL까지 검증된 허용 사이트만 표시
const sources=['manual_admin','admin_bulk','legacy_admin','candidate_registered','verified_test'];
const publishable=sites.filter(s=>!!s.permission&&sources.includes(s.registration_source)&&s.status!=='rejected'&&s.write_test_status==='passed'&&/^https?:\/\//.test(s.verified_post_url||''));
[['kwSiteFilter','전체 실게시 검증 사이트'],['wrSite','전체 실게시 검증 사이트'],['poolSiteFilter','전체 실게시 검증 사이트']].forEach(([id,label])=>{const sel=$(id);if(!sel)return;const cur=sel.value;sel.innerHTML='<option value="">'+label+' ('+publishable.length+'곳)</option>'+publishable.map(s=>`<option value="${esc(s.id)}">${esc(s.name||(s.site_url||'').slice(0,20))}</option>`).join('');if(publishable.some(s=>s.id===cur))sel.value=cur});
// 체크 상태 보존
const checked=new Set(getSiteIds());
if(!sites.length){$('siteList').innerHTML='<p style="color:var(--d);padding:30px;text-align:center">등록된 사이트가 없습니다</p>';return}
$('siteList').innerHTML='<table><thead><tr><th><input type="checkbox" id="allCb" onclick="document.querySelectorAll(\'.cb\').forEach(c=>c.checked=this.checked)"></th><th>이름</th><th>허용</th><th>URL</th><th>게시판</th><th>오늘</th><th>상태</th><th>동작</th></tr></thead><tbody>'+sites.map(siteRow).join('')+'</tbody></table>';
checked.forEach(id=>{const c=document.querySelector(`.cb[data-id="${id}"]`);if(c)c.checked=true})}

// ---- 통계/진행률 폴링 ----
async function poll(){const r=await api('/workers/stats','GET');if(!r)return;
$('q').textContent=r.queued||0;$('ok').textContent=r.success||0;$('fl').textContent=r.fail||0;$('sk').textContent=r.skipped||0;
const wstate=r.paused?'PAUSE':(r.active?'ON':'OFF');$('ws').textContent=wstate;$('ws').style.color=r.paused?'var(--y)':(r.active?'var(--g)':'var(--d)');
const total=r.total||0,done=r.done||0;
if(total>0){$('progCard').style.display='block';const pct=Math.round(done/total*100);$('progBar').style.width=pct+'%';$('progText').textContent=`${done} / ${total} (${pct}%)`+(r.skipped?` · 스킵 ${r.skipped}`:'')}else{$('progCard').style.display='none'}
if($('p-res').classList.contains('on'))renderHistory();if($('p-wlog').classList.contains('on')){renderWorkerLog();renderCaptchaTasks()}}

$('gContent').addEventListener('input',function(){$('gLen').textContent=this.value.length.toLocaleString()+'자'});
$('kwlist').addEventListener('input',function(){$('kwCount').textContent=parseList().length+'줄'});
renderSites();poll();loadPool();loadImages();loadImageFiles();loadWorkrooms();loadRegionTool();
setInterval(poll,2000);
setInterval(renderSites,4000);
</script>
</body></html>'''

def R(title,**kw):
    if title=='로그인': return render_template_string(HTML+LOGIN_HTML+'\n</body></html>',title=title,**kw)
    return render_template_string(HTML+DASH_HTML,title=title,**kw)

# ==================== Main ====================
def main():
    # 정상 환경 일치: 서버 타임존을 한국(Asia/Seoul)으로 (위장이 아니라 실제 운영지역과 맞춤)
    try:
        os.environ['TZ']='Asia/Seoul'; time.tzset()
    except Exception: pass
    cfg=load_config()
    host=os.environ.get('HOST','127.0.0.1'); port=int(os.environ.get('PORT','8888'))
    print(f'\n찌라시 마스터 v6 - 정직 발행 모드\nhttp://{host}:{port}\n브랜드: {cfg.get("brand","설정필요")}')
    if os.environ.get('CHIRASHI_PASSWORD') is None and cfg.get('password','admin1234')=='admin1234':
        print('⚠️  기본 비밀번호(admin1234) 사용 중 — 공개 서버라면 반드시 변경하세요!')
    # 재시작 복구: 미완료 작업 큐 복원
    try:
        n=recover_queue()
        if n: print(f'↻ 미완료 작업 {n}건 복구됨 (워커 시작 시 이어서 발행)')
    except Exception as e: print('복구 건너뜀:',e)
    # 시작 시 사이트 목록 최신화: 발행 막힌 사이트 자동 탈락
    try:
        dn=reconcile_sites()
        if dn: print(f'🧹 사이트 자동정리 {dn}건 (발행 불가 → 탈락)')
    except Exception as e: print('자동정리 건너뜀:',e)
    # 예약 발행 스케줄러 상시 가동
    try:
        threading.Thread(target=scheduler_loop,name='SCHED',daemon=True).start()
        print('⏰ 스케줄러 시작 (KST 기준)')
    except Exception as e: print('스케줄러 시작 실패:',e)
    # 지연 재시도 루프 / 텔레그램 명령 수신 / 발행글 생존 검증
    for fn,nm in [(retry_loop,'RETRY'),(telegram_loop,'TG'),(verify_loop,'VERIFY'),(member_scheduler_loop,'MSCHED'),(discover_loop,'DISCO')]:
        try: threading.Thread(target=fn,name=nm,daemon=True).start()
        except Exception as e: print(f'{nm} 시작 실패:',e)
    print('🔁 재시도·📱텔레그램·🔎검증 스레드 시작')
    # 헤드리스 서버(VPS)에서는 브라우저 자동 오픈 안 함
    if host in ('127.0.0.1','localhost') and os.environ.get('OPEN_BROWSER','1')!='0':
        try:
            import webbrowser; webbrowser.open(f'http://{host}:{port}')
        except: pass
    app.run(host=host,port=port,debug=False,threaded=True)

if __name__=='__main__':
    main()
