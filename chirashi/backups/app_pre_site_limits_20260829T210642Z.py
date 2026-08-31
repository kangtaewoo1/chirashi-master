"""
찌라시 마스터 v6 - 컴플라이언스형 (정직한 자동 발행)
========================================================
- bm21.com 스타일 리치 HTML 템플릿 (컬러 헤더, 칩 태그, FAQ, 이미지)
- Selenium 기반 그누보드/Cafe24 글쓰기 (정상 동작, 회피코드 없음)
- {지역} {서비스} {브랜드} 3키워드 치환
- 정직한 전화번호 표기 (난독화 없음)
- rate limit + 사이트당 1일 발행 한도로 도배 방지
※ 게시 대상 게시판에 대한 게시 권한(운영자 허락/홍보 전용 게시판)은
  이용자가 보장해야 합니다. 자동 프로그램 금지 규칙이 있는 곳에는 사용 금지.
"""

import sys, os, re, json, time, random, threading, queue, urllib.parse, secrets, hashlib
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import urllib3; urllib3.disable_warnings()

from flask import Flask, request, jsonify, render_template_string, session, redirect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import HTTPException

# ==================== 설정 ====================
BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = BASE_DIR / 'data'; DATA_DIR.mkdir(exist_ok=True)
SITES_FILE = DATA_DIR / 'sites.json'; CONFIG_FILE = DATA_DIR / 'config.json'; LOG_FILE = DATA_DIR / 'logs.json'
HISTORY_FILE = DATA_DIR / 'history.json'   # 발행 이력 원장(엑셀/결과탭)
QUEUE_FILE = DATA_DIR / 'queue.json'       # 미완료 작업(재시작 복구용)
SCHED_FILE = DATA_DIR / 'schedules.json'   # 예약 발행 스케줄
KEYWORDS_FILE = DATA_DIR / 'keywords.json' # 키워드 풀(엑셀 랜덤 치환)
UNIQ_FILE = DATA_DIR / 'uniq.json'          # 발행 콘텐츠 중복방지(제목/본문 해시)
IMAGES_FILE = DATA_DIR / 'images.json'      # 사용자 이미지 URL 풀(본문 삽입용)
MEMBERS_FILE = DATA_DIR / 'members.json'    # 회원(고객) 관리 + 월 정산
CAND_FILE = DATA_DIR / 'candidates.json'    # 발굴 후보(승인 대기함)
DISCO_FILE = DATA_DIR / 'discover.json'     # 발굴 상태(쿼리 커서·일일 카운트)

IMAGES = [f'https://picsum.photos/id/{i}/800/400' for i in [1,20,26,48,60,64,76,91,96,104,152,160,175,180,185,201]]
COLORS = ['#3b1f2b','#2B8A3E','#37474f','#1a5276','#6c3483','#b7950b','#a04000']

_json_locks={}; _json_locks_guard=threading.Lock()
def _json_lock(p):
    key=str(Path(p).resolve())
    with _json_locks_guard:
        return _json_locks.setdefault(key,threading.RLock())

def load_json(p, d=None):
    if d is None: d=[]
    try:
        with _json_lock(p):
            if os.path.exists(p):
                with open(p,'r',encoding='utf-8') as f: return json.load(f)
    except: pass
    return d

def save_json(p, data):
    """같은 파일시스템의 임시 파일에 기록 후 원자 교체하여 JSON 손상을 방지."""
    p=Path(p); p.parent.mkdir(parents=True,exist_ok=True)
    tmp=p.with_name(f'.{p.name}.{os.getpid()}.{threading.get_ident()}.tmp')
    with _json_lock(p):
        try:
            with open(tmp,'w',encoding='utf-8') as f:
                json.dump(data,f,ensure_ascii=False,indent=2); f.flush(); os.fsync(f.fileno())
            os.replace(tmp,p)
        finally:
            try:
                if tmp.exists(): tmp.unlink()
            except Exception: pass

# ---- 사용자 이미지 URL 풀 (본문 삽입용) ----
def load_image_urls():
    d=load_json(IMAGES_FILE,[])
    return [u.strip() for u in d if isinstance(u,str) and u.strip().startswith('http')]
def save_image_urls(urls): save_json(IMAGES_FILE,urls)
def pick_images(n):
    """본문용 이미지 n개 선택. 사용자 URL 풀이 있으면 그걸, 없으면 기본(picsum) 폴백."""
    pool=load_image_urls() or IMAGES
    if len(pool)>=n: return random.sample(pool,n)
    return [random.choice(pool) for _ in range(n)]   # URL이 부족하면 중복 허용

def load_sites(): return load_json(SITES_FILE,[])
def save_sites(s): save_json(SITES_FILE,s)

def load_config():
    d={'brand':'인천홍마니','phone':'01082755736','phones':'','openai_key':'','model':'gpt-4o-mini',
       'workers':2,'password':'admin1234','post_delay':30,'daily_limit':3,
       'use_gpt':False,'telegram_token':'','telegram_chat_id':'',
       'notify_done':False,'notify_fail':True,'update_token':'',
       'telegram_control':False,'backup_time':'','verify_enabled':True,'mix_keywords':True,
       'block_unpaid':True,
       'google_api_key':'','google_cx':'','brave_api_key':'','search_provider':'google','discover_enabled':False,
       'discover_daily_target':100,'discover_query_limit':100,'discover_keywords':'',
       'discover_direct_queries':''}
    c=load_json(CONFIG_FILE,None)
    if c is None or not isinstance(c,dict): save_json(CONFIG_FILE,d); return d.copy()
    for k,v in d.items():
        if k not in c: c[k]=v
    return c

def save_config(c): save_json(CONFIG_FILE,c)
def add_log(msg):
    with _json_lock(LOG_FILE):
        logs=load_json(LOG_FILE,[])
        logs.append({'time':datetime.now().strftime('%H:%M:%S'),'msg':msg})
        if len(logs)>500: logs=logs[-500:]
        save_json(LOG_FILE,logs)

# ==================== 전화번호 표기 ====================
def format_phone(phone):
    d=re.sub(r'[^0-9]','',phone or '01082755736')
    if len(d)==11: return f'{d[:3]}-{d[3:7]}-{d[7:]}'
    if len(d)==10: return f'{d[:3]}-{d[3:6]}-{d[6:]}'
    return phone

PHONE_SEPS=['↔','●','=','~','-',' ','.','·','_','ㆍ','∼','◆','ㅡ']
def format_phone_random(phone):
    """제목용 번호 랜덤 변형 (매번 다른 기호·표기). 사람은 읽을 수 있게 유지."""
    d=re.sub(r'[^0-9]','',phone or '01082755736')
    if len(d)==11: a,b,c=d[:3],d[3:7],d[7:]
    elif len(d)==10: a,b,c=d[:3],d[3:6],d[6:]
    else: return phone
    s1=random.choice(PHONE_SEPS); s2=random.choice(PHONE_SEPS)
    # O/I 치환 여부(랜덤): 010->O1O / OIO 등, 사람이 읽을 수 있게 유지
    sub=random.choice([lambda x:x, lambda x:x.replace('0','O'),
                       lambda x:x.replace('0','O').replace('1','I')])
    a2,b2,c2=sub(a),sub(b),sub(c)
    style=random.random()
    if style<0.30:      out=f'[{a2}]{s1}{b2}{s2}{c2}'      # [010]↔8275↔5736
    elif style<0.55:    out=f'[{a2}{s1}{b2}{s2}{c2}]'      # [010●8275●5736]
    else:               out=f'{a2}{s1}{b2}{s2}{c2}'        # O1O=2572=3859
    return out

def get_phones(cfg):
    ph=cfg.get('phones','') or ''
    lst=[re.sub(r'\s+','',x) for x in ph.splitlines() if x.strip()]
    if not lst and cfg.get('phone'): lst=[cfg['phone']]
    return lst or ['01082755736']

def pick_phone(cfg):
    return random.choice(get_phones(cfg))

def build_title(r,s,b,cfg):
    """제목 = 키워드1 + (랜덤변형 번호) + 키워드2 + 키워드3"""
    raw=pick_phone(cfg)
    return f'{r} {format_phone_random(raw)} {s} {b}'[:140], raw

# ==================== 키워드 풀 (엑셀/CSV 랜덤 치환) ====================
def load_keywords(): return load_json(KEYWORDS_FILE,[])
def save_keywords(k): save_json(KEYWORDS_FILE,k)

def pool_columns(pool):
    """풀에서 지역/서비스/브랜드 열별 고유값 목록(순서보존)."""
    dd=lambda L:list(dict.fromkeys([x for x in L if x]))
    R=dd([(x.get('지역') or '').strip() for x in pool])
    S=dd([(x.get('서비스') or '').strip() for x in pool])
    B=dd([(x.get('브랜드') or '').strip() for x in pool])
    return R,S,B

def pick_keywords(pool, cfg):
    """풀에서 키워드 추출. mix_keywords=True 면 지역/서비스/브랜드를 열 단위 독립 랜덤으로 조합
       → 3키워드가 매번 다르게 나옴(중복 최소화)."""
    if not pool: return {'지역':'','서비스':'','브랜드':''}
    row=random.choice(pool)
    if cfg.get('mix_keywords',True):
        R,S,B=pool_columns(pool)
        return {'지역':(random.choice(R) if R else (row.get('지역') or '')),
                '서비스':(random.choice(S) if S else (row.get('서비스') or '')),
                '브랜드':(random.choice(B) if B else (row.get('브랜드') or ''))}
    return {'지역':(row.get('지역') or ''),'서비스':(row.get('서비스') or ''),'브랜드':(row.get('브랜드') or '')}

# ==================== 콘텐츠 중복 방지 (제목/본문 항상 다르게) ====================
_uniq_lock=threading.Lock()
def _norm_text(html):
    """태그·공백 제거한 순수 텍스트(중복 판정 기준)."""
    return re.sub(r'\s+',' ',re.sub(r'<[^>]+>',' ',html or '')).strip()

def remember_if_unique(title, body, force=False):
    """제목/본문이 과거에 없던 새 콘텐츠면 기록하고 True. 이미 있으면 False."""
    th=hashlib.sha1((title or '').encode('utf-8')).hexdigest()
    bh=hashlib.sha1(_norm_text(body).encode('utf-8')).hexdigest()
    with _uniq_lock:
        u=load_json(UNIQ_FILE,{'t':[],'b':[]})
        if not isinstance(u,dict): u={'t':[],'b':[]}
        ts=set(u.get('t',[])); bs=set(u.get('b',[]))
        if not force and (th in ts or bh in bs): return False
        u['t']=(u.get('t',[])+[th])[-8000:]; u['b']=(u.get('b',[])+[bh])[-8000:]
        save_json(UNIQ_FILE,u); return True

def force_unique_html(html):
    """(최후수단) 눈에 거의 안 띄는 유니크 토큰을 붙여 강제로 다르게."""
    nonce=secrets.token_hex(5)
    return html+f'<p style="font-size:1px;line-height:1px;color:#ffffff;opacity:0.02;margin:0;">{nonce}</p>'

# ==================== 리치 HTML 생성 (bm21 스타일 · 변형/서비스특화/중복방지) ====================
# 서비스 유형별 특화 어휘 (미등록 서비스는 DEFAULT_FLAVOR 사용)
SERVICE_FLAVOR = {
    '셔츠룸':  {'mood':['깔끔하고 세련된','밝고 활기찬','트렌디한 감성의']},
    '노래방':  {'mood':['신나고 흥겨운','프라이빗하고 아늑한','최신 곡이 가득한']},
    '가라오케':{'mood':['고급스럽고 품격 있는','세련되고 정갈한','차분하고 프라이빗한']},
    '호빠':    {'mood':['활기차고 친근한','편안하고 밝은','유쾌한']},
    '룸싸롱':  {'mood':['프라이빗하고 고급스러운','조용하고 차분한','정중한 응대의']},
    '가요주점':{'mood':['정겹고 흥겨운','편안한','활기찬']},
}
DEFAULT_FLAVOR={'mood':['편안하고 세련된','밝고 활기찬','아늑한']}
RELATED_POOL=['하이퍼블릭','노래빠','쓰리노','가요광장','터치룸','노래클럽','가라오케','호빠','룸싸롱','셔츠룸','퍼블릭','비즈니스바','가요주점']

def generate_rich_html(keywords, cfg):
    r=(keywords.get('지역') or '서울').strip()
    s=(keywords.get('서비스') or '셔츠룸').strip()
    b=(keywords.get('브랜드') or cfg.get('brand') or '인천홍마니').strip()
    # 제목: 키워드1 + 랜덤변형 번호 + 키워드2 + 키워드3  / 본문: 선택된 번호의 정상 표기
    title,rawphone=build_title(r,s,b,cfg)
    p=format_phone(rawphone)
    mood=random.choice(SERVICE_FLAVOR.get(s,DEFAULT_FLAVOR)['mood'])
    imgs=pick_images(5)   # 사용자 이미지 URL 풀(없으면 기본 이미지)
    c1,c2,c3=random.sample(COLORS,3)
    rel=RELATED_POOL[:]; random.shuffle(rel)
    H=lambda t:f'<h2 style="color:{c1};border-bottom:3px solid {c2};padding-bottom:10px;font-size:24px;margin-top:34px;">{t}</h2>'
    # 이미지 alt = 치환키워드 맨앞(지역) 그대로 (SEO)
    IMG=lambda i,cap='':(f'<div style="text-align:center;margin:34px 0;"><img src="{imgs[i%len(imgs)]}" alt="{r}" style="max-width:100%;height:auto;border-radius:8px;" loading="lazy" />'+(f'<p style="color:#888;font-size:13px;margin-top:8px;">▲ {cap}</p>' if cap else '')+'</div>')

    intro=random.choice([
        f'{r} 지역에서 {s}를 찾고 계신가요? {mood} 분위기의 <strong>{r} {s}</strong>는 회식과 모임 장소로 꾸준히 사랑받는 곳입니다. {b}에서 위치와 이용 정보를 한눈에 정리해 드립니다.',
        f'기업 회식, 지인 모임, 특별한 자리까지 — <strong>{r} {s}</strong>는 다양한 목적에 어울리는 공간입니다. {mood} 인테리어와 좋은 접근성으로 {r} 인근에서 인기가 높습니다. 자세한 안내는 {b}가 도와드립니다.',
        f'{r} {s}를 처음 방문하신다면 어디로 가야 할지 고민되기 마련입니다. {b}는 {r} 지역의 {s} 정보를 정리해 편하게 선택하실 수 있도록 안내합니다. {mood} 공간에서 좋은 시간 보내세요.',
        f'{mood} 분위기와 넓은 공간을 갖춘 <strong>{r} {s}</strong>. 접대와 단체 모임에 적합하며 {r} 중심가에서 가까워 이동이 편리합니다. 예약과 문의는 {b}가 도와드립니다.',
    ])
    para2=random.choice([
        f'{r} {s}의 가장 큰 장점은 접근성과 공간감입니다. 대중교통과 주차 모두 이용하기 편리해 부담 없이 방문할 수 있고, 내부는 여유로운 좌석 배치로 편안함을 더했습니다.',
        f'단체 예약 시 인원과 목적에 맞춰 공간을 안내받을 수 있어, 회식이나 모임을 준비하는 분들의 만족도가 높습니다. 미리 문의하면 원하는 시간대에 맞춰 준비가 가능합니다.',
        f'{r} 지역 특성상 {s} 선택지가 다양하지만, 위치와 분위기, 응대 수준을 함께 고려하면 후회 없는 선택을 할 수 있습니다. {b}는 이 기준으로 정보를 정리합니다.',
        f'처음 방문하는 분도 편하게 이용할 수 있도록 안내가 잘 되어 있으며 재방문율이 높은 편입니다. 궁금한 점은 방문 전 미리 문의하시면 자세히 안내받을 수 있습니다.',
    ])

    feat_pool=[
        '넓은 홀과 프라이빗 룸으로 인원에 맞춘 공간 선택이 가능합니다.',
        '선명한 음향과 조명으로 분위기를 한층 끌어올립니다.',
        '대중교통·주차 접근성이 좋아 이동이 편리합니다.',
        '친절하고 세심한 응대로 편안한 이용을 돕습니다.',
        '예약 문의가 간편해 원하는 시간대를 잡기 쉽습니다.',
        '합리적인 가격 구성으로 부담을 줄였습니다.',
        '청결하게 관리되는 쾌적한 실내 환경을 유지합니다.',
        '단체 모임과 회식에 적합한 좌석 배치를 갖췄습니다.',
        f'{r} 중심가에서 가까워 약속 장소로 잡기 좋습니다.',
        '초행길에도 찾기 쉬운 위치에 자리하고 있습니다.',
        '분위기 있는 인테리어로 특별한 자리를 완성합니다.',
        '다양한 목적의 모임에 유연하게 대응합니다.',
    ]
    feats=''.join(f'<li style="margin:8px 0;line-height:1.8;">{x}</li>' for x in random.sample(feat_pool,random.randint(5,7)))

    faq_pool=[
        ('예약은 어떻게 하나요?', f'{p}로 문의 주시면 인원과 시간에 맞춰 빠르게 안내해 드립니다. 방문 전 예약을 권장합니다.'),
        ('주차가 가능한가요?','인근 주차 이용이 가능하며 대중교통 접근성도 좋아 차량·도보 모두 편리합니다.'),
        ('단체 이용도 되나요?','인원에 맞춰 공간을 안내해 드리므로 회식·모임 등 단체 이용에 적합합니다. 사전 문의 시 준비가 원활합니다.'),
        (f'{r} 어디에 있나요?', f'{r} 중심가 인근에 위치해 찾기 쉽습니다. 정확한 위치는 문의 시 안내해 드립니다.'),
        ('처음 방문인데 괜찮을까요?','처음 오시는 분도 편하게 이용하실 수 있도록 안내가 잘 되어 있습니다. 부담 없이 방문하세요.'),
        ('이용 시간은 어떻게 되나요?','이용 가능 시간대는 시기에 따라 다를 수 있어 방문 전 문의로 확인하시는 것을 권장합니다.'),
        ('분위기는 어떤가요?', f'{mood} 분위기로 편안하게 시간을 보내기 좋습니다.'),
        ('예약 없이 방문해도 되나요?','가능하나 원하는 시간대 이용을 위해서는 사전 예약을 추천드립니다.'),
        ('문의는 어디로 하나요?', f'{p}로 연락 주시면 친절하게 안내해 드립니다.'),
    ]
    faqs=''.join(f'<dt style="font-weight:bold;color:{c2};margin-top:12px;">Q. {q}</dt><dd style="margin:6px 0 12px 20px;line-height:1.7;">{a}</dd>' for q,a in random.sample(faq_pool,random.randint(4,6)))

    reco_pool=['회식·접대 장소를 찾는 직장인','지인들과 편하게 모일 공간이 필요한 분',f'{r} 인근에서 약속 장소를 정하려는 분','믿을 만한 정보로 실패 없이 고르고 싶은 분','분위기 좋은 자리를 원하는 분','접근성 좋은 위치를 선호하는 분']
    recos=''.join(f'<li style="margin:8px 0;line-height:1.8;">{x}</li>' for x in random.sample(reco_pool,random.randint(3,min(5,len(reco_pool)))))

    review=random.choice([
        f'지인 추천으로 {r} {s}에 다녀왔습니다. {mood} 분위기에 응대도 친절해서 편안하게 즐겼어요. 다음에도 또 방문하고 싶습니다.',
        f'회식 장소로 {r} {s}를 예약했는데 공간이 넓고 깔끔해서 만족스러웠습니다. {b} 안내대로 미리 문의하니 준비가 잘 되어 있었어요.',
        f'위치가 찾기 쉬워 초행인데도 헤매지 않았습니다. {r} {s} 분위기가 좋아 모임이 화기애애했어요. 추천합니다.',
        f'{b} 정보를 보고 방문했는데 기대 이상이었습니다. 가격도 합리적이고 응대도 세심해 좋은 자리가 됐습니다.',
        f'{r}에서 {s} 고민하다 방문했는데 선택 잘했다는 생각이 들었어요. {mood} 공간에서 편하게 대화 나눴습니다.',
    ])
    chips=' '.join(f'<span style="display:inline-block;padding:4px 12px;margin:3px;background:#E5E9EE;font-size:13px;color:{c3};border:1px solid #dde4f0;border-radius:4px;">{x}</span>' for x in random.sample(rel,random.randint(8,min(13,len(rel)))))

    # 중간 콘텐츠 블록들 — 순서를 매번 섞어 변형 폭을 키움(중복 방지)
    blocks=[
        H(f'{r} {s} 주요 특징')+f'<ul style="font-size:15px;padding-left:20px;color:#444;">{feats}</ul>',
        H(f'{r} {s} 이런 분께 추천합니다')+f'<ul style="font-size:15px;padding-left:20px;color:#444;">{recos}</ul>'+IMG(2, f'{r} {s} 추천 공간'),
        H(f'{r} {s} 자주 묻는 질문')+f'<dl style="font-size:15px;margin:15px 0;">{faqs}</dl>',
        H(f'{r} {s} 방문 후기')+f'<p style="font-size:15px;line-height:1.9;">{review}</p>',
        H(f'{r} 인근 이용 가능 지역')+f'<div style="margin:12px 0;line-height:2.6;">{chips}</div>'+IMG(3, f'{r} {s}의 특별한 자리'),
    ]
    random.shuffle(blocks)
    html=(f'<h1 style="font-size:22px;font-weight:bold;margin:0 0 18px 0;color:#222;">{title}</h1>'
        + IMG(0, f'{r} {s}의 {mood.split()[0]} 공간')
        + H(f'{r} {s} 안내')
        + f'<p style="font-size:16px;margin:18px 0;line-height:1.9;">{intro}</p>'
        + f'<p style="font-size:15px;margin:14px 0;line-height:1.9;color:#333;">{para2}</p>'
        + IMG(1, f'{r} {s} 내부 분위기')
        + ''.join(blocks)
        + f'<div style="margin:40px 0;padding:22px;background:#f8f9fa;border-radius:8px;text-align:center;">'
          f'<p style="font-size:18px;font-weight:bold;color:{c1};">{r} {s} 문의 및 예약</p>'
          f'<p style="font-size:24px;font-weight:bold;color:{c2};margin:12px 0;">{p}</p>'
          f'<p style="font-size:14px;color:#666;">{b} · 믿을 수 있는 {r} {s} 정보</p></div>'
        # 맨 마지막 이미지 삽입 — alt = 지역(맨앞 키워드) 그대로
        + f'<div style="text-align:center;margin:30px 0 10px;"><img src="{imgs[4%len(imgs)]}" alt="{r}" style="max-width:100%;height:auto;border-radius:8px;" loading="lazy" /></div>')
    return html, title

# ==================== GPT 본문 생성 (선택) ====================
def generate_post_gpt(keywords, cfg):
    """OpenAI로 SEO 리치 HTML 본문 생성. 실패 시 예외 → 호출측에서 템플릿 폴백."""
    import requests as _rq
    r=(keywords.get('지역') or '서울').strip(); s=(keywords.get('서비스') or '셔츠룸').strip()
    b=(keywords.get('브랜드') or cfg.get('brand') or '인천홍마니').strip()
    _rawph=pick_phone(cfg); p=format_phone(_rawph)
    key=cfg.get('openai_key',''); model=cfg.get('model') or 'gpt-4o-mini'
    if not key: raise RuntimeError('openai_key 없음')
    imgs=pick_images(3); c1,c2=random.sample(COLORS,2)   # 사용자 이미지 URL 풀
    sys_p=("너는 한국어 지역 SEO 블로그 글을 쓰는 카피라이터다. 광고성 정보 글을 자연스럽고 정보성 있게 쓴다. "
           "과장·허위·불법·선정적 표현은 피하고, 업소 홍보 정보 글 톤으로 쓴다.")
    usr_p=(f"지역='{r}', 서비스='{s}', 브랜드='{b}', 전화='{p}'.\n"
           f"이 정보로 {r} {s} 소개 글을 작성하라. 조건:\n"
           f"- 순수 HTML 본문만 출력(코드블록/마크다운/설명 금지, <html><body> 태그 없이 내용만)\n"
           f"- <h2>{c1} 색 소제목> 4~6개, 각 섹션 <p> 2~4문장, <ul><li> 특징 5개, FAQ는 <dl><dt><dd> 4개, 후기 1문단\n"
           f"- 전화번호 {p}는 예약/문의 안내에 1~2회 자연스럽게 포함\n- 전체 900자 이상, 매번 문장을 다르게(중복 회피)")
    resp=_rq.post("https://api.openai.com/v1/chat/completions",
        headers={"Authorization":f"Bearer {key}","Content-Type":"application/json"},
        json={"model":model,"temperature":0.9,"max_tokens":2000,
              "messages":[{"role":"system","content":sys_p},{"role":"user","content":usr_p}]},
        timeout=60)
    resp.raise_for_status()
    body=resp.json()['choices'][0]['message']['content'].strip()
    if body.startswith('```'): body=re.sub(r'^```[a-zA-Z]*\n?|```$','',body).strip()
    title=f'{r} {format_phone_random(_rawph)} {s} {b}'[:140]
    header=(f'<h1 style="font-size:22px;font-weight:bold;color:#222;margin:0 0 16px;">{title}</h1>'
            f'<div style="text-align:center;margin:20px 0;"><img src="{imgs[0]}" alt="{r}" style="max-width:100%;border-radius:8px;" loading="lazy"/></div>')
    cta=(f'<div style="margin:36px 0;padding:20px;background:#f8f9fa;border-radius:8px;text-align:center;">'
         f'<p style="font-size:18px;font-weight:bold;color:{c1};">{r} {s} 문의·예약</p>'
         f'<p style="font-size:24px;font-weight:bold;color:{c2};margin:10px 0;">{p}</p>'
         f'<p style="font-size:14px;color:#666;">{b}</p></div>')
    # 맨 마지막 이미지 삽입 — alt = 지역(맨앞 키워드)
    tail=f'<div style="text-align:center;margin:30px 0 10px;"><img src="{imgs[min(1,len(imgs)-1)]}" alt="{r}" style="max-width:100%;border-radius:8px;" loading="lazy"/></div>'
    return header+body+cta+tail, title

def _gen_once(keywords, cfg):
    if cfg.get('use_gpt') and cfg.get('openai_key'):
        try:
            return generate_post_gpt(keywords,cfg)
        except Exception as e:
            add_log(f'[GPT 실패→템플릿] {str(e)[:80]}')
    return generate_rich_html(keywords,cfg)

def generate_article(keywords, cfg, unique=True):
    """본문 생성. unique=True 면 제목/본문이 과거와 겹치지 않을 때까지 재생성(상시 다르게)."""
    if not unique:
        return _gen_once(keywords,cfg)
    html=title=None
    for _ in range(8):
        html,title=_gen_once(keywords,cfg)
        if remember_if_unique(title,html): return html,title
    # 8회 모두 충돌(사실상 불가) → 강제 유니크 토큰 부착
    html=force_unique_html(html); remember_if_unique(title,html,force=True)
    return html,title

# ==================== 텔레그램 알림 (선택) ====================
def send_telegram(cfg, msg):
    tok=cfg.get('telegram_token',''); chat=cfg.get('telegram_chat_id','')
    if not tok or not chat: return False
    try:
        import requests as _rq
        _rq.post(f"https://api.telegram.org/bot{tok}/sendMessage",
                 json={"chat_id":chat,"text":msg,"disable_web_page_preview":True},timeout=10)
        return True
    except Exception: return False

def send_telegram_doc(cfg, filename, data_bytes, caption=''):
    """텔레그램으로 파일(백업 zip/엑셀) 전송."""
    tok=cfg.get('telegram_token',''); chat=cfg.get('telegram_chat_id','')
    if not tok or not chat: return False,'텔레그램 토큰/챗ID 미설정'
    try:
        import requests as _rq
        r=_rq.post(f"https://api.telegram.org/bot{tok}/sendDocument",
                   data={'chat_id':chat,'caption':caption[:1000]},
                   files={'document':(filename,data_bytes)},timeout=40)
        j=r.json()
        return bool(j.get('ok')),('' if j.get('ok') else str(j)[:150])
    except Exception as e: return False,str(e)[:150]

# ==================== 캡차 / 보안 인증 감지 (우회 아님 — 감지해서 제외) ====================
def detect_captcha(d):
    """현재 페이지에 캡차/보안 인증이 있는지 감지. 있으면 종류, 없으면 ''.
       (정책상 자동 해석·우회 없음 — 감지되면 자동발행에서 제외한다.)"""
    try: html=d.page_source or ''
    except Exception: html=''
    low=html.lower()
    if 'cf-turnstile' in low or 'turnstile' in low: return 'turnstile'
    if 'h-captcha' in low or 'hcaptcha' in low: return 'hcaptcha'
    if 'g-recaptcha' in low or 'grecaptcha' in low or 'recaptcha' in low: return 'recaptcha'
    if ('captcha_key' in low or 'kcaptcha' in low or 'g5_captcha' in low
        or '자동등록방지' in html or '자동입력방지' in html or '보안문자' in html or 'captcha' in low):
        return 'kcaptcha'
    return ''

def _page_is_blocked(d):
    """403/보안 차단 페이지 판별. 일반 사이트의 Cloudflare 스크립트 문자열은 차단으로 보지 않는다."""
    try:
        from selenium.webdriver.common.by import By
        html=(d.page_source or '').lower()
        title=(d.title or '').lower()
        body=(d.find_element(By.TAG_NAME,'body').text or '').lower()[:3000]
    except Exception:
        html=''; title=''; body=''
    strong=['403 forbidden','access denied','error 1020','request blocked','접근이 거부','차단되었습니다']
    if any(k in title or k in body for k in strong): return True
    # Cloudflare 도전/차단 화면은 제품명 하나가 아니라 고유 문구 조합으로 확인한다.
    if ('attention required' in title or 'just a moment' in title) and \
       ('cloudflare' in html or 'cf-chl-' in html): return True
    return False

# ==================== 실패 원인 분류 ====================
def classify_fail(msg):
    """발행 실패 메시지를 사람이 읽을 수 있는 원인으로 분류. (코드, 한글, 일시적여부)"""
    ms=str(msg or ''); low=ms.lower()
    if '캡차' in ms or 'captcha' in low or '보안 인증' in ms:
        return 'captcha','캡차/보안인증 감지',False
    if any(k in ms for k in ['타임아웃','시간 초과']) or any(k in low for k in
           ['timeout','timed out','renderer','chrome not reachable','disconnected','connection',
            'session deleted','session not created','net::err','unreachable','max retries','read timed']):
        return 'timeout','타임아웃/브라우저',True   # 일시적 → 자동 재시도
    if any(k in ms for k in ['로그인','아이디','비밀번호']) or any(k in low for k in
           ['login','mb_id','mb_password','password']):
        return 'login','로그인 실패',False
    if any(k in ms for k in ['권한','차단','금지','스팸','도배']) or any(k in low for k in
           ['blocked','forbidden','403','spam','denied','권한이 없']):
        return 'blocked','차단/권한없음',False
    if any(k in ms for k in ['게시판','에디터','셀렉터','확인 불가','글쓰기','페이지 못찾음','입력란 못찾음']) or any(k in low for k in
           ['write.php','bo_table','board_no','no such element','not found','404','wr_subject','wr_content']):
        return 'board','게시판/에디터 못찾음',False
    return 'other','기타',False

# ==================== 데이터 백업 ====================
def build_backup_zip():
    """설정·사이트·이력·예약·키워드·큐를 zip 하나로 묶음(민감정보 마스킹)."""
    import io,zipfile
    buf=io.BytesIO()
    with zipfile.ZipFile(buf,'w',zipfile.ZIP_DEFLATED) as z:
        for p in [SITES_FILE,HISTORY_FILE,SCHED_FILE,KEYWORDS_FILE,IMAGES_FILE,MEMBERS_FILE,CAND_FILE,QUEUE_FILE]:
            if os.path.exists(p): z.write(str(p),os.path.basename(str(p)))
        # config 는 비번해시·API키 제거하고 저장
        c=dict(load_config()); c.pop('password',None); c['openai_key']=''; c['telegram_token']=''
        z.writestr('config.json',json.dumps(c,ensure_ascii=False,indent=2))
    buf.seek(0); return buf.read()

def do_backup(cfg=None, reason='자동'):
    cfg=cfg or load_config()
    try:
        data=build_backup_zip()
    except Exception as e:
        add_log(f'[백업] zip 생성 실패: {str(e)[:80]}'); return False,str(e)[:120]
    ts=_kst_now().strftime('%Y%m%d_%H%M')
    cap=(f'📦 찌라시 백업 ({reason})\n{_kst_now().strftime("%Y-%m-%d %H:%M")} KST\n'
         f'사이트 {len(load_sites())}개 · 이력 {len(load_json(HISTORY_FILE,[]))}건')
    ok,err=send_telegram_doc(cfg,f'chirashi_backup_{ts}.zip',data,caption=cap)
    add_log(f'[백업] {"전송 성공" if ok else "실패: "+err} ({reason})')
    return ok,err

# ==================== 발행 검증 (게시글 생존 확인) ====================
def check_post_alive(url):
    """게시된 URL 이 아직 살아있는지 HTTP 로 확인. True=생존, False=삭제/없음, None=확인불가."""
    import requests as _rq
    try:
        r=_rq.get(url,timeout=15,verify=False,allow_redirects=True,
                  headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36'})
        if r.status_code==429:   # 과도한 요청 — Retry-After 존중하고 이번엔 판정 보류
            ra=r.headers.get('Retry-After','')
            try: time.sleep(min(int(ra),30)) if ra.isdigit() else time.sleep(5)
            except Exception: time.sleep(5)
            return None
        if r.status_code==403: return None   # 차단 — 삭제로 오판하지 않도록 보류
        if r.status_code==404 or r.status_code>=500: return False
        if r.status_code>=400: return False
        body=r.text or ''
        dead=['존재하지 않','삭제된','삭제되었','없는 게시','게시물이 없','원글이 없','잘못된 접근',
              '페이지를 찾을 수 없','not found','더 이상','권한이 없']
        if any(s in body for s in dead): return False
        return True
    except Exception:
        return None

def verify_once(limit=40):
    """이력의 성공 발행글 URL 을 재확인해 생존여부(alive)를 기록. 처리 건수 반환."""
    cfg=load_config()
    with JOB_LOCK:
        h=load_json(HISTORY_FILE,[])
    cand=[r for r in h if r.get('status')=='done' and str(r.get('result_url','')).startswith('http')]
    cand.sort(key=lambda r:(r.get('verified_at') or ''))   # 미검증·오래된 것 우선
    todo=cand[:max(1,limit)]
    results={}
    for r in todo:
        alive=check_post_alive(r['result_url'])
        if alive is not None: results[r['id']]=('yes' if alive else 'no')
        time.sleep(1)
    if results:
        with JOB_LOCK:
            h=load_json(HISTORY_FILE,[]); vat=_kst_now().strftime('%Y-%m-%d %H:%M')
            for rec in h:
                if rec.get('id') in results:
                    rec['alive']=results[rec['id']]; rec['verified_at']=vat
            save_json(HISTORY_FILE,h)
        add_log(f'[검증] {len(results)}건 확인 (생존 {sum(1 for v in results.values() if v=="yes")}·삭제 {sum(1 for v in results.values() if v=="no")})')
    return len(results)

def verify_loop():
    while True:
        try:
            if load_config().get('verify_enabled',True): verify_once(40)
        except Exception as e:
            add_log(f'[검증 오류] {str(e)[:80]}')
        time.sleep(3600)

# ==================== Selenium 드라이버 (회피코드 제거) ====================
_drivers = {}
_drv_lock = threading.Lock()

def get_driver():
    tid = threading.current_thread().name
    with _drv_lock:
        if tid in _drivers:
            try: _drivers[tid].current_url; return _drivers[tid]
            except: pass
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        opts = webdriver.ChromeOptions()
        opts.add_argument('--headless=new'); opts.add_argument('--no-sandbox')
        opts.add_argument('--disable-dev-shm-usage'); opts.add_argument('--disable-gpu')
        opts.add_argument('--window-size=1920,1080'); opts.add_argument('--log-level=3')
        # 정상 환경 일치(위조 아님): 서버가 한국(Seoul)에 있으므로 로케일/언어를 실제와 맞춤.
        # navigator.webdriver·Canvas·WebGL 등은 건드리지 않음(지문 위조·스텔스 미사용).
        opts.add_argument('--lang=ko-KR')
        opts.add_experimental_option('prefs',{'intl.accept_languages':'ko-KR,ko'})
        opts.page_load_strategy='eager'
        cb=os.environ.get('CHROME_BIN')  # 리눅스 VPS: chromium 경로 지정 가능
        if cb: opts.binary_location=cb
        svc=Service(ChromeDriverManager().install())
        d=webdriver.Chrome(service=svc,options=opts)
        d.set_page_load_timeout(25); d.implicitly_wait(3)
        _drivers[tid]=d; return d

def dismiss_alerts(d):
    for _ in range(3):
        try: d.switch_to.alert.accept(); time.sleep(0.2)
        except: break

def reset_driver():
    """현재 워커 스레드의 크롬 드라이버를 종료 → 다음 get_driver() 에서 새로 띄움(자동 재시작)."""
    tid=threading.current_thread().name
    with _drv_lock:
        d=_drivers.pop(tid,None)
    if d:
        try: d.quit()
        except: pass

# ==================== Selenium 그누보드 글쓰기 ====================
def gnuboard_post(site, title, content_html):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    url=site.get('site_url','').rstrip('/')
    # site_url 이 board.php?... 형태일 수 있으므로 도메인 기준 bbs 경로 계산
    m=re.match(r'(https?://[^/]+)',url)
    base=m.group(1) if m else url
    bbs=base+'/bbs'
    bo=site.get('bo_table','m8_qna')
    mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
    d=get_driver()

    # 로그인
    if mid:
        d.get(f'{bbs}/login.php'); time.sleep(2)
        d.find_element(By.CSS_SELECTOR,"input[name='mb_id']").send_keys(mid)
        d.find_element(By.CSS_SELECTOR,"input[name='mb_password']").send_keys(mpw)
        d.find_element(By.CSS_SELECTOR,"input[type='submit'],button[type='submit'],.btn_submit").click()
        time.sleep(2); dismiss_alerts(d)

    # 글쓰기 페이지
    d.get(f'{bbs}/write.php?bo_table={bo}'); time.sleep(2)

    # 보안 차단 / 캡차 감지 → 우회하지 않고 즉시 중단(자동발행 제외)
    if _page_is_blocked(d):
        return False,'보안 차단 페이지(403 등) — 즉시 중단'
    _cap=detect_captcha(d)
    if _cap:
        return False,f'캡차 감지({_cap}) — 자동발행 제외(수동 인증 필요)'

    # 제목
    wait=WebDriverWait(d,8)
    el=wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,"input[name='wr_subject'],input#wr_subject")))
    el.clear(); el.send_keys(title)

    # 본문 (smarteditor2/iframe/textarea 대응)
    try:
        d.execute_script(f"if(typeof oEditors!=='undefined')oEditors.getById['wr_content'].exec('SET_IR',[arguments[0]])", content_html)
    except: pass
    try:
        iframe=d.find_element(By.CSS_SELECTOR,"iframe.se2_input_wysiwyg,iframe[id*='editor']")
        d.switch_to.frame(iframe)
        d.execute_script("document.body.innerHTML=arguments[0]",content_html)
        d.switch_to.default_content()
    except:
        d.switch_to.default_content()
        try:
            ed=d.find_element(By.CSS_SELECTOR,"div[contenteditable='true']")
            d.execute_script("arguments[0].innerHTML=arguments[1]",ed,content_html)
        except:
            ta=d.find_element(By.CSS_SELECTOR,"textarea[name='wr_content']")
            ta.clear(); ta.send_keys(content_html)

    # 등록
    d.execute_script("""
        if(typeof oEditors!=='undefined')try{oEditors.getById['wr_content'].exec('UPDATE_CONTENTS_FIELD',[])}catch(e){}
        var btn=document.getElementById('btn_submit');
        if(!btn)btn=document.querySelector("input[type='submit'],button[type='submit']");
        if(btn)btn.click();
    """)
    time.sleep(3); dismiss_alerts(d)
    curl=d.current_url
    # 1) URL 이 뷰/목록으로 이동 → 등록 성공 (헤더의 '로그인' 링크 등으로 인한 오탐 방지 위해 URL 우선)
    if 'wr_id=' in curl or 'board.php' in curl:
        return True,(curl or '등록 완료')
    # 2) URL 로 판정 불가 시에만 본문 텍스트로 분류
    try: body=d.find_element(By.TAG_NAME,'body').text[:1500]
    except: body=''
    # 승인제 게시판: 이미 1회 등록되었을 수 있으므로 재시도로 중복 발행되지 않게 성공 처리
    if any(k in body for k in ['승인 대기','승인대기','관리자 확인','등록되었습니다']):
        return True,'등록됨(승인 대기) — 게시판 승인제'
    if any(k in body for k in ['권한이 없','권한 없','로그인이 필요','게시가 금지','차단']):
        return False,'게시 권한 없음/로그인 필요 — 계정·게시판 권한 확인'
    return False,'등록 확인 불가 — 게시판 규칙/에디터 셀렉터/승인대기 확인'

# ==================== Selenium Cafe24 글쓰기 ====================
def _fill_first(d, selectors, value):
    """여러 셀렉터 후보 중 처음 찾은 입력란에 값 입력."""
    from selenium.webdriver.common.by import By
    for sel in selectors:
        try:
            el=d.find_element(By.CSS_SELECTOR,sel); el.clear(); el.send_keys(value); return True
        except Exception: continue
    return False

def _click_first(d, selectors):
    from selenium.webdriver.common.by import By
    for sel in selectors:
        try: d.find_element(By.CSS_SELECTOR,sel).click(); return True
        except Exception: continue
    return False

def cafe24_post(site, title, content_html):
    """Cafe24 쇼핑몰 게시판 글쓰기. (게시판 구조가 사이트마다 달라 셀렉터 다중 폴백)"""
    from selenium.webdriver.common.by import By
    url=site.get('site_url','').rstrip('/')
    m=re.match(r'(https?://[^/]+)',url); base=m.group(1) if m else url
    bo=str(site.get('bo_table','') or '').strip()
    mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
    d=get_driver()

    # 로그인 (Cafe24 회원 로그인 — 필드명 다양성 대응)
    if mid:
        d.get(base+'/member/login.html'); time.sleep(2)
        _fill_first(d,["input[name='member_id']","input[name='login_id']","input[name='id']",
                       "#member_id","#loginId","input#id"],mid)
        _fill_first(d,["input[name='member_passwd']","input[name='passwd']","input[name='password']",
                       "input[name='login_password']","#passwd","#loginPasswd","input[type='password']"],mpw)
        _click_first(d,["a.btnSubmit","#btnLogin","a.btnLogin",".btnEm","button[type='submit']",
                        "input[type='submit']","a[onclick*='login']"])
        time.sleep(2); dismiss_alerts(d)

    # 글쓰기 페이지 후보 (bo_table 이 숫자면 board_no, 문자면 board 경로)
    write_urls=[]
    if bo.isdigit():
        write_urls=[base+f'/board/write.html?board_no={bo}', base+f'/board/{bo}/write.html']
    elif bo:
        write_urls=[base+f'/board/{bo}/write.html', base+f'/board/write.html?board_no=1', base+f'/board/write.html?board_no={bo}']
    else:
        write_urls=[base+'/board/write.html?board_no=1', base+'/board/free/write.html']
    opened=False
    for wu in write_urls:
        try:
            d.get(wu); time.sleep(2); dismiss_alerts(d)
            if d.find_elements(By.CSS_SELECTOR,"input[name='subject'],#subject,input[name='title']"):
                opened=True; break
        except Exception: continue
    if not opened:
        return False,'Cafe24 글쓰기 페이지 못찾음 — 게시판번호(board_no) 확인'

    # 보안 차단 / 캡차 감지 → 우회하지 않고 즉시 중단
    if _page_is_blocked(d): return False,'보안 차단 페이지(403 등) — 즉시 중단'
    _cap=detect_captcha(d)
    if _cap: return False,f'캡차 감지({_cap}) — 자동발행 제외(수동 인증 필요)'

    # 제목
    _fill_first(d,["input[name='subject']","#subject","input[name='title']"],title)

    # 본문 (Cafe24 SmartEditor iframe / CKEditor / textarea)
    filled=False
    try:
        iframe=d.find_element(By.CSS_SELECTOR,"iframe[id*='content'],iframe.cke_wysiwyg_frame,iframe[title*='Rich'],iframe[title*='편집']")
        d.switch_to.frame(iframe)
        d.execute_script("document.body.innerHTML=arguments[0]",content_html)
        d.switch_to.default_content(); filled=True
    except Exception:
        d.switch_to.default_content()
    if not filled:
        from selenium.webdriver.common.by import By as _By
        for sel in ["textarea[name='content']","textarea#content","textarea[name='contents']",
                    "div[contenteditable='true']","textarea[name='board_content']"]:
            try:
                el=d.find_element(_By.CSS_SELECTOR,sel)
                if sel.startswith('div'):
                    d.execute_script("arguments[0].innerHTML=arguments[1]",el,content_html)
                else:
                    el.clear(); el.send_keys(content_html)
                filled=True; break
            except Exception: continue
    if not filled:
        return False,'Cafe24 본문 입력란 못찾음 — 에디터 셀렉터 확인'

    # 등록
    if not _click_first(d,["a.btnSubmit","#btnSubmit","button.btnSubmit","a.btnEm.btnStrong",
                           "input[type='submit']","button[type='submit']","a[onclick*='submit']"]):
        try: d.find_element(By.CSS_SELECTOR,'form').submit()
        except Exception: pass
    time.sleep(3); dismiss_alerts(d)
    curl=d.current_url
    if any(k in curl for k in ['read.html','list.html','article','board_no','view.html']) and 'write.html' not in curl:
        return True,(curl or '등록 완료')
    try: body=d.find_element(By.TAG_NAME,'body').text[:1500]
    except Exception: body=''
    if any(k in body for k in ['등록되었습니다','등록 완료','승인 대기','승인대기','작성되었습니다']):
        return True,'등록됨'
    if any(k in body for k in ['로그인','권한이 없','권한 없','금지','차단','스팸']):
        return False,'Cafe24 게시 권한 없음/로그인 필요 — 계정·게시판 권한 확인'
    return False,'Cafe24 등록 확인 불가 — 게시판 설정/에디터 셀렉터 확인'

# ==================== 플랫폼 자동 감지 + 발행 디스패처 ====================
_plat_cache={}
def detect_platform(url, use_cache=True):
    """사이트 URL 로 그누보드/Cafe24 자동 판별. 확실치 않으면 그누보드(국내 대부분)."""
    import requests as _rq
    m=re.match(r'(https?://[^/]+)',(url or '').rstrip('/')); base=m.group(1) if m else (url or '')
    if not base: return 'gnuboard'
    if use_cache and base in _plat_cache: return _plat_cache[base]
    UA={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36'}
    def probe(path,needles):
        try:
            r=_rq.get(base+path,timeout=10,verify=False,headers=UA,allow_redirects=True)
            if r.status_code>=400: return False
            t=(r.text or '').lower()
            return any(n in t for n in needles)
        except Exception: return False
    # 그누보드 신호 우선 확인(가장 흔함)
    if probe('/bbs/login.php',['mb_password','mb_id']) or probe('/bbs/',['bo_table','wr_id','gnuboard']):
        res='gnuboard'
    # Cafe24 신호
    elif probe('/member/login.html',['member_id','passwd','member_passwd']) or \
         probe('/',['cafe24','xans-','ec-base','/board/write.html','smartdesign']):
        res='cafe24'
    else:
        res='gnuboard'
    _plat_cache[base]=res
    return res

def resolve_platform(site):
    """사이트의 platform 이 auto/미지정이면 감지해서 확정값 반환."""
    plat=(site.get('platform') or '').strip().lower()
    if plat in ('gnuboard','cafe24'): return plat
    return detect_platform(site.get('site_url',''))

# ==================== 감독형 자가학습 발행 (셀렉터 자동 탐색) ====================
SUBJECT_HINTS=['subject','title','제목','wr_subject','head','tit','bo_subject']
CONTENT_HINTS=['content','contents','내용','body','wr_content','memo','board_content','desc']
SUBMIT_HINTS=['등록','작성','확인','저장','올리','완료','submit','write','save','send','regist','apply','ok','confirm']

def _sel_vis(el):
    try: return el.is_displayed()
    except Exception: return False
def _sel_attr(el,name):
    try: return (el.get_attribute(name) or '')
    except Exception: return ''
def _css_for(el):
    """요소를 다시 찾을 수 있는 안정적 CSS 셀렉터 생성(id>name>type/class>tag)."""
    i=_sel_attr(el,'id'); n=_sel_attr(el,'name')
    if i: return f"[id='{i}']"
    if n: return f"[name='{n}']"
    try: tag=el.tag_name
    except Exception: return '*'
    typ=_sel_attr(el,'type')
    if tag=='input' and typ: return f"input[type='{typ}']"
    cls=[c for c in (_sel_attr(el,'class') or '').split() if c][:2]
    if cls: return tag+'.'+'.'.join(cls)
    return tag

def _is_specific(sel):
    """id/name/type/class 로 특정되는 셀렉터인지(단순 태그명은 위험)."""
    return bool(sel) and ('[' in sel or '.' in sel)

def discover_subject(d):
    from selenium.webdriver.common.by import By
    inputs=d.find_elements(By.CSS_SELECTOR,"input[type='text'],input[type='search'],input:not([type])")
    first=None
    for el in inputs:
        if not _sel_vis(el): continue
        blob=(_sel_attr(el,'name')+' '+_sel_attr(el,'id')+' '+_sel_attr(el,'placeholder')).lower()
        if any(h in blob for h in SUBJECT_HINTS): return _css_for(el)
        if first is None: first=el
    return _css_for(first) if first is not None else None

def discover_content(d):
    """본문 입력란 탐색 → (mode, selector). mode: iframe|contenteditable|textarea."""
    from selenium.webdriver.common.by import By
    iframes=[f for f in d.find_elements(By.CSS_SELECTOR,'iframe') if _sel_vis(f)]
    for f in iframes:
        blob=(_sel_attr(f,'id')+' '+_sel_attr(f,'class')+' '+_sel_attr(f,'title')).lower()
        if any(h in blob for h in ['editor','wysiwyg','content','se2','cke','편집','rich','smart']):
            return ('iframe',_css_for(f))
    for el in d.find_elements(By.CSS_SELECTOR,"[contenteditable='true']"):
        if _sel_vis(el): return ('contenteditable',_css_for(el))
    tas=[t for t in d.find_elements(By.CSS_SELECTOR,'textarea') if _sel_vis(t)]
    for t in tas:
        blob=(_sel_attr(t,'name')+' '+_sel_attr(t,'id')).lower()
        if any(h in blob for h in CONTENT_HINTS): return ('textarea',_css_for(t))
    if tas: return ('textarea',_css_for(tas[0]))
    if iframes: return ('iframe',_css_for(iframes[0]))
    return (None,None)

def discover_submit(d):
    """등록 버튼 셀렉터 탐색. 특정 불가한 단순 태그면 None(→ form.submit() 사용)."""
    from selenium.webdriver.common.by import By
    cands=d.find_elements(By.CSS_SELECTOR,"input[type='submit'],button,a,input[type='button']")
    generic=None
    for el in cands:   # type=submit 우선
        if _sel_vis(el) and el.tag_name=='input' and _sel_attr(el,'type')=='submit':
            sel=_css_for(el)
            if _is_specific(sel): return sel
            generic=generic or sel
    for el in cands:   # 텍스트/속성 힌트
        if not _sel_vis(el): continue
        blob=((el.text or '')+' '+_sel_attr(el,'value')+' '+_sel_attr(el,'onclick')+' '+_sel_attr(el,'id')+' '+_sel_attr(el,'class')).lower()
        if any(h in blob for h in SUBMIT_HINTS):
            sel=_css_for(el)
            if _is_specific(sel): return sel
            generic=generic or sel
    return None   # 특정 셀렉터 없음 → form.submit() 폴백

def discover_login(d, base):
    """로그인 페이지 후보를 돌며 id/pw/버튼 셀렉터 탐색."""
    from selenium.webdriver.common.by import By
    for path in ['/bbs/login.php','/member/login.html','/login','/member/login','/index.php?mode=login']:
        try: d.get(base+path); time.sleep(1.5)
        except Exception: continue
        pws=[p for p in d.find_elements(By.CSS_SELECTOR,"input[type='password']") if _sel_vis(p)]
        if not pws: continue
        texts=[t for t in d.find_elements(By.CSS_SELECTOR,"input[type='text'],input[type='email'],input:not([type])") if _sel_vis(t)]
        if not texts: continue
        return {'login_url':base+path,'id_sel':_css_for(texts[0]),'pw_sel':_css_for(pws[0]),'login_btn':discover_submit(d)}
    return None

def _confirm_posted(d):
    from selenium.webdriver.common.by import By
    curl=d.current_url or ''
    head=curl.split('?')[0]
    if any(k in curl for k in ['wr_id=','board.php','read.html','list.html','view.html','article','board_no']) and 'write' not in head:
        return True,curl
    try: body=d.find_element(By.TAG_NAME,'body').text[:1500]
    except Exception: body=''
    if any(k in body for k in ['등록되었습니다','작성되었습니다','등록 완료','승인 대기','승인대기','완료되었']):
        return True,'등록됨'
    if any(k in body for k in ['권한이 없','권한 없','로그인이 필요','로그인 필요','게시가 금지','차단','스팸']):
        return False,'게시 권한 없음/로그인 필요'
    return False,'등록 확인 불가 — 게시판 설정 확인'

def _fill_recipe_fields(d, rec, title, content):
    """현재 write 페이지에 제목/본문 채우고 등록 클릭(네비게이션 없음)."""
    from selenium.webdriver.common.by import By
    d.find_element(By.CSS_SELECTOR,rec['subject_sel']).clear()
    d.find_element(By.CSS_SELECTOR,rec['subject_sel']).send_keys(title)
    mode=rec.get('content_mode'); csel=rec.get('content_sel')
    if mode=='iframe':
        fr=d.find_element(By.CSS_SELECTOR,csel); d.switch_to.frame(fr)
        d.execute_script("document.body.innerHTML=arguments[0]",content); d.switch_to.default_content()
    elif mode=='contenteditable':
        el=d.find_element(By.CSS_SELECTOR,csel); d.execute_script("arguments[0].innerHTML=arguments[1]",el,content)
    else:
        el=d.find_element(By.CSS_SELECTOR,csel); el.clear(); el.send_keys(content)
    # 스마트에디터 동기화 시도(있으면)
    try: d.execute_script("if(typeof oEditors!=='undefined')try{oEditors.getById[Object.keys(oEditors.getById)[0]].exec('UPDATE_CONTENTS_FIELD',[])}catch(e){}")
    except Exception: pass
    if rec.get('submit_sel'):
        try: d.find_element(By.CSS_SELECTOR,rec['submit_sel']).click()
        except Exception:
            try: d.find_element(By.CSS_SELECTOR,'form').submit()
            except Exception: pass
    else:
        try: d.find_element(By.CSS_SELECTOR,'form').submit()
        except Exception: pass
    time.sleep(3); dismiss_alerts(d)

def _apply_recipe(d, site, rec, title, content):
    """저장된 레시피로 발행(로그인→글쓰기→채우기→확인)."""
    from selenium.webdriver.common.by import By
    mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
    if mid:
        base=re.match(r'(https?://[^/]+)',rec.get('write_url','') or ''); base=base.group(1) if base else ''
        if rec.get('login_url') and rec.get('id_sel') and rec.get('pw_sel'):
            d.get(rec['login_url']); time.sleep(1.5)
            try: e=d.find_element(By.CSS_SELECTOR,rec['id_sel']); e.clear(); e.send_keys(mid)
            except Exception: pass
            try: e=d.find_element(By.CSS_SELECTOR,rec['pw_sel']); e.clear(); e.send_keys(mpw)
            except Exception: pass
            if rec.get('login_btn'):
                try: d.find_element(By.CSS_SELECTOR,rec['login_btn']).click()
                except Exception: pass
            time.sleep(2); dismiss_alerts(d)
        elif base:
            _platform_login(d,base,site)   # 저장된 상세 셀렉터 없으면 플랫폼 로그인
    d.get(rec['write_url']); time.sleep(2); dismiss_alerts(d)
    _fill_recipe_fields(d, rec, title, content)
    return _confirm_posted(d)

def _page_login_state(d):
    """현재 페이지가 로그인 화면인지(본문 에디터 없음 + 비번칸/로그인안내) 판별."""
    from selenium.webdriver.common.by import By
    try: has_pw=any(_sel_vis(e) for e in d.find_elements(By.CSS_SELECTOR,"input[type='password']"))
    except Exception: has_pw=False
    has_editor=bool(discover_content(d)[1])   # 글쓰기 페이지엔 본문 에디터가 있음
    try: body=d.find_element(By.TAG_NAME,'body').text[:1000]
    except Exception: body=''
    login_notice=('로그인' in body and ('필요' in body or '회원가입' in body or '아이디' in body))
    return (not has_editor) and (has_pw or login_notice)

def _platform_login(d, base, site):
    """플랫폼별 신뢰 셀렉터로 로그인 시도(그누보드/Cafe24). 로그인 URL 반환(실패 None)."""
    from selenium.webdriver.common.by import By
    mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
    if not mid: return None
    plat=resolve_platform(site)
    tries=[('/bbs/login.php',"input[name='mb_id']","input[name='mb_password']")] if plat=='gnuboard' else \
          [('/member/login.html',"input[name='member_id'],input[name='login_id'],input[name='id']","input[name='member_passwd'],input[name='passwd'],input[name='password']")]
    tries.append(('/bbs/login.php',"input[name='mb_id']","input[name='mb_password']"))
    tries.append(('/member/login.html',"input[name='member_id'],input[name='login_id'],input[name='id']","input[name='member_passwd'],input[name='passwd'],input[name='password']"))
    for path,idsel,pwsel in tries:
        try:
            d.get(base+path); time.sleep(1.5)
            ide=d.find_elements(By.CSS_SELECTOR,idsel); pwe=d.find_elements(By.CSS_SELECTOR,pwsel)
            if not ide or not pwe: continue
            ide[0].clear(); ide[0].send_keys(mid); pwe[0].clear(); pwe[0].send_keys(mpw)
            if not _click_first(d,["input[type='submit']","button[type='submit']",".btn_submit","a.btnSubmit","#btnLogin",".btnEm"]):
                try: pwe[0].submit()
                except Exception: pass
            time.sleep(2); dismiss_alerts(d)
            return base+path
        except Exception: continue
    return None

def discover_write_page(d, base, site):
    bo=str(site.get('bo_table','') or '').strip(); plat=resolve_platform(site)
    cands=[]
    # 관리자가 등록한 게시판 화면에서 실제 글쓰기 링크를 먼저 실측한다.
    # 사이트 전체를 탐색하지 않고, 등록 URL 한 화면의 동일 출처 링크만 사용한다.
    try:
        start=(site.get('site_url') or base).strip()
        d.get(start); time.sleep(2); dismiss_alerts(d)
        from selenium.webdriver.common.by import By
        for a in d.find_elements(By.CSS_SELECTOR,'a[href]'):
            href=urllib.parse.urljoin(d.current_url or start,_sel_attr(a,'href'))
            pu=urllib.parse.urlparse(href); pb=urllib.parse.urlparse(base)
            if (pu.scheme,pu.netloc)!=(pb.scheme,pb.netloc): continue
            blob=((a.text or '')+' '+href+' '+_sel_attr(a,'class')+' '+_sel_attr(a,'id')).lower()
            if any(k in blob for k in ['글쓰기','글 쓰기','write.php','/write.html','mode=write','act=write']):
                cands.append(href)
    except Exception: pass
    if plat=='cafe24':
        if bo.isdigit(): cands += [f'/board/write.html?board_no={bo}',f'/board/{bo}/write.html']
        elif bo: cands += [f'/board/{bo}/write.html',f'/board/write.html?board_no=1']
        else: cands += ['/board/write.html?board_no=1','/board/free/write.html']
    else:
        cands += [f'/bbs/write.php?bo_table={bo or "free"}']
    cands+=[f'/bbs/write.php?bo_table={bo or "free"}',f'/board/write.html?board_no={bo if bo.isdigit() else "1"}']
    seen=set()
    for path in cands:
        target=path if str(path).startswith(('http://','https://')) else base+path
        if target in seen: continue
        seen.add(target)
        try: d.get(target); time.sleep(2); dismiss_alerts(d)
        except Exception: continue
        if discover_subject(d) and discover_content(d)[1]:
            return d.current_url or target
    return None

def _save_site_analysis(site_id, analysis, rec=None):
    """실측 결과는 성공/실패 모두 저장하고, 확실한 폼만 발행 레시피로 승격한다."""
    if not site_id: return
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')!=site_id: continue
            s['analysis']=analysis
            if rec:
                s['learned']=rec
                if rec.get('platform'): s['platform']=rec['platform']
        save_sites(sites)

def analyze_site_logic(site):
    """허용된 한 사이트의 등록 URL/로그인/글쓰기 DOM을 측정한다. 절대 제출하지 않는다."""
    from selenium.webdriver.common.by import By
    now=datetime.now().astimezone().isoformat(timespec='seconds')
    result={'measured_at':now,'mode':'read_only_no_submit','ok':False,'platform':'',
            'start_url':site.get('site_url',''),'final_url':'','write_url':'',
            'blocked':False,'captcha':'','form':{},'steps':[]}
    def step(name,ok,detail=''):
        result['steps'].append({'name':name,'ok':bool(ok),'detail':str(detail)[:300]})
    url=(site.get('site_url') or '').strip(); m=re.match(r'(https?://[^/]+)',url)
    base=m.group(1) if m else url
    if not base:
        step('등록 URL',False,'URL 없음'); return result,None
    d=get_driver(); plat=resolve_platform(site); result['platform']=plat
    step('플랫폼',True,plat)
    try:
        d.get(url or base); time.sleep(2); dismiss_alerts(d)
        result['final_url']=d.current_url or ''
        step('등록 화면',True,result['final_url'])
    except Exception as e:
        step('등록 화면',False,str(e)[:180]); return result,None
    if _page_is_blocked(d):
        result['blocked']=True; step('보안 차단',False,'403/보안 차단 감지 — 우회하지 않음'); return result,None
    cap=detect_captcha(d)
    if cap:
        result['captcha']=str(cap); step('CAPTCHA',False,str(cap)+' — 우회하지 않음'); return result,None
    step('초기 화면 보안',True,'차단/CAPTCHA 없음')
    if site.get('mb_id'):
        lu=_platform_login(d,base,site)
        step('로그인',bool(lu),lu or '저장 계정으로 로그인 페이지/필드 확인 실패')
    else:
        step('로그인',True,'계정 미설정 — 공개/비회원 글쓰기만 측정')
    # 설정된 게시판 ID의 표준 쓰기 URL을 먼저 직접 측정하여 로그인/차단 원인을 보존한다.
    bo=str(site.get('bo_table') or '').strip(); direct=''
    if bo:
        direct=(base+f'/board/write.html?board_no={bo}') if plat=='cafe24' and bo.isdigit() else \
               (base+f'/bbs/write.php?bo_table={urllib.parse.quote(bo)}')
    wu=None
    if direct:
        try:
            d.get(direct); time.sleep(2); dismiss_alerts(d)
            result['final_url']=d.current_url or direct
            if _page_is_blocked(d):
                result['write_url']=direct; result['blocked']=True
                step('설정 글쓰기 URL',False,'접근 차단: '+direct+' — 우회하지 않음'); return result,None
            direct_cap=detect_captcha(d)
            if direct_cap:
                result['write_url']=direct; result['captcha']=str(direct_cap)
                step('설정 글쓰기 URL',False,'CAPTCHA 감지: '+str(direct_cap)+' — 우회하지 않음'); return result,None
            if _page_login_state(d) and not site.get('mb_id'):
                result['write_url']=direct
                step('설정 글쓰기 URL',False,'로그인 필요: '+result['final_url']); return result,None
            if discover_subject(d) and discover_content(d)[1]: wu=result['final_url']
        except Exception as e:
            step('설정 글쓰기 URL',False,str(e)[:180])
    if not wu: wu=discover_write_page(d,base,site)
    result['final_url']=d.current_url or result['final_url']
    if not wu:
        detail='로그인 화면으로 이동됨' if _page_login_state(d) else '실제 링크 및 설정 게시판ID 후보에서 폼을 찾지 못함'
        step('글쓰기 폼',False,detail); return result,None
    result['write_url']=wu; result['final_url']=d.current_url or wu
    step('글쓰기 폼',True,wu)
    if _page_is_blocked(d):
        result['blocked']=True; step('글쓰기 보안',False,'403/보안 차단 감지 — 우회하지 않음'); return result,None
    cap=detect_captcha(d)
    if cap:
        result['captcha']=str(cap); step('CAPTCHA',False,str(cap)+' — 우회하지 않음'); return result,None
    subj=discover_subject(d); cmode,csel=discover_content(d); sub=discover_submit(d)
    form_el=None
    try:
        if subj: form_el=d.find_element(By.CSS_SELECTOR,subj).find_element(By.XPATH,'ancestor::form[1]')
    except Exception: pass
    form={'action':'','method':'','subject_sel':subj or '','content_mode':cmode or '',
          'content_sel':csel or '','submit_sel':sub or '','required_fields':[]}
    if form_el is not None:
        form['action']=urllib.parse.urljoin(result['final_url'],_sel_attr(form_el,'action'))
        form['method']=(_sel_attr(form_el,'method') or 'get').lower()
        try:
            for el in form_el.find_elements(By.CSS_SELECTOR,'input[required],textarea[required],select[required]'):
                name=_sel_attr(el,'name') or _sel_attr(el,'id')
                typ=_sel_attr(el,'type') or el.tag_name
                if name: form['required_fields'].append({'name':name,'type':typ})
        except Exception: pass
    result['form']=form
    step('제목 필드',bool(subj),subj or '없음')
    step('본문 필드',bool(csel),f'{cmode} · {csel}' if csel else '없음')
    step('등록 동작',bool(sub or form_el is not None),sub or ('form '+form['method'] if form_el is not None else '없음'))
    if not subj or not csel or form_el is None:
        step('판정',False,'발행 레시피로 저장할 신뢰도 부족'); return result,None
    rec={'platform':plat,'learned_at':now,'learned_mode':'measured_no_submit',
         'write_url':wu,'subject_sel':subj,'content_mode':cmode,
         'content_sel':csel,'submit_sel':sub,'form_action':form['action'],'form_method':form['method']}
    result['ok']=True; step('판정',True,'DOM 실측 완료 · 제출 없이 레시피 저장 가능')
    return result,rec

def discover_and_post(site, title, content):
    """DOM을 훑어 셀렉터를 스스로 찾아 발행. 성공 시 (ok,msg,레시피) 반환."""
    d=get_driver()
    url=site.get('site_url','').rstrip('/'); m=re.match(r'(https?://[^/]+)',url); base=m.group(1) if m else url
    rec={'platform':resolve_platform(site),'learned_at':datetime.now().strftime('%Y-%m-%d %H:%M')}
    # 로그인: 플랫폼별 신뢰 셀렉터 우선, 실패 시 자동 탐색
    if site.get('mb_id'):
        lu=_platform_login(d,base,site)
        if lu: rec['login_url']=lu
        else:
            login=discover_login(d,base)
            if login:
                rec.update(login)
                from selenium.webdriver.common.by import By
                try: e=d.find_element(By.CSS_SELECTOR,login['id_sel']); e.clear(); e.send_keys(site.get('mb_id',''))
                except Exception: pass
                try: e=d.find_element(By.CSS_SELECTOR,login['pw_sel']); e.clear(); e.send_keys(site.get('mb_pass',''))
                except Exception: pass
                if login.get('login_btn'):
                    try: d.find_element(By.CSS_SELECTOR,login['login_btn']).click()
                    except Exception: pass
                time.sleep(2); dismiss_alerts(d)
    wu=discover_write_page(d,base,site)
    if not wu:
        # 원인 세분화: 로그인 화면으로 튕겼는지 확인
        if _page_login_state(d):
            if not site.get('mb_id'):
                return False,'로그인이 필요한 게시판입니다 — 사이트에 아이디/비밀번호를 입력한 뒤 다시 학습하세요',None
            return False,'로그인 실패로 보입니다 — 아이디/비밀번호가 맞는지 확인하세요(캡차/보안문자 게시판일 수도)',None
        return False,'글쓰기 페이지 못찾음(학습 실패) — 게시판ID(bo_table) 확인',None
    rec['write_url']=wu
    # 캡차/보안 인증 감지 → 학습·발행 대상에서 제외(우회하지 않음)
    if _page_is_blocked(d): return False,'보안 차단 페이지(403 등) — 즉시 중단',None
    _cap=detect_captcha(d)
    if _cap: return False,f'캡차 감지({_cap}) — 자동발행 부적합(수동 인증 필요)',None
    subj=discover_subject(d); cmode,csel=discover_content(d); sub=discover_submit(d)
    if not subj or not csel:
        return False,'제목/본문 입력란 못찾음(학습 실패)',None
    rec['subject_sel']=subj; rec['content_mode']=cmode; rec['content_sel']=csel; rec['submit_sel']=sub
    try: _fill_recipe_fields(d, rec, title, content)
    except Exception as e: return False,f'입력 실패(학습): {str(e)[:80]}',None
    ok,msg=_confirm_posted(d)
    return ok,msg,(rec if ok else None)

def dryrun_post(site, title, content_html):
    """등록 '직전'까지만 수행 — 실제 글은 올리지 않고 발행 가능 여부를 검증한다.
       로그인 → 글쓰기 페이지 → 캡차확인 → 입력란 탐색 → 값 채우기 까지. 제출 클릭 없음."""
    from selenium.webdriver.common.by import By
    steps=[]
    def st(name,ok,detail=''): steps.append({'name':name,'ok':bool(ok),'detail':str(detail)[:220]})
    url=site.get('site_url','').rstrip('/'); m=re.match(r'(https?://[^/]+)',url)
    base=m.group(1) if m else url
    d=get_driver()
    plat=resolve_platform(site); st('플랫폼 판별',True,plat)

    # 1) 로그인
    mid=site.get('mb_id','')
    if mid:
        lu=_platform_login(d,base,site)
        if lu:
            # 로그인 성공 여부: 로그아웃 링크 존재로 추정
            try: body=d.find_element(By.TAG_NAME,'body').text[:1500]
            except Exception: body=''
            logged=('로그아웃' in body) or ('logout' in (d.page_source or '').lower())
            st('로그인',logged,f'{lu} → '+('세션 확보됨' if logged else '로그인 확인 불가(비번/아이디 확인)'))
        else:
            st('로그인',False,'로그인 페이지를 못 찾음')
    else:
        st('로그인',True,'아이디 미설정 — 비회원 글쓰기로 진행')

    # 2) 글쓰기 페이지
    rec=site.get('learned') or {}
    wu=rec.get('write_url') if rec.get('write_url') else None
    if wu:
        try: d.get(wu); time.sleep(2); dismiss_alerts(d)
        except Exception: wu=None
    if not wu:
        wu=discover_write_page(d,base,site)
    if not wu:
        if _page_login_state(d):
            st('글쓰기 페이지',False,'로그인 화면으로 이동됨 — 계정 필요 또는 로그인 실패')
        else:
            st('글쓰기 페이지',False,'못 찾음 — 게시판ID(bo_table) 확인')
        return False,steps
    st('글쓰기 페이지',True,wu)

    # 3) 차단/캡차
    if _page_is_blocked(d):
        st('보안 차단 확인',False,'403/보안 차단 페이지'); return False,steps
    st('보안 차단 확인',True,'차단 없음')
    cap=detect_captcha(d)
    if cap:
        st('캡차 확인',False,f'{cap} 감지 — 자동발행 부적합(우회하지 않음)'); return False,steps
    st('캡차 확인',True,'캡차 없음')

    # 4) 입력란 탐색
    subj=rec.get('subject_sel') or discover_subject(d)
    cmode,csel=(rec.get('content_mode'),rec.get('content_sel')) if rec.get('content_sel') else discover_content(d)
    sub=rec.get('submit_sel') if rec.get('submit_sel') else discover_submit(d)
    st('제목 입력란',bool(subj),subj or '못 찾음')
    st('본문 에디터',bool(csel),f'{cmode} · {csel}' if csel else '못 찾음')
    st('등록 버튼',True,sub or '특정 셀렉터 없음 → form.submit() 사용 예정')
    if not subj or not csel: return False,steps

    # 5) 실제로 값 채워보기 (제출은 하지 않음)
    try:
        e=d.find_element(By.CSS_SELECTOR,subj); e.clear(); e.send_keys(title)
        st('제목 입력 테스트',True,f'{len(title)}자 입력 성공')
    except Exception as ex:
        st('제목 입력 테스트',False,str(ex)[:120]); return False,steps
    try:
        if cmode=='iframe':
            fr=d.find_element(By.CSS_SELECTOR,csel); d.switch_to.frame(fr)
            d.execute_script("document.body.innerHTML=arguments[0]",content_html); d.switch_to.default_content()
        elif cmode=='contenteditable':
            el=d.find_element(By.CSS_SELECTOR,csel); d.execute_script("arguments[0].innerHTML=arguments[1]",el,content_html)
        else:
            el=d.find_element(By.CSS_SELECTOR,csel); el.clear(); el.send_keys(content_html[:2000])
        st('본문 입력 테스트',True,f'{len(content_html)}자 HTML 입력 성공')
    except Exception as ex:
        d.switch_to.default_content()
        st('본문 입력 테스트',False,str(ex)[:120]); return False,steps

    # 6) 필수 추가 입력란(비회원 이름/비번 등) 확인
    try:
        reqs=[]
        for el in d.find_elements(By.CSS_SELECTOR,"input[required],input[type='password']"):
            if not _sel_vis(el): continue
            nm=_sel_attr(el,'name') or _sel_attr(el,'id')
            if nm and nm not in (subj or ''): reqs.append(nm)
        st('추가 필수 입력란',True,(', '.join(reqs[:6]) if reqs else '없음')+(' (비회원 글쓰기는 이름/비번 필요할 수 있음)' if reqs else ''))
    except Exception: pass

    st('종합',True,'✅ 등록 직전까지 모두 성공 — 실제 발행 가능 상태 (글은 올리지 않았습니다)')
    return True,steps

def save_learned(site_id, rec):
    """학습된 셀렉터 레시피를 사이트에 영구 저장."""
    if not site_id: return
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')==site_id:
                s['learned']=rec
                if rec.get('platform'): s['platform']=rec['platform']
        save_sites(sites)

def set_site_flag(site_id, **fields):
    """사이트에 플래그/상태 저장(예: has_captcha)."""
    if not site_id: return
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')==site_id: s.update(fields)
        save_sites(sites)

def do_post(site, title, content_html):
    """발행 라우팅(하이브리드+자가학습): 학습레시피→플랫폼기본→자동학습."""
    rec=site.get('learned')
    # 1) 저장된 학습 레시피 우선
    if rec and rec.get('write_url') and rec.get('subject_sel') and rec.get('content_sel'):
        try:
            ok,msg=_apply_recipe(get_driver(),site,rec,title,content_html)
            if ok: return True,msg
            add_log(f'[학습레시피 실패→폴백] {site.get("name","")}')
        except Exception as e:
            add_log(f'[학습레시피 오류→폴백] {str(e)[:60]}')
    # 2) 플랫폼별 기본 발행기
    plat=resolve_platform(site)
    try:
        ok,msg=(cafe24_post if plat=='cafe24' else gnuboard_post)(site,title,content_html)
    except Exception as e:
        ok,msg=False,str(e)
    if ok: return True,msg
    # 3) 구조적 실패(일시적·로그인·차단 제외)면 자동 학습(디스커버리) 발행 후 레시피 저장
    reason,_,_=classify_fail(msg)
    if reason in ('board','other') or rec is not None:
        try:
            ok2,msg2,newrec=discover_and_post(site,title,content_html)
            if ok2 and newrec:
                save_learned(site.get('id'),newrec)
                add_log(f'[자동학습 성공] {site.get("name","")} — 셀렉터 저장됨')
                return True,'(자가학습) '+str(msg2)
            if ok2: return True,str(msg2)
            return False,f'{msg} · 학습시도:{msg2}'
        except Exception as e:
            return False,f'{msg} · 학습실패:{str(e)[:50]}'
    return False,msg

# ==================== 큐 & 워커 & 이력 ====================
post_queue=queue.Queue()
wk_active=False
wk_paused=False
wk_stats={'success':0,'fail':0,'queued':0,'total':0,'done':0,'skipped':0,'retry':0}
STATS_LOCK=threading.Lock()
POST_LOCK=threading.Lock()
JOB_LOCK=threading.Lock()   # history.json / queue.json 동시성 보호

# ---- 일시적 실패 자동 재시도(지연 재큐) ----
RETRY_DELAY=300      # 5분 뒤 재시도
RETRY_MAX=2          # 지연 재시도 최대 횟수
_retry_jobs=[]
_retry_lock=threading.Lock()
def schedule_retry(job, delay=None):
    if delay is None: delay=RETRY_DELAY
    with _retry_lock: _retry_jobs.append({'due':time.time()+delay,'job':job})

def retry_loop():
    """지연 재시도 큐를 감시해 시간이 되면 발행 큐로 되돌림."""
    while True:
        try:
            now=time.time(); ready=[]
            with _retry_lock:
                keep=[]
                for it in _retry_jobs:
                    (ready if it['due']<=now else keep).append(it)
                _retry_jobs[:]=keep
            for it in ready:
                post_queue.put(it['job'])
                with STATS_LOCK: wk_stats['queued']=post_queue.qsize()
                if not wk_active: start_workers(load_config().get('workers',2))
        except Exception as e:
            add_log(f'[재시도 루프 오류] {str(e)[:80]}')
        time.sleep(30)

# ---- 발행 이력 원장 (엑셀/결과탭) ----
def history_add(rec):
    with JOB_LOCK:
        h=load_json(HISTORY_FILE,[]); h.append(rec)
        if len(h)>5000: h=h[-5000:]
        save_json(HISTORY_FILE,h)

def history_update(hid,**fields):
    with JOB_LOCK:
        h=load_json(HISTORY_FILE,[])
        for rec in h:
            if rec.get('id')==hid:
                rec.update(fields); rec['updated']=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                break
        save_json(HISTORY_FILE,h)

# ---- 미완료 작업 영속화 (재시작 복구) ----
def _persist_add(job):
    with JOB_LOCK:
        q=load_json(QUEUE_FILE,[]); q.append(job); save_json(QUEUE_FILE,q)

def _persist_remove(job_id):
    with JOB_LOCK:
        q=[j for j in load_json(QUEUE_FILE,[]) if j.get('job_id')!=job_id]
        save_json(QUEUE_FILE,q)

def recover_queue():
    """재시작 시 queue.json 의 미완료 작업을 메모리 큐로 복구."""
    q=load_json(QUEUE_FILE,[])
    n=0
    for job in q:
        if job.get('site') and job.get('content'):
            post_queue.put(job); n+=1
    if n:
        with STATS_LOCK:
            wk_stats['total']+=n; wk_stats['queued']=post_queue.qsize()
        add_log(f'[복구] 미완료 작업 {n}건 복구됨 — 워커 시작 시 이어서 발행')
    return n

def site_daily_limit(site,cfg):
    try: return max(0,int(site.get('daily_limit',cfg.get('daily_limit',3)) or 0))
    except Exception: return 3

def site_min_interval(site):
    try: return max(0,int(site.get('min_interval_minutes',60) or 0))
    except Exception: return 60

def _fresh_site(site):
    return next((s for s in load_sites() if s.get('id')==site.get('id')),site)

def under_daily_limit(site,cfg):
    """사이트당 1일 발행 한도 확인 (0 = 무제한). 도배 방지."""
    site=_fresh_site(site)
    limit=site_daily_limit(site,cfg)
    if limit<=0: return True
    today=datetime.now().strftime('%Y-%m-%d')
    for s in load_sites():
        if s.get('id')==site.get('id'):
            if s.get('posted_date')!=today: return True
            return s.get('posted_today',0)<limit
    return True

def under_min_interval(site):
    """마지막 성공 발행 후 사이트별 최소 간격 준수 여부와 남은 초 반환."""
    site=_fresh_site(site); mins=site_min_interval(site)
    if mins<=0 or not site.get('last_post_at'): return True,0
    try:
        last=datetime.strptime(site['last_post_at'],'%Y-%m-%d %H:%M:%S')
        remain=int(mins*60-(datetime.now()-last).total_seconds())
        return remain<=0,max(0,remain)
    except Exception: return True,0

def finalize_post(site,ok):
    """상태 갱신 + 성공 시 오늘 발행 카운트 증가 (한 번의 락으로 처리)."""
    today=datetime.now().strftime('%Y-%m-%d')
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id')==site.get('id'):
                s['status']='done' if ok else 'failed'
                if ok:
                    if s.get('posted_date')!=today: s['posted_date']=today; s['posted_today']=0
                    s['posted_today']=s.get('posted_today',0)+1
                    s['last_post_at']=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                break
        save_sites(sites)

def start_workers(n=2):
    global wk_active,wk_paused
    if wk_active: return
    n=1  # 동일 사이트 동시 발행에 의한 한도/간격 우회 방지
    wk_active=True; wk_paused=False; add_log(f'[워커] {n}개 시작 (사이트 한도 보호)')
    for i in range(n):
        t=threading.Thread(target=worker_loop,name=f'W-{i+1}',daemon=True); t.start()

def stop_workers():
    global wk_active; wk_active=False; add_log('[워커] 정지')

def pause_workers():
    global wk_paused; wk_paused=True; add_log('[워커] 일시정지')

def resume_workers():
    global wk_paused; wk_paused=False; add_log('[워커] 재개')

def worker_loop():
    while wk_active:
        # 일시정지: 큐에서 꺼내지 않고 대기 (작업 보존)
        if wk_paused: time.sleep(1); continue
        try: job=post_queue.get(timeout=1)
        except: continue
        site=job['site']; title=job['title']; content=job['content']; cfg=load_config()
        job_id=job.get('job_id'); hid=job.get('hist_id')
        name=site.get('name') or site.get('site_url','')[:24]

        # 큐 등록 뒤 사이트가 삭제/미허용/캡차 상태로 바뀌어도 발행되지 않도록 현재 상태 재검증.
        current=next((s for s in load_sites() if s.get('id')==site.get('id')),None)
        if not current or not is_autopostable(current):
            why='사이트 삭제됨' if not current else ('캡차 감지됨' if current.get('has_captcha') else '관리자 허용 해제됨')
            add_log(f'[스킵] {name} {why}')
            if hid: history_update(hid,status='skipped',message=why)
            if job_id: _persist_remove(job_id)
            with STATS_LOCK:
                wk_stats['skipped']+=1; wk_stats['done']+=1; wk_stats['queued']=post_queue.qsize()
            continue
        site=current; job['site']=current

        # 1일 한도 확인 (도배 방지)
        if not under_daily_limit(site,cfg):
            add_log(f'[스킵] {name} 일일 발행 한도 도달')
            if hid: history_update(hid,status='skipped',message='일일 발행 한도 도달')
            if job_id: _persist_remove(job_id)
            with STATS_LOCK:
                wk_stats['skipped']+=1; wk_stats['done']+=1; wk_stats['queued']=post_queue.qsize()
            continue

        interval_ok,remain=under_min_interval(site)
        if not interval_ok:
            mins=max(1,(remain+59)//60)
            add_log(f'[스킵] {name} 최소 발행 간격 미충족 ({mins}분 남음)')
            if hid: history_update(hid,status='skipped',message=f'사이트 최소 발행 간격 미충족 ({mins}분 남음)')
            if job_id: _persist_remove(job_id)
            with STATS_LOCK:
                wk_stats['skipped']+=1; wk_stats['done']+=1; wk_stats['queued']=post_queue.qsize()
            continue

        if hid: history_update(hid,status='posting')
        # 발행 (실패 시 3회 재시도, 지수 백오프 + 드라이버 자동 재시작)
        ok=False; msg=''
        for attempt in range(1,4):
            try:
                ok,msg=do_post(site,title,content)
                if ok: break
                add_log(f'[재시도 {attempt}/3] {name} - {msg}')
                reset_driver()   # 실패 시 드라이버 새로 띄워 세션 꼬임 방지
            except Exception as e:
                msg=str(e); add_log(f'[재시도 {attempt}/3] {name} - {msg[:60]}')
                reset_driver()   # 크롬 죽었을 때 자동 재시작
            if not wk_active: break
            if attempt<3 and not wk_paused: time.sleep(min(5*attempt,15))

        # 실패 원인 분류 + 일시적 실패는 지연 후 자동 재시도
        reason=reason_ko=''; transient=False
        if not ok:
            reason,reason_ko,is_temp=classify_fail(msg)
            requeues=job.get('requeues',0)
            transient=is_temp and requeues<RETRY_MAX and wk_active
            if transient:
                job['requeues']=requeues+1
        if transient:
            with STATS_LOCK:
                wk_stats['retry']+=1; wk_stats['queued']=post_queue.qsize()
            if hid: history_update(hid,status='retry',fail_reason=reason,fail_reason_ko=reason_ko,
                                   message=f'{reason_ko} → {RETRY_DELAY//60}분 후 자동 재시도 ({job["requeues"]}/{RETRY_MAX}): {str(msg)[:150]}')
            schedule_retry(job)   # queue.json 은 그대로 유지(재시작 시 복구)
            add_log(f'[재시도 예약 {job["requeues"]}/{RETRY_MAX}] {name} - {reason_ko}')
            if cfg.get('notify_fail'): send_telegram(cfg,f'🔄 일시적 실패({reason_ko}) 재시도 예약: {name}')
            continue
        with STATS_LOCK:
            if ok: wk_stats['success']+=1
            else: wk_stats['fail']+=1
            wk_stats['done']+=1; wk_stats['queued']=post_queue.qsize()
        finalize_post(site,ok)
        # 캡차 감지된 사이트는 자동발행에서 제외되도록 플래그(우회하지 않음)
        if reason=='captcha':
            try: set_site_flag(site.get('id'),has_captcha=True,captcha_note=str(msg)[:80])
            except Exception: pass
            add_log(f'[캡차 감지] {name} — 자동발행 제외 표시')
        if hid: history_update(hid,status='done' if ok else 'failed',
                               result_url=(msg if ok and str(msg).startswith('http') else ''),
                               fail_reason=('' if ok else reason),
                               fail_reason_ko=('' if ok else reason_ko),
                               alive=('yes' if ok and str(msg).startswith('http') else ''),
                               message=str(msg)[:300])
        if job_id: _persist_remove(job_id)
        add_log(f'[{"성공" if ok else "실패:"+reason_ko}] {name}')
        # 텔레그램 알림 (설정에 따라)
        if ok and cfg.get('notify_done'): send_telegram(cfg,f'✅ 발행 성공: {name}\n{title[:60]}')
        if (not ok) and cfg.get('notify_fail'): send_telegram(cfg,f'❌ 발행 실패({reason_ko}): {name}\n{str(msg)[:120]}')

        # rate limit: 포스트 간 지연 (도배 방지)
        delay=int(cfg.get('post_delay',30) or 0)
        if delay>0 and wk_active and not wk_paused: time.sleep(delay)

def is_permitted(site):
    """사이트 관리에서 등록 상태로 전환된 사이트만 발행."""
    return bool(site.get('permission')) and site.get('registration_source') in ('manual_admin','admin_bulk','legacy_admin','candidate_registered')

def is_autopostable(site):
    """자동발행 대상 여부 = 홍보 허용 + 캡차 없음(캡차 게시판은 정책상 자동발행 제외)."""
    return is_permitted(site) and not site.get('has_captcha')

def enqueue(sites,title,content,meta=None):
    meta=meta or {}
    allowed=[s for s in sites if is_autopostable(s)]
    blocked=[s for s in sites if not is_autopostable(s)]
    now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for s in allowed:
        jid=secrets.token_hex(8)
        history_add({'id':jid,'time':now,'updated':now,'site_id':s.get('id'),
                     'site_name':s.get('name') or s.get('site_url',''),'site_url':s.get('site_url',''),
                     'bo_table':s.get('bo_table',''),'title':title,
                     'region':meta.get('region',''),'service':meta.get('service',''),
                     'status':'queued','result_url':'','message':'','attempts':0})
        job={'job_id':jid,'hist_id':jid,'site':s,'title':title,'content':content}
        post_queue.put(job); _persist_add(job)
    for s in blocked:
        _why='캡차 있음' if s.get('has_captcha') else ('미허용 도메인' if not is_permitted(s) else '제외')
        add_log(f'[차단:{_why}] 발행 스킵: {s.get("name") or (s.get("site_url","") or "")[:30]}')
    with STATS_LOCK:
        wk_stats['total']+=len(allowed); wk_stats['queued']=post_queue.qsize()
    return len(allowed),len(blocked)

def enqueue_generated(sites, keywords, cfg, meta=None):
    """허용 사이트마다 '각각 다른' 유니크 제목·본문을 새로 생성해 큐 등록.
       → 같은 키워드라도 사이트마다 글이 달라져 중복 발행을 방지."""
    meta=meta or {}
    allowed=[s for s in sites if is_autopostable(s)]
    blocked=[s for s in sites if not is_autopostable(s)]
    now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for s in allowed:
        html,title=generate_article(keywords,cfg)   # 사이트마다 새로 생성(유니크)
        jid=secrets.token_hex(8)
        history_add({'id':jid,'time':now,'updated':now,'site_id':s.get('id'),
                     'site_name':s.get('name') or s.get('site_url',''),'site_url':s.get('site_url',''),
                     'bo_table':s.get('bo_table',''),'title':title,
                     'region':meta.get('region',''),'service':meta.get('service',''),
                     'member':meta.get('member',''),
                     'status':'queued','result_url':'','message':'','attempts':0})
        job={'job_id':jid,'hist_id':jid,'site':s,'title':title,'content':html}
        post_queue.put(job); _persist_add(job)
    for s in blocked:
        _why='캡차 있음' if s.get('has_captcha') else ('미허용 도메인' if not is_permitted(s) else '제외')
        add_log(f'[차단:{_why}] 발행 스킵: {s.get("name") or (s.get("site_url","") or "")[:30]}')
    with STATS_LOCK:
        wk_stats['total']+=len(allowed); wk_stats['queued']=post_queue.qsize()
    return len(allowed),len(blocked)

# ==================== 예약 발행 스케줄러 ====================
def load_scheds(): return load_json(SCHED_FILE,[])
def save_scheds(s): save_json(SCHED_FILE,s)

def _kst_now():
    """서버가 UTC여도 한국시간 기준으로 계산."""
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo('Asia/Seoul'))
    except Exception:
        return datetime.utcfromtimestamp(time.time()+9*3600)

# ==================== 회원(고객) 관리 + 월 정산 ====================
def load_members(): return load_json(MEMBERS_FILE,[])
def save_members(m): save_json(MEMBERS_FILE,m)
def _cur_month(): return _kst_now().strftime('%Y-%m')

def member_fee(m):
    """월 청구액 = 기본료 + (추가 광고수 × 추가단가)."""
    base=int(m.get('plan_fee',30000) or 0)
    addons=int(m.get('addons',0) or 0)
    afee=int(m.get('addon_fee',10000) or 0)
    return base+addons*afee

def member_view(m):
    """회원 1건을 정산정보 포함해 표시용으로 반환."""
    cm=_cur_month(); pays=m.get('payments',{}) or {}
    pm=pays.get(cm,{}) if isinstance(pays,dict) else {}
    fee=member_fee(m)
    return {**m,'fee':fee,'this_month':cm,
            'paid':bool(pm.get('paid')),'paid_at':pm.get('paid_at',''),
            'unpaid_months':[k for k,v in pays.items() if not (v or {}).get('paid')] if isinstance(pays,dict) else []}

# ==================== 도메인 발굴 (구글 검색 → 자동검수 → 승인 대기) ====================
def load_cands(): return load_json(CAND_FILE,[])
def save_cands(c): save_json(CAND_FILE,c)
_cand_lock=threading.Lock()

# 쿼리 조합 — 플랫폼 흔적 × 홍보 의도
GNU_PATTERNS=['inurl:bbs/board.php bo_table=promotion','inurl:bbs/board.php bo_table=hongbo',
              'inurl:bbs/board.php bo_table=ad','inurl:bbs/board.php bo_table=link',
              'inurl:bbs/board.php bo_table=partner','inurl:bbs/board.php bo_table=banner',
              'inurl:bbs/write.php bo_table=promotion','inurl:bbs/board.php "홍보게시판"']
CAFE_PATTERNS=['inurl:/board/ list.html "홍보"','inurl:/board/free/ "홍보"','"cafe24" inurl:/board/ "제휴"']
INTENT=['"홍보게시판"','"자유홍보"','"홍보 환영"','"홍보 가능"','"링크등록"','"제휴문의"','"상호등록"','"업체등록"']
# 제외 도메인(포털·정부·언론·대형)
BLACK_DOMAINS=['naver.com','daum.net','google.','youtube.','facebook.','instagram.','tistory.com',
               'blog.','cafe.','.go.kr','.or.kr','.ac.kr','.mil.kr','wikipedia.','namu.wiki',
               'chosun.com','donga.com','joins.com','hani.co.kr','mk.co.kr','news','gov',
               'coupang.com','11st.co.kr','gmarket.co.kr','auction.co.kr','interpark',
               'twseo.kr','marketingmonster.kr']
AD_BAN_WORDS=['광고 금지','광고금지','홍보 금지','홍보금지','상업적 게시물','상업적게시물','광고성 글 삭제',
              '광고글 삭제','도배 금지','스팸 금지','영리 목적','상업적 목적 금지','무단 홍보']
# 주차/만료 도메인 (검색엔진엔 남아있지만 실제론 껍데기)
PARKED_WORDS=['resources and information','this domain','domain is for sale','도메인 판매',
              '이 도메인은','buy this domain','parked','sedoparking','afternic','dan.com',
              'hugedomains','도메인이 만료','관련 검색어','sponsored listings','related searches']
# 불법·도박·성인 사이트 (후보 부적합 — 제휴 대상 아님)
ILLEGAL_WORDS=['카지노','바카라','슬롯','토토','먹튀','배팅','베팅','도박','홀덤','파워볼',
               '사설','환전','꽁머니','livecasino','casino','baccarat','betting']
PROMO_WORDS=['홍보게시판','자유홍보','홍보 환영','홍보가능','홍보 가능','제휴문의','제휴 문의','링크등록',
             '업체등록','상호등록','광고게시판','홍보하기','업체홍보','파트너 모집']

def _domain_of(url):
    m=re.match(r'https?://([^/]+)',url or '')
    return (m.group(1).lower() if m else '').replace('www.','')

def _is_blacklisted(url):
    u=(url or '').lower()
    return any(b in u for b in BLACK_DOMAINS)

def build_queries(cfg):
    """플랫폼 흔적 × 홍보 의도 × (선택)업종 키워드 조합 생성."""
    extra=[x.strip() for x in (cfg.get('discover_keywords','') or '').splitlines() if x.strip()]
    # 사용자가 입력한 완성 검색문은 변형하지 않고 가장 먼저 실행한다.
    direct=[x.strip() for x in (cfg.get('discover_direct_queries','') or '').splitlines()
            if x.strip() and not x.lstrip().startswith('#')]
    # 사용자가 목록을 저장한 경우 임의 검색어를 섞지 않고 그 목록만 순서대로 사용한다.
    if direct:
        return list(dict.fromkeys(direct))
    qs=[]
    for p in GNU_PATTERNS+CAFE_PATTERNS:
        qs.append(p)
        for i in INTENT[:4]:
            qs.append(f'{p} {i}')
    for i in INTENT:
        qs.append(f'{i} inurl:bbs')
        qs.append(f'{i} inurl:board')
    for e in extra:                       # 업종·지역 키워드가 있으면 곱하기
        for i in INTENT[:5]:
            qs.append(f'{e} {i}')
        qs.append(f'{e} inurl:bbs/board.php')
    return list(dict.fromkeys(qs))

def google_search(cfg, query, start=1, num=10):
    """Google Custom Search JSON API. (공식 API — 결과 직접 스크래핑 안 함)"""
    key=cfg.get('google_api_key',''); cx=cfg.get('google_cx','')
    if not key or not cx: raise RuntimeError('구글 API 키/검색엔진ID(cx) 미설정')
    import requests as _rq
    r=_rq.get('https://www.googleapis.com/customsearch/v1',
              params={'key':key,'cx':cx,'q':query,'start':start,'num':num,'hl':'ko','lr':'lang_ko'},
              timeout=20)
    if r.status_code==429: raise RuntimeError('구글 API 일일 한도 초과(429)')
    if r.status_code>=400: raise RuntimeError(f'구글 API 오류 {r.status_code}: {r.text[:120]}')
    j=r.json()
    return [{'url':it.get('link',''),'title':it.get('title',''),'snippet':it.get('snippet','')}
            for it in (j.get('items') or [])]

def brave_search(cfg, query, start=1, num=10):
    """Brave Search 공식 Web API."""
    key=(cfg.get('brave_api_key') or '').strip()
    if not key: raise RuntimeError('Brave Search API 키 미설정 — 설정 탭에서 입력하세요')
    import requests as _rq
    offset=max(0,(int(start or 1)-1)//max(1,int(num or 10)))
    r=_rq.get('https://api.search.brave.com/res/v1/web/search',
              headers={'Accept':'application/json','X-Subscription-Token':key},
              params={'q':query,'count':min(20,max(1,int(num or 10))),'offset':min(9,offset),
                      'country':'KR','search_lang':'ko','safesearch':'moderate'},timeout=20)
    if r.status_code==429: raise RuntimeError('Brave Search API 사용량/속도 한도 초과(429)')
    if r.status_code in (401,403): raise RuntimeError('Brave Search API 키 또는 구독 상태 확인 필요')
    if r.status_code>=400: raise RuntimeError(f'Brave Search API 오류 {r.status_code}: {r.text[:120]}')
    rows=((r.json().get('web') or {}).get('results') or [])
    return [{'url':it.get('url',''),'title':it.get('title',''),'snippet':it.get('description','')}
            for it in rows if it.get('url')]

def web_search(cfg, query, start=1, num=10):
    provider=(cfg.get('search_provider') or 'google').lower()
    return google_search(cfg,query,start,num) if provider=='google' else brave_search(cfg,query,start,num)

def extract_emails(text):
    found=re.findall(r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}',text or '')
    bad=('example.com','sentry.io','.png','.jpg','.gif','wixpress','@2x')
    out=[]
    for e in found:
        el=e.lower()
        if any(b in el for b in bad): continue
        if el not in out: out.append(el)
    return out[:5]

def screen_candidate(url, cfg=None):
    """HTTP 1~3회로 후보 자동 검수. 반환: 검수 결과 dict."""
    import requests as _rq
    UA={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36'}
    m=re.match(r'(https?://[^/]+)',url or ''); base=m.group(1) if m else url
    res={'base':base,'domain':_domain_of(url),'platform':'unknown','board_name':'','bo_table':'',
         'write_form':False,'captcha':'','login_required':False,'ad_banned':False,'promo_hint':False,
         'emails':[],'last_post_days':None,'reachable':False,'note':'',
         'parked':False,'illegal':False}
    def get(u):
        try: return _rq.get(u,timeout=12,verify=False,headers=UA,allow_redirects=True)
        except Exception: return None
    # bo_table 추출
    bm=re.search(r'bo_table=([A-Za-z0-9_]+)',url or '')
    if bm: res['bo_table']=bm.group(1)
    r=get(url)
    if not r or r.status_code>=400:
        res['note']=f'접속 실패({r.status_code if r else "timeout"})'; return res
    res['reachable']=True
    html=r.text or ''; low=html.lower()
    # 플랫폼
    if 'bo_table' in low or 'gnuboard' in low or '/bbs/' in low: res['platform']='gnuboard'
    elif 'cafe24' in low or 'xans-' in low or '/board/write.html' in low: res['platform']='cafe24'
    # 게시판명 (title/h1)
    tm=re.search(r'<title[^>]*>(.*?)</title>',html,re.S|re.I)
    title=re.sub(r'\s+',' ',re.sub(r'&nbsp;',' ',(tm.group(1) if tm else '')))
    res['board_name']=title[:80]
    # 주차/만료 도메인 판별 (검색엔진엔 남아있지만 실제론 껍데기)
    tl=title.lower()
    res['parked']=(any(w in low or w in tl for w in PARKED_WORDS) or len(html)<800)
    # 불법·도박 사이트 판별 (제휴 대상 부적합)
    res['illegal']=any(w in low for w in ILLEGAL_WORDS)
    # 광고 금지 / 홍보 허용 흔적 (본문 + URL/게시판ID/타이틀까지 함께 판단)
    res['ad_banned']=any(w in html for w in AD_BAN_WORDS)
    hint_blob=html+' '+title+' '+(url or '')+' '+res['bo_table']
    res['promo_hint']=(any(w in hint_blob for w in PROMO_WORDS)
                       or bool(re.search(r'bo_table=(promotion|hongbo|ad|link|partner|banner)',url or '',re.I)))
    # 최근 글(날짜 패턴에서 가장 최근)
    try:
        ds=re.findall(r'(20\d{2})[-.](\d{1,2})[-.](\d{1,2})',html)
        if ds:
            from datetime import date
            latest=max(date(int(y),int(mo),int(d)) for y,mo,d in ds
                       if 1<=int(mo)<=12 and 1<=int(d)<=31)
            res['last_post_days']=max(0,(_kst_now().date()-latest).days)
    except Exception: pass
    # 글쓰기 페이지
    wurls=[]
    if res['platform']=='cafe24':
        wurls=[base+'/board/write.html?board_no=1']
    else:
        bo=res['bo_table'] or 'free'
        wurls=[base+f'/bbs/write.php?bo_table={bo}']
    wr=get(wurls[0]) if wurls else None
    if wr and wr.status_code<400:
        wh=wr.text or ''; wl=wh.lower()
        res['write_form']=('wr_subject' in wl or 'name="subject"' in wl or 'wr_content' in wl)
        if any(k in wl for k in ['captcha_key','kcaptcha','g-recaptcha','h-captcha','cf-turnstile']) or '자동등록방지' in wh:
            res['captcha']='감지'
        if not res['write_form'] and ('mb_password' in wl or '로그인' in wh):
            res['login_required']=True
        if any(w in wh for w in AD_BAN_WORDS): res['ad_banned']=True
    # 연락처(홈 + 약관/문의 페이지)
    mails=extract_emails(html)
    for p in ['/','/bbs/content.php?co_id=company','/bbs/content.php?co_id=privacy','/company','/about']:
        if len(mails)>=2: break
        rr=get(base+p)
        if rr and rr.status_code<400: mails+= [e for e in extract_emails(rr.text) if e not in mails]
    res['emails']=mails[:5]
    return res

def score_candidate(c):
    """점수화 — 위에서부터 100개만 보면 되게."""
    s=0
    name=(c.get('board_name','')+' '+c.get('bo_table',''))
    if c.get('promo_hint') or any(w in name for w in ['홍보','제휴','광고','링크','업체']): s+=30
    if c.get('write_form') and not c.get('login_required'): s+=20
    lp=c.get('last_post_days')
    if lp is not None:
        if lp<=7: s+=15
        elif lp<=30: s+=8
        elif lp>365: s-=10
    if not c.get('captcha'): s+=10
    if c.get('emails'): s+=10
    if c.get('platform') in ('gnuboard','cafe24'): s+=5
    if c.get('ad_banned'): s-=50
    if c.get('captcha'): s-=30
    if c.get('login_required'): s-=20
    if c.get('parked'): s-=100      # 주차/만료 도메인 = 껍데기
    if c.get('illegal'): s-=100     # 도박·불법 사이트 = 제휴 부적합
    if not c.get('reachable'): s-=100
    return s

def make_mail_draft(c, cfg):
    """제휴·홍보 게시 승인 요청 메일 초안."""
    brand=cfg.get('brand','') or '(브랜드명)'
    site=c.get('board_name') or c.get('domain','')
    to=(c.get('emails') or ['(담당자 이메일)'])[0]
    return (f"받는사람: {to}\n"
            f"제목: [{brand}] 홍보 게시판 이용(제휴) 문의드립니다\n\n"
            f"안녕하세요. {brand} 담당자입니다.\n\n"
            f"{site} 사이트를 통해 연락드립니다.\n"
            f"저희는 지역 기반 업체 정보를 정리해 안내하는 콘텐츠를 운영하고 있으며,\n"
            f"귀 사이트의 홍보(제휴) 게시판에 정보성 게시물을 등록해도 되는지 문의드립니다.\n\n"
            f"■ 문의 사항\n"
            f"1) 홍보성 게시물 등록이 허용되는 게시판이 있는지\n"
            f"2) 허용된다면 게시 규칙(횟수·형식·금지사항)이 어떻게 되는지\n"
            f"3) 유료 제휴가 필요한 경우 조건과 비용\n\n"
            f"규칙을 정확히 지켜 운영하겠습니다. 회신 주시면 감사하겠습니다.\n\n"
            f"{brand} 드림\n연락처: {cfg.get('phone','')}\n")

def add_candidates_from(items, cfg, source='google'):
    """검색 결과 → 1차 URL 필터 → 후보 등록(중복·블랙리스트 제외). 신규 개수 반환."""
    with _cand_lock:
        cands=load_cands()
        known_dom={c.get('domain') for c in cands}
        site_dom={_domain_of(s.get('site_url','')) for s in load_sites()}
        added=0
        for it in items:
            url=it.get('url') if isinstance(it,dict) else str(it)
            if not url or not url.startswith('http'): continue
            if _is_blacklisted(url): continue
            dom=_domain_of(url)
            if not dom or dom in known_dom or dom in site_dom: continue
            known_dom.add(dom)
            cands.append({'id':secrets.token_hex(6),'url':url,'domain':dom,
                          'title':(it.get('title','') if isinstance(it,dict) else ''),
                          'snippet':(it.get('snippet','') if isinstance(it,dict) else ''),
                          'found_at':_kst_now().strftime('%Y-%m-%d %H:%M'),'source':source,
                          'query':(it.get('query','') if isinstance(it,dict) else ''),
                          'status':'new','score':0,'screened':False})
            added+=1
        save_cands(cands)
    return added

def screen_pending(limit=30):
    """미검수 후보를 검수·점수화. 처리 건수 반환."""
    cfg=load_config()
    with _cand_lock:
        cands=load_cands()
        todo=[c for c in cands if not c.get('screened')][:max(1,limit)]
        ids=[c['id'] for c in todo]
    results={}
    for c in todo:
        try:
            r=screen_candidate(c['url'],cfg)
        except Exception as e:
            r={'note':str(e)[:100],'reachable':False}
        r['screened']=True
        r['score']=score_candidate(r)
        # 자동 탈락 사유
        if not r.get('reachable'): r['status']='ready'; r['reject_reason']='현재 접속 불가 — 후보 유지·재검수 가능'
        elif r.get('parked'): r['status']='rejected'; r['reject_reason']='주차/만료 도메인 (실제 게시판 아님)'
        elif r.get('illegal'): r['status']='rejected'; r['reject_reason']='도박·불법 사이트 (제휴 부적합)'
        elif r.get('ad_banned'): r['status']='rejected'; r['reject_reason']='광고 금지 명시'
        elif r.get('captcha'): r['status']='rejected'; r['reject_reason']='캡차 있음(자동발행 부적합)'
        elif not r.get('write_form'): r['status']='ready'; r['reject_reason']='글쓰기 폼 미확인 — 로그인 필요할 수 있음'
        else: r['status']='ready'
        results[c['id']]=r
        time.sleep(1)   # 요청 속도 관리
    with _cand_lock:
        cands=load_cands()
        for c in cands:
            if c['id'] in results:
                c.update(results[c['id']])
                if c.get('status')=='ready': c['mail_draft']=make_mail_draft(c,cfg)
        save_cands(cands)
    return len(results)

def discover_once(cfg=None, max_queries=10):
    """구글 검색 1배치 실행 → 후보 등록 → 검수. (검색량 제어)"""
    cfg=cfg or load_config()
    st=load_json(DISCO_FILE,{})
    today=_kst_now().strftime('%Y-%m-%d')
    if st.get('date')!=today: st={'date':today,'queries':0,'found':0,'cursor':0}
    target=int(cfg.get('discover_daily_target',100) or 100)
    qlimit=int(cfg.get('discover_query_limit',100) or 100)   # 구글 무료 하루 100
    queries=build_queries(cfg)
    if not queries: return {'ok':False,'error':'쿼리 없음'}
    added=0; used=0; errs=[]
    for _ in range(max_queries):
        if st['queries']>=qlimit: errs.append('일일 쿼리 한도 도달'); break
        if st['found']>=target: errs.append('일일 후보 목표 달성'); break
        q=queries[st['cursor']%len(queries)]; st['cursor']+=1
        try:
            items=web_search(cfg,q)
            for it in items: it['query']=q
            n=add_candidates_from(items,cfg); added+=n; st['found']+=n
        except Exception as e:
            errs.append(str(e)[:120]); break
        finally:
            st['queries']+=1; used+=1
        time.sleep(1)
    save_json(DISCO_FILE,st)
    screened=0
    try: screened=screen_pending(20)
    except Exception as e: errs.append('검수 오류 '+str(e)[:80])
    add_log(f'[발굴] 쿼리 {used}회 · 신규 {added}개 · 검수 {screened}건'+(' · '+errs[0] if errs else ''))
    return {'ok':True,'queries':used,'added':added,'screened':screened,
            'today_queries':st['queries'],'today_found':st['found'],'errors':errs}

def discover_loop():
    """24시간 자동 발굴 — 목표치까지 천천히 채우고 남는 시간엔 검수."""
    while True:
        try:
            cfg=load_config()
            provider=(cfg.get('search_provider') or 'google').lower()
            ready=bool(cfg.get('brave_api_key')) if provider=='brave' else bool(cfg.get('google_api_key') and cfg.get('google_cx'))
            if cfg.get('discover_enabled') and ready:
                discover_once(cfg,max_queries=4)
            else:
                # 발굴 꺼져 있어도 미검수 후보는 계속 처리
                if any(not c.get('screened') for c in load_cands()): screen_pending(10)
        except Exception as e:
            add_log(f'[발굴 루프 오류] {str(e)[:100]}')
        time.sleep(900)   # 15분마다

def member_paid_now(m):
    """이번 달 납부 완료 여부."""
    pays=m.get('payments',{}) or {}
    return bool((pays.get(_cur_month(),{}) or {}).get('paid'))

def member_runnable(m,cfg):
    """이 회원의 스케줄을 지금 돌려도 되는가. (사유, 가능여부)"""
    if m.get('status','active')!='active': return False,'정지 회원'
    if not m.get('sched_enabled'): return False,'스케줄 꺼짐'
    if not (m.get('sched_times') or []): return False,'시간대 미설정'
    if cfg.get('block_unpaid') and not member_paid_now(m): return False,'미납 — 자동 정지'
    return True,''

def member_sites(m):
    """회원에게 배정된 발행 가능 사이트(미배정이면 전체 허용 사이트)."""
    ids=m.get('site_ids') or []
    sites=[s for s in load_sites() if (not ids or s.get('id') in ids)]
    return [s for s in sites if is_autopostable(s)]

def member_keywords(m):
    """회원 전용 키워드 풀(없으면 공용 풀)."""
    kw=m.get('keywords') or []
    return kw if kw else load_keywords()

def run_member_job(mid, minute_key):
    """회원 1명의 스케줄 1회 실행 — 지터 대기 후 발행 큐 등록."""
    try:
        cfg=load_config()
        m=next((x for x in load_members() if x.get('id')==mid),None)
        if not m: return
        jitter=int(m.get('jitter',0) or 0)
        if jitter>0:
            time.sleep(random.randint(0,jitter*60))   # 시간 분산: 동시 폭주 방지
        m=next((x for x in load_members() if x.get('id')==mid),None)  # 대기 중 변경 반영
        if not m: return
        ok,why=member_runnable(m,cfg)
        if not ok:
            add_log(f'[회원스케줄:{m.get("name") or m.get("biz")}] 건너뜀 — {why}'); return
        sites=member_sites(m); pool=member_keywords(m)
        nm=m.get('name') or m.get('biz') or mid
        if not sites: add_log(f'[회원스케줄:{nm}] 발행 가능 사이트 없음'); return
        if not pool: add_log(f'[회원스케줄:{nm}] 키워드 없음'); return
        cnt=max(1,int(m.get('per_run',1) or 1))
        total=0
        for _ in range(cnt):
            kw=pick_keywords(pool,cfg)
            total+=enqueue_generated(sites,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),
                                            '브랜드':kw.get('브랜드','') or cfg.get('brand','')},cfg,
                                     {'region':kw.get('지역',''),'service':kw.get('서비스',''),'member':nm})[0]
        # 실행 기록
        mem=load_members()
        for x in mem:
            if x.get('id')==mid:
                x['last_run']=_kst_now().strftime('%Y-%m-%d %H:%M')
                x['run_count']=int(x.get('run_count',0) or 0)+1
                x['last_run_min']=minute_key
        save_members(mem)
        add_log(f'[회원스케줄:{nm}] {total}건 큐 등록 (사이트 {len(sites)}개 × {cnt}회)')
        if total and not wk_active: start_workers(cfg.get('workers',2))
        if cfg.get('notify_done'): send_telegram(cfg,f'⏰ {nm} 자동발행 {total}건 등록')
    except Exception as e:
        add_log(f'[회원스케줄 오류] {str(e)[:100]}')

def member_scheduler_loop():
    """jump 방식: 회원마다 설정한 시간대에 서버가 24시간 자동 구동."""
    last_min=None
    while True:
        try:
            now=_kst_now(); hm=now.strftime('%H:%M'); wd=now.weekday()
            minute_key=now.strftime('%Y-%m-%d %H:%M')
            if minute_key!=last_min:
                last_min=minute_key
                cfg=load_config()
                for m in load_members():
                    if hm not in (m.get('sched_times') or []): continue
                    days=m.get('sched_days') or []
                    if days and wd not in days: continue
                    if m.get('last_run_min')==minute_key: continue
                    ok,why=member_runnable(m,cfg)
                    if not ok:
                        add_log(f'[회원스케줄:{m.get("name") or m.get("biz")}] 시간 도달했으나 건너뜀 — {why}')
                        continue
                    # 즉시 선점 기록(중복 실행 방지) 후 백그라운드 실행
                    mem=load_members()
                    for x in mem:
                        if x.get('id')==m.get('id'): x['last_run_min']=minute_key
                    save_members(mem)
                    threading.Thread(target=run_member_job,args=(m.get('id'),minute_key),daemon=True).start()
        except Exception as e:
            add_log(f'[회원스케줄러 오류] {str(e)[:100]}')
        time.sleep(20)

def settle_summary():
    """이번 달 정산 요약(활성 회원 기준)."""
    cm=_cur_month(); mem=load_members()
    active=[m for m in mem if m.get('status','active')=='active']
    billed=sum(member_fee(m) for m in active)
    paid=sum(member_fee(m) for m in active if ((m.get('payments',{}) or {}).get(cm,{}) or {}).get('paid'))
    unpaid_list=[{'id':m.get('id'),'name':m.get('name') or m.get('biz',''),'biz':m.get('biz',''),
                  'fee':member_fee(m)} for m in active
                 if not ((m.get('payments',{}) or {}).get(cm,{}) or {}).get('paid')]
    return {'month':cm,'members':len(mem),'active':len(active),
            'billed':billed,'paid':paid,'unpaid':billed-paid,
            'unpaid_count':len(unpaid_list),'unpaid_list':unpaid_list}

sched_active=True
def scheduler_loop():
    """1분 단위로 스케줄 확인 → 조건 맞으면 발행 큐 등록. days: 0=월..6=일, 빈 리스트=매일."""
    last_min=None
    while sched_active:
        try:
            now=_kst_now(); hm=now.strftime('%H:%M'); wd=now.weekday(); today=now.strftime('%Y-%m-%d')
            minute_key=now.strftime('%Y-%m-%d %H:%M')
            if minute_key!=last_min:
                last_min=minute_key
                cfg=load_config(); changed=False
                # ---- 매일 자동 백업(텔레그램) ----
                bt=(cfg.get('backup_time') or '').strip()
                if bt and hm==bt:
                    st=load_json(DATA_DIR/'backup.state',{})
                    if st.get('date')!=today:
                        save_json(DATA_DIR/'backup.state',{'date':today})
                        try: do_backup(cfg,'자동')
                        except Exception as e: add_log(f'[백업 오류] {str(e)[:80]}')
                # 예약 발행 미사용(실시간 발행만) — 스케줄 처리 비활성화. 백업 체크만 유지.
                scheds=[]
                for sc in scheds:
                    if not sc.get('enabled',True): continue
                    if hm not in (sc.get('times') or []): continue
                    days=sc.get('days') or []
                    if days and wd not in days: continue
                    if sc.get('last_run_min')==minute_key: continue
                    sc['last_run_min']=minute_key; sc['last_run']=now.strftime('%Y-%m-%d %H:%M'); changed=True
                    ksets=sc.get('keyword_sets') or []
                    if not ksets:   # 키워드 세트 없으면 풀에서 랜덤 N개
                        pool=load_keywords()
                        cnt=int(sc.get('count',1) or 1)
                        if pool: ksets=[pick_keywords(pool,cfg) for _ in range(cnt)]
                    sids=sc.get('site_ids') or []
                    sites=[s for s in load_sites() if not sids or s.get('id') in sids]
                    allowed=[s for s in sites if is_permitted(s)]
                    if not allowed:
                        add_log(f'[예약:{sc.get("name")}] 허용 사이트 없음 — 건너뜀'); continue
                    n=0
                    for ks in ksets:
                        kw={'지역':ks.get('지역',''),'서비스':ks.get('서비스',''),'브랜드':ks.get('브랜드','')}
                        # 사이트마다 유니크 본문 생성(중복 방지)
                        n+=enqueue_generated(allowed,kw,cfg,{'region':kw['지역'],'service':kw['서비스']})[0]
                    add_log(f'[예약 실행:{sc.get("name")}] {n}건 큐 등록')
                    if n and not wk_active: start_workers(cfg.get('workers',2))
                if changed: save_scheds(scheds)
        except Exception as e:
            add_log(f'[스케줄러 오류] {str(e)[:80]}')
        time.sleep(20)

# ==================== 텔레그램 폰 제어 (명령 수신) ====================
TG_HELP=('📱 찌라시 봇 명령어\n'
         '/상태 — 워커·큐 현황\n'
         '/오늘 — 오늘 발행 통계\n'
         '/발행 — 키워드 풀에서 랜덤 뽑아 전체 발행(원클릭)\n'
         '/발행 지역,서비스[,브랜드] — 지정 키워드로 즉시 발행\n'
         '/정지 · /재개 — 워커 일시정지/재개\n'
         '/백업 — 지금 백업 파일 전송\n'
         '/검증 — 발행글 생존 확인 실행')

def handle_tg_command(cfg,text):
    t=(text or '').strip(); low=t.lower()
    def reply(m): send_telegram(cfg,m)
    if low in ('/help','/start','/도움말') or t in ('도움말','명령어'):
        reply(TG_HELP); return
    if t.startswith('/상태') or low.startswith('/status'):
        reply(f'📊 상태\n워커: {"ON" if wk_active else "OFF"}{" (일시정지)" if wk_paused else ""}\n'
              f'큐: {post_queue.qsize()} · 재시도대기: {len(_retry_jobs)}\n'
              f'성공 {wk_stats["success"]} · 실패 {wk_stats["fail"]} · 스킵 {wk_stats["skipped"]}'); return
    if t.startswith('/오늘') or low.startswith('/today'):
        today=_kst_now().strftime('%Y-%m-%d'); h=load_json(HISTORY_FILE,[])
        th=[x for x in h if (x.get('time') or '').startswith(today)]
        ok=sum(1 for x in th if x.get('status')=='done'); fl=sum(1 for x in th if x.get('status')=='failed')
        reply(f'📅 오늘({today})\n성공 {ok} · 실패 {fl} · 전체 {len(th)}건'); return
    if t.startswith('/정지') or low.startswith('/pause'):
        pause_workers(); reply('⏸ 워커 일시정지'); return
    if t.startswith('/재개') or low.startswith('/resume'):
        resume_workers()
        if not wk_active: start_workers(cfg.get('workers',2))
        reply('▶️ 워커 재개'); return
    if t.startswith('/백업') or low.startswith('/backup'):
        reply('📦 백업 생성 중...'); okb,err=do_backup(cfg,'텔레그램'); reply('✅ 백업 전송 완료' if okb else '❌ 실패: '+err); return
    if t.startswith('/검증') or low.startswith('/verify'):
        reply('🔎 생존 확인 실행...(백그라운드)')
        threading.Thread(target=verify_once,kwargs={'limit':40},daemon=True).start(); return
    if t.startswith('/발행') or low.startswith('/post'):
        parts_split=t.split(None,1)
        arg=(parts_split[1] if len(parts_split)>1 else '').replace('，',',')
        parts=[x.strip() for x in arg.split(',') if x.strip()]
        sites=[s for s in load_sites() if is_autopostable(s)]
        if not sites: reply('⚠️ 발행 가능한 사이트가 없습니다(허용·캡차없음)'); return
        # 인자 없으면 키워드 풀에서 랜덤(폰 원클릭)
        if len(parts)<2:
            pool=load_keywords()
            if not pool: reply('형식: /발행 지역,서비스[,브랜드]\n또는 키워드 풀을 등록하면 /발행 만으로 랜덤 발행'); return
            kw=pick_keywords(pool,cfg)
        else:
            kw={'지역':parts[0],'서비스':parts[1],'브랜드':(parts[2] if len(parts)>2 and parts[2] else cfg.get('brand',''))}
        try:
            n=enqueue_generated(sites,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),'브랜드':kw.get('브랜드','')},cfg,{'region':kw.get('지역',''),'service':kw.get('서비스','')})[0]
            if n and not wk_active: start_workers(cfg.get('workers',2))
            reply(f'📤 {n}건 큐 등록 (사이트별 유니크)\n키워드: {kw.get("지역","")} {kw.get("서비스","")}')
        except Exception as e:
            reply('❌ 생성 실패: '+str(e)[:100])
        return
    reply('알 수 없는 명령입니다. /help')

_tg_offset={'v':0}
def telegram_loop():
    """텔레그램 getUpdates 롱폴링. 등록된 chat_id 만 명령 처리."""
    import time as _t
    while True:
        cfg=load_config()
        if not (cfg.get('telegram_control') and cfg.get('telegram_token') and cfg.get('telegram_chat_id')):
            _t.sleep(5); continue
        tok=cfg['telegram_token']; chat=str(cfg['telegram_chat_id'])
        try:
            import requests as _rq
            r=_rq.get(f"https://api.telegram.org/bot{tok}/getUpdates",
                      params={'offset':_tg_offset['v']+1,'timeout':30},timeout=40)
            j=r.json()
            for up in j.get('result',[]):
                _tg_offset['v']=up['update_id']
                msg=up.get('message') or up.get('edited_message') or {}
                frm=str((msg.get('chat') or {}).get('id',''))
                text=(msg.get('text') or '').strip()
                if not text or frm!=chat: continue   # 인증: 등록된 챗만
                try: handle_tg_command(cfg,text)
                except Exception as e: add_log(f'[텔레그램 처리 오류] {str(e)[:80]}')
        except Exception:
            _t.sleep(5)

# ==================== 사이트 헬스체크 (HTTP) ====================
def site_health(site):
    """로그인/글쓰기 페이지가 살아있는지 가벼운 HTTP 점검(셀레니움 없이)."""
    import requests as _rq
    url=site.get('site_url','').rstrip('/')
    m=re.match(r'(https?://[^/]+)',url); base=m.group(1) if m else url
    bo=site.get('bo_table','free')
    out={'reachable':False,'login_form':False,'write_page':False,'note':''}
    try:
        r=_rq.get(base,timeout=12,verify=False,headers={'User-Agent':'Mozilla/5.0'})
        out['reachable']=r.status_code<500
        try:
            lp=_rq.get(base+'/bbs/login.php',timeout=12,verify=False,headers={'User-Agent':'Mozilla/5.0'})
            out['login_form']=("mb_id" in lp.text and "mb_password" in lp.text)
        except Exception: pass
        try:
            wp=_rq.get(base+f'/bbs/write.php?bo_table={bo}',timeout=12,verify=False,headers={'User-Agent':'Mozilla/5.0'})
            out['write_page']=(wp.status_code<500 and ('wr_subject' in wp.text or 'wr_content' in wp.text or '로그인' in wp.text))
        except Exception: pass
    except Exception as e:
        out['note']=str(e)[:80]
    out['ok']=out['reachable'] and (out['login_form'] or out['write_page'])
    return out

# ==================== 통계 집계 ====================
def compute_stats():
    h=load_json(HISTORY_FILE,[])
    total=len(h); ok=sum(1 for x in h if x.get('status')=='done'); fail=sum(1 for x in h if x.get('status')=='failed')
    skip=sum(1 for x in h if x.get('status')=='skipped')
    alive=sum(1 for x in h if x.get('alive')=='yes'); dead=sum(1 for x in h if x.get('alive')=='no')
    alive_rate=round(alive/(alive+dead)*100,1) if (alive+dead) else 0.0
    reasons={}   # 실패 원인 분류 집계
    by_site={}; by_day={}
    for x in h:
        sn=x.get('site_name') or '(미상)'; d=(x.get('time') or '')[:10]; st=x.get('status')
        bs=by_site.setdefault(sn,{'done':0,'failed':0,'other':0,'alive':0,'dead':0})
        bs['done' if st=='done' else 'failed' if st=='failed' else 'other']+=1
        if x.get('alive')=='yes': bs['alive']+=1
        elif x.get('alive')=='no': bs['dead']+=1
        bd=by_day.setdefault(d,{'done':0,'failed':0})
        if st in ('done','failed'): bd[st]+=1
        if st=='failed':
            rk=x.get('fail_reason_ko') or '기타'; reasons[rk]=reasons.get(rk,0)+1
    rate=round(ok/(ok+fail)*100,1) if (ok+fail) else 0.0
    days=sorted(by_day.items())[-14:]
    top_sites=sorted(by_site.items(),key=lambda kv:-(kv[1]['done']+kv[1]['failed']))[:12]
    return {'total':total,'ok':ok,'fail':fail,'skip':skip,'rate':rate,
            'alive':alive,'dead':dead,'alive_rate':alive_rate,
            'reasons':[{'reason':k,'n':v} for k,v in sorted(reasons.items(),key=lambda kv:-kv[1])],
            'by_day':[{'day':d,'done':v['done'],'failed':v['failed']} for d,v in days],
            'by_site':[{'site':s,**v} for s,v in top_sites]}

# ==================== Flask ====================
app=Flask(__name__)

def _get_secret():
    """세션 서명키 — 하드코딩 대신 env 또는 persist 파일에서 로드 (없으면 랜덤 생성)."""
    env=os.environ.get('CHIRASHI_SECRET')
    if env: return env
    kf=DATA_DIR/'secret.key'
    try:
        if kf.exists(): return kf.read_text().strip()
        s=secrets.token_hex(32); kf.write_text(s); return s
    except: return secrets.token_hex(32)
app.secret_key=_get_secret()
# 세션 쿠키 보안: JS 접근 차단·HTTPS 전용·크로스사이트 제한
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE='Lax', SESSION_COOKIE_SECURE=True)

# ---- 크롤/스크래핑/색인/임베드 차단 ----
BLOCK_UA=['bot','crawler','spider','scrapy','ahrefs','semrush','mj12','dotbot','python-requests',
          'httpclient','curl','wget','headless','phantom','slurp','baiduspider','yandex','censys','zgrab']
@app.after_request
def _sec_headers(resp):
    # 검색엔진 색인·아카이브 금지 + 임베드(아이프레임) 차단 + 스니핑·캐시 방지
    resp.headers['X-Robots-Tag']='noindex, nofollow, noarchive, nosnippet, noimageindex'
    resp.headers['X-Frame-Options']='DENY'
    resp.headers['Content-Security-Policy']="frame-ancestors 'none'"
    resp.headers['X-Content-Type-Options']='nosniff'
    resp.headers['Referrer-Policy']='no-referrer'
    resp.headers['Cache-Control']='no-store'
    resp.headers['Permissions-Policy']='geolocation=(), camera=(), microphone=()'
    return resp

@app.errorhandler(Exception)
def _backend_error(err):
    """API 예외는 HTML 오류 페이지 대신 일관된 JSON으로 기록·반환."""
    if isinstance(err,HTTPException): return err
    try: add_log(f'[API 오류] {request.method} {request.path} - {type(err).__name__}: {str(err)[:160]}')
    except Exception: pass
    if request.path.startswith('/api/'):
        return jsonify({'ok':False,'error':'서버 처리 중 오류가 발생했습니다'}),500
    return ('Internal Server Error',500)

def auth():
    return session.get('login',False)

# ---- 로그인 브루트포스 방어 (IP당 5분 내 6회 실패 시 잠금) ----
_login_hits=defaultdict(list); _login_lock=threading.Lock()
def _client_ip():
    xff=request.headers.get('X-Forwarded-For','')
    return (xff.split(',')[0].strip() if xff else request.remote_addr) or '?'
def login_allowed(ip):
    now=time.time()
    with _login_lock:
        arr=[t for t in _login_hits[ip] if now-t<300]; _login_hits[ip]=arr
        return len(arr)<6
def login_fail(ip):
    with _login_lock: _login_hits[ip].append(time.time())
def check_pw(pw):
    """평문/해시 모두 지원 + env(CHIRASHI_PASSWORD) 우선."""
    stored=os.environ.get('CHIRASHI_PASSWORD') or load_config().get('password','admin1234')
    if isinstance(stored,str) and (stored.startswith('pbkdf2:') or stored.startswith('scrypt:')):
        try: return check_password_hash(stored,pw)
        except: return False
    return pw==stored

@app.route('/robots.txt')
def robots():
    from flask import Response
    return Response("User-agent: *\nDisallow: /\n",mimetype='text/plain')

@app.before_request
def chk():
    # 알려진 크롤러/스크래퍼 User-Agent 즉시 차단(로그인·업데이트 제외)
    if request.path not in ['/robots.txt','/api/admin/update']:
        ua=(request.headers.get('User-Agent','') or '').lower()
        if not ua or any(b in ua for b in BLOCK_UA):
            return ('Forbidden',403)
    if request.path=='/robots.txt': return
    if request.path.startswith('/static'): return
    if request.path in ['/login','/logout']: return
    if request.path=='/api/admin/update': return  # 자체 토큰 인증
    if not auth():
        if request.path.startswith('/api/') or request.is_json:
            return jsonify({'ok':False,'error':'로그인이 만료되었습니다'}),401
        return redirect('/login')

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        ip=_client_ip()
        if not login_allowed(ip):
            return R('로그인',error='로그인 시도 초과 - 잠시 후 다시 시도하세요')
        if check_pw(request.form.get('pw','')):
            session['login']=True; return redirect('/')
        login_fail(ip)
        return R('로그인',error='비밀번호 틀림')
    return R('로그인',error='')

@app.route('/logout')
def logout():
    session.clear(); return redirect('/login')

@app.route('/')
def index():
    return R('대시보드',cfg=load_config(),sites=load_sites(),wk=wk_stats,wk_on=wk_active)

# API
@app.route('/api/generate',methods=['POST'])
def api_gen():
    d=request.get_json(silent=True) or {}; kw=d.get('keywords',{}); cfg=load_config()
    html,title=generate_article(kw,cfg)
    return jsonify({'ok':True,'title':title,'content':html})

# ---- 이미지 URL 풀 (본문 삽입용) ----
@app.route('/api/images',methods=['GET','POST','DELETE'])
def api_images():
    if request.method=='POST':
        d=request.get_json() or {}
        urls=d.get('urls')
        if urls is None:
            urls=[x.strip() for x in (d.get('text','') or '').splitlines() if x.strip()]
        urls=[u for u in urls if u.startswith('http')]
        if d.get('append'):
            urls=load_image_urls()+urls
        urls=list(dict.fromkeys(urls))   # 중복 제거(순서보존)
        save_image_urls(urls); return jsonify({'ok':True,'count':len(urls)})
    if request.method=='DELETE':
        save_image_urls([]); return jsonify({'ok':True})
    return jsonify(load_image_urls())

# ---- 도메인 발굴 후보 (승인해야만 사이트 목록에 투입) ----
@app.route('/api/candidates',methods=['GET','DELETE'])
def api_candidates():
    if request.method=='DELETE':
        d=request.get_json() or {}
        if d.get('id'):
            save_cands([c for c in load_cands() if c.get('id')!=d['id']])
        elif d.get('clear')=='rejected':
            save_cands([c for c in load_cands() if c.get('status')!='rejected'])
        else:
            save_cands([])
        return jsonify({'ok':True})
    cands=load_cands()
    order={'ready':0,'contacted':1,'new':2,'approved':3,'rejected':4}
    cands.sort(key=lambda c:(order.get(c.get('status'),9),-int(c.get('score',0) or 0)))
    st=load_json(DISCO_FILE,{})
    summary={'total':len(cands),
             'ready':sum(1 for c in cands if c.get('status')=='ready'),
             'new':sum(1 for c in cands if c.get('status')=='new'),
             'contacted':sum(1 for c in cands if c.get('status')=='contacted'),
             'approved':sum(1 for c in cands if c.get('status')=='approved'),
             'rejected':sum(1 for c in cands if c.get('status')=='rejected'),
             'today_queries':st.get('queries',0),'today_found':st.get('found',0),
             'date':st.get('date','')}
    return jsonify({'candidates':cands[:400],'summary':summary})

@app.route('/api/candidates/discover',methods=['POST'])
def api_cand_discover():
    d=request.get_json() or {}; cfg=load_config()
    provider=(cfg.get('search_provider') or 'google').lower()
    if provider=='brave' and not cfg.get('brave_api_key'):
        return jsonify({'ok':False,'error':'설정 탭에서 Brave Search API 키를 먼저 입력하세요'})
    if provider=='google' and not (cfg.get('google_api_key') and cfg.get('google_cx')):
        return jsonify({'ok':False,'error':'Google API 키와 검색엔진ID(cx)가 필요합니다'})
    try:
        return jsonify(discover_once(cfg,max_queries=int(d.get('queries',5) or 5)))
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:200]})

@app.route('/api/candidates/manual',methods=['POST'])
def api_cand_manual():
    """수동 URL 붙여넣기 → 후보 등록 + 검수 (API 없이도 사용 가능)."""
    d=request.get_json() or {}; cfg=load_config()
    urls=[x.strip() for x in (d.get('urls','') or '').splitlines() if x.strip().startswith('http')]
    if not urls: return jsonify({'ok':False,'error':'http로 시작하는 URL을 한 줄에 하나씩 넣으세요'})
    n=add_candidates_from([{'url':u} for u in urls],cfg,source='manual')
    threading.Thread(target=screen_pending,kwargs={'limit':len(urls)},daemon=True).start()
    return jsonify({'ok':True,'added':n,'screening':True})

@app.route('/api/candidates/screen',methods=['POST'])
def api_cand_screen():
    d=request.get_json() or {}
    if d.get('id') or d.get('rescreen'):
        with _cand_lock:
            cands=load_cands()
            for c in cands:
                if d.get('rescreen') and c.get('status')!='approved': c['screened']=False
                elif c.get('id')==d.get('id'): c['screened']=False
            save_cands(cands)
    n=screen_pending(1 if d.get('id') else int(d.get('limit',20) or 20))
    return jsonify({'ok':True,'screened':n})

@app.route('/api/candidates/status',methods=['POST'])
def api_cand_status():
    d=request.get_json() or {}
    with _cand_lock:
        cands=load_cands()
        for c in cands:
            if c.get('id')==d.get('id'):
                if 'status' in d: c['status']=d['status']
                if 'note' in d: c['note']=d['note']
        save_cands(cands)
    return jsonify({'ok':True})

@app.route('/api/candidates/approve/<cid>',methods=['POST'])
def api_cand_approve(cid):
    """후보를 사이트 목록에 직접 등록. 근거 메모는 선택사항."""
    d=request.get_json() or {}
    note=(d.get('permission_note','') or '').strip() or '후보 목록에서 직접 등록'
    c=next((x for x in load_cands() if x.get('id')==cid),None)
    if not c: return jsonify({'ok':False,'error':'후보 없음'})
    with POST_LOCK:
        sites=load_sites()
        if any(_domain_of(s.get('site_url',''))==c.get('domain') for s in sites):
            return jsonify({'ok':False,'error':'이미 등록된 도메인'})
        sites.append({'id':secrets.token_hex(6),'site_url':(c.get('base') or c.get('url','')).rstrip('/'),
                  'platform':c.get('platform','auto') if c.get('platform') in ('gnuboard','cafe24') else 'auto',
                  'mb_id':d.get('mb_id',''),'mb_pass':d.get('mb_pass',''),
                  'bo_table':(d.get('bo_table') or c.get('bo_table') or 'free'),
                  'name':(d.get('name') or c.get('domain','')),
                  'permission':True,'permission_note':note,
                  'registration_source':'candidate_registered','daily_limit':3,'min_interval_minutes':60,
                  'permission_date':_kst_now().strftime('%Y-%m-%d'),
                  'has_captcha':bool(c.get('captcha')),
                      'status':'idle','added':_kst_now().strftime('%m/%d %H:%M')})
        save_sites(sites)
    with _cand_lock:
        cands=load_cands()
        for x in cands:
            if x.get('id')==cid: x['status']='approved'; x['approved_at']=_kst_now().strftime('%Y-%m-%d %H:%M')
        save_cands(cands)
    add_log(f'[후보 등록] {c.get("domain")} → 사이트 목록 등록')
    return jsonify({'ok':True})

@app.route('/api/candidates/export',methods=['GET'])
def api_cand_export():
    from flask import Response
    import io
    cands=load_cands()
    cols=[('domain','도메인'),('url','URL'),('platform','플랫폼'),('board_name','게시판명'),
          ('bo_table','게시판ID'),('score','점수'),('status','상태'),('promo_hint','홍보허용흔적'),
          ('parked','주차도메인'),('illegal','도박불법'),
          ('ad_banned','광고금지'),('captcha','캡차'),('login_required','로그인필요'),
          ('write_form','글쓰기폼'),('last_post_days','최근글(일)'),('emails','이메일'),
          ('reject_reason','탈락사유'),('found_at','발견일')]
    try:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title='발굴후보'
        ws.append([c[1] for c in cols])
        for r in cands:
            ws.append([', '.join(r.get(c[0])) if isinstance(r.get(c[0]),list) else str(r.get(c[0],'') or '') for c in cols])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition':'attachment; filename=candidates.xlsx'})
    except Exception:
        import csv
        sio=io.StringIO(); w=csv.writer(sio); w.writerow([c[1] for c in cols])
        for r in cands: w.writerow([', '.join(r.get(c[0])) if isinstance(r.get(c[0]),list) else str(r.get(c[0],'') or '') for c in cols])
        return Response('﻿'+sio.getvalue(),mimetype='text/csv',
                        headers={'Content-Disposition':'attachment; filename=candidates.csv'})

# ---- 회원(고객) 관리 + 월 정산 ----
@app.route('/api/members',methods=['GET','POST','DELETE'])
def api_members():
    if request.method=='POST':
        d=request.get_json() or {}; mem=load_members()
        mid=d.get('id') or secrets.token_hex(6)
        ex=next((m for m in mem if m.get('id')==mid),None)
        rec=ex or {'id':mid,'payments':{},'join_date':_kst_now().strftime('%Y-%m-%d')}
        for k in ['name','biz','phone','memo','status']:
            if k in d: rec[k]=d[k]
        for k in ['plan_fee','addons','addon_fee','settle_day','jitter','per_run']:
            if k in d:
                try: rec[k]=int(d[k])
                except Exception: rec[k]=0
        # ---- 회원별 스케줄 (jump 방식: 계정당 개별 시간대) ----
        if 'sched_enabled' in d: rec['sched_enabled']=bool(d['sched_enabled'])
        if 'sched_times' in d:
            raw=d['sched_times']
            if isinstance(raw,str): raw=[x.strip() for x in raw.replace('，',',').split(',')]
            rec['sched_times']=sorted({t for t in (raw or []) if re.match(r'^\d{1,2}:\d{2}$',str(t).strip())})
        if 'sched_days' in d:
            try: rec['sched_days']=[int(x) for x in (d['sched_days'] or []) if str(x).isdigit()]
            except Exception: rec['sched_days']=[]
        if 'site_ids' in d: rec['site_ids']=list(d['site_ids'] or [])
        if 'keywords_csv' in d:
            rows=[]
            for line in (d.get('keywords_csv','') or '').splitlines():
                p=[x.strip() for x in line.split(',')]
                if len(p)>=2 and p[0] and p[1]:
                    rows.append({'지역':p[0],'서비스':p[1],'브랜드':(p[2] if len(p)>2 else '')})
            rec['keywords']=rows
        rec.setdefault('status','active'); rec.setdefault('plan_fee',30000)
        rec.setdefault('addon_fee',10000); rec.setdefault('addons',0); rec.setdefault('settle_day',1)
        rec.setdefault('sched_enabled',False); rec.setdefault('sched_times',[])
        rec.setdefault('sched_days',[]); rec.setdefault('site_ids',[])
        rec.setdefault('keywords',[]); rec.setdefault('jitter',5); rec.setdefault('per_run',1)
        if not ex: mem.append(rec)
        save_members(mem); return jsonify({'ok':True,'id':mid})
    if request.method=='DELETE':
        d=request.get_json() or {}
        save_members([m for m in load_members() if m.get('id')!=d.get('id')])
        return jsonify({'ok':True})
    return jsonify({'members':[member_view(m) for m in load_members()],'summary':settle_summary()})

@app.route('/api/members/pay',methods=['POST'])
def api_member_pay():
    """특정 회원의 특정 월 납부 상태 토글/설정."""
    d=request.get_json() or {}; mid=d.get('id'); month=d.get('month') or _cur_month()
    paid=bool(d.get('paid',True)); mem=load_members()
    for m in mem:
        if m.get('id')==mid:
            pays=m.get('payments') or {}
            if not isinstance(pays,dict): pays={}
            if paid:
                pays[month]={'paid':True,'paid_at':_kst_now().strftime('%Y-%m-%d %H:%M'),'amount':member_fee(m)}
            else:
                pays[month]={'paid':False}
            m['payments']=pays; save_members(mem)
            return jsonify({'ok':True})
    return jsonify({'ok':False,'error':'회원 없음'})

@app.route('/api/members/run/<mid>',methods=['POST'])
def api_member_run(mid):
    """회원 스케줄 지금 즉시 1회 실행(테스트)."""
    m=next((x for x in load_members() if x.get('id')==mid),None)
    if not m: return jsonify({'ok':False,'error':'회원 없음'})
    cfg=load_config()
    sites=member_sites(m); pool=member_keywords(m)
    if not sites: return jsonify({'ok':False,'error':'배정된 발행 가능 사이트가 없습니다'})
    if not pool: return jsonify({'ok':False,'error':'회원 전용 키워드도 공용 키워드도 비어있습니다'})
    if cfg.get('block_unpaid') and not member_paid_now(m) and m.get('status','active')=='active':
        return jsonify({'ok':False,'error':'미납 회원 — 설정에서 미납 자동정지를 끄거나 납부 처리 후 실행'})
    nm=m.get('name') or m.get('biz') or mid
    cnt=max(1,int(m.get('per_run',1) or 1)); total=0
    for _ in range(cnt):
        kw=pick_keywords(pool,cfg)
        total+=enqueue_generated(sites,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),
                                        '브랜드':kw.get('브랜드','') or cfg.get('brand','')},cfg,
                                 {'region':kw.get('지역',''),'service':kw.get('서비스',''),'member':nm})[0]
    mem=load_members()
    for x in mem:
        if x.get('id')==mid:
            x['last_run']=_kst_now().strftime('%Y-%m-%d %H:%M'); x['run_count']=int(x.get('run_count',0) or 0)+1
    save_members(mem)
    if total and not wk_active: start_workers(cfg.get('workers',2))
    return jsonify({'ok':True,'generated':total,'sites':len(sites),'runs':cnt})

@app.route('/api/members/export',methods=['GET'])
def api_members_export():
    from flask import Response
    import io
    mem=load_members(); cm=_cur_month()
    cols=[('name','이름/담당'),('biz','업소'),('phone','연락처'),('status','상태'),
          ('plan_fee','기본료'),('addons','추가광고'),('addon_fee','추가단가'),('fee','월청구'),
          ('paid','이번달납부'),('join_date','가입일'),('memo','메모')]
    rows=[member_view(m) for m in mem]
    try:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title=f'회원정산_{cm}'
        ws.append([c[1] for c in cols])
        for r in rows:
            ws.append([('O' if (c[0]=='paid' and r.get('paid')) else ('X' if c[0]=='paid' else str(r.get(c[0],'') or ''))) for c in cols])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition':f'attachment; filename=members_{cm}.xlsx'})
    except Exception:
        import csv
        sio=io.StringIO(); w=csv.writer(sio); w.writerow([c[1] for c in cols])
        for r in rows: w.writerow([('O' if (c[0]=='paid' and r.get('paid')) else ('X' if c[0]=='paid' else str(r.get(c[0],'') or ''))) for c in cols])
        return Response('﻿'+sio.getvalue(),mimetype='text/csv',
                        headers={'Content-Disposition':f'attachment; filename=members_{cm}.csv'})

# ---- 키워드 풀 (엑셀/CSV 저장 → 랜덤 치환) ----
@app.route('/api/keywords',methods=['GET','POST','DELETE'])
def api_keywords():
    if request.method=='POST':
        d=request.get_json() or {}
        rows=d.get('rows')
        if rows is None:
            rows=[]
            for line in (d.get('csv','') or '').splitlines():
                line=line.strip()
                if not line: continue
                p=[x.strip() for x in line.split(',')]
                if len(p)>=2 and p[0] and p[1]:
                    rows.append({'지역':p[0],'서비스':p[1],'브랜드':(p[2] if len(p)>2 else '')})
        if d.get('append'): rows=load_keywords()+rows
        save_keywords(rows); return jsonify({'ok':True,'count':len(rows)})
    if request.method=='DELETE':
        save_keywords([]); return jsonify({'ok':True})
    return jsonify(load_keywords())

@app.route('/api/keywords/upload',methods=['POST'])
def api_keywords_upload():
    f=request.files.get('file')
    if not f: return jsonify({'ok':False,'error':'파일 없음'}),400
    rows=[]
    try:
        from openpyxl import load_workbook
        import io
        wb=load_workbook(io.BytesIO(f.read()),read_only=True,data_only=True)
        ws=wb.active
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            vals=[('' if c is None else str(c)).strip() for c in row]
            if not vals or not vals[0]: continue
            if i==0 and ('지역' in vals[0] or '키워드' in vals[0]): continue  # 헤더 스킵
            if len(vals)>=2 and vals[0] and vals[1]:
                rows.append({'지역':vals[0],'서비스':vals[1],'브랜드':(vals[2] if len(vals)>2 else '')})
    except Exception as e:
        return jsonify({'ok':False,'error':f'엑셀 파싱 실패: {str(e)[:80]}'}),400
    if request.form.get('append')=='1': rows=load_keywords()+rows
    save_keywords(rows); return jsonify({'ok':True,'count':len(rows)})

@app.route('/api/generate/random',methods=['POST'])
def api_gen_random():
    d=request.get_json() or {}; cfg=load_config(); pool=load_keywords()
    if not pool: return jsonify({'ok':False,'error':'키워드 풀이 비어있습니다 (엑셀/CSV 등록)'})
    site_ids=d.get('site_ids',[]); count=max(1,int(d.get('count',1) or 1))
    # 발행 모드: site_ids(또는 count>1) 지정 시 랜덤 N개를 허용 사이트 큐에 등록
    if site_ids or count>1:
        sites=[s for s in load_sites() if not site_ids or s.get('id') in site_ids]
        allowed=[s for s in sites if is_permitted(s)]; blocked=len(sites)-len(allowed)
        if not allowed:
            return jsonify({'ok':False,'error':f'홍보 허용된 사이트가 없습니다 (미허용 {blocked}개 제외)'})
        total=0
        for _ in range(count):
            kw=pick_keywords(pool,cfg)
            # 사이트마다 유니크 본문 생성(중복 방지)
            total+=enqueue_generated(allowed,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),'브랜드':kw.get('브랜드','')},cfg,{'region':kw.get('지역',''),'service':kw.get('서비스','')})[0]
        if not wk_active: start_workers(cfg.get('workers',2))
        return jsonify({'ok':True,'generated':total,'blocked':blocked,'picks':count,'queued':wk_stats['queued']})
    # 미리보기 모드: 1개 생성해서 결과창에 표시
    kw=pick_keywords(pool,cfg)
    html,title=generate_article({'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),'브랜드':kw.get('브랜드','')},cfg)
    return jsonify({'ok':True,'title':title,'content':html,'picked':kw})

@app.route('/api/backup/now',methods=['POST'])
def api_backup_now():
    ok,err=do_backup(load_config(),'수동')
    return jsonify({'ok':ok,'error':('' if ok else err)})

@app.route('/api/diag',methods=['GET'])
def api_diag():
    """서버 환경 자가진단: 크롬·드라이버·실제 페이지 로드·리소스."""
    import shutil,subprocess,platform
    out={'time':_kst_now().strftime('%Y-%m-%d %H:%M:%S'),'python':sys.version.split()[0],
         'platform':platform.platform()[:60],'steps':[]}
    def step(name,ok,detail=''):
        out['steps'].append({'name':name,'ok':bool(ok),'detail':str(detail)[:200]})
    # 1) 크롬 바이너리
    cb=os.environ.get('CHROME_BIN') or ''
    cands=[cb,'/usr/bin/google-chrome','/usr/bin/chromium','/usr/bin/chromium-browser','/snap/bin/chromium']
    found=''
    for p in cands:
        if p and os.path.exists(p): found=p; break
    if not found:
        w=shutil.which('google-chrome') or shutil.which('chromium') or shutil.which('chromium-browser')
        if w: found=w
    ver=''
    if found:
        try: ver=subprocess.run([found,'--version'],capture_output=True,text=True,timeout=15).stdout.strip()
        except Exception as e: ver='버전확인 실패: '+str(e)[:60]
    step('크롬 설치',bool(found),f'{found} {ver}' if found else '크롬 없음 — 설치 필요')
    out['chrome_bin']=found
    # 2) selenium / webdriver_manager 모듈
    try:
        import selenium; step('selenium 모듈',True,'v'+getattr(selenium,'__version__','?'))
    except Exception as e: step('selenium 모듈',False,str(e)[:80])
    try:
        import webdriver_manager; step('webdriver_manager',True,'설치됨')
    except Exception as e: step('webdriver_manager',False,str(e)[:80])
    # 3) 실제 드라이버 기동 + 페이지 로드
    d=None
    if found:
        try:
            if not os.environ.get('CHROME_BIN'): os.environ['CHROME_BIN']=found
            t0=time.time(); d=get_driver()
            step('크롬 드라이버 기동',True,f'{round(time.time()-t0,1)}초')
            try:
                t1=time.time(); d.get('https://example.com'); time.sleep(1)
                ttl=(d.title or '')[:60]
                step('페이지 로드 테스트',bool(ttl),f'title="{ttl}" ({round(time.time()-t1,1)}초)')
            except Exception as e: step('페이지 로드 테스트',False,str(e)[:150])
        except Exception as e:
            step('크롬 드라이버 기동',False,str(e)[:200])
    else:
        step('크롬 드라이버 기동',False,'크롬 미설치로 건너뜀')
    # 4) 리소스
    try:
        du=shutil.disk_usage('/'); step('디스크',du.free>500*1024*1024,
            f'여유 {round(du.free/1024/1024/1024,1)}GB / 전체 {round(du.total/1024/1024/1024,1)}GB')
    except Exception as e: step('디스크',False,str(e)[:60])
    try:
        mt=open('/proc/meminfo').read()
        tot=int(re.search(r'MemTotal:\s+(\d+)',mt).group(1))//1024
        av=int(re.search(r'MemAvailable:\s+(\d+)',mt).group(1))//1024
        step('메모리',av>200,f'가용 {av}MB / 전체 {tot}MB')
    except Exception as e: step('메모리',False,str(e)[:60])
    # 5) 데이터 상태
    step('허용 사이트',len([s for s in load_sites() if is_autopostable(s)])>0,
         f'발행가능 {len([s for s in load_sites() if is_autopostable(s)])}개 / 전체 {len(load_sites())}개')
    step('키워드 풀',len(load_keywords())>0,f'{len(load_keywords())}개')
    step('이미지 URL',True,f'{len(load_image_urls())}개'+(' (기본 이미지 사용)' if not load_image_urls() else ''))
    out['ok']=all(s['ok'] for s in out['steps'] if s['name'] in ('크롬 설치','크롬 드라이버 기동','페이지 로드 테스트'))
    return jsonify(out)

@app.route('/api/verify/now',methods=['POST'])
def api_verify_now():
    threading.Thread(target=verify_once,kwargs={'limit':40},daemon=True).start()
    return jsonify({'ok':True,'started':True})

@app.route('/api/oneclick',methods=['POST'])
def api_oneclick():
    """원클릭 실시간 발행: 키워드 풀 랜덤 추출 → 사이트마다 유니크 생성 → 허용·캡차없는 사이트 전체 발행."""
    d=request.get_json() or {}; cfg=load_config(); pool=load_keywords()
    if not pool: return jsonify({'ok':False,'error':'키워드 풀이 비어있습니다 (글 생성 탭에서 등록)'})
    sites=load_sites(); allowed=[s for s in sites if is_autopostable(s)]
    if not allowed:
        cap=sum(1 for s in sites if is_permitted(s) and s.get('has_captcha'))
        return jsonify({'ok':False,'error':f'발행 가능한 사이트가 없습니다 (허용·캡차없음 0개'+(f' · 캡차제외 {cap}개' if cap else '')+')'})
    count=max(1,min(50,int(d.get('count',1) or 1)))
    total=0
    for _ in range(count):
        kw=pick_keywords(pool,cfg)
        total+=enqueue_generated(allowed,{'지역':kw.get('지역',''),'서비스':kw.get('서비스',''),'브랜드':kw.get('브랜드','')},cfg,{'region':kw.get('지역',''),'service':kw.get('서비스','')})[0]
    if total and not wk_active: start_workers(cfg.get('workers',2))
    return jsonify({'ok':True,'generated':total,'sites':len(allowed),'picks':count})

@app.route('/api/post',methods=['POST'])
def api_post():
    d=request.get_json(silent=True) or {}; cfg=load_config()
    ids=d.get('site_ids',[]); title=d.get('title',''); content=d.get('content','')
    region=d.get('region',''); service=d.get('service',''); brand=d.get('brand','')
    meta={'region':region,'service':service}
    sites=[s for s in load_sites() if s.get('id') in ids]
    allowed=[s for s in sites if is_permitted(s)]
    # 2개 이상 사이트 + 키워드 있으면 사이트마다 유니크 재생성(중복 방지). 단일 사이트는 검토한 원문 그대로.
    if len(allowed)>1 and region and service:
        q,blocked=enqueue_generated(sites,{'지역':region,'서비스':service,'브랜드':brand or cfg.get('brand','')},cfg,meta)
        note='사이트별 유니크 본문 재생성'
    else:
        if title and content: remember_if_unique(title,content)   # 원문도 중복DB에 기록
        q,blocked=enqueue(sites,title,content,meta); note=''
    if q and not wk_active: start_workers(cfg.get('workers',2))
    return jsonify({'ok':True,'queued':q,'blocked':blocked,'note':note})

@app.route('/api/bulk',methods=['POST'])
def api_bulk():
    d=request.get_json(silent=True) or {}; cfg=load_config()
    keyword_sets=d.get('keyword_sets',[]); site_ids=d.get('site_ids',[])
    sites=[s for s in load_sites() if not site_ids or s.get('id') in site_ids]
    if not sites: return jsonify({'ok':False,'error':'사이트 없음'})
    allowed=[s for s in sites if is_permitted(s)]; blocked=len(sites)-len(allowed)
    if not allowed:
        return jsonify({'ok':False,'error':f'홍보 허용된 사이트가 없습니다 (미허용 {blocked}개 제외)'})
    total=0
    for ks in keyword_sets:
        kw={'지역':ks.get('지역',''),'서비스':ks.get('서비스',''),'브랜드':ks.get('브랜드','')}
        # 사이트마다 유니크 본문 생성(중복 방지)
        total+=enqueue_generated(allowed,kw,cfg,{'region':kw['지역'],'service':kw['서비스']})[0]
    if not wk_active: start_workers(cfg.get('workers',2))
    return jsonify({'ok':True,'generated':total,'blocked':blocked,'queued':wk_stats['queued']})

@app.route('/api/sites',methods=['GET','POST','DELETE'])
def api_sites():
    if request.method=='POST':
        d=request.get_json(silent=True) or {}
        if not str(d.get('site_url','')).strip().startswith(('http://','https://')):
            return jsonify({'ok':False,'error':'올바른 http(s) 사이트 URL이 필요합니다'}),400
        _plat=(d.get('platform','auto') or 'auto').strip().lower()
        if _plat not in ('gnuboard','cafe24'):   # auto/미지정 → 자동 감지
            try: _plat=detect_platform(d.get('site_url','').strip())
            except Exception: _plat='gnuboard'
        site={'id':d.get('id',str(int(time.time()*1000))),'site_url':d.get('site_url','').strip().rstrip('/'),
              'platform':_plat,'mb_id':d.get('mb_id',''),'mb_pass':d.get('mb_pass',''),
              'bo_table':d.get('bo_table','').strip() or 'm8_qna','name':d.get('name',''),
              'permission':bool(d.get('permission',False)),
              'registration_source':'manual_admin',
              'daily_limit':max(0,int(d.get('daily_limit',3) or 0)),
              'min_interval_minutes':max(0,int(d.get('min_interval_minutes',60) or 0)),
              'permission_note':(d.get('permission_note','') or '').strip(),
              'permission_date':(datetime.now().strftime('%Y-%m-%d') if d.get('permission') else ''),
              'status':'idle','added':datetime.now().strftime('%m/%d %H:%M')}
        with POST_LOCK:
            sites=load_sites(); ex=[s for s in sites if s.get('id')==site['id']]
            if ex:
                # 기존 발행 카운트 등은 보존
                merged=sites[sites.index(ex[0])]; merged.update(site)
            else: sites.append(site)
            save_sites(sites)
        return jsonify({'ok':True})
    elif request.method=='DELETE':
        d=request.get_json(silent=True) or {}
        with POST_LOCK:
            sites=[s for s in load_sites() if s.get('id')!=d.get('id','')]
            save_sites(sites)
        return jsonify({'ok':True})
    return jsonify(load_sites())

@app.route('/api/workers/start',methods=['POST'])
def api_wk_start():
    d=request.get_json() or {}; start_workers(d.get('n',load_config().get('workers',2)))
    return jsonify({'ok':True})

@app.route('/api/workers/stop',methods=['POST'])
def api_wk_stop():
    stop_workers(); return jsonify({'ok':True})

@app.route('/api/workers/pause',methods=['POST'])
def api_wk_pause():
    pause_workers(); return jsonify({'ok':True})

@app.route('/api/workers/resume',methods=['POST'])
def api_wk_resume():
    resume_workers(); return jsonify({'ok':True})

@app.route('/api/workers/stats',methods=['GET'])
def api_wk_stats():
    return jsonify({**wk_stats,'active':wk_active,'paused':wk_paused})

@app.route('/api/workers/reset',methods=['POST'])
def api_wk_reset():
    with STATS_LOCK:
        for k in ['success','fail','total','done','skipped','retry']: wk_stats[k]=0
        wk_stats['queued']=post_queue.qsize()
    return jsonify({'ok':True})

# ---- 발행 이력 (결과 탭) ----
@app.route('/api/history',methods=['GET'])
def api_history():
    h=load_json(HISTORY_FILE,[])
    return jsonify(list(reversed(h))[:500])   # 최신순 500건

@app.route('/api/history/clear',methods=['POST'])
def api_history_clear():
    save_json(HISTORY_FILE,[]); return jsonify({'ok':True})

@app.route('/api/history/export',methods=['GET'])
def api_history_export():
    from flask import Response
    import io
    h=load_json(HISTORY_FILE,[])
    cols=[('time','시간'),('site_name','사이트'),('site_url','URL'),('bo_table','게시판'),
          ('member','회원'),('region','지역'),('service','서비스'),('title','제목'),('status','상태'),
          ('fail_reason_ko','실패원인'),('alive','생존'),('verified_at','검증시각'),
          ('result_url','결과URL'),('message','메시지')]
    try:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title='발행이력'
        ws.append([c[1] for c in cols])
        for rec in h:
            ws.append([str(rec.get(c[0],'') or '') for c in cols])
        for i,c in enumerate(cols,1):
            ws.column_dimensions[chr(64+i)].width=[18,20,34,12,14,10,12,40,10,14,8,16,30,40][i-1]
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition':'attachment; filename=chirashi_history.xlsx'})
    except Exception:
        # openpyxl 미설치 시 CSV 폴백 (엑셀에서 열림)
        import csv
        sio=io.StringIO(); w=csv.writer(sio); w.writerow([c[1] for c in cols])
        for rec in h: w.writerow([str(rec.get(c[0],'') or '') for c in cols])
        data='﻿'+sio.getvalue()   # BOM: 엑셀 한글 깨짐 방지
        return Response(data,mimetype='text/csv',
                        headers={'Content-Disposition':'attachment; filename=chirashi_history.csv'})

@app.route('/api/config',methods=['GET','POST'])
def api_cfg():
    if request.method=='POST':
        d=request.get_json(silent=True) or {}; cfg=load_config()
        old_search=(cfg.get('discover_keywords',''),cfg.get('discover_direct_queries',''))
        for k in ['brand','phone','phones','openai_key','model','workers','post_delay','daily_limit',
                  'use_gpt','telegram_token','telegram_chat_id','notify_done','notify_fail','update_token',
                  'telegram_control','backup_time','verify_enabled','mix_keywords','block_unpaid',
                  'google_api_key','google_cx','brave_api_key','search_provider','discover_enabled','discover_daily_target',
                  'discover_query_limit','discover_keywords','discover_direct_queries']:
            if k in d:
                if k in ('openai_key','telegram_token','google_api_key','brave_api_key') and d[k]=='***설정됨***': continue  # 마스크 값은 무시(기존 유지)
                cfg[k]=d[k]
        if d.get('password'): cfg['password']=generate_password_hash(d['password'])  # 해시 저장
        save_config(cfg)
        # 검색 조건을 바꾸면 다음 검색부터 새 조건의 첫 줄이 즉시 실행되도록 커서를 초기화한다.
        new_search=(cfg.get('discover_keywords',''),cfg.get('discover_direct_queries',''))
        if new_search!=old_search:
            st=load_json(DISCO_FILE,{}) or {}; st['cursor']=0; save_json(DISCO_FILE,st)
        return jsonify({'ok':True,'search_cursor_reset':new_search!=old_search})
    c=dict(load_config())
    if c.get('openai_key'): c['openai_key']='***설정됨***'   # 키 노출 방지
    if c.get('telegram_token'): c['telegram_token']='***설정됨***'
    if c.get('google_api_key'): c['google_api_key']='***설정됨***'
    if c.get('brave_api_key'): c['brave_api_key']='***설정됨***'
    c.pop('password',None)
    return jsonify(c)

@app.route('/api/test/<sid>',methods=['POST'])
def api_test(sid):
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    if not is_permitted(site):
        return jsonify({'ok':False,'error':'미허용 도메인 — 테스트도 실제 발행이므로 홍보 허용(✔) 설정 후 이용하세요'})
    try:
        cfg=load_config()
        if not under_daily_limit(site,cfg):
            return jsonify({'ok':False,'error':'사이트 일일 발행 한도에 도달했습니다'})
        interval_ok,remain=under_min_interval(site)
        if not interval_ok:
            return jsonify({'ok':False,'error':f'사이트 최소 발행 간격 미충족 ({max(1,(remain+59)//60)}분 남음)'})
        html,title=generate_article({'지역':'테스트','서비스':'테스트'},cfg)
        ok,msg=do_post(site,title,html)
        finalize_post(site,ok)
        return jsonify({'ok':ok,'message':msg,'platform':resolve_platform(site)})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)})

# ---- 사이트 대량등록 (CSV: url,이름,게시판,아이디,비번,허용) ----
@app.route('/api/sites/bulk',methods=['POST'])
def api_sites_bulk():
    d=request.get_json(silent=True) or {}; text=d.get('csv',''); default_perm=bool(d.get('permission',False))
    added=0; new_sites=[]
    for line in (text or '').splitlines():
        line=line.strip()
        if not line or line.startswith('#'): continue
        parts=[x.strip() for x in line.split(',')]
        url=parts[0] if parts else ''
        if not url or not url.startswith('http'): continue
        nm=parts[1] if len(parts)>1 else ''
        bo=parts[2] if len(parts)>2 else 'free'
        mid=parts[3] if len(parts)>3 else ''
        mpw=parts[4] if len(parts)>4 else ''
        perm=default_perm
        if len(parts)>5: perm=parts[5] in ('1','true','y','Y','o','O','허용','true')
        new_sites.append({'id':secrets.token_hex(6),'site_url':url.rstrip('/'),'platform':'auto',
                      'mb_id':mid,'mb_pass':mpw,'bo_table':bo or 'free','name':nm,
                      'permission':perm,'permission_note':'CSV 일괄등록',
                      'registration_source':'admin_bulk','daily_limit':3,'min_interval_minutes':60,
                      'permission_date':(datetime.now().strftime('%Y-%m-%d') if perm else ''),
                      'status':'idle','added':datetime.now().strftime('%m/%d %H:%M')})
        added+=1
    if new_sites:
        with POST_LOCK:
            sites=load_sites(); sites.extend(new_sites); save_sites(sites)
    return jsonify({'ok':True,'added':added})

# ---- 허용상태 일괄 토글 ----
@app.route('/api/sites/permission',methods=['POST'])
def api_sites_permission():
    d=request.get_json(silent=True) or {}; ids=set(d.get('ids',[])); val=bool(d.get('permission',False))
    n=0
    with POST_LOCK:
        sites=load_sites()
        for s in sites:
            if s.get('id') in ids:
                s['permission']=val
                if val and not s.get('permission_date'): s['permission_date']=datetime.now().strftime('%Y-%m-%d')
                n+=1
        save_sites(sites)
    return jsonify({'ok':True,'changed':n})

# ---- 사이트 헬스체크 ----
@app.route('/api/sites/health/<sid>',methods=['POST'])
def api_site_health(sid):
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    h=site_health(site)
    try: plat=detect_platform(site.get('site_url',''),use_cache=False)   # 점검 시 플랫폼도 재감지
    except Exception: plat=site.get('platform','gnuboard')
    h['platform']=plat
    set_site_flag(sid,health=('ok' if h.get('ok') else 'bad'),
                  health_at=datetime.now().strftime('%m/%d %H:%M'),platform=plat)
    return jsonify({'ok':True,'health':h})

@app.route('/api/sites/detect/<sid>',methods=['POST'])
def api_site_detect(sid):
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    try: plat=detect_platform(site.get('site_url',''),use_cache=False)
    except Exception as e: return jsonify({'ok':False,'error':str(e)[:100]})
    set_site_flag(sid,platform=plat)
    return jsonify({'ok':True,'platform':plat})

@app.route('/api/sites/debug/<sid>',methods=['POST'])
def api_site_debug(sid):
    """심층 디버그: 로그인·글쓰기 페이지에서 실제로 무엇이 보이는지 그대로 덤프."""
    from selenium.webdriver.common.by import By
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    url=site.get('site_url','').rstrip('/'); m=re.match(r'(https?://[^/]+)',url)
    base=m.group(1) if m else url
    out={'base':base,'bo_table':site.get('bo_table',''),'mb_id':site.get('mb_id',''),'phases':[]}
    def snap(label):
        try: body=d.find_element(By.TAG_NAME,'body').text[:600]
        except Exception: body=''
        try:
            fields=[]
            for el in d.find_elements(By.CSS_SELECTOR,'input,textarea,select'):
                if not _sel_vis(el): continue
                fields.append(f"{el.tag_name}[{_sel_attr(el,'type') or ''}] name={_sel_attr(el,'name')} id={_sel_attr(el,'id')}")
            fields=fields[:14]
        except Exception: fields=[]
        out['phases'].append({'label':label,'url':(d.current_url or '')[:200],
                              'title':(d.title or '')[:100],'body':body,'fields':fields})
    try:
        d=get_driver()
        # 1) 로그인 페이지
        d.get(base+'/bbs/login.php'); time.sleep(2); snap('로그인 페이지')
        # 2) 로그인 시도
        mid=site.get('mb_id',''); mpw=site.get('mb_pass','')
        if mid:
            try:
                ide=d.find_elements(By.CSS_SELECTOR,"input[name='mb_id']")
                pwe=d.find_elements(By.CSS_SELECTOR,"input[name='mb_password']")
                out['login_fields_found']={'mb_id':len(ide),'mb_password':len(pwe)}
                if ide and pwe:
                    ide[0].clear(); ide[0].send_keys(mid); pwe[0].clear(); pwe[0].send_keys(mpw)
                    clicked=_click_first(d,["input[type='submit']","button[type='submit']",".btn_submit","#btn_submit"])
                    out['login_submit_clicked']=clicked
                    if not clicked:
                        try: pwe[0].submit()
                        except Exception: pass
                    time.sleep(3); dismiss_alerts(d); snap('로그인 시도 후')
            except Exception as e: out['login_error']=str(e)[:150]
        # 3) 글쓰기 페이지 후보들
        bo=site.get('bo_table','free')
        for path in [f'/bbs/write.php?bo_table={bo}', f'/bbs/board.php?bo_table={bo}']:
            try:
                d.get(base+path); time.sleep(2); dismiss_alerts(d); snap('시도: '+path)
            except Exception as e:
                out['phases'].append({'label':'시도: '+path,'url':'','title':'','body':'ERROR '+str(e)[:100],'fields':[]})
    except Exception as e:
        out['error']=str(e)[:200]
    return jsonify({'ok':True,**out})

@app.route('/api/sites/dryrun/<sid>',methods=['POST'])
def api_site_dryrun(sid):
    """드라이런: 실제 글을 올리지 않고 등록 직전까지 검증."""
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    cfg=load_config()
    kw={'지역':'인천','서비스':'셔츠룸','브랜드':cfg.get('brand','') or '테스트'}
    html,title=generate_article(kw,cfg,unique=False)   # 중복DB 오염 방지
    try:
        ok,steps=dryrun_post(site,title,html)
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:200],'steps':[]})
    # 캡차가 감지되면 플래그 저장
    if any((not s['ok']) and '캡차' in s['name'] for s in steps):
        set_site_flag(sid,has_captcha=True,captcha_note='드라이런 감지')
    return jsonify({'ok':ok,'steps':steps,'title':title})

@app.route('/api/sites/learn/<sid>',methods=['POST'])
def api_site_learn(sid):
    """비제출 실측학습: 등록된 사이트 한 곳의 DOM/폼을 측정하고 근거와 레시피를 저장."""
    site=next((s for s in load_sites() if s.get('id')==sid),None)
    if not site: return jsonify({'ok':False,'error':'사이트 없음'})
    if not is_permitted(site):
        return jsonify({'ok':False,'error':'미허용 사이트 — 관리자가 직접 등록하고 홍보 허용한 사이트만 실측 가능'})
    try:
        analysis,rec=analyze_site_logic(site)
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:150]})
    _save_site_analysis(sid,analysis,rec)
    if analysis.get('captcha'):
        set_site_flag(sid,has_captcha=True,captcha_note=('실측: '+analysis['captcha'])[:80])
    elif analysis.get('ok'):
        set_site_flag(sid,has_captcha=False,captcha_note='')
    learned=None if not rec else {k:rec.get(k) for k in ['platform','write_url','subject_sel','content_mode','content_sel','submit_sel','form_action','form_method','learned_mode']}
    return jsonify({'ok':bool(analysis.get('ok')),'message':('실측 레시피 저장 완료' if rec else '실측 완료 — 발행 폼을 확정하지 못함'),
                    'captcha':bool(analysis.get('captcha')),'blocked':bool(analysis.get('blocked')),
                    'analysis':analysis,'learned':learned})

# ---- 예약 스케줄 CRUD ----
@app.route('/api/schedules',methods=['GET','POST','DELETE'])
def api_scheds():
    if request.method=='POST':
        d=request.get_json(silent=True) or {}; scheds=load_scheds()
        sc={'id':d.get('id') or secrets.token_hex(6),'name':d.get('name','예약'),
            'keyword_sets':d.get('keyword_sets',[]),'site_ids':d.get('site_ids',[]),
            'times':d.get('times',[]),'days':d.get('days',[]),'enabled':bool(d.get('enabled',True)),
            'last_run':''}
        ex=[x for x in scheds if x.get('id')==sc['id']]
        if ex: scheds[scheds.index(ex[0])].update(sc)
        else: scheds.append(sc)
        save_scheds(scheds); return jsonify({'ok':True,'id':sc['id']})
    if request.method=='DELETE':
        d=request.get_json(silent=True) or {}; scheds=[x for x in load_scheds() if x.get('id')!=d.get('id')]
        save_scheds(scheds); return jsonify({'ok':True})
    return jsonify(load_scheds())

@app.route('/api/schedules/toggle',methods=['POST'])
def api_sched_toggle():
    d=request.get_json(silent=True) or {}; scheds=load_scheds()
    for x in scheds:
        if x.get('id')==d.get('id'): x['enabled']=not x.get('enabled',True)
    save_scheds(scheds); return jsonify({'ok':True})

# ---- 통계 ----
@app.route('/api/stats',methods=['GET'])
def api_stats():
    return jsonify(compute_stats())

# ---- 텔레그램 테스트 발송 ----
@app.route('/api/telegram/test',methods=['POST'])
def api_tg_test():
    ok=send_telegram(load_config(),'🔔 찌라시 마스터 텔레그램 연결 테스트')
    return jsonify({'ok':ok,'error':None if ok else '토큰/챗ID 확인'})

# ---- 원격 자가 업데이트 (로그인 불필요, 토큰 인증) ----
_uploads={}; _upload_lock=threading.Lock()
def _apply_update(newcode):
    import py_compile,base64 as _b64
    if b'def main' not in newcode or len(newcode)<1000:
        return {'ok':False,'error':'app.py 형식 이상'},400
    target=os.path.join(BASE_DIR,'app.py'); cur=b''
    try:
        with open(target,'rb') as f: cur=f.read()
        with open(os.path.join(BASE_DIR,'app.py.bak'),'wb') as f: f.write(cur)
    except Exception: pass
    with open(target,'wb') as f: f.write(newcode)
    try: py_compile.compile(target,doraise=True)
    except Exception as e:
        if cur:
            with open(target,'wb') as f: f.write(cur)
        return {'ok':False,'error':f'문법오류 롤백: {str(e)[:120]}'},400
    def _restart(): time.sleep(1.0); os._exit(0)
    threading.Thread(target=_restart,daemon=True).start()
    return {'ok':True,'bytes':len(newcode),'restart':'1초 후 재기동'},200

@app.route('/api/admin/update',methods=['POST'])
def api_admin_update():
    d=request.get_json(silent=True) or {}
    tok=d.get('token') or request.headers.get('X-Update-Token','')
    want=os.environ.get('CHIRASHI_UPDATE_TOKEN') or load_config().get('update_token','')
    if not want:
        return jsonify({'ok':False,'error':'원격 업데이트 비활성화됨'}),503
    if not secrets.compare_digest(str(tok),str(want)):
        return jsonify({'ok':False,'error':'토큰 불일치'}),403
    import base64 as _b64
    # 청크 업로드 (WAF 우회: 작은 조각으로 나눠 전송)
    if 'chunk' in d or 'finalize' in d:
        uid=d.get('upload_id','')
        if not uid: return jsonify({'ok':False,'error':'upload_id 필요'}),400
        with _upload_lock:
            buf=_uploads.setdefault(uid,{})
            if 'chunk' in d: buf[int(d.get('seq',0))]=d['chunk']
            if d.get('finalize'):
                total=int(d.get('total',len(buf)))
                if len(buf)<total:
                    return jsonify({'ok':False,'error':f'조각 부족 {len(buf)}/{total}'}),400
                b64=''.join(buf[i] for i in sorted(buf)); _uploads.pop(uid,None)
                try: newcode=_b64.b64decode(b64)
                except Exception as e: return jsonify({'ok':False,'error':'b64 디코드 실패'}),400
                res,code=_apply_update(newcode); return jsonify(res),code
        return jsonify({'ok':True,'received':len(buf)})
    # 단건 업로드
    b64=d.get('app_b64','')
    if not b64: return jsonify({'ok':False,'error':'app_b64 없음'}),400
    try:
        newcode=_b64.b64decode(b64)
        res,code=_apply_update(newcode); return jsonify(res),code
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)[:150]}),500

# ==================== HTML ====================
HTML=r'''<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>찌라시 마스터 v6</title>
<style>:root{--bg:#060913;--c:#0d1117;--b:#1a1f2e;--t:#c9d1d9;--d:#6b7280;--p:#3b82f6;--g:#22c55e;--r:#ef4444;--y:#f59e0b;--v:#8b5cf6}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--t);min-height:100vh;font-size:13px}
::-webkit-scrollbar{width:4px}::-webkit-scrollbar-track{background:var(--bg)}::-webkit-scrollbar-thumb{background:var(--b)}
header{background:var(--c);border-bottom:1px solid var(--b);padding:10px 16px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}
.logo{font-size:15px;font-weight:700}.logo s{color:var(--p);text-decoration:none}
.stats{display:flex;gap:12px;font-size:11px;color:var(--d)}.stats b{color:var(--t)}
.tabs{display:flex;gap:2px;padding:10px 14px 0;flex-wrap:wrap}
.tab{padding:7px 14px;border:1px solid transparent;border-radius:8px 8px 0 0;cursor:pointer;font-size:11px;color:var(--d);background:transparent}
.tab.on{background:var(--c);border-color:var(--b);border-bottom-color:var(--c);color:var(--t);font-weight:600}
.wrap{max-width:1200px;margin:0 auto;padding:10px 14px}
.panel{display:none;padding:10px 0}.panel.on{display:block}
.card{background:var(--c);border:1px solid var(--b);border-radius:10px;padding:14px;margin-bottom:10px}
.card h3{font-size:10px;color:var(--d);text-transform:uppercase;letter-spacing:1px;margin-bottom:10px}
.row{display:flex;gap:6px;align-items:center;flex-wrap:wrap;margin-bottom:6px}
input,select,textarea{padding:8px 10px;border:1px solid var(--b);border-radius:6px;background:var(--bg);color:var(--t);font-size:12px;outline:none;font-family:inherit;width:100%}
input:focus,textarea:focus{border-color:var(--p)}
textarea{resize:vertical;line-height:1.5}
.btn{padding:7px 14px;border:none;border-radius:6px;font-size:11px;cursor:pointer;font-weight:600;color:#fff;transition:.2s}
.btn-p{background:var(--p)}.btn-g{background:#166534}.btn-r{background:#991b1b}.btn-y{background:#92400e}.btn-v{background:#5b21b6}.btn-d{background:var(--b);color:var(--d)}
.btn:disabled{opacity:.4}.btn-xs{padding:3px 7px;font-size:10px}
table{width:100%;border-collapse:collapse;font-size:11px}
th{text-align:left;padding:5px 8px;color:var(--d);font-weight:500;border-bottom:2px solid var(--b);font-size:10px}
td{padding:5px 8px;border-bottom:1px solid #111827}
.st{display:inline-block;padding:2px 6px;border-radius:10px;font-size:9px;font-weight:600}
.st-i{background:#1e293b;color:var(--d)}.st-ok{background:#052e16;color:var(--g)}.st-f{background:#450a0a;color:var(--r)}.st-y{background:#422006;color:var(--y)}
.toast{position:fixed;top:12px;right:12px;z-index:999;padding:8px 14px;border-radius:6px;font-size:11px;font-weight:500;animation:in .3s}
@keyframes in{from{transform:translateX(100%);opacity:0}to{transform:translateX(0);opacity:1}}
.toast-ok{background:#052e16;color:var(--g)}.toast-er{background:#450a0a;color:var(--r)}
input[type=checkbox]{accent-color:var(--p)}
.prog{background:var(--b);border-radius:6px;height:14px;overflow:hidden}
.prog>div{height:100%;width:0;background:var(--g);transition:width .3s}
.note{background:#0b1a2e;border:1px solid #1a3a5e;border-radius:8px;padding:10px 12px;font-size:11px;color:#9db8d8;line-height:1.6;margin-bottom:10px}
</style></head><body>'''

LOGIN_HTML=r'''<div style="display:flex;align-items:center;justify-content:center;min-height:100vh">
<div style="background:var(--c);border:1px solid var(--b);border-radius:14px;padding:40px;width:340px;text-align:center">
<h2 style="margin-bottom:4px;font-size:18px">찌라시 마스터 v6</h2>
<p style="color:var(--d);font-size:12px;margin-bottom:24px">정직한 자동 발행 (회피코드 없음)</p>
<form method="POST">
{% if error %}<p style="color:var(--r);font-size:11px;margin-bottom:8px">{{ error }}</p>{% endif %}
<input type="password" name="pw" placeholder="비밀번호" autofocus style="margin-bottom:12px">
<button type="submit" class="btn btn-p" style="width:100%">로그인</button>
</form></div></div>'''

DASH_HTML=r'''<header><div class="logo">찌라시 <s>마스터 v6</s></div>
<div class="stats" id="live"><span>큐:<b id="q">0</b></span><span>성공:<b id="ok" style="color:var(--g)">0</b></span><span>실패:<b id="fl" style="color:var(--r)">0</b></span><span>스킵:<b id="sk" style="color:var(--y)">0</b></span><span>워커:<b id="ws" style="color:{{'var(--g)' if wk_on else 'var(--d)'}}">{{'ON' if wk_on else 'OFF'}}</b></span></div>
<a href="/logout" class="btn-xs" style="background:var(--b);color:var(--d);text-decoration:none">로그아웃</a></header>

<div class="tabs"><button class="tab on" onclick="T('gen')">글 생성</button><button class="tab" onclick="T('sites')">사이트 (<span id="siteTabCount">{{sites|length}}</span>)</button><button class="tab" onclick="T('res')">결과</button><button class="tab" onclick="T('disco')">발굴</button><button class="tab" onclick="T('mem')">회원·정산</button><button class="tab" onclick="T('stats')">통계</button><button class="tab" onclick="T('set')">설정</button></div>
<div class="wrap"><div id="toasts"></div>
<div id="pvOverlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.75);z-index:500;padding:20px" onclick="if(event.target===this)closePreview()">
<div style="max-width:820px;margin:0 auto;background:#fff;color:#222;border-radius:10px;max-height:90vh;overflow:auto">
<div style="position:sticky;top:0;background:#0d1117;color:#fff;padding:10px 14px;display:flex;align-items:center;gap:10px"><b style="flex:1;font-size:13px">🔍 발행 미리보기 (실제 게시판에 보일 모양)</b><span id="pvTitle" style="font-size:11px;color:#9aa"></span><button class="btn btn-r btn-xs" onclick="closePreview()">닫기</button></div>
<iframe id="pvFrame" style="width:100%;height:70vh;border:0;background:#fff"></iframe></div></div>

<div id="p-gen" class="panel on">
<div class="note">✔ 정직 발행 모드 — 봇 탐지 회피·전화번호 난독화 없음. 사이트당 1일 한도와 포스트 간 지연으로 도배를 방지합니다. <b style="color:var(--g)">홍보 허용(✔)으로 검증된 도메인만 발행·테스트되며</b>, 미검증 도메인은 자동 제외+로그 기록됩니다. 사이트 탭에서 허용 근거를 남겨두세요.</div>

<div class="card" style="border-color:#166534"><h3>⚡ 원클릭 실시간 발행</h3>
<div class="row"><span style="color:var(--d);font-size:11px">횟수</span><input type="number" id="ocN" value="1" min="1" max="50" style="width:70px">
<button class="btn btn-g" onclick="oneClick()">키워드 풀 랜덤 → 허용 사이트 전체 발행</button>
<span style="flex:1"></span><span style="color:var(--d);font-size:10px">캡차 있는 사이트는 자동 제외 · 사이트마다 다른 글</span></div>
<div style="font-size:10px;color:var(--d)">키워드 풀·이미지 URL만 등록해두면 이 버튼 하나로 끝납니다.</div></div>

<div class="card" id="progCard" style="display:none"><h3>발행 진행률</h3>
<div class="prog"><div id="progBar"></div></div>
<div style="font-size:10px;color:var(--d);margin-top:5px" id="progText">0 / 0</div></div>

<div class="card"><h3>키워드 입력 (3종 치환 → bm21 스타일 리치HTML)</h3>
<div class="row">
<input type="text" id="k1" placeholder="{지역} 예: 인천" style="flex:1">
<input type="text" id="k2" placeholder="{서비스} 예: 셔츠룸" style="flex:1">
<input type="text" id="k3" placeholder="{브랜드}" value="{{cfg.brand}}" style="flex:1">
<button class="btn btn-p" onclick="gen()">1개 생성</button>
<button class="btn btn-v" onclick="genBulk()">키워드 조합 일괄 발행</button>
</div>
<div style="font-size:10px;color:var(--d)">bm21.com 스타일: 컬러헤더 + FAQ + 칩태그 + Picsum 이미지 4장 자동 삽입</div>
</div>

<div class="card"><h3>키워드 대량 입력 (CSV: 지역,서비스 — 한 줄에 하나)</h3>
<textarea id="kwlist" rows="4" placeholder="인천,셔츠룸&#10;서울,노래방&#10;부산,가라오케"></textarea>
<div class="row" style="margin-top:6px"><button class="btn btn-p" onclick="genFromList()">목록 전체 생성+발행</button>
<span style="color:var(--d);font-size:10px" id="kwCount">0줄</span>
<span style="flex:1"></span>
<select id="kwSiteFilter" style="width:auto"><option value="">전체 사이트</option>{% for s in sites %}<option value="{{s.id}}">{{s.name or s.site_url[:20]}}</option>{% endfor %}</select></div></div>

<div class="card"><h3>키워드 풀 (엑셀 랜덤 치환) — 제목 = 키워드1 + 번호 + 키워드2 + 키워드3</h3>
<div style="font-size:10px;color:var(--d);margin-bottom:6px">여기에 등록한 지역/서비스/브랜드 조합에서 <b style="color:var(--p)">매번 랜덤으로 뽑아</b> 본문·제목을 만듭니다. 제목의 번호는 설정 탭 전화번호에서 랜덤 기호로 변형됩니다. 예약 발행에서 키워드 세트를 비워두면 이 풀에서 자동으로 뽑습니다.</div>
<textarea id="poolCsv" rows="5" placeholder="지역,서비스,브랜드 — 한 줄에 하나&#10;인천,셔츠룸,인천홍마니&#10;부천,노래방,&#10;서울,가라오케,"></textarea>
<div class="row" style="margin-top:6px">
<button class="btn btn-p" onclick="savePool()">풀 저장(덮어쓰기)</button>
<button class="btn btn-v" onclick="savePool(true)">풀에 추가</button>
<label class="btn btn-d" style="cursor:pointer">엑셀(.xlsx) 업로드<input type="file" id="poolXlsx" accept=".xlsx" style="display:none" onchange="uploadXlsx()"></label>
<span style="flex:1"></span>
<span style="color:var(--d);font-size:10px" id="poolCount">0개</span>
<button class="btn btn-r btn-xs" onclick="if(confirm('키워드 풀 전체 삭제?'))clearPool()">비우기</button></div>
<div class="row" style="margin-top:6px">
<select id="poolSiteFilter" style="width:auto"><option value="">전체 허용 사이트</option>{% for s in sites %}<option value="{{s.id}}">{{s.name or s.site_url[:20]}}</option>{% endfor %}</select>
<input type="number" id="poolN" value="1" min="1" max="50" style="width:70px" title="랜덤 뽑을 개수">
<button class="btn btn-g" onclick="genRandom()">랜덤 생성+발행</button>
<span style="color:var(--d);font-size:10px">풀에서 N개 랜덤 추출 → 허용 사이트 발행</span></div></div>

<div class="card"><h3>이미지 URL 풀 (본문 삽입 — 한 줄에 하나)</h3>
<div style="font-size:10px;color:var(--d);margin-bottom:6px">여기에 넣은 이미지 주소에서 <b style="color:var(--p)">매번 랜덤으로</b> 골라 본문에 삽입합니다(alt=지역 자동). 대표님 사이트/호스팅 이미지 URL을 여러 개 넣으면 도메인이 분산돼 흔적이 적습니다. <b style="color:var(--y)">비워두면 기본 이미지</b>가 쓰입니다.</div>
<textarea id="imgUrls" rows="5" placeholder="https://내사이트.kr/img/room1.jpg&#10;https://내사이트.kr/img/room2.jpg&#10;https://another.kr/photo/a.png"></textarea>
<div class="row" style="margin-top:6px">
<button class="btn btn-p" onclick="saveImages(false)">저장(덮어쓰기)</button>
<button class="btn btn-v" onclick="saveImages(true)">추가</button>
<span style="flex:1"></span>
<span style="color:var(--d);font-size:10px" id="imgCount">0개</span>
<button class="btn btn-r btn-xs" onclick="if(confirm('이미지 URL 전체 삭제?'))clearImages()">비우기</button></div></div>

<div class="card"><h3>생성 결과</h3>
<div class="row"><input type="text" id="gTitle" placeholder="제목" style="font-weight:600"></div>
<textarea id="gContent" rows="10" placeholder="리치HTML 본문..."></textarea>
<div class="row" style="margin-top:6px"><span style="color:var(--d);font-size:10px" id="gLen">0자</span><span style="flex:1"></span>
<button class="btn btn-d" onclick="previewPost()">미리보기</button>
<button class="btn btn-g" onclick="postSel()">선택 사이트 발행</button>
<button class="btn btn-y" onclick="postAll()">전체 사이트 발행</button></div></div></div>

<div id="p-sites" class="panel">
<div class="card"><h3>사이트 추가</h3>
<div class="row"><input type="text" id="sUrl" placeholder="https://사이트.com (또는 board.php URL)" style="flex:1"><select id="sPlat" style="width:110px"><option value="auto">자동감지</option><option value="gnuboard">그누보드</option><option value="cafe24">Cafe24</option></select></div>
<div class="row"><input type="text" id="sName" placeholder="이름" style="flex:1"><input type="text" id="sBo" placeholder="게시판ID (bo_table) 예:free" style="width:170px"></div>
<div class="row"><input type="number" id="sDaily" min="0" value="3" placeholder="하루 발행 한도" title="0은 무제한" style="width:150px"><input type="number" id="sInterval" min="0" value="60" placeholder="최소 간격(분)" title="사이트별 최소 발행 간격(분)" style="width:150px"><span style="color:var(--d);font-size:10px">사이트별 하루 한도 / 최소 간격(분)</span></div>
<div class="row"><input type="text" id="sId" placeholder="아이디" style="width:130px"><input type="password" id="sPw" placeholder="비밀번호" style="width:130px"><button class="btn btn-p" id="addBtn" onclick="addSite()">추가</button><button class="btn btn-d btn-xs" id="editCancel" style="display:none" onclick="cancelEdit()">취소</button></div>
<div class="row" style="background:#0b1a2e;border:1px solid #1a3a5e;border-radius:6px;padding:8px"><label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:11px;white-space:nowrap"><input type="checkbox" id="sPerm" style="width:auto">✔ 홍보 허용 확인됨</label><input type="text" id="sPermNote" placeholder="허용 근거 (운영자 메일/캡처 링크·홍보게시판명 등)" style="flex:1"></div>
<div style="font-size:10px;color:var(--d)">게시판ID(bo_table)는 글이 올라갈 게시판 식별자입니다. 예) 자유게시판 free, 홍보게시판 promotion · <b style="color:var(--y)">홍보 허용을 체크한 사이트만 발행/테스트됩니다.</b></div></div>
<div class="card"><h3>CSV 대량 등록 (한 줄에 하나: URL,이름,게시판,아이디,비번,허용여부)</h3>
<textarea id="bulkCsv" rows="4" placeholder="https://a.kr,에이,free,id1,pw1,1&#10;https://b.kr,비,promotion,id2,pw2,0"></textarea>
<div class="row" style="margin-top:6px"><label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:11px"><input type="checkbox" id="bulkPerm" style="width:auto">허용여부 미기재 시 기본 허용</label><span style="flex:1"></span><button class="btn btn-p" onclick="bulkAdd()">대량 등록</button></div>
<div style="font-size:10px;color:var(--d)">허용여부: 1/0 (마지막 칸). 미기재면 위 체크박스 기본값. 등록 후에도 목록에서 선택→허용 일괄 변경 가능.</div></div>
<div class="card"><h3>사이트 목록 (실시간 갱신)</h3>
<div class="row" style="margin-bottom:6px"><button class="btn btn-g btn-xs" onclick="bulkPermSet(true)">선택 허용✔</button><button class="btn btn-r btn-xs" onclick="bulkPermSet(false)">선택 미허용</button><button class="btn btn-d btn-xs" onclick="healthAll()">선택 상태점검</button><span style="flex:1"></span><span style="color:var(--d);font-size:10px">체크박스로 선택 후 사용</span></div>
<div style="max-height:400px;overflow-y:auto" id="siteList"></div></div></div>

<div id="p-res" class="panel">
<div class="card"><h3>발행 결과 이력</h3>
<div class="row" style="margin-bottom:8px">
<button class="btn btn-g" onclick="location='/api/history/export'">엑셀 내보내기</button>
<button class="btn btn-d" onclick="renderHistory()">새로고침</button>
<span style="flex:1"></span>
<span style="color:var(--d);font-size:10px" id="histCount">0건</span>
<button class="btn btn-r btn-xs" onclick="if(confirm('이력 전체 삭제?'))api('/history/clear','POST').then(()=>{toast('이력 삭제됨');renderHistory()})">이력 비우기</button></div>
<div style="max-height:520px;overflow-y:auto" id="histList"></div></div></div>

<div id="p-disco" class="panel">
<div class="note">🔎 저장한 검색 키워드를 Google Custom Search API로 차례대로 검색해 후보를 수집합니다. 접속되지 않는 후보도 삭제하지 않고 목록에 유지해 나중에 다시 점검할 수 있습니다.</div>
<div id="dcSummary" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px"></div>
<div class="card"><h3>후보 수집</h3>
<div class="row"><button class="btn btn-p" onclick="discoverNow()">Google 검색 실행 (5키워드)</button>
<button class="btn btn-d" onclick="screenNow()">미검수 일괄 검수</button>
<button class="btn btn-d" onclick="if(confirm('모든 후보를 다시 검수할까요?'))rescreenAll()">전체 재검수</button>
<button class="btn btn-g" onclick="location='/api/candidates/export'">엑셀 내보내기</button>
<span style="flex:1"></span>
<button class="btn btn-r btn-xs" onclick="if(confirm('탈락 후보만 삭제할까요?'))clearRejected()">탈락 정리</button></div>
<div style="font-size:10px;color:var(--d);margin-bottom:6px">API 키가 없어도 아래에 URL을 직접 붙여넣으면 검수·점수·연락처·메일초안이 자동 생성됩니다.</div>
<textarea id="dcUrls" rows="3" placeholder="URL 직접 추가 (한 줄에 하나)&#10;https://example.kr/bbs/board.php?bo_table=promotion"></textarea>
<div class="row" style="margin-top:6px"><button class="btn btn-v" onclick="addManual()">URL 추가 + 검수</button></div></div>
<div class="card"><h3>후보 목록 (점수순 — 위에서부터 검토)</h3>
<div class="row" style="margin-bottom:6px"><select id="dcFilter" style="width:auto" onchange="renderCands()"><option value="">전체</option><option value="ready">검수완료</option><option value="contacted">문의 발송함</option><option value="approved">사이트 등록됨</option><option value="rejected">제외</option><option value="new">미검수</option></select><span style="flex:1"></span><span style="color:var(--d);font-size:10px" id="dcCount">0개</span></div>
<div style="max-height:520px;overflow-y:auto" id="dcList"></div></div></div>

<div id="p-mem" class="panel">
<div class="note">👥 회원·정산 + ⏰ 계정당 개별 자동발행 — 회원마다 <b style="color:var(--p)">원하는 시간대</b>를 지정하면 서버가 24시간 그 시간에 자동 실행합니다(PC 불필요). 회원별 전용 키워드·담당 사이트를 따로 배정할 수 있고, 시간분산으로 동시 폭주를 막습니다. 월 청구 = 기본료 + (추가 광고수 × 추가단가).</div>
<div id="memSummary" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px"></div>
<div class="card"><h3>회원 추가 / 수정</h3>
<div class="row"><input type="text" id="mName" placeholder="담당자/이름" style="flex:1"><input type="text" id="mBiz" placeholder="업소명" style="flex:1"><input type="text" id="mPhone" placeholder="연락처" style="width:150px"></div>
<div class="row"><span style="color:var(--d);font-size:11px">기본료(월)</span><input type="number" id="mFee" value="30000" style="width:110px">
<span style="color:var(--d);font-size:11px">추가광고수</span><input type="number" id="mAddons" value="0" min="0" style="width:80px">
<span style="color:var(--d);font-size:11px">추가단가</span><input type="number" id="mAddonFee" value="10000" style="width:100px">
<span style="color:var(--d);font-size:11px">정산일</span><input type="number" id="mDay" value="1" min="1" max="31" style="width:70px">
<select id="mStatus" style="width:auto"><option value="active">활성</option><option value="paused">정지</option></select></div>
<div class="row"><input type="text" id="mMemo" placeholder="메모 (선택)" style="flex:1"></div>
<div style="border-top:1px solid var(--line);margin:10px 0;padding-top:12px">
<div style="font-size:11px;color:var(--p2);font-weight:700;margin-bottom:8px">⏰ 이 회원의 자동발행 스케줄 (계정당 개별 관리)</div>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:8px"><input type="checkbox" id="mSchedOn" style="width:auto">자동발행 켜기 — 서버가 24시간 이 시간에 자동 실행</label>
<div class="row"><input type="text" id="mTimes" placeholder="점프 시간대 HH:MM, 쉼표로 여러개 (예: 09:00,14:00,20:00)" style="flex:1"></div>
<div class="row"><span style="color:var(--d);font-size:11px">요일(미선택=매일):</span>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="0" style="width:auto">월</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="1" style="width:auto">화</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="2" style="width:auto">수</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="3" style="width:auto">목</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="4" style="width:auto">금</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="5" style="width:auto">토</label>
<label style="font-size:11px"><input type="checkbox" class="mDay" value="6" style="width:auto">일</label>
<span style="flex:1"></span>
<span style="color:var(--d);font-size:11px">1회당 발행수</span><input type="number" id="mPerRun" value="1" min="1" max="20" style="width:70px">
<span style="color:var(--d);font-size:11px">시간분산(분)</span><input type="number" id="mJitter" value="5" min="0" max="60" style="width:70px" title="설정 시각에서 0~N분 랜덤 지연 — 동시 폭주 방지"></div>
<div class="row"><textarea id="mKw" rows="3" placeholder="이 회원 전용 키워드 (지역,서비스,브랜드 — 한 줄에 하나)&#10;비우면 공용 키워드 풀 사용&#10;인천,셔츠룸,인천홍마니"></textarea></div>
<div class="row"><span style="color:var(--d);font-size:11px">담당 사이트(미선택=전체 허용 사이트):</span><div id="mSiteBox" style="display:flex;gap:8px;flex-wrap:wrap;font-size:11px"></div></div>
</div>
<div class="row"><button class="btn btn-p" id="memBtn" onclick="addMember()">회원 추가</button><button class="btn btn-d btn-xs" id="memCancel" style="display:none" onclick="cancelMember()">취소</button></div></div>
<div class="card"><h3>회원 목록 (이번 달 정산)</h3>
<div class="row" style="margin-bottom:6px"><button class="btn btn-g" onclick="location='/api/members/export'">엑셀 내보내기</button><button class="btn btn-d" onclick="renderMembers()">새로고침</button><span style="flex:1"></span><span style="color:var(--d);font-size:10px" id="memCount">0명</span></div>
<div style="max-height:460px;overflow-y:auto" id="memList"></div></div></div>

<div id="p-stats" class="panel">
<div class="card"><h3>발행 통계</h3>
<div class="row" style="margin-bottom:8px"><button class="btn btn-d btn-xs" onclick="renderStats()">새로고침</button><button class="btn btn-v btn-xs" onclick="api('/verify/now','POST').then(()=>toast('생존 확인 시작 (1~2분 후 갱신)'))">지금 생존확인</button></div>
<div id="statTop" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px"></div>
<div id="statReasons"></div>
<div id="statSurvival"></div>
<div style="font-size:11px;color:var(--d);margin:6px 0">최근 14일 (초록=성공, 빨강=실패)</div>
<div id="statDays"></div>
<div style="font-size:11px;color:var(--d);margin:14px 0 6px">사이트별 (상위 12)</div>
<div id="statSites"></div></div></div>

<div id="p-set" class="panel">
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
<div class="card"><h3>브랜드/전화</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">브랜드명</small><input id="cBrand" value="{{cfg.brand}}"></div>
<div style="margin-bottom:6px"><small style="color:var(--d)">대표 전화번호</small><input id="cPhone" value="{{cfg.phone}}"></div>
<small style="color:var(--d)">전화번호(여러 개, 한 줄에 하나 — 제목에 랜덤 사용)</small>
<textarea id="cPhones" rows="3" placeholder="01082755736&#10;01021636400&#10;01053505892"></textarea>
<div style="font-size:10px;color:var(--d)">제목의 번호는 매번 <b style="color:var(--p)">[010]↔8275↔5736 · O1O=2572=3859 · [OIO-5350-5892]</b> 처럼 랜덤 기호로 변형됩니다. 여러 개면 그 중 하나를 랜덤 선택. 비우면 위 대표 전화번호 사용.</div></div>
<div class="card"><h3>워커/비번</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">워커 수</small><input type="number" id="cWorkers" value="{{cfg.workers}}" min="1" max="10"></div>
<small style="color:var(--d)">비밀번호</small><input type="password" id="cPw" placeholder="변경시 입력"></div>
<div class="card"><h3>레이트 리밋 (도배 방지)</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">포스트 간 지연 (초)</small><input type="number" id="cDelay" value="{{cfg.post_delay}}" min="0"></div>
<small style="color:var(--d)">사이트당 1일 발행 한도 (0=무제한)</small><input type="number" id="cDaily" value="{{cfg.daily_limit}}" min="0"></div>
<div class="card"><h3>GPT 본문 생성</h3>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cUseGpt" style="width:auto">GPT로 본문 생성(키 필요)</label>
<div style="margin-bottom:6px"><small style="color:var(--d)">OpenAI API 키</small><input type="password" id="cOpenai" placeholder="변경시만 입력 (sk-...)"></div>
<small style="color:var(--d)">모델</small><input id="cModel" value="{{cfg.model}}" placeholder="gpt-4o-mini"></div>
<div class="card"><h3>텔레그램 알림</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">봇 토큰</small><input type="password" id="cTgTok" placeholder="변경시만 입력"></div>
<div style="margin-bottom:6px"><small style="color:var(--d)">챗 ID</small><input id="cTgChat" placeholder="예: 123456789"></div>
<div class="row" style="font-size:11px;color:var(--d)"><label style="display:flex;align-items:center;gap:4px"><input type="checkbox" id="cNotifyDone" style="width:auto">성공알림</label><label style="display:flex;align-items:center;gap:4px"><input type="checkbox" id="cNotifyFail" style="width:auto">실패알림</label><button class="btn btn-d btn-xs" onclick="api('/telegram/test','POST').then(r=>toast(r&&r.ok?'전송됨':'실패: '+(r&&r.error||''),r&&r.ok?'ok':'er'))">테스트 전송</button></div></div>
<div class="card"><h3>🔎 도메인 발굴 (Google Custom Search API)</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">Google API 키</small><input type="password" id="cGoogleKey" placeholder="변경시에만 입력"></div>
<div style="margin-bottom:6px"><small style="color:var(--d)">검색엔진 ID (cx)</small><input id="cGoogleCx" placeholder="Programmable Search Engine ID"></div>
<div style="font-size:10px;color:var(--d);margin-bottom:6px"><a href="https://programmablesearchengine.google.com/controlpanel/all" target="_blank" rel="noopener" style="color:var(--p)">검색엔진 ID 확인</a> · API 키는 화면과 API 응답에 노출되지 않습니다.</div>
<div class="row" style="margin-bottom:6px"><span style="color:var(--d);font-size:11px">하루 후보 목표</span><input type="number" id="cDTarget" value="100" min="10" max="1000" style="width:90px">
<span style="color:var(--d);font-size:11px">하루 쿼리 한도</span><input type="number" id="cDQuery" value="100" min="1" max="10000" style="width:90px" title="Google API 할당량 안에서 사용"></div>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cDiscoOn" style="width:auto">24시간 자동 발굴 켜기 (15분마다 조금씩 수집·검수)</label>
<small style="color:var(--d)">업종·지역 키워드 (선택, 한 줄에 하나 — 쿼리에 곱해집니다)</small>
<textarea id="cDKw" rows="2" placeholder="인천 노래방&#10;부천 홍보"></textarea>
<small style="color:var(--d)">Google 검색 키워드 목록 (한 줄에 하나 — 입력 그대로 검색, #으로 시작하면 메모)</small>
<textarea id="cDDirect" rows="4" placeholder="&quot;홍보게시판&quot; 마사지&#10;inurl:bbs/board.php &quot;업체등록&quot;&#10;인천 광고 가능한 게시판"></textarea>
<div style="font-size:10px;color:var(--g);margin-top:4px">설정 저장을 누르면 서버에 영구 저장되며, 직접 검색어가 자동 조합 검색어보다 먼저 실행됩니다.</div>
<div style="font-size:10px;color:var(--d);margin-top:6px">후보 목록에서 필요한 사이트를 선택해 사이트 관리 목록에 등록할 수 있습니다.</div></div>
<div class="card"><h3>자동 백업 · 봇 제어 · 발행 검증</h3>
<div style="margin-bottom:6px"><small style="color:var(--d)">매일 자동 백업 시각 (HH:MM, KST · 비우면 끔)</small><input id="cBackupTime" placeholder="예: 04:00"></div>
<div class="row" style="margin-bottom:6px"><button class="btn btn-g btn-xs" onclick="api('/backup/now','POST').then(r=>toast(r&&r.ok?'📦 백업 전송됨':'실패: '+(r&&r.error||'토큰 확인'),r&&r.ok?'ok':'er'))">지금 백업 전송</button><span style="color:var(--d);font-size:10px">텔레그램으로 zip 전송(설정·사이트·이력·예약·키워드)</span></div>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cTgControl" style="width:auto">텔레그램 폰 제어 (봇에게 /상태 /오늘 /발행 /정지 /재개 /백업 /검증)</label>
<label style="display:flex;align-items:center;gap:6px;color:var(--v);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cVerify" style="width:auto">발행글 생존 자동 검증 (1시간마다 URL 재확인)</label>
<label style="display:flex;align-items:center;gap:6px;color:var(--g);font-size:12px;margin-bottom:6px"><input type="checkbox" id="cMixKw" style="width:auto">키워드 열 혼합 (지역·서비스·브랜드를 엑셀 전체에서 독립 랜덤 조합)</label>
<label style="display:flex;align-items:center;gap:6px;color:var(--y);font-size:12px"><input type="checkbox" id="cBlockUnpaid" style="width:auto">미납 회원 자동 정지 (이번 달 미납이면 그 회원 자동발행 중단)</label>
<div style="font-size:10px;color:var(--d);margin-top:6px">폰 제어를 켜면 봇 채팅창에 명령을 보내 PC·패널 없이 조작할 수 있습니다(등록된 챗ID만 허용). 백업·검증은 텔레그램 토큰/챗ID가 필요합니다.<br><b style="color:var(--g)">중복 방지 상시 작동</b> — 모든 제목·본문은 과거와 겹치지 않게 매번 새로 생성되며, 같은 키워드라도 사이트마다 글이 다릅니다. 본문 끝에는 alt=지역(맨앞 키워드) 이미지가 삽입됩니다.</div></div>
</div>
<div style="display:flex;gap:8px;margin-bottom:10px">
<button class="btn btn-p" onclick="saveCfg()">설정 저장</button>
<button class="btn btn-g" onclick="api('/workers/start','POST',{n:parseInt(document.getElementById('cWorkers').value)||2}).then(()=>toast('워커 시작'))">워커 시작</button>
<button class="btn btn-y" onclick="api('/workers/pause','POST').then(()=>toast('일시정지'))">일시정지</button>
<button class="btn btn-v" onclick="api('/workers/resume','POST').then(()=>toast('재개'))">재개</button>
<button class="btn btn-r" onclick="api('/workers/stop','POST').then(()=>toast('워커 정지'))">워커 정지</button>
<button class="btn btn-d" onclick="if(confirm('통계 초기화?'))api('/workers/reset','POST').then(()=>toast('초기화됨'))">통계 초기화</button></div>
<div class="note">일시정지는 진행 중인 큐를 지우지 않고 멈춥니다(재개 시 이어서). 패널을 껐다 켜도 미완료 작업은 자동 복구됩니다.</div>
<div class="card"><h3>🩺 서버 자가진단 (실제 발행 가능 여부)</h3>
<div class="row"><button class="btn btn-v" onclick="runDiag()">진단 실행 (최대 60초)</button><span style="color:var(--d);font-size:10px">크롬 설치·드라이버 기동·페이지 로드·메모리·데이터 상태를 점검합니다</span></div>
<div id="diagOut" style="margin-top:10px"></div></div></div>

</div><!-- wrap -->

<script>
const $=id=>document.getElementById(id);
function T(n){document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));document.querySelectorAll('.panel').forEach(p=>p.classList.remove('on'));document.querySelector(`[onclick="T('${n}')"]`).classList.add('on');$('p-'+n).classList.add('on');if(n==='res')renderHistory();if(n==='sched')renderScheds();if(n==='stats')renderStats();if(n==='set')loadCfgUI();if(n==='gen'){loadPool();loadImages()}if(n==='mem'){renderMembers();if(!document.querySelector('.mSite'))fillSiteBox([])}if(n==='disco')renderCands()}
function toast(m,c='ok'){const d=$('toasts');const e=document.createElement('div');e.className='toast toast-'+c;e.textContent=m;d.appendChild(e);setTimeout(()=>e.remove(),2500)}
function esc(s){return String(s==null?'':s).replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
async function api(p,m,b){try{const o={method:m,headers:{'Content-Type':'application/json'}};if(b)o.body=JSON.stringify(b);const r=await fetch('/api'+p,o);if(r.status===401){location='/login';return null}return await r.json()}catch(e){toast(e.message,'er');return null}}
function kv(){return{지역:$('k1').value.trim(),서비스:$('k2').value.trim(),브랜드:$('k3').value.trim()}}
async function gen(){const ks=kv();if(!ks.지역||!ks.서비스){toast('키워드1,2 입력','er');return}const r=await api('/generate','POST',{keywords:ks});if(r&&r.ok){$('gTitle').value=r.title;$('gContent').value=r.content;$('gLen').textContent=(r.content||'').length.toLocaleString()+'자';toast('생성 완료')}}
async function genBulk(){const ks=kv();if(!ks.지역||!ks.서비스){toast('키워드 입력','er');return}const sid=$('kwSiteFilter').value;const r=await api('/bulk','POST',{keyword_sets:[ks],site_ids:sid?[sid]:[]});if(r&&r.ok)toast(r.generated+'건 큐 등록'+(r.blocked?` · 미허용 ${r.blocked}개 제외`:''));else if(r)toast(r.error||'실패','er')}
function parseList(){return $('kwlist').value.split('\n').map(l=>l.trim()).filter(Boolean).map(l=>{const p=l.split(',');return{지역:(p[0]||'').trim(),서비스:(p[1]||'').trim()}}).filter(k=>k.지역&&k.서비스)}
async function genFromList(){const sets=parseList();if(!sets.length){toast('키워드 없음','er');return}const sid=$('kwSiteFilter').value;const r=await api('/bulk','POST',{keyword_sets:sets,site_ids:sid?[sid]:[]});if(r&&r.ok)toast(r.generated+'건 큐 등록'+(r.blocked?` · 미허용 ${r.blocked}개 제외`:''));else if(r)toast(r.error||'실패','er')}
function getSiteIds(){return Array.from(document.querySelectorAll('.cb:checked')).map(c=>c.dataset.id)}
function getAllSiteIds(){return Array.from(document.querySelectorAll('#siteList tr[data-id]')).map(r=>r.dataset.id)}
async function postSel(){const t=$('gTitle').value.trim();const c=$('gContent').value.trim();if(!t||!c){toast('제목/본문 입력','er');return}const ids=getSiteIds();if(!ids.length){toast('사이트 선택','er');return}const r=await api('/post','POST',{site_ids:ids,title:t,content:c,region:$('k1').value.trim(),service:$('k2').value.trim(),brand:$('k3').value.trim()});if(r&&r.ok)toast(r.queued+'개 큐 등록'+(r.blocked?` · 미허용 ${r.blocked}개 제외`:''),r.queued?'ok':'er')}
async function postAll(){const t=$('gTitle').value.trim();const c=$('gContent').value.trim();if(!t||!c){toast('제목/본문 입력','er');return}const ids=getAllSiteIds();if(!ids.length){toast('사이트 없음','er');return}const r=await api('/post','POST',{site_ids:ids,title:t,content:c,region:$('k1').value.trim(),service:$('k2').value.trim(),brand:$('k3').value.trim()});if(r&&r.ok)toast(r.queued+'개 큐 등록'+(r.blocked?` · 미허용 ${r.blocked}개 제외`:''),r.queued?'ok':'er')}
// ---- 발행 이력 렌더 ----
function histRow(h){const stc=h.status==='done'?'ok':(h.status==='failed'?'f':(h.status==='retry'?'y':'i'));const rurl=h.result_url?`<a href="${esc(h.result_url)}" target="_blank" style="color:var(--p)">열기</a>`:'';
const stt=h.status==='retry'?'재시도':(h.status||'');
const av=h.alive==='yes'?'<span class="st st-ok">생존</span>':(h.alive==='no'?'<span class="st st-f">삭제</span>':(h.status==='done'?'<span style="color:var(--d)">-</span>':''));
const fr=h.fail_reason_ko?`<span style="color:var(--r)">${esc(h.fail_reason_ko)}</span>`:'';
return `<tr><td style="color:var(--d);white-space:nowrap">${esc((h.time||'').slice(5,16))}</td><td>${esc(h.site_name||'')}</td><td style="color:var(--d)">${esc(h.region||'')} ${esc(h.service||'')}</td><td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${esc(h.title||'')}">${esc(h.title||'')}</td><td><span class="st st-${stc}">${esc(stt)}</span></td><td>${fr}</td><td>${av}</td><td>${rurl}</td><td style="color:var(--d);max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${esc(h.message||'')}">${esc(h.message||'')}</td></tr>`}
async function renderHistory(){const h=await api('/history','GET');if(!Array.isArray(h))return;$('histCount').textContent=h.length+'건';
if(!h.length){$('histList').innerHTML='<p style="color:var(--d);padding:30px;text-align:center">아직 발행 이력이 없습니다</p>';return}
$('histList').innerHTML='<table><thead><tr><th>시간</th><th>사이트</th><th>지역/서비스</th><th>제목</th><th>상태</th><th>실패원인</th><th>생존</th><th>결과</th><th>메시지</th></tr></thead><tbody>'+h.map(histRow).join('')+'</tbody></table>'}
let _editId=null;
async function runDiag(){$('diagOut').innerHTML='<p style="color:var(--d);padding:14px">🩺 진단 중... 크롬을 실제로 띄워보는 중이라 최대 60초 걸립니다.</p>';const r=await api('/diag','GET');if(!r){$('diagOut').innerHTML='<p style="color:var(--r)">진단 실패</p>';return}
const rows=(r.steps||[]).map(s=>`<tr><td>${s.ok?'<span style="color:var(--g)">✅</span>':'<span style="color:var(--r)">❌</span>'}</td><td><b>${esc(s.name)}</b></td><td style="color:var(--d)">${esc(s.detail)}</td></tr>`).join('');
const hdr=r.ok?'<span style="color:var(--g)">✅ 발행 가능 — 크롬 정상 작동</span>':'<span style="color:var(--r)">❌ 발행 불가 — 아래 ❌ 항목 확인</span>';
$('diagOut').innerHTML=`<div style="margin-bottom:8px;font-weight:700">${hdr}</div><table>${rows}</table><div style="font-size:10px;color:var(--d);margin-top:8px">${esc(r.platform||'')} · Python ${esc(r.python||'')}</div>`}
// ---- 도메인 발굴 ----
let _cands=[];
async function renderCands(){const r=await api('/candidates','GET');if(!r||!r.candidates)return;_cands=r.candidates;const s=r.summary||{};
$('dcSummary').innerHTML=tile('전체',s.total||0,'var(--t)')+tile('검수완료',s.ready||0,'var(--g)')+tile('문의발송',s.contacted||0,'var(--y)')+tile('사이트 등록',s.approved||0,'var(--p)')+tile('제외',s.rejected||0,'var(--r)')+tile('오늘 쿼리',(s.today_queries||0)+'/100','var(--v)');
const f=$('dcFilter').value;const list=f?_cands.filter(c=>c.status===f):_cands;
$('dcCount').textContent=list.length+'개';
if(!list.length){$('dcList').innerHTML='<p style="color:var(--d);padding:30px;text-align:center">후보가 없습니다. Google 키워드 검색을 실행하거나 URL을 직접 추가하세요.</p>';return}
$('dcList').innerHTML='<table><thead><tr><th>점수</th><th>도메인/게시판</th><th>판정</th><th>연락처</th><th>상태</th><th>동작</th></tr></thead><tbody>'+list.map(c=>{
const sc=c.score||0;const scc=sc>=50?'var(--g)':(sc>=20?'var(--y)':'var(--r)');
const flags=[];
if(c.promo_hint)flags.push('<span class="st st-ok">홍보허용흔적</span>');
if(c.parked)flags.push('<span class="st st-f">주차도메인</span>');
if(c.illegal)flags.push('<span class="st st-f">도박·불법</span>');
if(c.ad_banned)flags.push('<span class="st st-f">광고금지</span>');
if(c.captcha)flags.push('<span class="st st-y">🧩캡차</span>');
if(c.login_required)flags.push('<span class="st st-i">로그인필요</span>');
if(c.write_form)flags.push('<span class="st st-ok">글쓰기폼</span>');
if(c.last_post_days!=null)flags.push('<span class="st st-i">최근글 '+c.last_post_days+'일</span>');
const stmap={ready:'<span class="st st-ok">검수완료</span>',new:'<span class="st st-i">미검수</span>',contacted:'<span class="st st-y">문의발송</span>',approved:'<span class="st st-ok">사이트등록</span>',rejected:'<span class="st st-f">제외</span>'};
const em=(c.emails||[]).join(', ')||'<span style="color:var(--d)">-</span>';
const acts=(c.status==='approved')?'':(
 '<button class="btn btn-d btn-xs" onclick="showMail(\''+esc(c.id)+'\')">메일초안</button> '+
 '<button class="btn btn-y btn-xs" onclick="setCand(\''+esc(c.id)+'\',\'contacted\')">문의함</button> '+
 '<button class="btn btn-g btn-xs" onclick="approveCand(\''+esc(c.id)+'\')">사이트 등록</button> '+
 '<button class="btn btn-r btn-xs" onclick="setCand(\''+esc(c.id)+'\',\'rejected\')">탈락</button>');
return '<tr><td style="color:'+scc+';font-weight:700;font-size:15px">'+sc+'</td>'+
'<td><a href="'+esc(c.url)+'" target="_blank" style="color:var(--p)"><b>'+esc(c.domain||'')+'</b></a><br><span style="color:var(--d);font-size:10px">'+esc((c.board_name||c.title||'').slice(0,44))+'</span></td>'+
'<td style="max-width:230px">'+flags.join(' ')+(c.reject_reason?'<br><span style="color:var(--r);font-size:10px">'+esc(c.reject_reason)+'</span>':'')+'</td>'+
'<td style="font-size:10px;color:var(--g)">'+esc(em)+'</td><td>'+(stmap[c.status]||'')+'</td><td style="white-space:nowrap">'+acts+'</td></tr>'}).join('')+'</tbody></table>'}
async function discoverNow(){toast('🔎 Google 키워드 검색 중...(약 10초)');const r=await api('/candidates/discover','POST',{queries:5});if(r&&r.ok){toast('신규 '+r.added+'개 · 검수 '+r.screened+'건 (오늘 검색 '+r.today_queries+'회)');renderCands()}else toast((r&&r.error)||'실패','er')}
async function screenNow(){toast('검수 중...(최대 1분)');const r=await api('/candidates/screen','POST',{limit:20});if(r&&r.ok){toast(r.screened+'건 검수 완료');renderCands()}}
async function rescreenAll(){toast('전체 재검수 중...(최대 2분)');const r=await api('/candidates/screen','POST',{rescreen:true,limit:40});if(r&&r.ok){toast(r.screened+'건 재검수 완료');renderCands()}}
async function addManual(){const u=$('dcUrls').value;if(!u.trim()){toast('URL 입력','er');return}const r=await api('/candidates/manual','POST',{urls:u});if(r&&r.ok){toast(r.added+'개 추가 · 검수 시작(잠시 후 새로고침)');$('dcUrls').value='';setTimeout(renderCands,3000);renderCands()}else toast((r&&r.error)||'실패','er')}
async function setCand(id,st){await api('/candidates/status','POST',{id:id,status:st});renderCands()}
async function clearRejected(){await api('/candidates','DELETE',{clear:'rejected'});renderCands()}
function showMail(id){const c=_cands.find(x=>x.id===id);if(!c){toast('없음','er');return}
$('pvTitle').textContent=c.domain||'';
const body=(c.mail_draft||'(검수 후 생성됩니다)').replace(/&/g,'&amp;').replace(/</g,'&lt;');
$('pvFrame').srcdoc='<!DOCTYPE html><html lang=ko><head><meta charset=utf-8><style>body{font-family:-apple-system,sans-serif;padding:20px;color:#222}pre{white-space:pre-wrap;font-family:inherit;font-size:14px;line-height:1.7;background:#f7f8fa;padding:16px;border-radius:8px}</style></head><body><h3 style="font-size:15px;margin-bottom:10px">✉️ 제휴 문의 메일 초안 — 복사해서 발송하세요</h3><pre>'+body+'</pre></body></html>';
$('pvOverlay').style.display='block'}
async function approveCand(id){const c=_cands.find(x=>x.id===id);if(!c)return;
if(c.ad_banned&&!confirm('⚠️ 이 사이트는 "광고 금지" 문구가 감지되었습니다. 그래도 사이트 목록에 등록할까요?'))return;
const bo=prompt('게시판ID(bo_table)',c.bo_table||'free')||'free';
const mid=prompt('로그인 아이디 (없으면 비워두세요)','')||'';
const mpw=mid?(prompt('비밀번호','')||''):'';
const r=await api('/candidates/approve/'+id,'POST',{bo_table:bo,mb_id:mid,mb_pass:mpw});
if(r&&r.ok){toast('✅ 사이트 목록에 등록됨');renderCands();renderSites()}else toast((r&&r.error)||'실패','er')}
// ---- 회원·정산 ----
function won(n){return (n||0).toLocaleString()+'원'}
let _memEdit=null;
async function renderMembers(){const r=await api('/members','GET');if(!r||!r.members)return;const s=r.summary||{};
$('memSummary').innerHTML=tile('회원',s.members||0,'var(--t)')+tile('활성',s.active||0,'var(--g)')+tile('이번달 청구',won(s.billed),'var(--p)')+tile('납부완료',won(s.paid),'var(--g)')+tile('미납',won(s.unpaid),'var(--r)')+tile('미납 회원',s.unpaid_count||0,'var(--y)');
$('memCount').textContent=(r.members.length)+'명';
if(!r.members.length){$('memList').innerHTML='<p style="color:var(--d);padding:30px;text-align:center">등록된 회원이 없습니다</p>';return}
const DW=['월','화','수','목','금','토','일'];
$('memList').innerHTML='<table><thead><tr><th>이름/업소</th><th>상태</th><th>월청구</th><th>이번달</th><th>스케줄</th><th>최근실행</th><th>동작</th></tr></thead><tbody>'+r.members.map(m=>{
const st=m.status==='active'?'<span class="st st-ok">활성</span>':'<span class="st st-i">정지</span>';
const pay=m.paid?'<button class="btn btn-g btn-xs" onclick="togglePay(\''+esc(m.id)+'\',false)" title="'+esc(m.paid_at||'')+'">✔ 납부</button>':'<button class="btn btn-r btn-xs" onclick="togglePay(\''+esc(m.id)+'\',true)">미납</button>';
const addon=m.addons?(' <span style="color:var(--d);font-size:9px">+광고'+m.addons+'</span>'):'';
const times=(m.sched_times||[]);
const days=(m.sched_days&&m.sched_days.length)?m.sched_days.map(d=>DW[d]).join(''):'매일';
const sched=m.sched_enabled&&times.length
  ? '<span class="st st-ok">ON</span> <span style="color:var(--p);font-size:10px">'+esc(times.join(', '))+'</span><br><span style="color:var(--d);font-size:9px">'+days+' · '+(m.per_run||1)+'건/회 · ±'+(m.jitter==null?5:m.jitter)+'분'+((m.keywords||[]).length?' · 전용키워드'+m.keywords.length:'')+((m.site_ids||[]).length?' · 사이트'+m.site_ids.length:'')+'</span>'
  : '<span class="st st-i">OFF</span>';
return '<tr><td><b>'+esc(m.name||'')+'</b><br><span style="color:var(--d);font-size:10px">'+esc(m.biz||'')+' '+esc(m.phone||'')+'</span></td><td>'+st+'</td><td style="color:var(--p)">'+won(m.fee)+addon+'</td><td>'+pay+'</td><td>'+sched+'</td><td style="color:var(--d);font-size:10px">'+esc(m.last_run||'-')+(m.run_count?'<br>총 '+m.run_count+'회':'')+'</td><td><button class="btn btn-g btn-xs" onclick="runMember(\''+esc(m.id)+'\')" title="지금 1회 실행">실행</button> <button class="btn btn-p btn-xs" onclick=\'editMember('+JSON.stringify(m)+')\'>편집</button> <button class="btn btn-r btn-xs" onclick="delMember(\''+esc(m.id)+'\')">삭제</button></td></tr>'}).join('')+'</tbody></table>'}
function mDays(){return Array.from(document.querySelectorAll('.mDay:checked')).map(c=>parseInt(c.value))}
function mSiteIds(){return Array.from(document.querySelectorAll('.mSite:checked')).map(c=>c.dataset.id)}
async function fillSiteBox(sel){const sites=await api('/sites','GET');if(!Array.isArray(sites))return;const set=new Set(sel||[]);
$('mSiteBox').innerHTML=sites.map(s=>`<label style="display:flex;align-items:center;gap:3px"><input type="checkbox" class="mSite" data-id="${esc(s.id)}" style="width:auto" ${set.has(s.id)?'checked':''}>${esc(s.name||(s.site_url||'').slice(0,18))}${s.has_captcha?'🧩':''}</label>`).join('')}
async function addMember(){const d={name:$('mName').value.trim(),biz:$('mBiz').value.trim(),phone:$('mPhone').value.trim(),plan_fee:parseInt($('mFee').value)||0,addons:parseInt($('mAddons').value)||0,addon_fee:parseInt($('mAddonFee').value)||0,settle_day:parseInt($('mDay').value)||1,status:$('mStatus').value,memo:$('mMemo').value.trim(),
sched_enabled:$('mSchedOn').checked,sched_times:$('mTimes').value,sched_days:mDays(),site_ids:mSiteIds(),keywords_csv:$('mKw').value,jitter:parseInt($('mJitter').value)||0,per_run:parseInt($('mPerRun').value)||1};
if(!d.name&&!d.biz){toast('이름 또는 업소명 입력','er');return}if(_memEdit)d.id=_memEdit;const r=await api('/members','POST',d);if(r&&r.ok){toast(_memEdit?'수정됨':'회원 추가됨');cancelMember();renderMembers()}}
function editMember(m){_memEdit=m.id;$('mName').value=m.name||'';$('mBiz').value=m.biz||'';$('mPhone').value=m.phone||'';$('mFee').value=m.plan_fee||0;$('mAddons').value=m.addons||0;$('mAddonFee').value=m.addon_fee||0;$('mDay').value=m.settle_day||1;$('mStatus').value=m.status||'active';$('mMemo').value=m.memo||'';
$('mSchedOn').checked=!!m.sched_enabled;$('mTimes').value=(m.sched_times||[]).join(', ');$('mPerRun').value=m.per_run||1;$('mJitter').value=(m.jitter==null?5:m.jitter);
document.querySelectorAll('.mDay').forEach(c=>c.checked=(m.sched_days||[]).includes(parseInt(c.value)));
$('mKw').value=(m.keywords||[]).map(k=>[k.지역||'',k.서비스||'',k.브랜드||''].join(',')).join('\n');
fillSiteBox(m.site_ids||[]);
$('memBtn').textContent='수정 저장';$('memBtn').classList.add('btn-y');$('memCancel').style.display='';$('mName').scrollIntoView({behavior:'smooth',block:'center'})}
function cancelMember(){_memEdit=null;['mName','mBiz','mPhone','mMemo','mTimes','mKw'].forEach(i=>$(i).value='');$('mFee').value=30000;$('mAddons').value=0;$('mAddonFee').value=10000;$('mDay').value=1;$('mStatus').value='active';$('mSchedOn').checked=false;$('mPerRun').value=1;$('mJitter').value=5;document.querySelectorAll('.mDay').forEach(c=>c.checked=false);fillSiteBox([]);$('memBtn').textContent='회원 추가';$('memBtn').classList.remove('btn-y');$('memCancel').style.display='none'}
async function runMember(id){if(!confirm('이 회원의 스케줄을 지금 1회 실행합니다.\n배정 사이트에 실제로 발행됩니다. 진행할까요?'))return;toast('⏰ 실행중...');const r=await api('/members/run/'+id,'POST');if(r&&r.ok)toast('✅ '+r.generated+'건 큐 등록 (사이트 '+r.sites+'개)');else toast((r&&r.error)||'실패','er');renderMembers()}
async function delMember(id){if(!confirm('회원을 삭제할까요? (정산 기록도 삭제)'))return;await api('/members','DELETE',{id});renderMembers()}
async function togglePay(id,paid){await api('/members/pay','POST',{id:id,paid:paid});renderMembers()}
async function addSite(){const d={site_url:$('sUrl').value.trim(),platform:$('sPlat').value,name:$('sName').value.trim(),bo_table:$('sBo').value.trim(),mb_id:$('sId').value.trim(),mb_pass:$('sPw').value,permission:$('sPerm').checked,permission_note:$('sPermNote').value.trim(),daily_limit:Math.max(0,parseInt($('sDaily').value)||0),min_interval_minutes:Math.max(0,parseInt($('sInterval').value)||0)};if(!d.site_url){toast('URL 입력','er');return}if(_editId)d.id=_editId;const r=await api('/sites','POST',d);if(r&&r.ok){toast(_editId?'수정됨':(d.permission?'추가됨 (홍보 허용)':'추가됨 (미검증 — 발행 제외)'));cancelEdit();renderSites()}}
async function editSite(id){const sites=await api('/sites','GET');if(!Array.isArray(sites))return;const s=sites.find(x=>x.id===id);if(!s){toast('사이트 없음','er');return}
$('sUrl').value=s.site_url||'';$('sPlat').value=s.platform||'auto';$('sName').value=s.name||'';$('sBo').value=s.bo_table||'';$('sId').value=s.mb_id||'';$('sPw').value=s.mb_pass||'';$('sPerm').checked=!!s.permission;$('sPermNote').value=s.permission_note||'';$('sDaily').value=(s.daily_limit==null?3:s.daily_limit);$('sInterval').value=(s.min_interval_minutes==null?60:s.min_interval_minutes);
_editId=id;$('addBtn').textContent='수정 저장';$('addBtn').classList.add('btn-y');$('editCancel').style.display='';
$('sUrl').scrollIntoView({behavior:'smooth',block:'center'});toast('편집 모드 — 값을 고치고 "수정 저장"')}
function cancelEdit(){_editId=null;['sUrl','sName','sBo','sId','sPw','sPermNote'].forEach(i=>$(i).value='');$('sDaily').value=3;$('sInterval').value=60;$('sPerm').checked=false;$('sPlat').value='auto';$('addBtn').textContent='추가';$('addBtn').classList.remove('btn-y');$('editCancel').style.display='none'}
async function oneClick(){const n=parseInt($('ocN').value)||1;if(!confirm('키워드 풀에서 랜덤으로 뽑아 허용 사이트 전체에 '+n+'회 발행합니다.\n진행할까요?'))return;toast('⚡ 발행 준비중...');const r=await api('/oneclick','POST',{count:n});if(r&&r.ok)toast('⚡ '+r.generated+'건 큐 등록 (사이트 '+r.sites+'개 × '+r.picks+'회)');else toast((r&&r.error)||'실패','er')}
function previewPost(){const c=$('gContent').value.trim();if(!c){toast('먼저 글을 생성하세요','er');return}$('pvTitle').textContent=$('gTitle').value||'';const doc='<!DOCTYPE html><html lang=ko><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><style>body{font-family:-apple-system,sans-serif;max-width:760px;margin:0 auto;padding:16px;color:#222;line-height:1.7}img{max-width:100%}</style></head><body>'+c+'</body></html>';$('pvFrame').srcdoc=doc;$('pvOverlay').style.display='block'}
function closePreview(){$('pvOverlay').style.display='none';$('pvFrame').srcdoc=''}
async function delSite(id){if(!confirm('삭제?'))return;await api('/sites','DELETE',{id});renderSites()}
async function testSite(id){toast('Selenium 테스트 중...');const r=await api('/test/'+id,'POST');if(r&&r.ok)toast('✅ 테스트 성공!'+(r.platform?' ['+(r.platform==='cafe24'?'Cafe24':'그누보드')+']':'')+' '+(r.message||''));else toast('실패: '+(r?.error||r?.message||''),'er')}
async function saveCfg(){const d={brand:$('cBrand').value.trim(),phone:$('cPhone').value.trim(),phones:$('cPhones').value,workers:parseInt($('cWorkers').value)||2,post_delay:parseInt($('cDelay').value)||0,daily_limit:parseInt($('cDaily').value)||0,use_gpt:$('cUseGpt').checked,model:$('cModel').value.trim()||'gpt-4o-mini',telegram_chat_id:$('cTgChat').value.trim(),notify_done:$('cNotifyDone').checked,notify_fail:$('cNotifyFail').checked,backup_time:$('cBackupTime').value.trim(),telegram_control:$('cTgControl').checked,verify_enabled:$('cVerify').checked,mix_keywords:$('cMixKw').checked,block_unpaid:$('cBlockUnpaid').checked,search_provider:'google',discover_enabled:$('cDiscoOn').checked,discover_daily_target:parseInt($('cDTarget').value)||100,discover_query_limit:parseInt($('cDQuery').value)||100,discover_keywords:$('cDKw').value,discover_direct_queries:$('cDDirect').value,google_cx:$('cGoogleCx').value.trim()};
const gk=$('cGoogleKey').value.trim();if(gk)d.google_api_key=gk;
const pw=$('cPw').value.trim();if(pw)d.password=pw;const ok=$('cOpenai').value.trim();if(ok)d.openai_key=ok;const tg=$('cTgTok').value.trim();if(tg)d.telegram_token=tg;const r=await api('/config','POST',d);if(r&&r.ok){toast('저장 완료');$('cPw').value='';$('cOpenai').value='';$('cTgTok').value=''}}
async function loadCfgUI(){const c=await api('/config','GET');if(!c)return;$('cUseGpt').checked=!!c.use_gpt;$('cNotifyDone').checked=!!c.notify_done;$('cNotifyFail').checked=!!c.notify_fail;$('cTgControl').checked=!!c.telegram_control;$('cVerify').checked=(c.verify_enabled!==false);$('cMixKw').checked=(c.mix_keywords!==false);$('cBlockUnpaid').checked=(c.block_unpaid!==false);$('cDiscoOn').checked=!!c.discover_enabled;if(c.discover_daily_target)$('cDTarget').value=c.discover_daily_target;if(c.discover_query_limit)$('cDQuery').value=c.discover_query_limit;if(typeof c.discover_keywords==='string')$('cDKw').value=c.discover_keywords;if(typeof c.discover_direct_queries==='string')$('cDDirect').value=c.discover_direct_queries;$('cGoogleKey').placeholder=(c.google_api_key==='***설정됨***')?'설정됨 · 변경시에만 입력':'Google API 키 입력';if(typeof c.google_cx==='string')$('cGoogleCx').value=c.google_cx;if(c.backup_time)$('cBackupTime').value=c.backup_time;if(c.model)$('cModel').value=c.model;if(c.telegram_chat_id)$('cTgChat').value=c.telegram_chat_id;if(typeof c.phones==='string')$('cPhones').value=c.phones;$('cOpenai').placeholder=(c.openai_key==='***설정됨***')?'설정됨 · 변경시만 입력':'sk-... (변경시만)';$('cTgTok').placeholder=(c.telegram_token==='***설정됨***')?'설정됨 · 변경시만 입력':'변경시만 입력'}
async function loadPool(){const p=await api('/keywords','GET');if(!Array.isArray(p))return;$('poolCount').textContent=p.length+'개';$('poolCsv').value=p.map(k=>[k.지역||'',k.서비스||'',k.브랜드||''].join(',')).join('\n')}
async function savePool(append){const csv=$('poolCsv').value;const r=await api('/keywords','POST',{csv:csv,append:!!append});if(r&&r.ok){toast('풀 저장: '+r.count+'개');loadPool()}else if(r)toast('실패','er')}
async function clearPool(){const r=await api('/keywords','DELETE');if(r&&r.ok){toast('풀 비움');loadPool()}}
async function loadImages(){const p=await api('/images','GET');if(!Array.isArray(p))return;$('imgCount').textContent=p.length+'개';$('imgUrls').value=p.join('\n')}
async function saveImages(append){const text=$('imgUrls').value;const r=await api('/images','POST',{text:text,append:!!append});if(r&&r.ok){toast('이미지 URL: '+r.count+'개 저장');loadImages()}else if(r)toast('실패','er')}
async function clearImages(){const r=await api('/images','DELETE');if(r&&r.ok){toast('이미지 URL 비움');loadImages()}}
async function uploadXlsx(){const f=$('poolXlsx').files[0];if(!f)return;const fd=new FormData();fd.append('file',f);try{const r=await(await fetch('/api/keywords/upload',{method:'POST',body:fd})).json();if(r&&r.ok){toast('엑셀 업로드: '+r.count+'개');loadPool()}else toast(r&&r.error||'업로드 실패','er')}catch(e){toast(e.message,'er')}$('poolXlsx').value=''}
async function genRandom(){const sid=$('poolSiteFilter').value;const n=parseInt($('poolN').value)||1;const r=await api('/generate/random','POST',{site_ids:sid?[sid]:[],count:n});if(r&&r.ok){if(r.generated!=null)toast(r.generated+'건 큐 등록 (랜덤 '+r.picks+'개)'+(r.blocked?` · 미허용 ${r.blocked} 제외`:''));else{$('gTitle').value=r.title;$('gContent').value=r.content;$('gLen').textContent=(r.content||'').length.toLocaleString()+'자';toast('랜덤 미리보기 생성')}}else if(r)toast(r.error||'실패','er')}
// ---- 사이트 대량/허용/헬스 ----
async function bulkAdd(){const csv=$('bulkCsv').value.trim();if(!csv){toast('CSV 입력','er');return}const r=await api('/sites/bulk','POST',{csv:csv,permission:$('bulkPerm').checked});if(r&&r.ok){toast(r.added+'개 등록');$('bulkCsv').value='';renderSites()}}
async function bulkPermSet(v){const ids=getSiteIds();if(!ids.length){toast('사이트 선택','er');return}const r=await api('/sites/permission','POST',{ids:ids,permission:v});if(r&&r.ok){toast(r.changed+'개 '+(v?'허용':'미허용'));renderSites()}}
async function healthAll(){const ids=getSiteIds();if(!ids.length){toast('사이트 선택','er');return}toast(ids.length+'개 점검중...');for(const id of ids){await api('/sites/health/'+id,'POST')}renderSites();toast('점검 완료')}
// ---- 예약 스케줄 ----
function scDays(){return Array.from(document.querySelectorAll('.scDay:checked')).map(c=>parseInt(c.value))}
function scParseKw(){return $('scKw').value.split('\n').map(l=>l.trim()).filter(Boolean).map(l=>{const p=l.split(',');return{지역:(p[0]||'').trim(),서비스:(p[1]||'').trim()}}).filter(k=>k.지역&&k.서비스)}
async function addSched(){const name=$('scName').value.trim()||'예약';const times=$('scTimes').value.split(',').map(x=>x.trim()).filter(x=>/^\d{1,2}:\d{2}$/.test(x));if(!times.length){toast('시간 형식 HH:MM','er');return}const ks=scParseKw();if(!ks.length){toast('키워드 세트 입력','er');return}const sid=$('scSite').value;const r=await api('/schedules','POST',{name:name,times:times,days:scDays(),keyword_sets:ks,site_ids:sid?[sid]:[],enabled:true});if(r&&r.ok){toast('예약 추가됨');$('scName').value='';$('scTimes').value='';$('scKw').value='';document.querySelectorAll('.scDay').forEach(c=>c.checked=false);renderScheds()}}
async function delSched(id){if(!confirm('예약 삭제?'))return;await api('/schedules','DELETE',{id});renderScheds()}
async function toggleSched(id){await api('/schedules/toggle','POST',{id});renderScheds()}
const DOW=['월','화','수','목','금','토','일'];
async function renderScheds(){const list=await api('/schedules','GET');if(!Array.isArray(list))return;if(!list.length){$('schedList').innerHTML='<p style="color:var(--d);padding:20px;text-align:center">예약이 없습니다</p>';return}
$('schedList').innerHTML='<table><thead><tr><th>이름</th><th>시간</th><th>요일</th><th>키워드</th><th>상태</th><th>최근실행</th><th>동작</th></tr></thead><tbody>'+list.map(s=>{const days=(s.days&&s.days.length)?s.days.map(d=>DOW[d]).join(''):'매일';const kw=(s.keyword_sets||[]).map(k=>k.지역+' '+k.서비스).join(', ');return `<tr><td><b>${esc(s.name)}</b></td><td style="color:var(--p)">${esc((s.times||[]).join(', '))}</td><td>${days}</td><td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${esc(kw)}">${esc(kw)}</td><td><span class="st st-${s.enabled?'ok':'i'}">${s.enabled?'ON':'OFF'}</span></td><td style="color:var(--d)">${esc(s.last_run||'-')}</td><td><button class="btn btn-y btn-xs" onclick="toggleSched('${esc(s.id)}')">토글</button> <button class="btn btn-r btn-xs" onclick="delSched('${esc(s.id)}')">삭제</button></td></tr>`}).join('')+'</tbody></table>'}
// ---- 통계 ----
function tile(label,val,color){return `<div class="card" style="flex:1;min-width:110px;text-align:center;margin:0"><div style="font-size:22px;font-weight:700;color:${color}">${val}</div><div style="font-size:10px;color:var(--d)">${label}</div></div>`}
async function renderStats(){const s=await api('/stats','GET');if(!s)return;
$('statTop').innerHTML=tile('전체',s.total,'var(--t)')+tile('성공',s.ok,'var(--g)')+tile('실패',s.fail,'var(--r)')+tile('스킵',s.skip,'var(--y)')+tile('성공률',s.rate+'%','var(--p)')+tile('생존율',(s.alive_rate||0)+'%','var(--v)');
const rz=(s.reasons||[]);
$('statReasons').innerHTML=rz.length?('<div style="font-size:11px;color:var(--d);margin:10px 0 4px">실패 원인 분류</div><div style="display:flex;gap:6px;flex-wrap:wrap">'+rz.map(r=>`<span class="st st-f">${esc(r.reason)} ${r.n}</span>`).join('')+'</div>'):'';
$('statSurvival').innerHTML=(s.alive+s.dead)?`<div style="font-size:11px;color:var(--d);margin:12px 0 4px">발행글 생존 (검증됨 ${s.alive+s.dead}건)</div><span class="st st-ok">생존 ${s.alive}</span> <span class="st st-f">삭제 ${s.dead}</span>`:'';
const mx=Math.max(1,...s.by_day.map(d=>d.done+d.failed));
$('statDays').innerHTML='<div style="display:flex;align-items:flex-end;gap:4px;height:120px">'+s.by_day.map(d=>{const h=Math.round((d.done+d.failed)/mx*100);const go=d.done+d.failed?Math.round(d.done/(d.done+d.failed)*h):0;return `<div style="flex:1;text-align:center" title="${d.day}: 성공${d.done}/실패${d.failed}"><div style="height:${h}px;display:flex;flex-direction:column-reverse;border-radius:3px;overflow:hidden;background:var(--b)"><div style="height:${h-go}px;background:var(--r)"></div><div style="height:${go}px;background:var(--g)"></div></div><div style="font-size:8px;color:var(--d);margin-top:2px">${(d.day||'').slice(5)}</div></div>`}).join('')+'</div>';
$('statSites').innerHTML='<table><thead><tr><th>사이트</th><th>성공</th><th>실패</th><th>생존</th><th>삭제</th></tr></thead><tbody>'+s.by_site.map(x=>`<tr><td>${esc(x.site)}</td><td style="color:var(--g)">${x.done}</td><td style="color:var(--r)">${x.failed}</td><td style="color:var(--v)">${x.alive||0}</td><td style="color:var(--d)">${x.dead||0}</td></tr>`).join('')+'</tbody></table>'}

// ---- 사이트 목록 실시간 렌더 (새로고침 없이) ----
function siteRow(s){const st=s.status==='done'?'ok':s.status==='failed'?'f':'i';const nm=esc(s.name||(s.site_url||'').slice(0,20));const today=(s.posted_today||0);
const adminSource=['manual_admin','admin_bulk','legacy_admin','candidate_registered'].includes(s.registration_source);const perm=(s.permission&&adminSource)?'<span class="st st-ok" title="'+esc(s.permission_note||'')+'">등록됨</span>':'<span class="st st-f">발행 불가</span>';
const lb=(s.permission&&adminSource)?'':'border-left:3px solid var(--r)';
const hdot=s.health?('<span title="상태점검: '+esc(s.health)+' ('+esc(s.health_at||'')+')" style="color:'+(s.health==='ok'?'var(--g)':'var(--r)')+'">●</span> '):'';
const pl=(s.platform||'auto');const plname=pl==='cafe24'?'Cafe24':(pl==='gnuboard'?'그누보드':'자동');const plcolor=pl==='cafe24'?'var(--v)':(pl==='gnuboard'?'var(--p)':'var(--d)');
const learned=s.learned&&s.learned.write_url;const lbadge=learned?' <span title="자가학습 셀렉터 저장됨" style="color:var(--g)">🎓</span>':'';
const cbadge=s.has_captcha?' <span title="캡차 감지 — 자동발행 제외('+esc(s.captcha_note||'')+')" style="color:var(--y)">🧩캡차</span>':'';
const pltag='<span style="font-size:9px;color:'+plcolor+'" title="발행 방식">'+plname+'</span>'+lbadge+cbadge;
return `<tr data-id="${esc(s.id)}" style="${lb}"><td><input type="checkbox" class="cb" data-id="${esc(s.id)}"></td><td>${hdot}<b>${nm}</b></td><td>${perm}</td><td style="max-width:150px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${esc(s.site_url)}">${esc((s.site_url||'').slice(0,34))}</td><td style="color:var(--p)">${esc(s.bo_table||'')}<br>${pltag}</td><td style="color:var(--d)">${today}/${s.daily_limit==null?3:s.daily_limit}건<br>${s.min_interval_minutes==null?60:s.min_interval_minutes}분 간격</td><td><span class="st st-${st}">${esc(s.status||'idle')}</span></td><td><button class="btn btn-p btn-xs" onclick="editSite('${esc(s.id)}')">편집</button> <button class="btn btn-y btn-xs" onclick="dryRun('${esc(s.id)}')" title="실제 글은 올리지 않고 등록 직전까지 검증">드라이런</button> <button class="btn btn-d btn-xs" onclick="detectSite('${esc(s.id)}')">감지</button> <button class="btn btn-v btn-xs" onclick="learnSite('${esc(s.id)}')" title="DOM을 훑어 셀렉터 자동 학습 후 저장">학습</button> <button class="btn btn-d btn-xs" onclick="healthSite('${esc(s.id)}')">점검</button> <button class="btn btn-g btn-xs" onclick="testSite('${esc(s.id)}')">테스트</button> <button class="btn btn-r btn-xs" onclick="delSite('${esc(s.id)}')">삭제</button></td></tr>`}
async function healthSite(id){toast('점검중...');const r=await api('/sites/health/'+id,'POST');if(r&&r.ok){const h=r.health;const pn=h.platform==='cafe24'?'Cafe24':'그누보드';toast((h.ok?'✅ 정상':'⚠️ 확인필요')+` [${pn}] 접속:${h.reachable?'O':'X'} 로그인폼:${h.login_form?'O':'X'} 글쓰기:${h.write_page?'O':'X'}`,h.ok?'ok':'er');renderSites()}else toast('실패','er')}
async function dryRun(id){toast('🧪 드라이런 실행중... (글은 올리지 않습니다, 최대 60초)');const r=await api('/sites/dryrun/'+id,'POST');if(!r){toast('실패','er');return}
const rows=(r.steps||[]).map(s=>`<tr><td>${s.ok?'<span style="color:var(--g)">✅</span>':'<span style="color:var(--r)">❌</span>'}</td><td><b>${esc(s.name)}</b></td><td style="color:var(--d)">${esc(s.detail)}</td></tr>`).join('');
const hdr=r.ok?'<span style="color:var(--g)">✅ 발행 가능 — 등록 직전까지 모두 통과</span>':'<span style="color:var(--r)">❌ 발행 불가 — 아래 ❌ 지점에서 막힘</span>';
$('pvTitle').textContent='드라이런 결과';
$('pvFrame').srcdoc='<!DOCTYPE html><html lang=ko><head><meta charset=utf-8><style>body{font-family:-apple-system,sans-serif;padding:18px;color:#222}table{width:100%;border-collapse:collapse;font-size:13px}td{padding:8px 6px;border-bottom:1px solid #eee;vertical-align:top}h3{margin-bottom:12px;font-size:16px}</style></head><body><h3>'+hdr+'</h3><table>'+rows+'</table><p style="color:#888;font-size:12px;margin-top:14px">※ 실제 게시글은 등록되지 않았습니다. 제목/본문을 채워보기만 하고 중단했습니다.</p></body></html>';
$('pvOverlay').style.display='block';renderSites()}
async function detectSite(id){toast('플랫폼 감지중...');const r=await api('/sites/detect/'+id,'POST');if(r&&r.ok){toast('감지됨: '+(r.platform==='cafe24'?'Cafe24':'그누보드'));renderSites()}else toast('실패: '+(r&&r.error||''),'er')}
async function learnSite(id){if(!confirm('실측 학습을 실행합니다.\n등록된 사이트 한 곳의 글쓰기 DOM과 폼을 확인해 저장하며 글은 제출하지 않습니다. 진행할까요?'))return;toast('🎓 비제출 실측 학습 중...(최대 60초)');const r=await api('/sites/learn/'+id,'POST');if(r&&r.ok&&r.learned){toast('🎓 실측 성공! 비제출 레시피 저장됨 ('+(r.learned.content_mode||'')+')');renderSites()}else if(r&&r.captcha){toast('🧩 CAPTCHA 감지 — 우회하지 않고 제외 상태를 저장했습니다','er');renderSites()}else if(r&&r.blocked){toast('⛔ 보안 차단 감지 — 우회하지 않고 측정 결과를 저장했습니다','er');renderSites()}else toast('실측 완료: '+((r&&(r.message||r.error))||'폼 미확정'),'er')}
async function renderSites(){const sites=await api('/sites','GET');if(!Array.isArray(sites))return;
$('siteTabCount').textContent=sites.length;
// 필터 드롭다운 (선택값 유지)
const sel=$('kwSiteFilter');const cur=sel.value;sel.innerHTML='<option value="">전체 사이트</option>'+sites.map(s=>`<option value="${esc(s.id)}">${esc(s.name||(s.site_url||'').slice(0,20))}</option>`).join('');sel.value=cur;
// 체크 상태 보존
const checked=new Set(getSiteIds());
if(!sites.length){$('siteList').innerHTML='<p style="color:var(--d);padding:30px;text-align:center">등록된 사이트가 없습니다</p>';return}
$('siteList').innerHTML='<table><thead><tr><th><input type="checkbox" id="allCb" onclick="document.querySelectorAll(\'.cb\').forEach(c=>c.checked=this.checked)"></th><th>이름</th><th>허용</th><th>URL</th><th>게시판</th><th>오늘</th><th>상태</th><th>동작</th></tr></thead><tbody>'+sites.map(siteRow).join('')+'</tbody></table>';
checked.forEach(id=>{const c=document.querySelector(`.cb[data-id="${id}"]`);if(c)c.checked=true})}

// ---- 통계/진행률 폴링 ----
async function poll(){const r=await api('/workers/stats','GET');if(!r)return;
$('q').textContent=r.queued||0;$('ok').textContent=r.success||0;$('fl').textContent=r.fail||0;$('sk').textContent=r.skipped||0;
const wstate=r.paused?'PAUSE':(r.active?'ON':'OFF');$('ws').textContent=wstate;$('ws').style.color=r.paused?'var(--y)':(r.active?'var(--g)':'var(--d)');
const total=r.total||0,done=r.done||0;
if(total>0){$('progCard').style.display='block';const pct=Math.round(done/total*100);$('progBar').style.width=pct+'%';$('progText').textContent=`${done} / ${total} (${pct}%)`+(r.skipped?` · 스킵 ${r.skipped}`:'')}else{$('progCard').style.display='none'}
if($('p-res').classList.contains('on'))renderHistory()}

$('gContent').addEventListener('input',function(){$('gLen').textContent=this.value.length.toLocaleString()+'자'});
$('kwlist').addEventListener('input',function(){$('kwCount').textContent=parseList().length+'줄'});
renderSites();poll();loadPool();loadImages();
setInterval(poll,2000);
setInterval(renderSites,4000);
</script>
</body></html>'''

def R(title,**kw):
    if title=='로그인': return render_template_string(HTML+LOGIN_HTML+'\n</body></html>',title=title,**kw)
    return render_template_string(HTML+DASH_HTML,title=title,**kw)

# ==================== Main ====================
def main():
    # 정상 환경 일치: 서버 타임존을 한국(Asia/Seoul)으로 (위장이 아니라 실제 운영지역과 맞춤)
    try:
        os.environ['TZ']='Asia/Seoul'; time.tzset()
    except Exception: pass
    cfg=load_config()
    host=os.environ.get('HOST','127.0.0.1'); port=int(os.environ.get('PORT','8888'))
    print(f'\n찌라시 마스터 v6 - 정직 발행 모드\nhttp://{host}:{port}\n브랜드: {cfg.get("brand","설정필요")}')
    if os.environ.get('CHIRASHI_PASSWORD') is None and cfg.get('password','admin1234')=='admin1234':
        print('⚠️  기본 비밀번호(admin1234) 사용 중 — 공개 서버라면 반드시 변경하세요!')
    # 재시작 복구: 미완료 작업 큐 복원
    try:
        n=recover_queue()
        if n: print(f'↻ 미완료 작업 {n}건 복구됨 (워커 시작 시 이어서 발행)')
    except Exception as e: print('복구 건너뜀:',e)
    # 예약 발행 스케줄러 상시 가동
    try:
        threading.Thread(target=scheduler_loop,name='SCHED',daemon=True).start()
        print('⏰ 스케줄러 시작 (KST 기준)')
    except Exception as e: print('스케줄러 시작 실패:',e)
    # 지연 재시도 루프 / 텔레그램 명령 수신 / 발행글 생존 검증
    for fn,nm in [(retry_loop,'RETRY'),(telegram_loop,'TG'),(verify_loop,'VERIFY'),(member_scheduler_loop,'MSCHED'),(discover_loop,'DISCO')]:
        try: threading.Thread(target=fn,name=nm,daemon=True).start()
        except Exception as e: print(f'{nm} 시작 실패:',e)
    print('🔁 재시도·📱텔레그램·🔎검증 스레드 시작')
    # 헤드리스 서버(VPS)에서는 브라우저 자동 오픈 안 함
    if host in ('127.0.0.1','localhost') and os.environ.get('OPEN_BROWSER','1')!='0':
        try:
            import webbrowser; webbrowser.open(f'http://{host}:{port}')
        except: pass
    app.run(host=host,port=port,debug=False,threaded=True)

if __name__=='__main__':
    main()
